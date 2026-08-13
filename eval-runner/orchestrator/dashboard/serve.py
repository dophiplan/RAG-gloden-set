#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
serve.py — 관제 대시보드 로컬 서버 (사양서 v1.3 §10)

- state.json + ledger.jsonl **읽기 전용 뷰** (원장 편집 UI 없음 — 스크롤만)
- 검수큐 카드의 approve/reject 는 pipeline CLI 를 그대로 호출 (사유 필수·원장 기록)
- 로컬 전용: 127.0.0.1 바인딩

실행: python3 dashboard/serve.py [--port 8791]
"""
import argparse
import json
import re
import subprocess
import sys
import unicodedata
from collections import Counter
from http.server import HTTPServer, SimpleHTTPRequestHandler
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "tools"))


def N(s):
    return unicodedata.normalize("NFC", str(s)) if s is not None else ""


def j(o):
    return json.dumps(o, ensure_ascii=False).encode("utf-8")


# ── 제품 1뎁스 구조 (U1) ──────────────────────────────────
# 사용자에게 보이는 건 제품(RV·RC). 내부 트랙 코드(RV2 등)는 세대 구현 상세.
PRODUCT_META = {
    "RV2": {"display": "RV", "product_name": "리모트뷰", "gen": "골든셋 v2.0 — 현역(구축 중)"},
    "RC2": {"display": "RC", "product_name": "리모트콜", "gen": "골든셋 v2.0 — 현역(구축 중)"},
}
LEGACY_GENS = {   # display → 이전 세대 (데이터 폴더 코드, 라벨) — 은퇴일: 원장 GOLDENSET_RETIRED
    "RV": [{"code": "RV", "gen": "골든셋 v1 · 806문항 — 2026-07-20 은퇴(정답키 공개, 참고용)"}],
    "RC": [{"code": "RC", "gen": "골든셋 v1.1 · 891문항 — 2026-07-20 은퇴(정답키 공개, 참고용)"}],
}
HIDDEN_CODES = {"EE"}   # 자동 테스트 전용 제품 — 화면에서 숨김 (E2E가 쓰고 지나가는 자리)


def display_of(code):
    return PRODUCT_META.get(code, {}).get("display", code)


def _worker_alive(code):
    """제품별 실행 주체 생존 — 상태가 RUNNING인데 실행 주체가 없으면 '유령 RUNNING'.
    [수리 2026-08-13 난희 지적] 멈춰 있는데 돌고 있는 것처럼 보이던 문제.
    [수리 2026-08-13 2차] auto_run pid 파일만 보면 nohup으로 도는 보조 작업(구멍 메우기 등)을
    '유령'으로 오판 — 진행 파일이 최근(10분 내) 갱신됐으면 실행 주체가 있는 것."""
    pf = ROOT / "results" / f"_run_{code}.pid"
    if pf.exists():
        try:
            pid = int(pf.read_text().strip())
            st = subprocess.run(["ps", "-o", "stat=", "-p", str(pid)],
                                capture_output=True, text=True).stdout.strip()
            if st and not st.startswith("Z"):   # 좀비(Z)는 죽은 것
                return True
        except Exception:
            pass
    # 폴백: 진행 파일 신선도 — 배치마다 갱신되므로 최근 갱신 = 뛰는 중
    pg = ROOT / "results" / f"_progress_{code}.json"
    try:
        import time
        return (time.time() - pg.stat().st_mtime) < 600
    except Exception:
        return False


def _qa_ver(p):
    """외부QA 파일의 vN — 최신 선택은 반드시 이 숫자로 (import_qa._ver_of와 동일 규칙 [P1-2])"""
    m = re.search(r"_v(\d+)\.xlsx$", p.name)
    return int(m.group(1)) if m else 0


def api_state():
    import os
    st = json.loads((ROOT / "state.json").read_text(encoding="utf-8"))
    # 제품 1뎁스 뷰모델: display 제품 → {현역 트랙 코드, 세대 라벨, 이전 세대들}
    st["_products"] = {}
    for code, ps in st["products"].items():
        if code in HIDDEN_CODES:
            continue
        meta = PRODUCT_META.get(code, {"display": code, "product_name": code, "gen": "현역"})
        st["_products"][meta["display"]] = {
            "code": code, "name": meta["product_name"], "gen": meta["gen"],
            "legacy": LEGACY_GENS.get(meta["display"], []),
        }
        # G20: 외부 Q&A 별도 트랙 현황 — 최신 = 버전 숫자 기준 (문자열 정렬은 문항 수 자릿수에 속음 [P1-2])
        papers = list((ROOT / "data" / code / "external_qa").glob(f"외부QA_시험지_{code}_*문항_v*.xlsx"))
        latest = max(papers, key=_qa_ver) if papers else None
        pm = re.search(r"_(\d+)문항_", latest.name) if latest else None
        ps["_qa"] = {"paper": N(latest.name), "n": int(pm.group(1)) if pm else None} if latest else None
        ps["_worker"] = _worker_alive(code)   # RUNNING 표시의 진위 판정용
    # 모델 모드 부가
    try:
        from model_adapter import detect_mode, effective_recheck_rate
        m = detect_mode()
        if os.environ.get("ORCH_MOCK") == "1":
            label, cls = "연습 모드 — 가짜 AI (공짜)", "practice"
        elif m["mode"] == 0:
            label, cls = "모델 미연결 — 실전모드_켜기 실행 필요", "off"
        else:
            parts = []
            if m["have"].get("generator"): parts.append("출제 claude")
            if m["have"].get("judge"): parts.append("채점 Kimi" if m["mode"] >= 2 else "채점 claude(신규세션)")
            if m["have"].get("reviewer"): parts.append("교차 codex")
            label, cls = "실전 모드 — " + " · ".join(parts), "live"
        st["_mode"] = {"mode": m["mode"], "label": label, "cls": cls,
                       "plan": m["plan"], "recheck": effective_recheck_rate()}
    except Exception as e:
        st["_mode"] = {"mode": "?", "label": "상태 확인 실패", "cls": "off", "error": str(e)}
    # 킬스위치: HALTED 제품 수
    # 의도적 잠금(정책 보류)은 사고 카운트에서 제외 — 헤더 '사고 N'은 진짜 사고만
    st["_halt_count"] = sum(1 for p in st["products"].values()
                            if p["status"] == "HALTED"
                            and not str(p.get("halt_reason") or "").startswith("의도적 잠금"))
    return st


def api_ledger(n=60):
    """원장 테일 — 화면 표시용. 테스트 전용 제품(EE·ZZ) 행은 숨김 [P1-5]
    (원장 파일 자체는 append-only 그대로 — 표시만 거른다. 과거 누적 테스트 소음이 원장의 60%)"""
    p = ROOT / "ledger.jsonl"
    if not p.exists():
        return []
    hidden = HIDDEN_CODES | {"ZZ", "TT"}
    out = []
    for x in reversed(p.read_text(encoding="utf-8").strip().splitlines()):
        try:
            r = json.loads(x)
        except Exception:
            continue   # 손상 행 1개로 화면 전체가 백지 되지 않게 (전수검수 F3 계열)
        if r.get("product") in hidden:
            continue
        out.append(r)
        if len(out) >= n:
            break
    return out


def api_queue():
    import time
    q = ROOT / "검수큐"
    cards = []
    # 열린 게이트 명세 — 이미 닫힌(승인/반려 완료) 게이트의 유령 카드를 청소하기 위해
    try:
        _st = json.loads((ROOT / "state.json").read_text(encoding="utf-8"))
        open_ids = {g["id"] for ps in _st["products"].values() for g in ps.get("open_gates", [])}
    except Exception:
        open_ids = None
    if q.is_dir():
        for f in sorted(q.glob("*.md")):
            body = f.read_text(encoding="utf-8")
            # 유령 카드 청소 — 게이트는 닫혔는데 카드만 남으면(늦은 소견 재생성 등) 사람이
            # 죽은 카드에 반려를 눌러 사유가 소실되는 사고. 발행 직후 레이스 방지: 2분 유예.
            if (open_ids is not None and f.name.startswith("GATE_")
                    and f.stem.removeprefix("GATE_") not in open_ids
                    and time.time() - f.stat().st_mtime > 120):
                done = q / "완료" / f.name
                done.parent.mkdir(exist_ok=True)
                if done.exists() and "## 설계본부 소견" in body:
                    # [P3] 2회차 재생성 소견도 보존 — 완료본에 (구)소견이 있어도 새 소견이
                    # 다른 내용이면 이어붙인다 (기존: 구 소견 존재 시 새 소견 무통보 폐기)
                    seg = body[body.index("## 설계본부 소견"):].strip()
                    dtext = done.read_text(encoding="utf-8")
                    if seg not in dtext:
                        done.write_text(dtext.rstrip() + "\n\n" + seg + "\n", encoding="utf-8")
                elif not done.exists():
                    done.write_text(body, encoding="utf-8")
                f.unlink()
                continue
            kind = "GATE" if f.name.startswith("GATE_") else "INPUT"
            gate_id = f.stem.removeprefix("GATE_")
            m = re.search(r"제품: (\w+)", body)
            acks = re.findall(r"- \[ \] (?:ack: )?(.+)", body)
            prod_s = m.group(1) if m else "?"
            uploaded = None
            if kind == "INPUT":   # 코퍼스처럼 여러 파일을 나눠 올리는 카드 — 현황을 화면에
                sub = "corpus" if "CORPUS" in f.stem else "08_scoring"
                d = ROOT / "data" / prod_s / sub
                uploaded = sorted(p.name for p in d.glob("*")
                                  if p.is_file() and not p.name.startswith(".")) if d.is_dir() else []
            cards.append({"file": N(f.name), "kind": kind, "id": gate_id,
                          "product": prod_s,
                          "title": N(body.splitlines()[0].lstrip("# ")),
                          "body": body, "acks": acks if kind == "GATE" else [],
                          "uploaded": uploaded})
    return cards


def api_scores():
    """회차 성적 미니보드 — results/score_<P>_<r>/score_report.json 실측 집계"""
    out = {}
    for d in sorted(list((ROOT / "results").glob("score_*_r*"))
                    + list((ROOT / "results").glob("score_*_base*"))):
        # 정식 회차(rN) + 기준선(baseN — 우리가 미리 잰 참고분, 화면에서 별도 줄로 표시)
        m = re.fullmatch(r"score_([A-Z0-9]+)_(r\d+|base\d+)", d.name)   # 병기(_v12 등) 제외
        if not m:
            continue
        prod, rnd = m.group(1), m.group(2)
        rp = d / "score_report.json"
        if not rp.exists():
            continue
        try:
            rep = json.loads(rp.read_text(encoding="utf-8"))
        except Exception:
            continue   # 쓰는 중/손상 회차 하나로 성적판 전체가 500 되지 않게
        c = Counter(r.get("검색") for r in rep)
        g = Counter(r.get("생성") for r in rep)
        top1 = c.get("hit_top1", 0)
        # 검색축만 회차 감지 — 회차 리포트에 '검색축만' 명기 여부로 판단 (생성 수치를 미응시로 표기)
        search_only = any("검색축만" in md.read_text(encoding="utf-8")[:2500]
                          for md in d.glob("*리포트*.md"))
        # 지표 툴팁용 분석(엄격/결측 제외/합집합) — report_gen 이 회차마다 생성
        anal = None
        ap = d / "analysis.json"
        if ap.exists():
            try:
                anal = json.loads(ap.read_text(encoding="utf-8"))
            except Exception:
                pass
        out.setdefault(prod, {})[rnd] = {
            "top1": top1, "top5": top1 + c.get("hit_top5", 0),
            "pass": ("미응시" if search_only else g.get("pass", 0)),
            "partial": ("미응시" if search_only else g.get("partial", 0)),
            "검색축만": search_only, "분석": anal,
            "E환각": sum(1 for r in rep if r.get("E형환각")),
            "E거절": sum(1 for r in rep if r.get("E형거절")),
            "n": len(rep),
            # G20: 외부 Q&A 별도 트랙은 기계 내용 대조 채점 — 라벨로 구분 (골든셋 채점기 아님)
            "scorer": ("내용 대조(Q&A 트랙)" if any(d.glob("외부QA_r*_리포트.md"))
                       else "run_score_v11"),
        }
    # 문서 기록 이관 — 로컬 재채점본이 없는 회차를 인수인계 보고서 수치로 병기 (출처 라벨)
    for f in (ROOT / "results").glob("기록이관_*.json"):
        prod = f.stem.split("_", 1)[1]
        try:
            rec = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        for rnd, v in rec.items():
            if rnd.startswith("_") or rnd in out.get(prod, {}):
                continue   # 로컬 실측이 항상 우선 — 이관본은 빈 회차만 채움
            out.setdefault(prod, {})[rnd] = {**v, "scorer": "기록 이관",
                "분석": {"top1": v.get("top1"), "top5": v.get("top5"), "n": v.get("n"),
                         "note": v.get("note"), "이관": True}}
    return out


def _scan_stage_files(code, sm, canon, gen_label=None):
    d = ROOT / "data" / code / sm["dir"]
    files = []
    if d.is_dir():
        fl = sorted((p for p in d.iterdir() if p.is_file() and not p.name.startswith(".")),
                    key=lambda p: p.name)

        def vkey(p):
            m = re.search(r"_v(\d+(?:_\d+)+)", p.name)
            return tuple(int(x) for x in m.group(1).split("_")) if m else (0,)
        best = max(fl, key=vkey).name if fl else None
        import datetime
        for p in fl:
            st_ = p.stat()
            files.append({"file": N(p.name), "version": "",
                          "canonical": canon.get(N(p.name), p.name == best),
                          "size": st_.st_size, "gen": gen_label,
                          "mtime": datetime.datetime.fromtimestamp(st_.st_mtime).strftime("%m-%d %H:%M"),
                          "path": f"{code}/{sm['dir']}/{p.name}"})
    return files


def api_catalog():
    """제품(1뎁스) → 단계 → 파일. 현역 세대 + 이전 세대(라벨 병기) 통합 스캔 (U1·U6)."""
    meta = json.loads((ROOT / "catalog" / "stage_meta.json").read_text(encoding="utf-8"))
    canon = {}
    mf = ROOT / "catalog" / "manifest.json"
    if mf.exists():
        for r in json.loads(mf.read_text(encoding="utf-8"))["files"]:
            if r.get("canonical"):
                canon[r["file"]] = True
    st = json.loads((ROOT / "state.json").read_text(encoding="utf-8"))
    out = {"products": {}}
    for code in st["products"]:
        if code in HIDDEN_CODES:
            continue
        disp = display_of(code)
        stages = []
        for sm in meta["stages"]:
            files = _scan_stage_files(code, sm, canon)                      # 현역 세대
            for lg in LEGACY_GENS.get(disp, []):                            # 이전 세대 (참고)
                files += _scan_stage_files(lg["code"], sm, canon, gen_label="v1·은퇴")
            stages.append({"no": sm["no"], "name": sm["name"], "gate": sm["gate"],
                           "what": sm["what"], "consumes": sm["consumes"],
                           "produces": sm["produces"], "tool": sm["tool"],
                           "file_count": len(files), "files": files})
        out["products"][disp] = {"stages": stages,
                                 "total_files": sum(s["file_count"] for s in stages)}
    return out


_MODEL_CACHE = {"ts": 0, "ids": None}


def _moonshot_models():
    """Kimi 가용 모델 목록 (1시간 캐시) — 단종·신모델 감지용. 실패 시 None(판단 보류)"""
    import time
    import urllib.request
    if time.time() - _MODEL_CACHE["ts"] < 3600:
        return _MODEL_CACHE["ids"]
    try:
        import os
        key = os.environ.get("JUDGE_KEY", "")
        req = urllib.request.Request("https://api.moonshot.ai/v1/models",
                                     headers={"Authorization": f"Bearer {key}"})
        with urllib.request.urlopen(req, timeout=10) as r:
            ids = [m["id"] for m in json.loads(r.read().decode()).get("data", [])]
        _MODEL_CACHE.update(ts=time.time(), ids=ids)
    except Exception:
        _MODEL_CACHE.update(ts=time.time(), ids=None)
    return _MODEL_CACHE["ids"]


def _kimi_ver(mid):
    m = re.search(r"kimi-k([\d.]+)$", mid or "")
    return float(m.group(1)) if m else None


def _model_alert_card(engine, alert):
    """모델 단종·신모델 감지 → 검수큐 카드 발행 (같은 내용이면 중복 발행 안 함)"""
    import datetime
    qdir = ROOT / "검수큐"
    f = qdir / f"MODELALERT_{engine}.md"
    if f.exists() and alert in f.read_text(encoding="utf-8"):
        return
    # 사람이 닫은(완료로 이동) 알림은 같은 내용으로 재발행 안 함 — 새 모델이 나오면 내용이 달라져 다시 뜸
    done = qdir / "완료" / f"MODELALERT_{engine}.md"
    if done.exists() and alert in done.read_text(encoding="utf-8"):
        return
    qdir.mkdir(exist_ok=True)
    f.write_text(
        f"# MODELALERT_{engine} — 모델 교체 검토 필요\n"
        f"- 발행: {datetime.datetime.now().isoformat(timespec='seconds')} · 제품: 공통 · 단계: 운영\n\n"
        f"## 무엇을\n{alert}\n\n"
        f"## 처리 방법\n"
        f"- 교체하기로 하면: 저(Claude)에게 말하면 config 변경 + 규칙 C(캘리브레이션 재시험) 자동 처리\n"
        f"- 유지하기로 하면: 이 카드를 검수큐/완료 로 옮기면 끝 (성적 비교 보전)\n",
        encoding="utf-8")
    sys.path.insert(0, str(ROOT / "tools"))
    from olib import ledger_append
    ledger_append("MAINTENANCE", "MODEL_ALERT", "script:serve",
                  evidence={"engine": engine, "알림": alert})


def api_ai_status():
    """AI 팀 현황 — 역할별 엔진·연결 상태(🟢/⚪) + 제품별 투입 선택(ai_use 체크박스)"""
    import os
    from olib import load_config
    cfg = load_config()
    mock = os.environ.get("ORCH_MOCK") == "1"
    have = {}
    try:
        from model_adapter import detect_mode
        have = detect_mode(cfg, os.environ)["have"]
    except Exception:
        pass
    # role_ko = 본업. 앙상블에선 전원이 '추출'도 겸한다 — 라벨 앞머리는 화면(전략별)에서 붙임
    ROLE_META = [("generator", "병합·출제 대표", "claude"),
                 ("judge", "채점", "Kimi"),
                 ("reviewer", "교차 검토", "codex")]
    engines = []
    for role, role_ko, default_eng in ROLE_META:
        m = cfg.get("models", {}).get(role) or {}
        cmd = m.get("command", [])
        raw = " ".join([str(m.get("model", "")), m.get("provider", ""),
                        " ".join(cmd) if isinstance(cmd, list) else str(cmd)]).lower()
        if "kimi" in raw or "moonshot" in raw:
            eng = "Kimi"
        elif "codex" in raw or "gpt" in raw:
            eng = "codex"
        elif "claude" in raw or "anthropic" in raw:
            eng = "claude"
        else:
            eng = default_eng
        connected = bool(mock or have.get(role))
        if mock:
            how = "가짜 AI (연습 모드)"
        elif not m:
            how = "미설정 — config.yaml 에 없음"
        elif m.get("provider") == "cli":
            how = "구독 계정 (CLI)" if connected else "구독 CLI 미설치 — 계정 대기"
        else:
            how = "API 키" if connected else f"API 키 없음 ({m.get('api_key_env', '?')})"
        entry = {"role": role, "role_ko": role_ko, "engine": eng,
                 "connected": connected, "how": how, "model": m.get("model", "")}
        # 최신성 감시 (Kimi) — 자동 교체는 안 함: 채점 모델 교체 = 규칙 C(캘리브 재시험), 사람 결정
        if m.get("provider") == "moonshot" and connected and not mock:
            ids = _moonshot_models()
            if ids is not None:
                if m.get("model") not in ids:
                    entry["alert"] = f"⚠ {m.get('model')} 단종 — 목록에 없음, 교체 필요"
                else:
                    cur_v = _kimi_ver(m.get("model"))
                    top = max((v for v in (_kimi_ver(i) for i in ids) if v), default=None)
                    if cur_v and top and top > cur_v:
                        entry["alert"] = f"✨ kimi-k{top:g} 출시 — 교체는 사람 결정 (규칙 C: 교체 시 캘리브레이션 재시험)"
            if entry.get("alert"):
                _model_alert_card(eng, entry["alert"])   # 검수큐 카드로 사람 호출
        engines.append(entry)
    st = json.loads((ROOT / "state.json").read_text(encoding="utf-8"))
    use = {}
    for code, ps in st["products"].items():
        if code in HIDDEN_CODES:
            continue
        use[display_of(code)] = {
            "code": code,
            "ai_use": ps.get("ai_use") or {"generator": True, "judge": True, "reviewer": True},
            "strategy": ps.get("strategy", "ensemble"),
        }
    return {"mock": mock, "engines": engines, "products": use}


def api_runlog(code, n=8):
    """실행 기록 요약 — 원문은 파일(results/_run_*.log)에 저장, 화면엔 '챕터 전환'급만.
    이정표: 단계 결과(→)·재개·중단·완료·사고. 배치 카운트 잡음은 제외(진행선이 담당)."""
    if not re.fullmatch(r"[A-Z0-9]{1,8}", code or ""):
        return {"lines": []}
    p = ROOT / "results" / f"_run_{code}.log"
    if not p.exists():
        return {"lines": []}
    try:
        lines = p.read_text(encoding="utf-8", errors="replace").strip().splitlines()
    except Exception:
        return {"lines": []}
    KEEP = ("→", "체크포인트", "⏸", "⛔", "✅", "🤖", "🔁", "재개", "중단", "완료", "HALT")
    marks = [l for l in lines if any(k in l for k in KEEP) and "배치…" not in l]
    return {"lines": marks[-n:]}


def _extractor_rows(code, prog=None):
    """추출자별 현황 — 체크포인트가 진실. 진행 파일(뛰는 중에만 갱신)이 없어도 읽힌다.
    [난희 요청 2026-08-10] 멈춤·사고 상태에서도 '누가 어디까지' 보여야 한다.
    [수리 2026-08-13 난희 실측] 구멍 메우기(_gapfaq 등 태그 체크포인트)가 도는데 패널이 안 보임 —
    본 추출 파일(_ckpt_coverage_CI.json)만 읽어서. 태그 파일 전부 합산한다."""
    roles = (prog or {}).get("roles") or {}
    tt_any = max([(v or {}).get("total") or 0 for v in roles.values()] or [0])
    names = {"generator": "claude", "judge": "Kimi", "reviewer": "codex"}
    agg = {}   # role → {done,total,units,fails} (여러 구간 합산)
    for ck in sorted((ROOT / "results").glob(f"_ckpt_coverage_{code}*.json")):
        try:
            c = json.loads(ck.read_text(encoding="utf-8"))
        except Exception:
            continue
        tagged = ck.stem != f"_ckpt_coverage_{code}"   # 태그 파일 = 구멍 메우기 등 부가 구간
        for role in names:
            v = c.get(role)
            if not isinstance(v, dict):
                continue
            a = agg.setdefault(role, {"done": 0, "total": 0, "units": 0, "fails": 0})
            a["done"] += v.get("done", 0)
            a["total"] += int(v.get("n_batches") or (0 if tagged else tt_any))
            a["units"] += len(v.get("units", []))
            a["fails"] += len(v.get("fails") or [])
    rows, tot = [], 0
    for role, label in names.items():
        if role not in agg:
            continue
        a = agg[role]
        tot += a["units"]
        rows.append({"who": label, "role": role, "done": a["done"],
                     "total": a["total"] or tt_any, "units": a["units"], "fails": a["fails"]})
    return rows, tot


def api_progress(code):
    """③ 등 장시간 작업의 실시간 진행/막힘 — 엔진이 배치마다 갱신하는 파일을 그대로.
    진행 파일이 없어도(멈춤·사고) 체크포인트 기반 현황은 항상 반환한다."""
    if not re.fullmatch(r"[A-Z0-9]{1,8}", code or ""):
        return {"active": False}
    try:
        _st = json.loads((ROOT / "state.json").read_text(encoding="utf-8"))["products"][code]
    except Exception:
        _st = {}
    p = ROOT / "results" / f"_progress_{code}.json"
    d = {}
    if p.exists():
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            d = {}
    stale_stage = bool(d) and d.get("stage", "COVERAGE_MAP") != _st.get("stage")
    if not d or stale_stage:
        # 진행 파일이 없거나 낡음 — 체크포인트만으로 '멈춘 자리' 보고 (active=False, snapshot=True)
        rows, tot = _extractor_rows(code)
        if not rows:
            return {"active": False}
        return {"active": False, "snapshot": True, "stage": _st.get("stage"),
                "status": _st.get("status"), "worker_alive": _worker_alive(code),
                "extractors": rows, "units_total": tot,
                "phase": "멈춘 자리 (체크포인트 보존)"}
    d["active"] = True
    d["status"] = _st.get("status")
    d["worker_alive"] = _worker_alive(code)
    rows, tot = _extractor_rows(code, d)
    if rows:
        d["extractors"] = rows
        d["units_total"] = tot
    d["overall"] = _overall_coverage(code, d)
    return d


def _overall_coverage(code, prog):
    """전체 지도 완성도 % — [난희 지적 2026-08-13] 구멍 메우기 %는 '이번 조각' 진도라
    전체가 얼마나 됐는지 안 보임. 맥락 파일(전체 청크·기커버·이번 범위)로 환산한다.
    범위 내 진행은 배치 비율 근사(청크 크기 편차 있음) — 화면에 '약'을 붙인다."""
    cf = ROOT / "results" / f"_gapctx_{code}.json"
    if not cf.exists():
        return None
    try:
        ctx = json.loads(cf.read_text(encoding="utf-8"))
    except Exception:
        return None
    tot = int(ctx.get("total_chunks") or 0)
    if not tot:
        return None
    before = int(ctx.get("covered_before") or 0)
    scope = int(ctx.get("scope_chunks") or 0)
    # 이번 범위 내 진행률 — 진행 파일의 배치 비율 (여러 주자면 최대치: 같은 범위를 나눠 뛰므로)
    frac = 0.0
    for v in (prog or {}).get("roles", {}).values():
        if (v or {}).get("total"):
            frac = max(frac, min(1.0, v.get("batch", 0) / v["total"]))
    covered_now = before + int(scope * frac)
    remain = {k: int(n) for k, n in (ctx.get("pending_scopes") or {}).items()}
    return {"before_pct": round(before / tot * 100),
            "now_pct": round(covered_now / tot * 100),
            "after_scope_pct": round((before + scope) / tot * 100),
            "total_chunks": tot, "covered_now": covered_now,
            "scope": ctx.get("scope"), "pending": remain}


def api_xlsx(relpath, max_rows=400, name=None, prod=None):
    """xlsx 미리보기 — 엑셀 안 열고 툴에서 본다 (data/ 하위만, 숨김 시트=봉인은 비노출).
    name+prod 로 부르면 파일명만으로 최신 실물을 찾는다 (게이트 카드 → 실물 팝업)."""
    import openpyxl
    base = (ROOT / "data").resolve()
    if name and prod:
        hits = sorted((ROOT / "data" / prod).rglob(name)) if (ROOT / "data" / prod).is_dir() else []
        if not hits:
            return {"error": f"파일 못 찾음: {name}"}
        p = hits[-1].resolve()
    else:
        p = (base / (relpath or "")).resolve()
    try:
        p.relative_to(base)   # [P3] prefix 문자열 비교는 data2/ 같은 형제 디렉토리를 통과시킴
    except ValueError:
        return {"error": "미리보기 불가 (data 하위 xlsx만)"}
    if not p.exists() or p.suffix != ".xlsx":
        return {"error": "미리보기 불가 (data 하위 xlsx만)"}
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    sheets = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue   # 숨김 시트 = 봉인(블라인드) — 노출 금지
        rows = []
        for r in ws.iter_rows(max_row=max_rows + 1, values_only=True):
            rows.append([("" if c is None else str(c))[:400] for c in (r or [])[:14]])
        sheets[N(sn)] = {"rows": rows, "truncated": (ws.max_row or 0) > max_rows + 1,
                         "total": ws.max_row or 0}
    wb.close()
    return {"file": N(p.name), "path": N(str(p.relative_to(base))), "sheets": sheets}


def _s2_ledger_file(prod):
    d = ROOT / "data" / prod / "07_stage2"
    c = sorted(d.glob(f"{prod}_본판정_판정대장_*.xlsx")) if d.is_dir() else []
    return c[-1] if c else None


def api_s2diff(prod):
    """⑦ 이중 판정 불일치 검토 — 갈린 문항만 팝업에서 클릭으로 확정 (엑셀 왕복 금지)"""
    import openpyxl
    f = _s2_ledger_file(prod)
    if not f:
        return {"rows": []}
    # 질문·정답은 통합 대장에서 조인
    led = sorted((ROOT / "data" / prod / "05_unified_ledger").glob("*통합대장*.xlsx"))
    qmap = {}
    if led:
        lw = openpyxl.load_workbook(led[-1], read_only=True, data_only=True).active
        lh = [N(c) for c in next(lw.iter_rows(max_row=1, values_only=True))]
        qi = lh.index("질문") if "질문" in lh else 3
        ai = next((i for i, h in enumerate(lh) if h.startswith("정답")), 4)
        for r in lw.iter_rows(min_row=2, values_only=True):
            qmap[N(r[0])] = (N(r[qi]), N(r[ai]))
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    hdr = [N(c) for c in next(ws.iter_rows(max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(hdr)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if N(r[col.get("불일치", 6)]) != "✚":
            continue
        iid = N(r[0])
        q, a = qmap.get(iid, ("", ""))
        rows.append({"id": iid, "q": q, "a": a,
                     "kimi": N(r[col.get("판정(채점관 Kimi)", 2)]),
                     "kimi_why": N(r[col.get("판정문(전건 보존)", 3)])[:400],
                     "claude": N(r[col.get("검토 판정(claude 새 세션)", 4)]),
                     "claude_why": N(r[col.get("검토 판정문", 5)])[:400],
                     "final": N(r[col["최종 판정"]]) if "최종 판정" in col and col["최종 판정"] < len(r) else ""})
    wb.close()
    return {"rows": rows, "file": N(f.name)}


def api_qa_handoff(prod):
    """G20 · 외부 Q&A 별도 트랙 전달 꾸러미 — 시험지(질문만) + 안내문 + 응답로그 예시 zip."""
    import io
    import zipfile
    import openpyxl
    import datetime
    d = ROOT / "data" / prod / "external_qa"
    papers = list(d.glob(f"외부QA_시험지_{prod}_*문항_v*.xlsx"))
    if not papers:
        return None, "Q&A 시험지 없음 — 외부 Q&A 분류 카드 승인 후 이용 가능"
    pub = max(papers, key=_qa_ver)   # [P1-2] 최신 = 버전 숫자
    ws = openpyxl.load_workbook(pub, read_only=True).active
    first = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    qid0 = str(first[0]) if first else f"{prod}-Q001"
    n = max(0, (ws.max_row or 1) - 1)
    name = "리모트콜" if prod.startswith("RC") else ("리모트뷰" if prod.startswith("RV") else prod)
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    guide = f"""# {name} 외부 Q&A 세트 응시 요청 (별도 트랙 — 골든셋 아님)

> 꾸러미 생성: {stamp} · 시험지: {N(pub.name)} ({n}문항)
> **검색축만**: 답변 생성(LLM 호출)은 생략하고 검색 결과(hits)만 기록해 주세요.

## 이 시험지는 무엇이 다른가
- 외부에서 제작된 질문·답변 세트 중 코퍼스에 근거가 실재하는 문항만 추린 것입니다.
- 기존 골든셋 회차와는 **성적이 분리 집계**됩니다 (별도 트랙).

## 부탁드리는 것
각 질문을 RAG 시스템에 넣고, 검색 hits(rank 순, 본문 포함)를 json 1개로 회신 부탁드립니다.
hits 항목에 **본문(content) 텍스트가 꼭 포함**돼야 합니다 — 채점이 내용 대조 방식이라 URL만으로는 판정이 안 됩니다.

## 응답 로그 형식 (예시 파일 동봉)
{{
  "responses": [
    {{ "id": "{qid0}", "hits": [ {{"rank": 1, "url": "…", "content": "…청크 본문…"}} ], "answer": null }}
  ]
}}
"""
    example = json.dumps({"responses": [
        {"id": qid0, "hits": [{"rank": 1, "url": "https://…", "content": "…검색된 청크 본문…"},
                              {"rank": 2, "url": "https://…", "content": "…"}], "answer": None}]},
        ensure_ascii=False, indent=1)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(pub, N(pub.name))
        z.writestr("응시_안내문.md", guide)
        z.writestr("응답로그_예시.json", example)
    ledger_append_safe("EXTERNAL_QA", "QA_HANDOFF_DOWNLOADED",
                       evidence={"시험지": N(pub.name), "문항": n}, product=prod)
    return buf.getvalue(), f"외부QA_꾸러미_{prod}_{n}문항.zip"


def ledger_append_safe(stage, action, evidence=None, product=None):
    try:
        sys.path.insert(0, str(ROOT / "tools"))
        from olib import ledger_append
        ledger_append(stage, action, "사람:대시보드", evidence=evidence, product=product)
    except Exception:
        pass


def api_handoff(prod, scope="full"):
    """⑧ 팀장님 전달 꾸러미 — 발행본 + 응시 안내문을 zip 한 방에 (툴에서 직접 다운로드).
    scope: full=전체 응시(검색+생성) / search=검색축만(top1·top5, answer:null) — 안내문이 달라진다."""
    import io
    import zipfile
    import openpyxl
    import datetime
    pubs = sorted((ROOT / "data" / prod / "08_scoring").glob("*질문셋_발행본*.xlsx"))
    if not pubs:
        return None, "발행본 없음 — ⑧ 도달 후 이용 가능"
    pub = pubs[-1]
    ws = openpyxl.load_workbook(pub, read_only=True).active
    n = max(0, (ws.max_row or 1) - 1)
    name = "리모트콜" if prod.startswith("RC") else ("리모트뷰" if prod.startswith("RV") else prod)
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    search = scope == "search"
    scope_ko = "검색축만 (top1·top5 히트율 — answer 미제출)" if search else "전체 (검색축 + 생성축)"
    ans_line = ('"answer": null  ← 이번 회차는 전 문항 null로 통일해 주세요'
                if search else '"answer": "…시스템 답변…"')
    ask = ("RAG 시스템에 각 질문을 넣되, **답변 생성(LLM 호출)은 생략**하고 검색 결과(hits)만 기록해 주세요. "
           "answer는 전 문항 null로 통일합니다 (일부만 null이면 결손으로 반려됩니다)."
           if search else
           "RAG 시스템에 각 질문을 그대로 넣고, 응답 로그를 json 1개로 회신 부탁드립니다.")
    tail = ("- 채점: 검색축(top1·top5)만 산출 — 생성축·E형은 리포트에 '미응시' 표기\n"
            if search else
            "- hits: 검색 근거(rank 순) · answer: 최종 생성 답변\n")
    guide = f"""# {name} RAG 평가 질문셋 응시 요청 (골든셋 v2)

> 꾸러미 생성: {stamp} · 발행본: {N(pub.name)} ({n}문항)
> **이번 회차 응시 범위: {scope_ko}**
> ※ 꾸러미는 요청 시점의 최신판으로 자동 조립됩니다 — 재요청 시 관제판에서 버튼 한 번 더.

## 파일
- {N(pub.name)} — {name} {n}문항 (문항ID · 질문 2컬럼, 정답 비공개)

## 부탁드리는 것
{ask}

## 응답 로그 형식
{{
  "meta": {{ "corpus_version": "…(인입 코퍼스 버전 — 문서 N건·청크 M건 표기)" }},
  "responses": [
    {{ "id": "{prod}-001", "hits": [ {{"rank":1, "source":"…"}} ], {ans_line} }}
  ]
}}

- responses는 전 문항(빠짐없이), id는 발행본의 문항ID 그대로
{tail}- 받는 즉시 자동 채점 → 성적 리포트로 회신드립니다.
"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(pub, N(pub.name))
        z.writestr("응시_안내.md", guide)
    try:
        sys.path.insert(0, str(ROOT / "tools"))
        from olib import ledger_append
        ledger_append("SCORING", "HANDOFF_DOWNLOADED", "사람:난희",
                      evidence={"발행본": N(pub.name), "응시 범위": scope_ko, "생성": stamp},
                      product=prod)
    except Exception:
        pass
    tag = "검색축만" if search else "전체응시"
    return buf.getvalue(), f"{prod}_전달꾸러미_{tag}_{datetime.date.today():%m%d}.zip"


def api_report(prod, rnd):
    """회차 리포트 꾸러미 — 리포트 md + 원자료(json/xlsx)를 zip으로 (툴에서 직접 다운로드)"""
    import io
    import re as _re
    import zipfile
    if not _re.fullmatch(r"[A-Z0-9]{1,8}", prod) or not _re.fullmatch(r"r\d{1,3}", rnd):
        return None, "인자 오류"
    d = ROOT / "results" / f"score_{prod}_{rnd}"
    if not d.is_dir():
        return None, f"{prod} {rnd} 채점 산출물 없음"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for p in sorted(d.iterdir()):
            if p.is_file() and not p.name.startswith("."):
                z.write(p, N(p.name))
    sys.path.insert(0, str(ROOT / "tools"))
    from olib import ledger_append
    ledger_append("SCORING", "REPORT_DOWNLOADED", "사람:난희",
                  evidence={"회차": rnd, "구성": [N(p.name) for p in sorted(d.iterdir()) if p.is_file()]},
                  product=prod)
    return buf.getvalue(), f"{prod}_{rnd}_리포트꾸러미.zip"


def _calin_file(prod):
    d = ROOT / "data" / prod / "06_calibration"
    c = sorted(d.glob("*판정30*.xlsx")) if d.is_dir() else []
    return c[-1] if c else None


def api_calin(prod):
    """⑥ 채점관 면접 답안지 — 카드 안에서 바로 판정하도록 기입 시트를 화면에 노출.
    블라인드: judge 판정(숨김 시트)은 절대 내보내지 않는다."""
    import openpyxl
    f = _calin_file(prod)
    if not f:
        return {"rows": []}
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    sn = next((s for s in wb.sheetnames if "기입" in N(s)), None)
    if not sn:
        wb.close()
        return {"rows": []}
    rows = []
    for r in wb[sn].iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        rows.append({"id": N(r[0]), "type": N(r[1]), "q": N(r[2]), "a": N(r[3]),
                     "crit": N(r[4]), "v": N(r[5]) if len(r) > 5 and r[5] else ""})
    wb.close()
    return {"rows": rows, "file": N(f.name)}


def api_action(payload):
    """approve / reject / onboard / resume / run — pipeline CLI 경유 (원장 기록 보장)"""
    cmd = payload.get("cmd")
    args = [sys.executable, str(ROOT / "tools" / "pipeline.py")]
    if cmd == "approve":
        args += ["approve", payload["gate_id"], "--actor", payload.get("actor", "난희")]
        if payload.get("ack_all"):
            args += ["--ack-all"]
    elif cmd == "reject":
        if not payload.get("reason"):
            return {"ok": False, "out": "반려는 사유 필수"}
        args += ["reject", payload["gate_id"], "--reason", payload["reason"],
                 "--actor", payload.get("actor", "난희")]
    elif cmd == "resume":
        if not payload.get("reason"):
            return {"ok": False, "out": "HALT 해제는 사유 필수"}
        args += ["resume", "--after-fix", payload["product"], "--reason", payload["reason"],
                 "--actor", payload.get("actor", "난희")]
    elif cmd == "onboard":
        args += ["onboard", "--product", payload["product"], "--name", payload.get("name", ""),
                 "--base", payload.get("base", "blank"), "--actor", payload.get("actor", "난희"),
                 "--start", payload.get("start", "full"),
                 "--strategy", payload.get("strategy", "ensemble")]
    elif cmd == "set-strategy":
        args += ["set-strategy", "--product", payload["product"],
                 "--strategy", payload["strategy"], "--actor", payload.get("actor", "난희")]
    elif cmd == "set-members":
        args += ["set-members", "--product", payload["product"],
                 "--use", payload.get("use", "generator"), "--actor", payload.get("actor", "난희")]
    elif cmd == "new-round":
        args += ["new-round", "--product", payload["product"], "--actor", payload.get("actor", "난희")]
    elif cmd == "expand":
        args += ["expand", "--product", payload["product"], "--actor", payload.get("actor", "난희")]
    elif cmd == "qa-import":
        # 외부 Q&A 인입 — 업로드 직후 자동 대조·분류 (G19)
        args += ["qa-import", "--product", payload["product"], "--actor", payload.get("actor", "난희")]
    elif cmd == "close-ensemble":
        # ③ 앙상블 조기 마감 — 완주한 추출자 기준으로 병합 진행 (난희 요청)
        args += ["close-ensemble", "--product", payload["product"], "--actor", payload.get("actor", "난희")]
        if payload.get("force"):
            args += ["--force"]
    elif cmd == "qa-score":
        # 외부 Q&A 별도 트랙 채점 — 응답로그 업로드 직후 (G20)
        args += ["qa-score", "--product", payload["product"], "--actor", payload.get("actor", "난희")]
        if payload.get("log"):
            args += ["--log", payload["log"]]
    elif cmd == "s2diff-set":
        # 불일치 문항 확정 클릭 → 판정대장 '최종 판정' 컬럼에 즉시 기록
        import openpyxl
        prod, iid, final = payload["product"], N(payload.get("id", "")), N(payload.get("final", ""))
        f = _s2_ledger_file(prod)
        if not f:
            return {"ok": False, "out": "판정대장 없음"}
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        hdr = [N(c.value) for c in ws[1]]
        if "최종 판정" not in hdr:
            ws.cell(1, len(hdr) + 1).value = "최종 판정"
            ws.cell(1, len(hdr) + 2).value = "사람 개입"
            hdr += ["최종 판정", "사람 개입"]
        fc, hc = hdr.index("최종 판정") + 1, hdr.index("사람 개입") + 1
        hit = False
        for r in range(2, ws.max_row + 1):
            if N(ws.cell(r, 1).value) == iid:
                ws.cell(r, fc).value = final
                ws.cell(r, hc).value = "○"
                hit = True
                break
        if not hit:
            return {"ok": False, "out": f"문항 없음: {iid}"}
        wb.save(f)
        sys.path.insert(0, str(ROOT / "tools"))
        from olib import ledger_append
        ledger_append("STAGE2", "S2DIFF_HUMAN_FINAL", "사람:난희",
                      evidence={"문항": iid, "최종": final}, product=prod)
        return {"ok": True, "out": f"{iid} → 최종 {final}"}
    elif cmd == "calin-set-bulk":
        # 체크한 문항들 일괄 판정 — 파일 1회 열고 한 번에 기록
        import openpyxl
        prod, v = payload["product"], N(payload.get("verdict", ""))
        ids = {N(i) for i in payload.get("ids", [])}
        if v not in ("합격", "부분", "0점", "불합격", "") or not ids:   # 불합격=구형 시트 호환
            return {"ok": False, "out": "판정 값/대상 오류"}
        f = _calin_file(prod)
        if not f:
            return {"ok": False, "out": "판정지 없음"}
        wb = openpyxl.load_workbook(f)
        ws = wb[next(s for s in wb.sheetnames if "기입" in N(s))]
        hit = done = total = 0
        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            if N(row[0].value) in ids:
                row[5].value = v
                hit += 1
            total += 1
            done += 1 if N(row[5].value or "") else 0
        wb.save(f)
        return {"ok": True, "hit": hit, "filled": done, "total": total,
                "out": f"{hit}건 일괄 {v or '해제'} · {done}/{total}"}
    elif cmd == "calin-set":
        # 카드 안 판정 클릭 → 기입 시트에 즉시 기록 (엑셀 파일이 단일 원장 — 채점기와 동일 소스)
        import openpyxl
        prod, iid, v = payload["product"], N(payload.get("id", "")), N(payload.get("verdict", ""))
        if v not in ("합격", "부분", "0점", "불합격", ""):   # 불합격=구형 시트 호환
            return {"ok": False, "out": "판정 값 오류"}
        f = _calin_file(prod)
        if not f:
            return {"ok": False, "out": "판정지 없음"}
        wb = openpyxl.load_workbook(f)
        sn = next((s for s in wb.sheetnames if "기입" in N(s)), None)
        ws = wb[sn]
        hit, done, total = False, 0, 0
        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            if N(row[0].value) == iid:
                row[5].value = v
                hit = True
            total += 1
            done += 1 if N(row[5].value or "") else 0
        if not hit:
            return {"ok": False, "out": f"문항 없음: {iid}"}
        wb.save(f)
        return {"ok": True, "filled": done, "total": total, "out": f"기록됨 · {done}/{total}"}
    elif cmd == "dismiss-card":
        # 알림형 카드 닫기 — 완료로 이동 (MODELALERT 전용, 경로 이탈 차단)
        name = N(payload.get("file", "")).replace("/", "").replace("..", "")
        if not name.startswith("MODELALERT_") or not name.endswith(".md"):
            return {"ok": False, "out": "닫기는 알림형(MODELALERT) 카드만 가능"}
        src = ROOT / "검수큐" / name
        if not src.exists():
            return {"ok": False, "out": "카드 없음"}
        dst = ROOT / "검수큐" / "완료"
        dst.mkdir(exist_ok=True)
        (dst / name).write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
        src.unlink()
        sys.path.insert(0, str(ROOT / "tools"))
        from olib import ledger_append
        ledger_append("MAINTENANCE", "ALERT_DISMISSED", "사람:난희",
                      evidence={"card": name, "처리": "확인 후 닫음 — 같은 내용 재발행 억제"})
        return {"ok": True, "out": "알림 닫음 — 같은 내용으로는 다시 안 떠요"}
    elif cmd == "run":
        # 장시간 AI 작업(③ 등) — 백그라운드 실행. 서버(단일 스레드)와 화면이 얼지 않게.
        import os
        prod = payload["product"]
        # 출제 AI 미연결 가드 — 돌려봤자 즉시 실패할 실행을 친절하게 차단
        try:
            sys.path.insert(0, str(ROOT / "tools"))
            from model_adapter import detect_mode
            from olib import load_config as _lc
            if os.environ.get("ORCH_MOCK") != "1" and not detect_mode(_lc(), os.environ)["have"].get("generator"):
                return {"ok": False,
                        "out": "⛔ 출제 AI(claude)가 연결돼 있지 않아 실행할 수 없어요. "
                               "터미널에서 claude 로그인(구독) 또는 키설정.txt에 API 키를 넣고 "
                               "실전모드_켜기.command 를 다시 실행한 뒤 시도해 주세요."}
        except ImportError:
            pass   # 어댑터 로드 실패 시 가드는 건너뛴다 (실행 자체를 막지 않음)
        pidf = ROOT / "results" / f"_run_{prod}.pid"
        pidf.parent.mkdir(exist_ok=True)
        # 중복 실행 방지 — 이미 도는 프로세스가 있으면 새로 안 띄운다 (재개 버튼 연타 등)
        if pidf.exists():
            try:
                old = int(pidf.read_text())
                stat = subprocess.run(["ps", "-o", "stat=", "-p", str(old)],
                                      capture_output=True, text=True).stdout.strip()
                if stat and not stat.startswith("Z"):   # 좀비(Z)는 죽은 것 — 실행 중 오인 금지
                    return {"ok": True, "out": "이미 실행 중이에요 — 진행선에서 상태를 확인하세요 (중복 실행 안 함)."}
            except ValueError:
                pass                      # 깨진 pid 파일 — 무시하고 새로 시작
        # 무인 자동 진행 워커 — 한도로 멈춰도 스스로 재개 (사람이 밤새 버튼 누를 필요 없음)
        worker = [sys.executable, str(ROOT / "tools" / "auto_run.py"), "--product", prod]
        # 원문 로그 보관 체계: 제품별 폴더 + 날짜_시각_시작파트 파일명 — 나중에 찾을 수 있게
        import datetime
        try:
            stage = json.loads((ROOT / "state.json").read_text(encoding="utf-8"))["products"][prod]["stage"]
        except Exception:
            stage = "run"
        ldir = ROOT / "results" / "logs" / prod
        ldir.mkdir(parents=True, exist_ok=True)
        logf = ldir / f"{datetime.datetime.now():%Y-%m-%d_%H%M}_{stage}.log"
        cur = ROOT / "results" / f"_run_{prod}.log"
        cur.unlink(missing_ok=True)
        cur.symlink_to(logf)          # '_run_<제품>.log' = 항상 최신 로그를 가리키는 별칭
        p = subprocess.Popen(worker, cwd=str(ROOT), env={**os.environ},
                             stdout=open(logf, "ab"), stderr=subprocess.STDOUT)
        pidf.write_text(str(p.pid))
        return {"ok": True, "out": "⏳ 실행 시작 — 한도로 멈춰도 자동으로 이어가요(사람 개입 불필요). 진행은 트랙 아래 진행선에서."}
    else:
        return {"ok": False, "out": f"미지원: {cmd}"}
    import os
    p = subprocess.run(args, capture_output=True, text=True, cwd=str(ROOT),
                       env={**os.environ})   # ORCH_MOCK 등 서버 환경 승계
    return {"ok": p.returncode == 0, "out": (p.stdout + p.stderr).strip()}


class H(SimpleHTTPRequestHandler):
    def __init__(self, *a, **kw):
        super().__init__(*a, directory=str(ROOT), **kw)

    def log_message(self, *a):
        pass

    def _send(self, code, body, ctype="application/json; charset=utf-8"):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        if self.path == "/" or self.path.startswith("/?") or self.path.startswith("/dashboard/index.html"):
            # 대시보드 HTML은 항상 최신으로 — 브라우저 캐시가 옛 화면을 보여주는 사고 방지
            body = (ROOT / "dashboard" / "index.html").read_bytes()
            return self._send(200, body, "text/html; charset=utf-8")
        if self.path.startswith("/api/state"):
            return self._send(200, j(api_state()))
        if self.path.startswith("/api/ledger"):
            return self._send(200, j(api_ledger()))
        if self.path.startswith("/api/queue"):
            return self._send(200, j(api_queue()))
        if self.path.startswith("/api/scores"):
            return self._send(200, j(api_scores()))
        if self.path.startswith("/api/catalog"):
            return self._send(200, j(api_catalog()))
        if self.path.startswith("/api/ai_status"):
            return self._send(200, j(api_ai_status()))
        if self.path.startswith("/api/progress"):
            from urllib.parse import urlparse, parse_qs
            q = parse_qs(urlparse(self.path).query)
            return self._send(200, j(api_progress(N(q.get("product", [""])[0]))))
        if self.path.startswith("/api/calin"):
            from urllib.parse import urlparse, parse_qs
            q = parse_qs(urlparse(self.path).query)
            return self._send(200, j(api_calin(N(q.get("product", [""])[0]))))
        if self.path.startswith("/api/xlsx"):
            from urllib.parse import urlparse, parse_qs, unquote
            q = parse_qs(urlparse(self.path).query)
            return self._send(200, j(api_xlsx(N(unquote(q.get("path", [""])[0]))or None,
                                              name=N(unquote(q.get("name", [""])[0])) or None,
                                              prod=N(q.get("product", [""])[0]) or None)))
        if self.path.startswith("/api/qa-handoff"):
            from urllib.parse import urlparse, parse_qs, quote
            q = parse_qs(urlparse(self.path).query)
            data, fname = api_qa_handoff(N(q.get("product", [""])[0]))
            if data is None:
                return self._send(404, j({"ok": False, "out": fname}))
            self.send_response(200)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(fname)}")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if self.path.startswith("/api/handoff"):
            from urllib.parse import urlparse, parse_qs
            q = parse_qs(urlparse(self.path).query)
            data, fname = api_handoff(N(q.get("product", [""])[0]),
                                      N(q.get("scope", ["full"])[0]) or "full")
            if data is None:
                return self._send(404, j({"ok": False, "out": fname}))
            self.send_response(200)
            self.send_header("Content-Type", "application/zip")
            from urllib.parse import quote
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(fname)}")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if self.path.startswith("/api/report"):
            from urllib.parse import urlparse, parse_qs
            q = parse_qs(urlparse(self.path).query)
            data, fname = api_report(N(q.get("product", [""])[0]), N(q.get("round", [""])[0]))
            if data is None:
                return self._send(404, j({"ok": False, "out": fname}))
            from urllib.parse import quote
            self.send_response(200)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(fname)}")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if self.path.startswith("/api/s2diff"):
            from urllib.parse import urlparse, parse_qs
            q = parse_qs(urlparse(self.path).query)
            return self._send(200, j(api_s2diff(N(q.get("product", [""])[0]))))
        if self.path.startswith("/api/runlog"):
            from urllib.parse import urlparse, parse_qs
            q = parse_qs(urlparse(self.path).query)
            return self._send(200, j(api_runlog(N(q.get("product", [""])[0]))))
        return super().do_GET()

    def do_POST(self):
        # [P2-6] CSRF 방어 — 브라우저의 타 사이트 페이지가 localhost로 승인/업로드 POST를
        # 쏘는 것 차단. Origin 없는 요청(CLI·봇·curl)은 로컬 도구라 허용.
        origin = self.headers.get("Origin")
        if origin and origin not in ("http://localhost:8791", "http://127.0.0.1:8791"):
            return self._send(403, j({"ok": False, "out": f"차단된 출처: {origin}"}))
        if self.path.startswith("/api/action"):
            n = int(self.headers.get("Content-Length", 0))
            payload = json.loads(self.rfile.read(n) or b"{}")
            return self._send(200, j(api_action(payload)))
        if self.path.startswith("/api/upload"):
            return self._send(200, j(self._upload()))
        return self._send(404, j({"ok": False}))

    UPLOAD_DIRS = {"CORPUS": "corpus", "SCORING": "08_scoring",
                   "COVERAGE": "03_coverage_map", "UNIFIED": "05_unified_ledger",
                   "CALIBRATION": "06_calibration",
                   "QALOG": "external_qa/로그",   # Q&A 응답로그 — 원본과 섞이면 재대조가 오인 (QA보다 먼저 매칭돼야 함)
                   "QA": "external_qa"}

    def _upload(self):
        """INPUT 카드용 파일 업로드 — 쿼리: product, target(카드 종류), name. 본문 = 파일 원바이트.
        저장 후 입구 검사는 pipeline run 이 수행 ('받으면 무조건 실측부터')."""
        from urllib.parse import urlparse, parse_qs
        q = parse_qs(urlparse(self.path).query)
        prod = N(q.get("product", [""])[0])
        target = N(q.get("target", ["CORPUS"])[0]).upper()
        name = N(q.get("name", ["upload.bin"])[0])   # [P3] parse_qs가 이미 디코드 — unquote 중복 금지
        name = name.replace("/", "_").replace("\\", "_").replace("..", "_")  # 경로 이탈 차단
        sub = next((d for k, d in self.UPLOAD_DIRS.items() if k in target), "corpus")
        # 채점만 모드: SCORING 카드에 골든셋(xlsx)과 로그(json)를 같이 올려도 자동 분류
        if "SCORING" in target and name.lower().endswith(".xlsx"):
            sub = "05_unified_ledger"
        if not re.fullmatch(r"[A-Z0-9]{1,8}", prod):
            return {"ok": False, "out": f"제품 코드 오류: {prod}"}
        dest_dir = ROOT / "data" / prod / sub
        dest_dir.mkdir(parents=True, exist_ok=True)
        n = int(self.headers.get("Content-Length", 0))
        if n <= 0 or n > 500 * 1024 * 1024:
            return {"ok": False, "out": f"크기 오류: {n}B"}
        dest = dest_dir / name
        with open(dest, "wb") as f:
            remaining = n
            while remaining > 0:
                chunk = self.rfile.read(min(1 << 20, remaining))
                if not chunk:
                    break
                f.write(chunk)
                remaining -= len(chunk)
        if remaining > 0:
            # [P3] 부분 수신 = 잘린 파일 — 저장·원장 기록 금지 (손상 코퍼스가 입구 검사로 흘러들던 것)
            dest.unlink(missing_ok=True)
            return {"ok": False, "out": f"업로드 중단 감지 — {n-remaining:,}/{n:,}B만 수신. 다시 올려주세요."}
        # 응시 범위 선언 사이드카 — 형식 게이트가 실물(answer null 여부)과 대조
        scope = N(q.get("scope", [""])[0])
        if scope in ("search", "full") and name.lower().endswith(".json") and "08_scoring" in sub:
            (dest_dir / f"{name}.scope").write_text(scope, encoding="utf-8")
        sys.path.insert(0, str(ROOT / "tools"))
        from olib import ledger_append
        ledger_append("INPUT", "FILE_UPLOADED", "사람:대시보드",
                      evidence={"file": name, "size": n, "dest": f"data/{prod}/{sub}/",
                                **({"응시 범위 선언": "검색축만" if scope == "search" else "전체"}
                                   if scope in ("search", "full") else {})},
                      product=prod)
        return {"ok": True, "out": f"업로드 완료: {name} ({n:,}B) → data/{prod}/{sub}/ — 입구 검사를 실행합니다"}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--port", type=int, default=8791)
    a = ap.parse_args()
    srv = HTTPServer(("127.0.0.1", a.port), H)
    print(f"관제 대시보드: http://localhost:{a.port}/  (Ctrl+C 종료)")
    srv.serve_forever()


if __name__ == "__main__":
    main()
