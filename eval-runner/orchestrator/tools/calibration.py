#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
calibration.py — ⑥ 캘리브레이션 게이트 + 규칙 C (사양서 v1.3 §4-⑥, 규칙 C)

- 30건 이중 판정(LLM vs 사람) 대조표에서 일치율을 **실측 재계산** (기록된 '일치 여부'를
  믿지 않고 원컬럼 쌍으로 다시 센다 — P-02 선언≠실측. 기록과 다르면 경고).
- 일치율 ≥ threshold(0.90) → 사람 판정측 게이트 카드 발행 (승인 시 calibration_passed=true).
- 미달 → 기준서·프롬프트 개정 후 재측정 권고 카드.
- 규칙 C: 모델 구성(models 섹션) 또는 기준서 버전이 바뀌면 calibration_passed=false 리셋
  → 본판정 잠금. 지문(fingerprint) = sha256(models JSON + 기준서 파일 해시들).

사용:
  python3 tools/calibration.py measure --product RV
  python3 tools/calibration.py check-config --product RV     # 규칙 C 감시
  python3 tools/calibration.py selftest
"""
import argparse
import hashlib
import json
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from olib import (ROOT, N, load_config, load_state, save_state,
                  ledger_append, issue_gate_card, issue_input_card)

DATA = ROOT / "data"


def find_table(prod):
    d = DATA / prod / "06_calibration"
    if not d.is_dir():
        return None
    cands = (sorted(d.glob("*대조표*완료*.xlsx")) or sorted(d.glob("*대조표*.xlsx"))
             or sorted(d.glob("*판정30*.xlsx")))
    return cands[-1] if cands else None


def measure_agreement(path):
    """대조표에서 (judge, 사람) 판정 쌍 실측 재계산."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "대조표" in N(s)), wb.sheetnames[0])
    ws = wb[sheet]
    hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    ji = next((i for i, h in enumerate(hdr) if "judge" in h.lower() and "판정" in h), None)
    hi = next((i for i, h in enumerate(hdr) if "사람" in h and "판정" in h), None)
    ri = next((i for i, h in enumerate(hdr) if "일치" in h), None)
    ci = next((i for i, h in enumerate(hdr) if "사유" in h), None)
    if ji is None or hi is None:
        wb.close()
        raise ValueError(f"대조표 컬럼 인식 실패: {hdr}")
    pairs, recorded_match, causes = [], 0, {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        j = N(r[ji]) if ji < len(r) else ""
        h = N(r[hi]) if hi < len(r) else ""
        if not j or not h:
            continue
        pairs.append((j, h))
        if ri is not None and ri < len(r) and "일치" == N(r[ri]):
            recorded_match += 1
        if ci is not None and ci < len(r) and N(r[ci]):
            causes[N(r[ci])] = causes.get(N(r[ci]), 0) + 1
    wb.close()
    measured = sum(1 for j, h in pairs if j == h)
    return {"total": len(pairs), "measured_match": measured,
            "recorded_match": recorded_match, "agreement": measured / len(pairs) if pairs else 0.0,
            "mismatch_causes": causes, "sheet": sheet}


# ── 규칙 C 지문 ─────────────────────────────────────────────
def _cli_version(cmd0):
    """[FIX-06] CLI 버전 문자열 — 회귀용 오버라이드(ORCH_CLI_VERSION_OVERRIDE) 지원.
    한계: 클라이언트 업그레이드만 감지 — 서버 쪽 구독 기본모델 교체는 못 잡는다.
    1차 방어는 config에 model 명시 고정 (지문은 클라이언트까지만 본다)."""
    import os
    import subprocess
    ov = os.environ.get("ORCH_CLI_VERSION_OVERRIDE")
    if ov:
        return ov
    try:
        p = subprocess.run([cmd0, "--version"], capture_output=True, text=True, timeout=20)
        v = (p.stdout or p.stderr).strip()[:120]
        return v or "unknown"
    except Exception:
        print(f"⚠ {cmd0} --version 실패 — 지문에 'unknown' 반영 (버전 변경 감지 불가 상태)")
        return "unknown"


def fingerprint(prod, cfg=None):
    cfg = cfg or load_config()
    h = hashlib.sha256()
    h.update(json.dumps(cfg.get("models", {}), sort_keys=True, ensure_ascii=False).encode())
    # [FIX-06] CLI provider 는 클라이언트 버전 문자열도 지문에 포함
    for role, m in sorted((cfg.get("models") or {}).items()):
        if m and m.get("provider") == "cli":
            cmd = m.get("command") or ["claude"]
            cmd0 = cmd[0] if isinstance(cmd, list) else str(cmd).split()[0]
            h.update(f"{role}:{_cli_version(cmd0)}".encode())
    # 기준서(judge 프롬프트) 파일들 — 07_stage2 의 프롬프트 md
    for p in sorted((DATA / prod / "07_stage2").glob("*프롬프트*")):
        h.update(p.read_bytes())
    return h.hexdigest()


def check_config(prod, cfg=None, actor="script:rule_c"):
    """모델 구성/기준서 변경 감시 — 변경 시 calibration_passed=false 리셋 (규칙 C)"""
    cfg = cfg or load_config()
    fp = fingerprint(prod, cfg)
    st = load_state()
    ps = st["products"].get(prod)
    if ps is None:
        return False
    stored = ps.get("judge_fingerprint")
    if stored is None:
        ps["judge_fingerprint"] = fp
        save_state(st)
        return False
    if stored != fp:
        was = ps["calibration_passed"]
        ps["calibration_passed"] = False
        ps["judge_fingerprint"] = fp
        save_state(st)
        ledger_append("CALIBRATION", "RULE_C_RESET", actor,
                      evidence={"calibration_was": was},
                      reason="모델 구성 또는 기준서 버전 변경 — 캘리브레이션 게이트 재통과 전 본판정 차단 "
                             "(채점관을 교체하면 면접부터 다시 본다)",
                      product=prod)
        print(f"⚠ 규칙 C 발동 — {prod} calibration_passed=false 리셋 (원장 기록)")
        return True
    return False


# ── pipeline 인터페이스 ──────────────────────────────────────
def run(prod, cfg):
    st = load_state()
    if st["products"][prod].get("calibration_passed"):
        return "DONE", {"캘리브레이션": "이미 통과 (규칙 C 지문 유효)"}
    t = find_table(prod)
    if not t:
        # 대조표 자체가 없으면 judge 30건 판정을 실행해 만든다 (⑥ 실행부)
        import judge_run
        outcome, ev = judge_run.run_calibration_judging(prod, cfg)
        if outcome != "DONE":
            return outcome, ev
        t = find_table(prod)
    m = measure_agreement(t)
    if m["total"] == 0:
        # judge 판정은 있으나 사람 블라인드 판정 미기입 — 사람 게이트로 대기
        issue_gate_card(prod, "CALIBRATION", f"CALIN_{prod}",
                        what_stopped=f"사람 블라인드 판정 30건 기입 대기 — {N(t.name)} 시트 '2_대조표_판정완료'의 "
                                     f"'사람 판정' 컬럼을 기입한 뒤 승인 (기입 전 승인 금지)",
                        evidence={"파일": N(t.name), "judge 판정": "완료·보존", "사람 판정": "0/30 기입"},
                        recommendation="블라인드 원칙 — judge 판정 열람 전 독립 판정")
        return "WAITING_HUMAN", {"사람 판정": "미기입"}
    thr = cfg["pipeline"].get("calibration_threshold", 0.90)
    ev = {"대조표": N(t.name), "일치율(실측)": f"{m['measured_match']}/{m['total']} = {m['agreement']:.1%}",
          "임계": f"≥{thr:.0%}", "불일치 분류": m["mismatch_causes"] or "—"}
    warn = ""
    if m["recorded_match"] and m["recorded_match"] != m["measured_match"]:
        warn = (f"⚠ 기록된 일치({m['recorded_match']}) ≠ 실측({m['measured_match']}) — "
                f"assert의 대조 기준 자체를 검증하라")
        ev["경고"] = warn
    if m["agreement"] >= thr:
        issue_gate_card(prod, "CALIBRATION", f"CAL_{prod}",
                        what_stopped=f"캘리브레이션 임계 통과 — 사람 판정측 최종 확인 게이트 (승인 시 본판정 잠금 해제)",
                        evidence=ev,
                        recommendation="일치율 임계 충족. 불일치 건 3분류 → 기준서 개정 반영 여부 확인 후 승인.")
        return "WAITING_HUMAN", ev
    else:
        issue_gate_card(prod, "CALIBRATION", f"CAL_{prod}",
                        what_stopped=f"캘리브레이션 임계 미달 ({m['agreement']:.1%} < {thr:.0%}) — 기준서·프롬프트 개정 후 재측정",
                        evidence=ev,
                        recommendation="불일치 사유 3분류(기준서 공백/문항 모호/판정 오류)별 개정 → 재측정. 승인 금지 권고.")
        return "WAITING_HUMAN", ev


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    s = sub.add_parser("measure"); s.add_argument("--product", required=True)
    s = sub.add_parser("check-config"); s.add_argument("--product", required=True)
    sub.add_parser("selftest")
    a = ap.parse_args()
    cfg = load_config()
    if a.cmd == "measure":
        t = find_table(a.product)
        if not t:
            sys.exit("대조표 없음")
        m = measure_agreement(t)
        thr = cfg["pipeline"]["calibration_threshold"]
        print(f"대조표: {N(t.name)}")
        print(f"일치율(실측): {m['measured_match']}/{m['total']} = {m['agreement']:.1%} (임계 ≥{thr:.0%})")
        print(f"기록된 일치: {m['recorded_match']} {'(실측과 일치)' if m['recorded_match']==m['measured_match'] else '⚠ 실측과 다름!'}")
        print(f"불일치 분류: {m['mismatch_causes']}")
        print(f"판정: {'통과' if m['agreement'] >= thr else '미달 — 개정 후 재측정'}")
    elif a.cmd == "check-config":
        changed = check_config(a.product, cfg)
        if not changed:
            print(f"{a.product} — 구성 변경 없음 (지문 일치)")
    elif a.cmd == "selftest":
        selftest(cfg)


def selftest(cfg):
    ok = True
    # 1) 실데이터: RV 대조표 → 27/30 = 90.0% ≥ 90%
    m = measure_agreement(find_table("RV"))
    good = m["total"] == 30 and m["measured_match"] == 27 and abs(m["agreement"] - 0.90) < 1e-9
    ok &= good
    print(f"  {'✅' if good else '❌'} RV 대조표 실측: {m['measured_match']}/{m['total']} = {m['agreement']:.1%} (기대 27/30=90%)")
    good = m["agreement"] >= cfg["pipeline"]["calibration_threshold"]
    ok &= good
    print(f"  {'✅' if good else '❌'} 임계 {cfg['pipeline']['calibration_threshold']:.0%} 게이트 → 통과")
    # 2) 규칙 C: 지문 안정성 (같은 입력 → 같은 지문)
    good = fingerprint("RV", cfg) == fingerprint("RV", cfg)
    ok &= good
    print(f"  {'✅' if good else '❌'} 지문 결정성")
    # 3) 규칙 C: 모델 구성 변경 → 지문 변화
    import copy
    cfg2 = copy.deepcopy(cfg)
    cfg2["models"]["judge"]["model"] = "다른-채점관"
    good = fingerprint("RV", cfg) != fingerprint("RV", cfg2)
    ok &= good
    print(f"  {'✅' if good else '❌'} 모델 교체 → 지문 변화 (리셋 트리거)")
    print(f"selftest: {'전건 통과' if ok else '실패 있음'}")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
