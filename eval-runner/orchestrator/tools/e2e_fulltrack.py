#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
e2e_fulltrack.py — 신규 제품 전 트랙 관통 테스트 (①→⑧)

"코퍼스 하나 넣으면 커버리지맵 → 골든셋 → 캘리브레이션 → 본판정 → 질문셋 발행 →
 응답 로그 → 형식 게이트 → 채점"까지 오케스트레이터가 실제로 해내는지 증명한다.

- ORCH_MOCK=1: 결정적 mock 모델 (키 없이 기계 전체 검증). 키 설정 시 같은 코드가 실모델.
- 사람 게이트는 스크립트가 사람 역할을 시뮬레이션 (블라인드 판정 기입 포함).
- 성공 기준: 최종 상태 = SCORING·WAITING_HUMAN(성적표 확정 게이트) + 각 단계 산출물 실재.

사용: ORCH_MOCK=1 python3 tools/e2e_fulltrack.py [--keep]
"""
import json
import os
import shutil
import subprocess
import sys
import unicodedata
from pathlib import Path

os.environ.setdefault("ORCH_MOCK", "1")
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "tools"))
PROD = "EE"   # 자동 테스트 전용 (TT는 난희 연습장 — 건드리지 않음)
DATA = ROOT / "data" / PROD


def N(s):
    return unicodedata.normalize("NFC", str(s)) if s is not None else ""


def sh(*args, ok_codes=(0,)):
    p = subprocess.run([sys.executable, str(ROOT / "tools" / "pipeline.py"), *args],
                       capture_output=True, text=True, cwd=str(ROOT),
                       env={**os.environ, "ORCH_MOCK": "1"})
    out = p.stdout + p.stderr
    if p.returncode not in ok_codes:
        print(out)
        raise SystemExit(f"pipeline {args} → exit {p.returncode}")
    return out


def state():
    return json.loads((ROOT / "state.json").read_text(encoding="utf-8"))["products"][PROD]


def gates():
    return [g["id"] for g in state()["open_gates"]]


CORPUS_DOCS = {
    "manual": [
        "티티뷰어는 원격지 PC 화면을 실시간으로 표시하는 기능을 제공한다. 최대 4대까지 동시 모니터링이 가능하다.",
        "화면 잠금 기능을 켜면 원격 제어 중 로컬 모니터가 검게 표시된다. 개인정보 보호를 위해 권장된다.",
        "파일 전송은 드래그 앤 드롭으로 동작하며 1회 최대 2GB까지 지원한다.",
        "모바일 앱은 iOS 15 이상과 Android 10 이상에서 설치할 수 있다.",
    ],
    "policy": [
        "무료 평가판은 14일 동안 모든 기능을 제한 없이 사용할 수 있다.",
        "라이선스는 동시 접속 세션 수 기준으로 과금되며 연간 계약이 기본이다.",
        "개인정보는 수집일로부터 3년 후 자동 파기된다. 파기 내역은 관리 콘솔에서 확인 가능하다.",
        "장애 보상은 월 가용률 99.5퍼센트 미만일 때 이용료의 10퍼센트를 크레딧으로 지급한다.",
    ],
    "notice": [
        "2026년 6월 정기 점검은 매월 둘째 주 화요일 새벽 2시부터 4시까지 진행된다.",
        "신규 버전 9.2는 클립보드 동기화 성능이 개선되었고 다국어 입력 버그가 수정되었다.",
        "구버전 8.x 지원은 2026년 12월 31일에 종료된다. 사전에 업그레이드가 필요하다.",
        "보안 패치 KB-2026-07은 원격 인쇄 모듈의 취약점을 수정한다. 즉시 적용을 권장한다.",
    ],
}


def approve_all(max_rounds=40):
    """사람 역할 시뮬: 열린 게이트를 전부 승인하며 run 반복. 특수 게이트는 사람 작업 수행."""
    log = []
    for _ in range(max_rounds):
        ps = state()
        if ps["status"] == "HALTED":
            raise SystemExit(f"HALT: {ps['halt_reason']}")
        gs = gates()
        if not gs:
            out = sh("run", "--product", PROD)
            log.append(out.strip().splitlines()[-1] if out.strip() else "")
            ps = state()
            if ps["stage"] == "SCORING" and ps["status"] == "WAITING_HUMAN":
                return log  # 성적표 확정 게이트 도달 = 목표
            if ps["status"] in ("DONE",):
                return log
            if ps["status"] == "WAITING_INPUT":
                return log  # 호출측에서 입력물 공급
            if not gates() and ps["status"] == "WAITING_HUMAN":
                raise SystemExit("게이트 없는 WAITING_HUMAN — 교착")
            continue
        for g in gs:
            if g.startswith("CALIN_"):
                fill_human_judgments()
            sh("approve", g, "--ack-all", "--actor", "난희(시뮬)")
            log.append(f"승인: {g}")
    raise SystemExit("최대 라운드 초과 — 교착 의심")


def fill_human_judgments():
    """사람 블라인드 판정 시뮬: judge 판정 복사 + 1건만 의도적 불일치 (29/30=96.7%)"""
    import openpyxl
    f = sorted((DATA / "06_calibration").glob("*판정30*.xlsx"))[-1]
    wb = openpyxl.load_workbook(f)
    ws = wb["2_대조표_판정완료"]
    first = True
    for row in ws.iter_rows(min_row=2):
        j = N(row[1].value)
        if not j:
            continue
        h = j
        if first and j == "합격":
            h = "부분"     # 의도적 불일치 1건 — 대조 경로 검증
            first = False
        row[2].value = h
        row[3].value = "일치" if h == j else "불일치"
        if h != j:
            row[4].value = "판정 오류(모의) — E2E 검증용 의도적 불일치"
    wb.save(f)
    print(f"  [사람 시뮬] 블라인드 판정 30건 기입 ({N(f.name)})")


def make_mock_response_log():
    """팀장님 시스템 응시 시뮬: 질문셋 → 응답 로그 json (일부 정답·일부 오답·rank 연속)"""
    import openpyxl
    import hashlib
    qs = sorted((DATA / "08_scoring").glob("*질문셋*발행본*.xlsx"))[-1]
    wb = openpyxl.load_workbook(qs, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [(N(r[0]), N(r[1])) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    wb.close()
    # 정답 원문 참조를 위해 골든셋 로드
    import scoring
    g = scoring.golden_xlsx(PROD)
    gwb = openpyxl.load_workbook(g, read_only=True, data_only=True)
    gws = gwb[gwb.sheetnames[0]]
    hdr = [N(c) for c in next(gws.iter_rows(min_row=1, max_row=1, values_only=True))]
    acol = "정답" if "정답" in hdr else "정답 (필수 포함 요소)"
    ii, ai, si = hdr.index("ID"), hdr.index(acol), hdr.index("근거 출처")
    answers = {}
    for r in gws.iter_rows(min_row=2, values_only=True):
        answers[N(r[ii])] = (N(r[ai]), N(r[si]))
    gwb.close()
    responses = []
    for qid, q in rows:
        h = int(hashlib.sha256(qid.encode()).hexdigest(), 16)
        ans, src = answers.get(qid, ("", ""))
        correct = h % 10 < 7           # 70% 정답 시뮬
        m = __import__("re").search(r"https?://\S+?/([\w.-]+\.md)", src)
        fname = m.group(1) if m else "manual.md"
        hits = [{"rank": k,
                 "source_url": f"https://corpus.local/{fname}" if (correct and k <= 2) else f"https://other.local/etc{k}.md",
                 "source_name": fname if (correct and k <= 2) else f"etc{k}.md",
                 "chunk_id": f"c{k}",
                 "snippet": ans[:60] if correct and k == 1 else f"관련 문서 발췌 {k}"}
                for k in range(1, 6)]
        if qid.split("-")[1].startswith("E"):
            answer = "확인할 수 없습니다 — 관련 자료가 없습니다." if correct \
                else "양자암호 전송은 설정 메뉴에서 켤 수 있습니다."   # 환각 시뮬
        else:
            answer = ans if correct else "잘 모르겠습니다. 다른 자료를 확인해 주세요."
        responses.append({"id": qid, "hits": hits, "answer": answer})
    log = {"meta": {"run_date": "2026-07-16", "system_version": "mock@e2e",
                    "corpus_version": f"{PROD} corpus — 문서 {len(CORPUS_DOCS)}건·청크 "
                                      f"{sum(len(v) for v in CORPUS_DOCS.values())}건"},
           "responses": responses}
    out = DATA / "08_scoring" / f"{PROD}_응답로그_r1.json"
    out.write_text(json.dumps(log, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"  [팀장님 시뮬] 응답 로그 {len(responses)}건 → {N(out.name)}")


def main():
    keep = "--keep" in sys.argv
    # ── 초기화: 이전 TT 흔적 제거
    if DATA.exists():
        shutil.rmtree(DATA)
    td = ROOT / "terrain.d" / f"{PROD}.yaml"
    if td.exists():
        td.unlink()
    st_path = ROOT / "state.json"
    st = json.loads(st_path.read_text(encoding="utf-8"))
    st["products"].pop(PROD, None)
    st_path.write_text(json.dumps(st, ensure_ascii=False, indent=2), encoding="utf-8")
    for f in (ROOT / "검수큐").glob(f"*_{PROD}*.md"):
        f.unlink()

    print("═══ E2E 전 트랙: 신규 제품 TT (mock 모델) ═══")
    # 1) 온보딩
    sh("onboard", "--product", PROD, "--name", "자동테스트", "--base", "RC")
    assert state()["stage"] == "CORPUS_AUDIT" and state()["status"] == "WAITING_INPUT"
    print("① 온보딩 → WAITING_INPUT (INPUT_CORPUS 카드) ✅")

    # 2) 코퍼스 투입 (사람: 파일 배치)
    cdir = DATA / "corpus"
    cdir.mkdir(parents=True, exist_ok=True)
    for doc, chunks in CORPUS_DOCS.items():
        (cdir / f"{doc}.md").write_text("\n\n\n".join(chunks), encoding="utf-8")
    print(f"② 코퍼스 투입: 문서 {len(CORPUS_DOCS)} · 청크 {sum(len(v) for v in CORPUS_DOCS.values())}")

    # 3) 전 트랙 자동 진행 (게이트는 사람 시뮬 승인)
    approve_all()
    ps = state()
    print(f"③ 1차 정지점: {ps['stage']} · {ps['status']}")
    # ⑧ 응답로그 대기면 — 팀장님 시뮬 투입 후 재진행
    if ps["stage"] == "SCORING" and ps["status"] == "WAITING_INPUT":
        make_mock_response_log()
        approve_all()
        ps = state()

    # ── 검증: 산출물 실재 + 최종 상태
    print("\n═══ 산출물 실측 ═══")
    checks = [
        ("커버리지맵", list((DATA / "03_coverage_map").glob("*.xlsx"))),
        ("골든셋 배치", list((DATA / "04_goldenset_batch").glob("*.xlsx"))),
        ("통합 대장", list((DATA / "05_unified_ledger").glob("*.xlsx"))),
        ("캘리브 판정30", list((DATA / "06_calibration").glob("*.xlsx"))),
        ("판정대장", list((DATA / "07_stage2").glob("*판정대장*.xlsx"))),
        ("질문셋 발행본", list((DATA / "08_scoring").glob("*발행본*.xlsx"))),
        ("성적표", list((ROOT / "results").glob(f"score_{PROD}_*/score_report.json"))),
    ]
    ok = True
    for name, files in checks:
        good = len(files) > 0
        ok &= good
        print(f"  {'✅' if good else '❌'} {name}: {[N(f.name) for f in files][:2]}")
    final_ok = ps["stage"] == "SCORING" and ps["status"] == "WAITING_HUMAN" \
        and ps["calibration_passed"]
    ok &= final_ok
    print(f"  {'✅' if final_ok else '❌'} 최종 상태: {ps['stage']} · {ps['status']} · cal={ps['calibration_passed']} (기대: SCORING·WAITING_HUMAN·True)")
    # 원장에 재검 시드 기록 확인
    led = (ROOT / "ledger.jsonl").read_text(encoding="utf-8")
    seed_ok = '"STAGE2_JUDGED"' in led and '"recheck_ids"' in led
    ok &= seed_ok
    print(f"  {'✅' if seed_ok else '❌'} 재검 시드+추출 목록 원장 기록")

    if not keep:
        pass  # 산출물 보존 (재실행 시 자동 초기화)
    print(f"\nE2E: {'전 구간 관통 ✅' if ok else '실패 ❌'}")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
