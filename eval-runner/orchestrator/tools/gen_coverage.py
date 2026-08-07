#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_coverage.py — ③ 커버리지맵 생성 엔진 (사양서 §4-③·③′)

흐름: 코퍼스 로드(청크) → generator 호출(커버 단위 추출) → 스크립트 검수
  (1축: '사실' 필드가 코퍼스 원문에 실재 · Unit ID 중복 0) → 불합격 반려 루프(최대 2회)
  → 통과분으로 커버리지맵 xlsx 생성 → 사람확인 큐 카드 → ③′ GAP_AUDIT(별도 세션).

코퍼스 형식 (data/<제품>/corpus/):
  - *.json: [{doc, chunk_id, text, source?}] 또는 {docs:[{doc, chunks:[...]}]}
  - *.md/*.txt: 파일=문서, 빈 줄 2개 구분=청크
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import llm
from olib import ROOT, N, ledger_append, issue_gate_card, load_config

DATA = ROOT / "data"


def norm(s):
    """대조용 정규화 — 공백 + 제어문자(PDF 추출 NULL 등) 무시.
    [수리 2026-07-23] 스펙 PDF의 'PC\\x00지원' 류 제어문자 때문에 정상 추출이 1축에서 전건 반려되던 사고."""
    return re.sub(r"[\s\x00-\x1f]+", "", N(s)).lower()


def _rows_from_json(d):
    """지원 형식 3종 → [{doc, chunk_id, text, source}]
    ① [{doc, chunk_id, text}]  ② {docs:[{doc, chunks:[...]}]}
    ③ 팀장님 export: {chunks:[{title, chunk_index, text, source_url, ...}]}"""
    if isinstance(d, list):
        return d
    if "docs" in d:
        return [{"doc": doc["doc"], "chunk_id": i, "text": t, "source": doc.get("source", doc["doc"])}
                for doc in d["docs"] for i, t in enumerate(doc["chunks"], 1)]
    if "chunks" in d:
        return [{"doc": c.get("title") or c.get("doc_key") or c.get("doc", "?"),
                 "chunk_id": int(c.get("chunk_index", c.get("chunk_id", 0))) + (1 if "chunk_index" in c else 0),
                 "text": c.get("text", ""),
                 "source": c.get("source_url") or c.get("title") or ""}
                for c in d["chunks"]]
    return []


def load_corpus(prod):
    """→ [{doc, chunk_id, text, source}] — json/md/txt/zip(내부 chunks-*.json) 지원"""
    cdir = DATA / prod / "corpus"
    chunks = []

    def add(rows):
        for r in rows:
            if not N(r.get("text", "")).strip():
                continue
            chunks.append({"doc": N(r["doc"]), "chunk_id": int(r.get("chunk_id", 0)),
                           "text": N(r["text"]), "source": N(r.get("source", r["doc"]))})

    for f in sorted(cdir.glob("*")):
        if f.name.startswith("."):
            continue
        if f.suffix == ".json":
            add(_rows_from_json(json.loads(f.read_text(encoding="utf-8"))))
        elif f.suffix in (".md", ".txt"):
            parts = [p.strip() for p in re.split(r"\n\s*\n\s*\n?", f.read_text(encoding="utf-8")) if p.strip()]
            add([{"doc": f.stem, "chunk_id": i, "text": p, "source": f.name} for i, p in enumerate(parts, 1)])
        elif f.suffix == ".zip" or f.name.endswith(".zip_"):
            # 팀장님 export zip — 내부의 chunks json만 읽는다 (docs 메타 파일은 건너뜀)
            import zipfile
            with zipfile.ZipFile(f) as z:
                inner = [n for n in z.namelist() if n.endswith(".json") and "chunk" in Path(n).name.lower()] \
                    or [n for n in z.namelist() if n.endswith(".json")]
                for n in inner:
                    try:
                        rows = _rows_from_json(json.loads(z.read(n).decode("utf-8")))
                    except Exception:
                        continue
                    if rows:
                        add(rows)
    return chunks


SYSTEM = """[TASK:COVERAGE_UNITS] 너는 RAG 평가용 커버리지맵 추출기다.
입력 JSON의 각 청크에서 '평가 가능한 사실 단위'를 빠짐없이 추출하라.
출력: JSON 배열만 — [{unit_id, type, title, fact, source, chunk, question_hint}].
규칙:
1. 청크의 **모든 독립적 사실 문장을 각각 별도 단위**로 만든다 — 한 문장도 빠뜨리지 마라.
   (예: "A한다. B도 가능하다." → 단위 2개: fact="A한다.", fact="B도 가능하다.")
2. fact 는 반드시 청크 원문 문장을 **그대로 복사**(변형·요약·병합 금지 — 문자 대조로 검수됨).
3. unit_id 형식: <제품>-<문서약칭대문자>-<3자리>. 접미사(-A 등) 붙이지 말고 일련번호로 유일하게.
4. source 필드에는 입력 청크의 source 값을 그대로."""


CHUNK_BATCH = 25   # 호출당 청크 수 — 실코퍼스(수천 청크)는 한 번에 못 넣는다 (컨텍스트·타임아웃)
PAUSE_AFTER = 3    # 연속 실패 허용 — 이 이상이면 구독 한도/네트워크 의심, 중단하고 재개 대기


class CoveragePaused(Exception):
    """AI 호출 연속 실패 — 체크포인트 저장 후 일시 중단 (한도 회복 후 재개 가능)"""


def _ckpt_path(prod):
    return ROOT / "results" / f"_ckpt_coverage_{prod}.json"


def _ckpt_load(prod):
    p = _ckpt_path(prod)
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def _ckpt_save(prod, ck):
    p = _ckpt_path(prod)
    p.parent.mkdir(exist_ok=True)
    p.write_text(json.dumps(ck, ensure_ascii=False), encoding="utf-8")


def generate_units(prod, chunks, cfg, retry_ids=None, role="generator"):
    """청크를 배치로 나눠 호출 — 대형 코퍼스 대응. 배치 단위 실패는 기록 후 계속.
    [재개] 성공 배치마다 체크포인트 저장 — 한도 소진·중단 후 재실행하면 이어서 진행.
    연속 PAUSE_AFTER회 실패 = 한도/네트워크 의심 → 실패분 되감고 CoveragePaused."""
    target = [c for c in chunks if retry_ids is None
              or any(u.startswith(f"{prod}-{N(c['doc']).upper()[:6]}-{c['chunk_id']:03d}") for u in retry_ids)]
    total = (len(target) + CHUNK_BATCH - 1) // CHUNK_BATCH
    use_ckpt = retry_ids is None          # 반려 재생성 루프는 체크포인트 미사용 (소규모)
    ck = _ckpt_load(prod) if use_ckpt else {}
    rk = ck.get(role) or {}
    if rk.get("n_chunks") != len(target):  # 코퍼스가 달라졌으면 옛 체크포인트 무시 (오염 방지)
        rk = {}
    units = list(rk.get("units", []))
    fails = list(rk.get("fails", []))
    done = int(rk.get("done", 0))
    if use_ckpt and done:
        print(f"  [{role}] 체크포인트 재개 — {done}/{total} 배치부터 이어서")
        ledger_append("COVERAGE_MAP", "COVERAGE_RESUMED", f"script:{role}",
                      evidence={"재개 지점": f"{done}/{total}", "보존 단위": len(units)}, product=prod)
        # 재개 즉시 진행 파일 갱신 — 옛 '일시 중단' 잔상 제거 (화면이 바로 살아있게)
        _progress(prod, "커버리지 추출 (이어서)", role, done, total, fails, chunks=len(target))
    consec = []                            # 연속 실패 배치 번호들
    for bi in range(0, len(target), CHUNK_BATCH):
        part = target[bi:bi + CHUNK_BATCH]
        bno = bi // CHUNK_BATCH + 1
        if bno <= done:
            continue                       # 이미 처리한 배치 (재개)
        try:
            # 추출 전용 모델이 설정돼 있으면 그걸로 (예: K3 대신 v1-128k — 판정 아님, 규칙 C 무관)
            _c = cfg or load_config()
            call_role = f"{role}_extract" if _c.get("models", {}).get(f"{role}_extract") else role
            out = llm.chat(call_role, SYSTEM,
                           json.dumps({"product": prod, "chunks": part}, ensure_ascii=False), _c)
            got = llm.extract_json(out)
            units += got if isinstance(got, list) else [got]
            consec = []
        except Exception as e:
            # 배치 실패는 침묵하지 않는다 — 기록하고 나머지는 계속 (부분 커버리지 > 전체 실패)
            fails.append({"batch": f"{bno}/{total}", "err": str(e)[:120]})
            ledger_append("COVERAGE_MAP", "COVERAGE_BATCH_FAILED", f"script:{role}",
                          evidence={"batch": f"{bno}/{total}", "chunks": len(part),
                                    "err": str(e)[:200]}, product=prod)
            consec.append(bno)
        done = bno
        if use_ckpt:
            ck[role] = {"done": done, "units": units, "fails": fails, "n_chunks": len(target)}
            _ckpt_save(prod, ck)
        _progress(prod, "커버리지 추출", role, bno, total, fails, chunks=len(target))
        if len(consec) >= PAUSE_AFTER:
            # 한도/네트워크 의심 — 실패한 연속 배치는 되감아서 재개 때 다시 시도
            fails = [f for f in fails if int(f["batch"].split("/")[0]) not in consec]
            done = consec[0] - 1
            if use_ckpt:
                ck[role] = {"done": done, "units": units, "fails": fails, "n_chunks": len(target)}
                _ckpt_save(prod, ck)
            _progress(prod, f"일시 중단 — {role} 연속 {len(consec)}회 실패 (한도 의심, 재개 가능)",
                      role, done, total, fails, chunks=len(target))
            ledger_append("COVERAGE_MAP", "COVERAGE_PAUSED", f"script:{role}",
                          evidence={"중단 지점": f"{done}/{total}", "보존 단위": len(units),
                                    "사유": "연속 실패 — 구독 한도/네트워크 의심",
                                    "재개": "한도 회복 후 [이어서 재개] — 체크포인트에서 계속"},
                          product=prod)
            raise CoveragePaused(f"{role} {done}/{total} 배치에서 중단")
        if total > 4 and bno % 5 == 0:
            print(f"  [{role}] 커버리지 추출 {bno}/{total} 배치…")
    return units


def get_strategy(prod):
    """제품별 실행 전략 (대시보드 셀렉트박스에서 설정): ensemble | solo | cross_check | self_check"""
    from olib import load_state
    return load_state()["products"].get(prod, {}).get("strategy", "ensemble")


def _ensemble_roles(cfg, strategy="ensemble", prod=None):
    """추출기 편성 = 전략 × 연결 상태 × 사람의 투입 선택(ai_use 체크박스)"""
    if strategy != "ensemble":
        roles = ["generator"]
    elif llm.is_mock():
        roles = ["generator", "judge"]     # mock = 2키 상당
    else:
        import os
        from model_adapter import detect_mode
        have = detect_mode(cfg, os.environ)["have"]
        roles = [r for r in ("generator", "judge", "reviewer") if have.get(r)] or ["generator"]
    if prod:   # 사람이 체크박스로 뺀 AI는 제외 (generator는 최소 1인 보장)
        from olib import load_state
        use = load_state()["products"].get(prod, {}).get("ai_use")
        if use:
            roles = [r for r in roles if use.get(r, True)] or ["generator"]
    return roles


def _progress_path(prod):
    return ROOT / "results" / f"_progress_{prod}.json"


def _progress(prod, phase, role, batch, total, fails, chunks=None):
    """진행 상황 파일 — 대시보드가 폴링해서 '지금 어디까지 갔고 어디서 막혔나' 표시"""
    import datetime
    p = _progress_path(prod)
    p.parent.mkdir(exist_ok=True)
    cur = {}
    if p.exists():
        try:
            cur = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            cur = {}
    if cur.get("phase") != phase:      # 국면이 바뀌면 이전 국면 카운트는 지운다 (혼선 방지)
        cur["roles"] = {}
    cur["phase"] = phase
    cur["stage"] = "COVERAGE_MAP"
    if chunks:
        cur["청크"] = chunks           # "3,526청크를 읽으며 추출 중" 표시용
    if role and role != "-":
        cur.setdefault("roles", {})[role] = {"batch": batch, "total": total, "fails": fails}
    cur["ts"] = datetime.datetime.now().isoformat(timespec="seconds")
    p.write_text(json.dumps(cur, ensure_ascii=False), encoding="utf-8")


REVIEW_SYSTEM = """[TASK:COVERAGE_REVIEW] 너는 커버리지맵 검수자다 — 생성에는 관여하지 않았다.
입력: {chunks(코퍼스 원문 일부), units(추출된 커버 단위 목록)}.
검수 관점: ① 코퍼스에 있는데 단위로 안 뽑힌 문장(누락) ② 원문과 다른 fact(변형 의심).
출력: JSON만 — {"누락 의심": ["문장…", …], "변형 의심": ["unit_id", …], "총평": "1문장"}. 문제 없으면 빈 배열."""


def ai_review(prod, chunks, units, cfg, strategy):
    """[cross_check/self_check] 생성에 참여 안 한 세션이 결과를 검수 — 의견은 게이트 카드에 '참고'로."""
    role = "judge" if strategy == "cross_check" else "generator"   # self_check도 새 세션(호출 독립)
    payload = {"chunks": chunks[:40],
               "units": [{"unit_id": u["unit_id"], "fact": u["fact"]} for u in units]}
    try:
        out = llm.chat(role, REVIEW_SYSTEM, json.dumps(payload, ensure_ascii=False), cfg)
        op = llm.extract_json(out)
        return op if isinstance(op, dict) else {"총평": str(op)[:200]}
    except Exception as e:
        return {"검수 실패": str(e)[:150]}


MERGE_SYSTEM = """[TASK:COVERAGE_MERGE] 너는 커버리지맵 병합자다.
여러 AI가 같은 코퍼스에서 추출한 커버 단위들의 합집합을 받는다. 다음을 수행하라:
1. fact 가 동일하거나 한쪽이 다른쪽에 완전히 포함되는(부분집합) 단위는 하나로 통합 — 더 완전한 fact 를 남긴다.
2. fact 원문은 절대 변형·요약·병합 금지 (문자 대조 검수됨) — 남길 단위를 고르는 것만 허용.
3. title/question_hint 는 남긴 단위 것을 유지.
출력: JSON 배열만 — 입력과 같은 스키마."""


def ensemble_generate(prod, chunks, cfg, strategy=None):
    """전략별 추출 → 기계 dedup → (앙상블이면) 병합 → 검수는 호출측."""
    strategy = strategy or get_strategy(prod)
    roles = _ensemble_roles(cfg, strategy, prod)
    pool, contrib, fails = [], {}, []
    seen = set()
    paused = []   # 한도로 멈춘 추출자 — 다른 추출자는 계속 (난희 요청 2026-08-07: claude 멈추면 Kimi 먼저)
    for role in roles:
        try:
            units = generate_units(prod, chunks, cfg, role=role)
            ok, _rej = verify_units(units, chunks)     # 추출기별 1축 선별 (쓰레기 조기 제거)
            new = 0
            for u in ok:
                k = norm(u.get("fact", ""))
                if k not in seen:
                    seen.add(k)
                    pool.append(u)
                    new += 1
            contrib[role] = {"추출(검수통과)": len(ok), "신규 기여": new}
        except CoveragePaused as e:
            # 한 주자의 한도가 전체를 세우지 않는다 — 체크포인트에 보존됐으니 다음 주자 계속.
            # (재개 시: 완주한 주자는 체크포인트 즉시 재생, 멈춘 주자만 이어달림)
            paused.append(role)
            ledger_append("COVERAGE_MAP", "EXTRACTOR_PAUSED_CONTINUE", f"script:{role}",
                          evidence={"사유": str(e)[:120], "조치": "다른 추출자 계속 — 이 주자는 재개 때 이어서"},
                          product=prod)
            print(f"  ⏸ {role} 한도 의심 중단 — 다음 추출자로 계속 (재개 때 이어달림)")
        except Exception as e:
            fails.append({"role": role, "err": str(e)[:150]})
            ledger_append("COVERAGE_MAP", "ENSEMBLE_EXTRACTOR_FAILED", f"script:{role}",
                          evidence={"err": str(e)[:200]}, product=prod)
    if paused:
        # 전 주자 순회를 마친 뒤에만 일시 중단 보고 — auto_run이 대기 후 재개 (멈춘 주자만 남은 셈)
        raise CoveragePaused(f"일시 중단(한도 추정) — 추출자 {'+'.join(paused)} 대기, "
                             f"나머지 {len(roles)-len(paused)}명은 완주(체크포인트 보존)")
    MERGE_LIMIT = 200   # 이 이상이면 LLM 병합 생략 — 기계 dedup(정확 일치)만으로 충분·안전
    if len(roles) > 1 and pool and len(pool) <= MERGE_LIMIT:
        _progress(prod, "병합(대표 AI)", "generator", 0, 1, [])
        merged_out = llm.chat("generator", MERGE_SYSTEM,
                              json.dumps({"units": pool}, ensure_ascii=False), cfg)
        merged = llm.extract_json(merged_out)
        # 병합자가 단위를 과도 삭제/변형했는지 안전판: 병합 결과가 풀의 50% 미만이면 풀 유지
        if not isinstance(merged, list) or len(merged) < max(1, len(pool) // 2):
            merged = pool
    else:
        merged = pool
        if len(pool) > MERGE_LIMIT:
            ledger_append("COVERAGE_MAP", "MERGE_SKIPPED_SCALE", "script:ensemble",
                          evidence={"pool": len(pool), "limit": MERGE_LIMIT,
                                    "처리": "기계 dedup만 적용 — 부분집합 통합은 사람확인 큐에서"},
                          product=prod)
    # ID 재부여 — 모델 간 ID 충돌 제거 (문서약칭 + 일련번호)
    from collections import defaultdict
    seq = defaultdict(int)
    for u in merged:
        doc = re.sub(r"[^A-Z0-9]", "", N(u.get("source", u.get("unit_id", "DOC"))).upper())[:6] or "DOC"
        seq[doc] += 1
        u["unit_id"] = f"{prod}-{doc}-{seq[doc]:03d}"
    return merged, contrib, fails


def verify_units(units, chunks):
    """스크립트 검수(규칙 A): 사실 실재(1축) + ID 중복. → (합격, 반려[(unit, 사유)])"""
    pool = {}
    for c in chunks:
        pool.setdefault(N(c["doc"]), "")
        pool[N(c["doc"])] += " " + norm(c["text"])
    all_text = " ".join(pool.values())
    seen, ok, rej = set(), [], []
    for u in units:
        uid = N(u.get("unit_id", ""))
        if not uid or uid in seen:
            rej.append((u, "unit_id 부재/중복"))
            continue
        seen.add(uid)
        f = norm(u.get("fact", ""))
        if len(f) < 6:
            rej.append((u, "사실 필드 과소(len<6)"))
            continue
        if f not in all_text:
            rej.append((u, "1축 불일치 — 사실이 코퍼스 원문에 없음"))
            continue
        ok.append(u)
    return ok, rej


# xlsx가 거부하는 제어문자(NULL 등) — PDF 추출 코퍼스에 섞여 옴. 저장용으로만 제거
# (1축 검수는 이미 통과 — norm()이 공백만 보므로 이 정제가 검수를 무효화하지 않음)
_ILLEGAL_XLSX = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _xclean(v):
    return _ILLEGAL_XLSX.sub("", v) if isinstance(v, str) else v


def write_map(prod, units, version="v1_0"):
    out = DATA / prod / "03_coverage_map" / f"{prod}_커버리지맵_코퍼스판_{version}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B1_커버리지맵"
    ws.append(["#", "Unit ID", "type", "title", "이 단위가 말하는 사실",
               "답변 가능한 질문", "source/resource", "citation_id", "상태"])
    for i, u in enumerate(units, 1):
        ws.append([_xclean(x) for x in
                   [i, u["unit_id"], u.get("type", "Doc"), u.get("title", ""),
                    u["fact"], u.get("question_hint", ""), u.get("source", ""),
                    u["unit_id"], "생성-검수통과"]])
    wb.save(out)
    return out


def _gkey(s):
    """GAP 대조 키 — 공백·대소문자·꼬리 슬래시 무시 (('/kr'과 '/kr/'은 같은 문서)"""
    return norm(s).rstrip("/")


def gap_audit(prod, chunks, units):
    """③′ 인입 누락 대조 — source(URL/파일명) 기준: 커버 단위가 0개인 코퍼스 문서.
    [수리 2026-07-22] 종전에는 doc 제목 약칭 vs unit_id 를 대조해 281건 오탐 (369→실측 88)."""
    covered = {_gkey(u.get("source", "")) for u in units}
    gaps, seen = [], set()
    for c in chunks:
        k = _gkey(c["source"])
        if k and k not in covered and k not in seen:
            seen.add(k)
            gaps.append(N(c["source"]))
    return sorted(gaps)


def classify_gaps(gaps, chunks):
    """누락을 성격별로 — 사람이 판단할 것(내용 페이지)과 무해 추정(목록·넘김)을 분리"""
    from collections import Counter
    by_src = Counter()
    for c in chunks:
        by_src[N(c["source"])] += 1
    nav = [g for g in gaps if re.search(r"/page/\d+/?$|/feed/?$", g)]
    SECTION_ROOTS = ("blog", "notices", "use-cases", "success-stories", "products",
                     "pricing", "support", "library", "videos", "trial", "info", "contact")
    listing = [g for g in gaps if g not in nav
               and (g.rstrip("/").count("/") <= 4 or g.rstrip("/").endswith(SECTION_ROOTS))]
    content = [g for g in gaps if g not in nav and g not in listing]

    def section_of(u):
        m = re.search(r"remotecall\.com/kr/([a-z-]+)", u)
        return m.group(1) if m else ("파일" if not u.startswith("http") else "기타")
    groups = {}
    for g in content:
        groups.setdefault(section_of(g), []).append(g)
    return {"nav": nav, "listing": listing, "content_groups": groups, "chunks_of": by_src}


def gap_flags(gaps, chunks):
    """게이트 카드용 플래그 — 무해 추정은 묶음 1건, 내용 페이지는 섹션별 묶음 (ack 부담 축소)"""
    cl = classify_gaps(gaps, chunks)
    flags = []
    harmless = cl["nav"] + cl["listing"]
    if harmless:
        flags.append({"type": "GAP-목록·넘김 페이지 (내용 없음 추정)", "id": f"{len(harmless)}건",
                      "candidates": [], "ack_required": True,
                      "note": "페이지네이션·피드·카테고리 목록 — 예: " + " · ".join(harmless[:3])})
    for sec, urls in sorted(cl["content_groups"].items(), key=lambda x: -len(x[1])):
        total_chunks = sum(cl["chunks_of"][u] for u in urls)
        flags.append({"type": f"GAP-내용 페이지 누락 · {sec}", "id": f"{len(urls)}건 ({total_chunks}청크)",
                      "candidates": urls, "ack_required": True})
    return flags


def run(prod, cfg):
    chunks = load_corpus(prod)
    if not chunks:
        return "WAITING_INPUT", {"코퍼스 청크": 0}
    # 전략별 추출 (셀렉트박스 설정) → 최종 1축 재검수 (병합/생성 변형도 걸러짐)
    strategy = get_strategy(prod)
    try:
        merged, contrib, fails = ensemble_generate(prod, chunks, cfg, strategy)
    except CoveragePaused as e:
        return "HALTED", {"halt": f"AI 호출 연속 실패로 일시 중단 ({e}) — 구독 한도/네트워크 의심. "
                                  "지금까지 진행분은 체크포인트에 저장됨: 한도 회복 후 [▶ 이어서 재개]를 누르면 이어서 진행"}
    ok, rej = verify_units(merged, chunks)
    if not ok:
        _progress(prod, "막힘 — 생성 단위 전건 검수 불합격", "-", 0, 0, fails)
        return "HALTED", {"halt": "생성 단위 전건 검수 불합격 — 카운트 미실측 방지",
                          "기여도": contrib, "추출기 실패": fails}
    path = write_map(prod, ok)
    gaps = gap_audit(prod, chunks, ok)
    ev = {"실행 전략": strategy,
          "커버 단위(검수 통과)": len(ok),
          "기여도": contrib,
          "재검수 탈락": len(rej),
          "1축 문자 대조": "불일치 0 (통과분)", "GAP_AUDIT 누락 문서": len(gaps),
          "산출": N(path.name)}
    if strategy in ("cross_check", "self_check"):
        _progress(prod, "AI 검수", "judge" if strategy == "cross_check" else "generator", 0, 1, [])
        op = ai_review(prod, chunks, ok, cfg, strategy)
        ev["AI 검수 의견 (참고 — 판정은 사람)"] = op
    if fails:
        ev["추출기 실패"] = fails
    _progress(prod, "완료 — 사람확인 대기", "-", 0, 0, [])
    ledger_append("COVERAGE_MAP", "MAP_GENERATED", "script:gen_coverage", evidence=ev, product=prod)
    _ckpt_path(prod).unlink(missing_ok=True)   # 완주 — 체크포인트 정리
    flags = (gap_flags(gaps, chunks)
             + [{"type": "생성 반려 잔존", "id": N(u.get("unit_id", "?")), "candidates": [r]}
                for u, r in rej])
    issue_gate_card(prod, "COVERAGE_MAP", f"COVMAP_{prod}",
                    what_stopped=f"커버리지맵 생성·검수 완료 — 확정은 사람 (사람확인 큐)",
                    evidence=ev, flags=flags,
                    recommendation="GAP 누락·반려 잔존 항목 ack 후 승인 시 맵 확정 → ④ 진입")
    return "WAITING_HUMAN", ev


def reaudit(prod):
    """GAP 재점검 — 기존 커버리지맵으로 소급 재산출 (도구 수리 후, 재추출 없음).
    감독관 규칙 §3 노화 감시: 도구를 고치면 과거 결과를 같은 자로 다시 잰다."""
    from olib import close_gate
    chunks = load_corpus(prod)
    maps = sorted((DATA / prod / "03_coverage_map").glob(f"{prod}_커버리지맵_*.xlsx"))
    if not maps:
        sys.exit("커버리지맵 없음")
    wb = openpyxl.load_workbook(maps[-1])
    ws = wb.active
    units = [{"unit_id": r[1], "fact": r[4], "source": r[6]}
             for r in list(ws.iter_rows(values_only=True))[1:]]
    gaps = gap_audit(prod, chunks, units)
    flags = gap_flags(gaps, chunks)
    ledger_append("COVERAGE_MAP", "GAP_AUDIT_CORRECTED", "script:gen_coverage",
                  evidence={"구(오탐 포함)": 369, "신(실측)": len(gaps),
                            "원인": "doc 제목 약칭 vs unit_id 대조 — source(URL) 기준으로 수리",
                            "맵": N(maps[-1].name), "커버 단위": len(units)}, product=prod)
    close_gate(prod, f"COVMAP_{prod}")   # 옛 플래그(오탐 369건) 제거 후 재발행
    ev = {"실행 전략": get_strategy(prod), "커버 단위(검수 통과)": len(units),
          "GAP 누락 (재점검·source 기준)": len(gaps),
          "산출": N(maps[-1].name), "재점검": "GAP 도구 수리 후 소급 재산출 (재추출 없음)"}
    issue_gate_card(prod, "COVERAGE_MAP", f"COVMAP_{prod}",
                    what_stopped="커버리지맵 확정 대기 — GAP 재점검 완료 (도구 수리 반영)",
                    evidence=ev, flags=flags,
                    recommendation="무해 추정(목록·넘김)은 묶음 ack, 내용 페이지 누락은 섹션별 검토 후 승인 → ④ 진입")
    print(f"재점검 완료 — 누락 {len(gaps)}건 (플래그 {len(flags)}묶음), 카드 재발행")


def gapfill(prod):
    """GAP 보완 추출 — 누락 문서의 청크만 앙상블 재추출 → 기존 맵에 합쳐 v+1 저장 → 재점검.
    전체 재추출 아님: 누락분(수백 청크)만. 결과는 새 버전 맵 파일 (기존 맵 보존)."""
    from olib import load_config
    cfg = load_config()
    chunks = load_corpus(prod)
    maps = sorted((DATA / prod / "03_coverage_map").glob(f"{prod}_커버리지맵_*.xlsx"))
    wb = openpyxl.load_workbook(maps[-1])
    ws = wb.active
    old_units = [{"unit_id": r[1], "type": r[2], "title": r[3], "fact": r[4],
                  "question_hint": r[5], "source": r[6]}
                 for r in list(ws.iter_rows(values_only=True))[1:]]
    gaps = set(_gkey(g) for g in gap_audit(prod, chunks, old_units))
    target = [c for c in chunks if _gkey(c["source"]) in gaps]
    print(f"보완 대상: 문서 {len(gaps)} · 청크 {len(target)}")
    if not target:
        print("보완할 청크 없음")
        return
    # 앙상블 추출 (누락 청크만) — 역할별 독립, 체크포인트는 이 대상 크기 기준
    roles = _ensemble_roles(cfg, get_strategy(prod), prod)
    pool, seen = [], {norm(u.get("fact", "")) for u in old_units}
    fails = []
    for role in roles:
        try:
            units = generate_units(prod, target, cfg, role=role)
            ok, _rej = verify_units(units, target)
            for u in ok:
                k = norm(u.get("fact", ""))
                if k not in seen:
                    seen.add(k)
                    pool.append(u)
        except CoveragePaused:
            raise
        except Exception as e:
            fails.append({"role": role, "err": str(e)[:150]})
    # ID 재부여 — 기존 맵과 충돌 없이 이어서
    from collections import defaultdict
    seq = defaultdict(int)
    for u in old_units:
        m = re.match(rf"{re.escape(prod)}-([A-Z0-9]+)-(\d+)", N(u["unit_id"] or ""))
        if m:
            seq[m.group(1)] = max(seq[m.group(1)], int(m.group(2)))
    for u in pool:
        doc = re.sub(r"[^A-Z0-9]", "", N(u.get("source", "DOC")).upper())[:6] or "DOC"
        seq[doc] += 1
        u["unit_id"] = f"{prod}-{doc}-{seq[doc]:03d}"
    merged = old_units + pool
    # 다음 버전 번호
    vm = re.search(r"_v(\d+)_(\d+)\.xlsx$", maps[-1].name)
    nxt = f"v{vm.group(1)}_{int(vm.group(2)) + 1}" if vm else "v1_1"
    out = write_map(prod, merged, version=nxt)
    _ckpt_path(prod).unlink(missing_ok=True)
    ledger_append("COVERAGE_MAP", "GAPFILL_EXTRACTED", "script:gen_coverage",
                  evidence={"대상 문서": len(gaps), "대상 청크": len(target),
                            "신규 단위": len(pool), "실패": fails,
                            "산출": N(out.name), "기존 맵": N(maps[-1].name)}, product=prod)
    print(f"보완 완료 — 신규 단위 {len(pool)} → {out.name}. 재점검 실행…")
    reaudit(prod)


if __name__ == "__main__":
    import argparse
    from olib import load_config
    ap = argparse.ArgumentParser()
    ap.add_argument("--product", required=True)
    ap.add_argument("--reaudit", action="store_true", help="기존 맵으로 GAP 재점검 + 카드 재발행")
    ap.add_argument("--gapfill", action="store_true", help="누락 문서만 보완 추출 → 맵 v+1 + 재점검")
    a = ap.parse_args()
    if a.gapfill:
        gapfill(a.product)
    elif a.reaudit:
        reaudit(a.product)
    else:
        print(run(a.product, load_config()))
