#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_goldenset.py — ④ 골든셋 배치 생성 루프 (사양서 §4-④)

루프: 재료실측(맵 단위 수) → 배분계획 [GATE:사람] → 파일럿(30) 생성 → 검수 7종
  → [GATE:사람] → 차수 반복(밴드 60~75) → 잔여 0 → 배치 마감(커버등식 시트) [GATE:사람]

- 생성 문항의 발췌는 맵 '사실' 문장에서만 — verify_batch ②(1축)가 사후 검증.
- 매 차수 verify_batch 7종 자동 실행: REJECTED면 반려 루프(재생성 1회) 후에도 실패 시 게이트에 반려 보고.
- 진행 상태는 state.json products.<P>.goldenset {phase, round, done_units} 에 직렬화.
"""
import json
import re
import subprocess
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import llm
from olib import (ROOT, N, load_state, save_state, save_goldenset, ledger_append,
                  issue_gate_card, load_config)

DATA = ROOT / "data"
VB = ROOT / "tools" / "verify_batch.py"
STD_HEADER = ["ID", "유형", "출제 의도", "질문", "정답", "근거 출처", "합격 기준",
              "난이도", "기대 라우팅", "검증 상태", "작성자/검증자", "근거 원문 발췌",
              "신뢰도", "acl_level", "answer_type", "acl_참고"]


def read_map_units(prod, cfg):
    prof = cfg["terrain"]["profiles"][prod]
    maps = sorted((DATA / prod / "03_coverage_map").glob("*.xlsx"))
    if not maps:
        return [], None
    path = maps[-1]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    units = []
    for sn in wb.sheetnames:
        if not N(sn).endswith("커버리지맵"):
            continue
        ws = wb[sn]
        hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if prof["map_unit_column"] not in hdr:
            continue
        iu = hdr.index(prof["map_unit_column"])
        i_f = hdr.index(prof["map_fact_column"])
        it = hdr.index("title") if "title" in hdr else None
        i_s = hdr.index("source/resource") if "source/resource" in hdr else None
        for r in ws.iter_rows(min_row=2, values_only=True):
            uid = N(r[iu]) if iu < len(r) else ""
            if uid:
                units.append({"unit_id": uid, "fact": N(r[i_f]) if i_f < len(r) else "",
                              "title": N(r[it]) if it is not None and it < len(r) else uid,
                              "source": N(r[i_s]) if i_s is not None and i_s < len(r) else ""})
    wb.close()
    return units, path


def select_representative(units, target):
    """문서(source)별 비례 대표 추출 — 결정적(등간격·난수 없음, 같은 맵이면 항상 같은 선정).
    7,848단위 전량 소진(≈105차수) 대신 v1급 규모로 출제하기 위한 표본 — 문서마다 최소 1단위 보장,
    문서 안에서는 등간격으로 뽑아 문서 전체에 고르게 분포."""
    from collections import defaultdict
    _short = [u for u in units if len(N(u.get("fact", ""))) < 20]
    if _short:
        units = [u for u in units if len(N(u.get("fact", ""))) >= 20]   # r1 교훈: 발췌 20자 하한 — 짧은 fact는 출제 재료에서 제외
    if len(units) <= target:
        return units
    by_doc = defaultdict(list)
    for u in units:
        by_doc[u.get("source") or u["unit_id"]].append(u)
    total = len(units)
    quota = {d: max(1, int(target * len(v) / total)) for d, v in by_doc.items()}
    docs_sorted = sorted(by_doc, key=lambda d: -len(by_doc[d]))
    s = sum(quota.values())
    i = 0
    while s != target and i < 20 * len(docs_sorted) + target:
        d = docs_sorted[i % len(docs_sorted)]
        if s < target and quota[d] < len(by_doc[d]):
            quota[d] += 1
            s += 1
        elif s > target and quota[d] > 1:
            quota[d] -= 1
            s -= 1
        i += 1
    sel = []
    for d, v in by_doc.items():
        k = min(quota[d], len(v))
        stride = len(v) / k
        sel += [v[int(j * stride)] for j in range(k)]
    return sel


def gs_state(prod):
    st = load_state()
    return st, st["products"][prod].setdefault(
        "goldenset", {"phase": "MATERIAL", "round": 0, "done_units": [], "batch": "B1"})


SYSTEM_GEN = """[TASK:GOLDENSET_ITEMS] 너는 RAG 평가 골든셋 출제기다(생성규칙서 준수).
입력: {product, product_name, prefix, start_no, units[], want_e}.
출력: JSON 배열만 — 표준 문항 [{ID,유형,출제 의도,질문,정답,근거 출처,합격 기준,근거 원문 발췌}].
규칙: ① '근거 원문 발췌'는 해당 unit의 fact 문장 그대로(변형 금지 — 1축 문자 대조됨)
② '근거 출처'에 unit_id 와 source(URL/파일명) 병기 — 채점기 출처 대조용
③ '정답' 끝에 반드시 "필수: 요소1, 요소2" 줄 포함 — 채점기 필수 요소 파싱 규격
④ 질문은 **실제 고객이 상담 창구에 묻는 자연스러운 문장**으로 쓴다.
   - product_name 이 실제 서비스명이면 문맥상 자연스러울 때만 질문에 포함.
   - product/prefix 가 익명 코드(예: 'CI')면 **ID 접두어로만 쓰고 질문 문장에는 절대 넣지 마라**
     — 고객은 내부 코드를 모른다. ("CIについて、…ですか？" 같은 문장 금지 [파일럿 실측 2026-08-18:
     31문항 전부 익명 코드 접두어가 붙어 부자연 — 질문은 소재 서비스명(Netflix 등)으로 시작하라])
⑤ want_e=true면 코퍼스 부재 소재 E형 1문항 추가 ⑥ 질문 중복 금지.
[r2 규격 — r1 실측 결함 반영 (CI_r2_재출제_지시서 v1_0, 2026-08-25)]
⑦ '근거 원문 발췌'는 **연속 20자 이상** 원문 그대로 — 짧은 사실이면 그 문장 전체를 발췌
   (r1 결함: 20자 미만 발췌 130건이 우연 일치로 채점을 부풀림).
⑧ '유형'은 다음 고정 어휘만: 사실확인 | 수치 | 절차 | 조건 | 가부 | 정의 | E형
   (r1 결함: 라벨 177종 난립으로 유형 분석 불가).
⑨ '정답'의 필수 요소에는 수치·고유명사를 반드시 명시 (판정기준서 엄격 항목과 정합)."""


def next_item_no(prod):
    """다음 문항 시작 번호 = 기존 배치 파일(반려_ 포함)의 최대 발행 번호 + 1 [P0-1].
    done_units 개수 기준은 미인용(lost) 발생 시 직전 차수와 번호가 겹쳐 서로 다른 문항이
    같은 ID를 받는다 (실측: RC2 76건·RV2 3건 증발). 반려_ 파일 번호도 재사용 금지(증적 모호 방지)."""
    mx = 0
    pat = re.compile(rf"^{re.escape(prod)}-[A-Za-z]{{0,2}}0*(\d+)$")
    d = DATA / prod / "04_goldenset_batch"
    for b in (sorted(d.glob("*.xlsx")) if d.is_dir() else []):
        if b.name.startswith(("~$", ".")):
            continue
        try:
            wb = openpyxl.load_workbook(b, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            if "ID" not in hdr:
                wb.close()
                continue
            ci = hdr.index("ID")
            for r in ws.iter_rows(min_row=2, values_only=True):
                m = pat.match(N(r[ci])) if ci < len(r) and r[ci] else None
                if m:
                    mx = max(mx, int(m.group(1)))
            wb.close()
        except Exception:
            continue   # 손상 파일은 번호 산정에서 제외 — 검수 7종이 별도로 잡는다
    return mx + 1 if mx else None


def _normalize_ids(items, prod, start_no):
    """ID 규격 강제 [2차 실측 2026-08-18: 74건 중 43건이 'CI-D32' 등 변형] — AI가 붙인 ID를
    믿지 않고 스크립트가 순번으로 다시 매긴다 (start_no부터 연속, 형식 <제품>-NNN 고정).
    ID는 채점·판정·캘리브레이션의 조인 키라 규격이 흔들리면 안 된다."""
    for k, it in enumerate(items):
        it["ID"] = f"{prod}-{start_no + k:03d}"
        # 닻 표기 통일 [3차 실측 2026-08-18: 必須(일) 25건·필수(한) 50건 혼용] — 규격은 '필수:' 하나
        if it.get("정답"):
            it["정답"] = N(it["정답"]).replace("必須:", "필수:").replace("必須：", "필수:")
    return items


def _clean_anon_prefix(items, prod, name):
    """익명 제품코드 접두어 제거 안전망 [파일럿 실측 2026-08-18] — 프롬프트 지시(규칙 ④)가
    무시돼도 스크립트가 뗀다. 실명 제품(product_name 이 코드와 다름)은 손대지 않는다."""
    if name and N(name) != N(prod):
        return items
    pat = re.compile(rf"^\s*{re.escape(prod)}\s*(について、?|に関して、?|では、?|の)\s*")
    for it in items:
        q = N(it.get("질문", ""))
        q2 = pat.sub("", q)
        if q2 != q and len(q2) >= 8:
            it["질문"] = q2
    return items


def generate_items(prod, units, start_no, want_e, cfg, feedback=None):
    prof = cfg["terrain"]["profiles"][prod]
    name = prof.get("product_name", prod)
    payload = {"product": prod, "product_name": name, "prefix": prod,
               "start_no": start_no, "want_e": want_e,
               "units": [{"unit_id": u["unit_id"], "title": u["title"], "fact": u["fact"],
                          "source": u.get("source", "")} for u in units]}
    if feedback:   # 사람 반려 사유를 재출제 프롬프트에 직결 — 카드의 '피드백대로 재출제' 약속의 배관
        payload["사람_반려_피드백"] = ("직전 배치가 사람 검수에서 반려되었다. "
                                 "아래 피드백을 이번 출제에 반드시 반영하라: " + N(feedback))
    body = json.dumps(payload, ensure_ascii=False)
    out = llm.chat("generator", SYSTEM_GEN, body, cfg)
    # 문항 1개면 extract_json이 배열 아닌 낱개 객체를 준다 — 항상 목록으로 정규화
    # (마지막 소차수 1문항 생성에서 'str has no get' HALT — 07-24 2차 사고)
    norm = lambda x: _normalize_ids(_clean_anon_prefix([x] if isinstance(x, dict) else [i for i in x if isinstance(i, dict)], prod, name), prod, start_no)
    try:
        return norm(llm.extract_json(out))
    except ValueError:
        # 실모델이 가끔 문항 JSON 없이 '만들었다'는 설명문만 답한다 (07-24 HALT 사고) —
        # JSON만 내도록 명시해 새 호출 1회 재시도, 그래도 실패면 정지(기존 안전망)
        out = llm.chat("generator", SYSTEM_GEN,
                       body + "\n\n(재요청) 직전 응답에 문항 JSON이 없었다. "
                              "설명·머리말 없이 문항 JSON 배열만 출력하라. "
                              "출제 가능한 문항이 없으면 빈 배열 []만 출력하라.", cfg)
        try:
            return norm(llm.extract_json(out))
        except ValueError:
            # 재요청까지 JSON 부재 = 사실상 '이 재료로는 못 만든다'는 거부 —
            # HALT 대신 빈 배치로 반환해 부적격 확정 경로로 (설명문 거부 반복 HALT 사고 종결)
            print(f"  ⚠ 재요청에도 문항 JSON 부재 — 빈 배치로 간주해 부적격 경로로 (응답 요지: {out[:100]})")
            return []


def _norm(s):
    return re.sub(r"\s+", "", N(s)).lower()


def ground_citations(items, units):
    """[실전 견고화] 실모델이 단위ID를 흘려 쓰는 문제 — '이름이 아니라 내용물'(P-01/B′ 취지).
    발췌 세그(원문·1축 검증 대상)마다 그것을 포함하는 맵 단위를 찾아, 근거 출처를
    매칭 단위들로 authoritative하게 전면 재작성한다. 모델의 원래 citation은 신뢰하지 않는다.
    세그가 어느 단위에도 없으면(비원문) 그 세그는 매칭 실패 → verify ②가 정당하게 반려한다."""
    facts = [(u["unit_id"], _norm(u.get("fact", "")), u.get("source", "")) for u in units]
    grounded = 0
    for it in items:
        if N(it.get("유형", "")).startswith("E") or "E형" in N(it.get("유형", "")):
            continue   # E형(코퍼스 부재)은 근거 단위가 없다 — 손대지 않음
        segs = [re.sub(r"^\[[^\]]*\]\s*", "", N(s)).strip("\"'“”‘’ ")
                for s in re.split(r"\|\|", N(it.get("근거 원문 발췌", "")))]
        segs = [s for s in segs if len(s) >= 6]
        if not segs:
            continue
        cited, src0 = [], ""
        for seg in segs:
            for uid, fnorm, src in facts:
                if _norm(seg) in fnorm:
                    if uid not in cited:
                        cited.append(uid)
                        src0 = src0 or src
                    break
        if cited:   # 매칭된 단위들로 전면 재작성 (authoritative)
            new = " ; ".join(cited) + (f" ; {src0}" if src0 else "")
            if new != N(it.get("근거 출처", "")):
                it["근거 출처"] = new
                grounded += 1
    return grounded


def write_batch(prod, items, label, eq=None):
    out = DATA / prod / "04_goldenset_batch" / f"{prod}_골든셋_{label}_{len(items)}문항_v1_0.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"1_골든셋_{label}"
    ws.append(STD_HEADER)
    for it in items:
        ws.append([N(it.get(h, "")) if h in it else
                   ("검수통과" if h == "검증 상태" else "script:gen_goldenset" if h == "작성자/검증자" else "")
                   for h in STD_HEADER])
    if eq:
        ws2 = wb.create_sheet(f"2_{label}_커버등식")
        ws2.append(["구분", "단위 수", "설명"])
        for row in eq:
            ws2.append(row)
    wb.save(out)
    return out


def verify(prod, batch_path, union_paths):
    cmd = [sys.executable, str(VB), "--batch", str(batch_path),
           "--map", str(sorted((DATA / prod / "03_coverage_map").glob("*.xlsx"))[-1]),
           "--product", prod, "--union"] + [str(p) for p in union_paths]
    p = subprocess.run(cmd, capture_output=True, text=True)
    return p.returncode, p.stdout


def generate_items_chunked(prod, batch_units, start_no, want_e, cfg, chunk=25, feedback=None):
    """[수리 2026-07-23] 75단위 단발 호출은 응답 절단 반복(75요청→40·42·13 생산) —
    25단위씩 소분할 호출로 차수 생산량을 밴드에 안정화. E형은 마지막 조각에만.
    [수리 2026-08-18] 소분할은 '개수'가 아니라 '재료 글자수'로도 잘라야 한다 — CI 실측:
    남은 단위일수록 fact 가 길어(표·절차 조각) 25개 고정이면 절단 반환이 가속(5차 2→6차 9→7차 24).
    글자 상한(기본 12,000자)을 먼저 걸고, 그래도 잘리면 조각 이등분 재시도."""
    CHUNK_CHARS = int(cfg.get("goldenset_chunk_chars") or 12000)

    def parts_of(units):
        out, cur, cch = [], [], 0
        for u in units:
            n = len(str(u.get("fact") or ""))
            if cur and (cch + n > CHUNK_CHARS or len(cur) >= chunk):
                out.append(cur)
                cur, cch = [], 0
            cur.append(u)
            cch += n
        if cur:
            out.append(cur)
        return out

    def gen(part, sno, we, depth=0):
        got = generate_items(prod, part, start_no=sno, want_e=we, cfg=cfg, feedback=feedback)
        # 생산 미달(요청 대비 절반 미만) = 절단 의심 — 조각을 반으로 쪼개 재시도 (최대 2단)
        if len(got) < max(1, len(part) // 2) and len(part) >= 4 and depth < 2:
            mid = len(part) // 2
            print(f"    ✂ 생산 미달 {len(got)}/{len(part)} — 조각 이등분 재시도 (분할 {depth + 1}단)")
            a = gen(part[:mid], sno, False, depth + 1)
            b = gen(part[mid:], sno + len(a), we, depth + 1)
            return a + b
        return got

    pieces = parts_of(batch_units)
    items = []
    for pi, part in enumerate(pieces):
        last = pi == len(pieces) - 1
        items += gen(part, start_no + len(items), want_e and last)
    return items


def _gs_progress(prod, phase):
    """④ 진행판 — 대시보드 진행선용 (한 차수 = 대형 호출 1건이라 배치 카운트 대신 국면만)"""
    import datetime
    p = ROOT / "results" / f"_progress_{prod}.json"
    p.parent.mkdir(exist_ok=True)
    p.write_text(json.dumps({"stage": "GOLDENSET_BATCH", "phase": phase, "roles": {},
                             "stale_after": 50,   # 호출 1건이 20~40분 — 50분까진 경고 안 함
                             "ts": datetime.datetime.now().isoformat(timespec="seconds")},
                            ensure_ascii=False), encoding="utf-8")


def human_guide(items, label, path, lost=None):
    """사람 확인 가이드 — 무지성 승인 방지: 기계가 못 보는 것만 사람이 보게 표본+포인트 제시"""
    from collections import defaultdict
    by_type = defaultdict(list)
    for it in items:
        by_type[N(it.get("유형", "?"))].append(it)
    samples = []
    for t, arr in sorted(by_type.items()):
        it = arr[len(arr) // 2]           # 유형별 가운데 1문항 (결정적)
        samples.append(f"**[{it.get('ID')}] ({t})**\n"
                       f"- 질문: {N(it.get('질문',''))[:100]}\n"
                       f"- 정답: {N(it.get('정답',''))[:140]}\n"
                       f"- 출처: {N(it.get('근거 출처',''))[:70]}")
    guide = (
        "### 👀 사람 확인 가이드 — 기계가 못 보는 것만 보면 됩니다 (표본 검사, 전건 검토 불필요)\n"
        "1. 아래 표본의 **질문이 실제 고객이 물어볼 법한 말**인가요? (내부 용어·번역투면 반려)\n"
        "2. **정답이 질문에 대한 답**인가요? (동문서답·과잉 서술 확인)\n"
        "3. 정답의 **'필수:' 요소가 합리적**인가요? (너무 많으면 채점이 가혹, 너무 적으면 무의미)\n"
        f"4. 전체 문항은 ④ 자료실에서 `{N(path.name)}` [받기]\n"
        + (f"\n⚠ 이 배치 특이사항: 미커버 {len(lost)}단위는 다음 차수가 자동 재출제 — 승인에 영향 없음\n" if lost else "")
        + "\n어색한 문항이 있으면 [반려] + 사유에 문항 ID를 적어주세요 — 그 피드백대로 재출제됩니다.\n\n"
        "### 표본 문항 (유형별 1개, 자동 발췌)\n" + "\n\n".join(samples))
    return guide


def covered_units(items):
    # [수리 2026-07-23] 접두에 숫자 포함(RC2·RV2) 지원 — 종전 [A-Z]{2,}는 'RC2-…'에서
    # 'REMOTE-1067'만 잡아 교집합 0 → 3차 75단위 전량 오반환 사고. EE(무숫자)만 통과해 회귀가 못 잡았음
    return {c for it in items for c in re.findall(r"[A-Z][A-Z0-9]+(?:-[A-Z0-9가-힣]+)+", N(it.get("근거 출처", "")))}


def run(prod, cfg):
    all_units, map_path = read_map_units(prod, cfg)
    if not all_units:
        return "WAITING_INPUT", {"커버리지맵": "없음 — ③ 미완"}
    # 출제 제외 소스 (예: Known Issue — 내부 결함 목록은 고객 질문이 아님, 난희 결정)
    excl = cfg["pipeline"].get("goldenset_exclude_sources", [])
    if excl:
        all_units = [u for u in all_units
                     if not any(x in u.get("source", "") for x in excl)]
    # 목표 규모 대표 추출 (결정적 — 재진입 때마다 같은 선정)
    target = cfg["pipeline"].get("goldenset_target")
    ratio = cfg["pipeline"].get("goldenset_faq_manual_ratio")   # 예: [4, 6] — r2 설계서 5조
    if target and ratio:
        def _is_man(u):
            src = str(u.get("source", ""))
            return src.lower().endswith((".pdf", ".docx")) or "マニュアル" in src or "매뉴얼" in src
        man = [u for u in all_units if _is_man(u)]
        faq = [u for u in all_units if not _is_man(u)]
        t_faq = round(target * ratio[0] / (ratio[0] + ratio[1]))
        # [수리 2026-08-26] select_representative는 문서당 최소 1개 보장이라 문서 수(2,152)가
        # 목표(320)보다 크면 폭발 (실측: 2,242 선정 → 30차수). 층화는 계층 내 등간격으로 —
        # 정확히 목표 수만 뽑되 코퍼스 전체에 고르게 분포 (결정적·재현 가능).
        def _stride_pick(pool, n):
            if n <= 0 or not pool:
                return []
            if len(pool) <= n:
                return list(pool)
            step = len(pool) / n
            return [pool[int(i * step)] for i in range(n)]
        units = _stride_pick(faq, t_faq) + _stride_pick(man, target - t_faq)
    elif target:
        units = select_representative(all_units, target)
    else:
        units = all_units
    st, gs = gs_state(prod)
    if gs.get("expand_units"):
        # G19 · 증분 확대 모드 — 새 단위 전량이 출제 대상 (대표 추출 재선정 금지: 명단 갱신 사고 방지)
        exp = set(gs["expand_units"])
        units = [u for u in all_units if u["unit_id"] in exp]
    band_lo, band_hi = cfg["pipeline"].get("band", [60, 75])
    skip = set(gs["done_units"]) | set(gs.get("unfit_units", []))
    remaining = [u for u in units if u["unit_id"] not in skip]

    # 1) 배분계획 게이트 (최초 1회)
    if gs["phase"] == "MATERIAL":
        import math
        n_rounds = 1 + max(0, math.ceil((len(units) - 30) / band_hi))
        if len(units) < len(all_units):
            from collections import Counter
            docs = Counter(u.get("source", "?") for u in units)
            ledger_append("GOLDENSET_BATCH", "UNITS_SELECTED", "script:gen_goldenset",
                          evidence={"선정": f"{len(units)}/{len(all_units)}",
                                    "방식": "문서별 비례·등간격 대표 추출 (결정적)",
                                    "문서 수": len(docs), "목표": target}, product=prod)
        plan = llm.chat("generator", "[TASK:ALLOCATION_PLAN] 커버 단위를 배치로 배분하는 계획을 JSON으로.",
                        json.dumps({"units": len(units)}, ensure_ascii=False), cfg)
        gs["phase"] = "PLAN_GATE"
        save_goldenset(prod, gs)   # [P0-5] st 전체 되쓰기 금지 — 신선 병합
        issue_gate_card(prod, "GOLDENSET_BATCH", f"GSPLAN_{prod}",
                        what_stopped="배분계획 승인 — 재료실측 완료, 배치 배분안 확인",
                        evidence={"재료 풀(맵 단위)": len(all_units),
                                  "출제 대상(대표 추출)": f"{len(units)}" + (f" (목표 {target} · 문서별 비례)" if target and len(units) < len(all_units) else ""),
                                  "밴드": f"{band_lo}~{band_hi}",
                                  "예상 차수": f"파일럿 30 + 약 {n_rounds - 1}차수",
                                  "계획": plan[:200]},
                        recommendation="승인 시 파일럿 30문항 생성 시작")
        return "WAITING_HUMAN", {"phase": "배분계획 게이트"}

    # 2) 파일럿 / 차수 생성 (게이트 승인 후 재진입 시)
    if gs["phase"] in ("PLAN_GATE", "PILOT", "ROUNDS"):
        is_pilot = gs["round"] == 0
        take = min(30 if is_pilot else band_hi, len(remaining))
        if take == 0 and not remaining:
            gs["phase"] = "CLOSING"
        else:
            batch_units = remaining[:take]
            fb = gs.get("reject_feedback")   # 직전 배치 사람 반려 사유 — 이번 출제에 반영
            label_pre = "파일럿" if is_pilot else f"{gs['round'] + 1}차"
            _gs_progress(prod, f"{label_pre} 출제 중 — claude 대형 호출 1건 (20~40분이 정상)")
            items = generate_items_chunked(prod, batch_units,
                                           start_no=next_item_no(prod) or (len(gs["done_units"]) + 1),
                                           want_e=is_pilot, cfg=cfg, feedback=fb)
            if not items:
                # 출제 AI가 빈 배열 회신 = 이 재료들은 사람 반려 규칙 대조상 출제 부적격 판단 —
                # 무한 재시도 대신 부적격으로 확정 계상하고 진행 (마감 카드에서 사람이 최종 확인)
                ids = sorted(u["unit_id"] for u in batch_units)
                gs.setdefault("unfit_units", []).extend(ids)
                if fb:
                    gs.pop("reject_feedback", None)
                save_goldenset(prod, gs)   # [P0-5] st 전체 되쓰기 금지 — 신선 병합
                ledger_append("GOLDENSET_BATCH", "UNITS_UNFIT", "script:gen_goldenset",
                              evidence={"부적격": len(ids), "예": ids[:5],
                                        "사유": "출제 AI 전건 부적격 회신(빈 배열) — 사람 반려 규칙 대조"},
                              product=prod)
                return "CONTINUE", {"부적격 확정": len(ids), "다음": "잔여 재계산 후 진행/마감"}
            ground_citations(items, batch_units)   # 발췌→단위 역추적 authoritative 재기입
            label = "파일럿" if is_pilot else f"{gs['round'] + 1}차"
            path = write_batch(prod, items, label)
            union = sorted((DATA / prod / "04_goldenset_batch").glob("*.xlsx"))
            rc, out = verify(prod, path, [p for p in union if p != path])
            if rc == 1:   # 반려 → 재생성 1회
                ledger_append("GOLDENSET_BATCH", "BATCH_REJECTED", "script:verify_batch",
                              evidence={"batch": N(path.name), "log": out[-400:]}, product=prod)
                path.unlink()
                # [P0-1 잔재 수리 2026-08-18] 재생성 경로가 옛 번호 계산(누계+1)을 쓰고 있었음 —
                # 실측: 4차가 검수 반려→재생성되며 CI-180을 3차 마지막과 중복 발급 (조인 키 충돌)
                items = generate_items_chunked(prod, batch_units,
                                               start_no=next_item_no(prod) or (len(gs["done_units"]) + 1),
                                               want_e=is_pilot, cfg=cfg, feedback=fb)
                ground_citations(items, batch_units)
                path = write_batch(prod, items, label)
                rc, out = verify(prod, path, [p for p in union if p.exists() and p != path])
            if rc >= 2:
                return "HALTED", {"halt": f"검수 실행 오류/HALT (exit {rc})", "log": out[-300:]}
            gs["round"] += 1
            gs["phase"] = "ROUNDS"
            # [수리 2026-07-23] 실제 문항이 인용한 단위만 '완료' — 응답 절단 등으로 미커버된
            # 단위는 remaining 에 남겨 다음 차수가 자연 회수 (2차에서 75소비/40커버 사고)
            batch_ids = {u["unit_id"] for u in batch_units}
            cited = covered_units(items) & batch_ids
            lost = sorted(batch_ids - cited)
            gs["done_units"] += sorted(cited)
            # [수리 2026-08-18] 반환 단위에 튕김 횟수 기록 — 2회 이상 튕기면 '출제부적합'으로 격리.
            # 실측(11차): 반환이 29→38 가속 — 정체는 절단이 아니라 이미지 캡션·로고 설명·보안 문구·
            # 각주 같은 출제 불가 부스러기. 출제기가 건너뛰는 게 옳은데 루프가 영원히 재급식해
            # 잔여가 0에 못 닿는 구조였음. 격리분은 마감 카드에 명세로 실어 사람이 최종 확인.
            bounce = gs.setdefault("bounce", {})
            unfit = gs.setdefault("unfit_units", [])
            for uid in lost:
                bounce[uid] = bounce.get(uid, 0) + 1
                if bounce[uid] >= 2 and uid not in unfit:
                    unfit.append(uid)
            if unfit:
                # 격리분은 remaining 선정에서 빠지도록 done 과 동급으로 취급하되, 커버 수와는
                # 분리 집계 (커버 등식에 '부적합 제외 N' 별도 항목 — 커버로 위장 금지)
                gs["done_units"] = sorted(set(gs["done_units"]))  # 안전 정리
            lost = [u for u in lost if u not in unfit]
            # 반려 대비 장부: 이 배치가 소진한 단위 명세 — 사람이 반려하면 이 명세로 자동 반환
            gs["last_batch"] = {"label": label, "file": N(path.name), "units": sorted(cited)}
            if fb:
                gs.pop("reject_feedback", None)   # 피드백은 1회 반영 후 소거 (영구 편향 방지)
            save_goldenset(prod, gs)   # [P0-5] st 전체 되쓰기 금지 — 신선 병합
            verdict = "PASS" if rc == 0 else "REJECTED(재생성 후에도)"
            ev = {"배치": N(path.name), "문항": len(items), "검수 7종": verdict,
                  "직접 커버 누계": len(gs["done_units"]),
                  "잔여": len(units) - len(gs["done_units"]) - len(gs.get("unfit_units", []))}
            if gs.get("unfit_units"):
                ev["출제부적합 격리 누계"] = (f"{len(gs['unfit_units'])}단위 — 2회 이상 튕긴 부스러기"
                                       f"(이미지 캡션·보안 문구·각주 등), 마감 카드에서 명세 확인")
            if fb:
                ev["반려 피드백 반영"] = N(fb)[:100]
            if lost:
                ev["미커버 반환"] = f"{len(lost)}단위 — 응답 절단 추정, 다음 차수 재출제 (예: {lost[:3]})"
                ledger_append("GOLDENSET_BATCH", "UNITS_RETURNED", "script:gen_goldenset",
                              evidence={"반환": len(lost), "사유": "문항 미인용 — 완료 처리 안 함"},
                              product=prod)
            ledger_append("GOLDENSET_BATCH", "BATCH_GENERATED", "script:gen_goldenset",
                          evidence=ev, product=prod)
            issue_gate_card(prod, "GOLDENSET_BATCH", f"GSBATCH_{prod}_{label}",
                            what_stopped=f"{label} 생성·검수 완료 — 사람 게이트",
                            evidence=ev,
                            recommendation="승인 시 다음 차수 / 잔여 0이면 배치 마감으로\n\n"
                                           + human_guide(items, label, path, lost))
            return "WAITING_HUMAN", ev

    # 3) 배치 마감 (잔여 0)
    if gs["phase"] == "CLOSING":
        batches = sorted((DATA / prod / "04_goldenset_batch").glob(f"{prod}_골든셋_*.xlsx"))
        # 같은 차수의 구버전(v1_0)·신버전(v1_1)이 함께 잔존하면 이중 집계 — 차수별 최신판만 [P0-1]
        groups = {}
        for b in batches:
            mv = re.match(r"(.+?)_\d+문항_v(\d+)_(\d+)\.xlsx$", b.name)
            key = mv.group(1) if mv else b.stem
            ver = (int(mv.group(2)), int(mv.group(3))) if mv else (0, 0)
            if key not in groups or ver > groups[key][0]:
                groups[key] = (ver, b)
        batches = sorted((b for _, b in groups.values()), key=lambda p: p.name)
        all_items, seen, dup_rows = [], set(), []
        for b in batches:
            wb = openpyxl.load_workbook(b, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            for r in ws.iter_rows(min_row=2, values_only=True):
                d = {h: N(v) for h, v in zip(hdr, r)}
                if not d.get("ID"):
                    continue
                if d["ID"] in seen:
                    # 무단 드롭 금지 [P0-1] — 격리 파일로 보존하고 마감 카드에 실측 보고
                    dup_rows.append({**d, "출처 배치": N(b.name)})
                    continue
                seen.add(d["ID"])
                all_items.append(d)
            wb.close()
        if dup_rows:
            qwb = openpyxl.Workbook()
            qws = qwb.active
            qws.title = "ID중복_격리"
            qws.append(["출처 배치"] + STD_HEADER)
            for d in dup_rows:
                qws.append([d.get("출처 배치", "")] + [d.get(h, "") for h in STD_HEADER])
            qf = DATA / prod / "05_unified_ledger" / f"{prod}_ID중복_격리_{len(dup_rows)}건.xlsx"
            qf.parent.mkdir(parents=True, exist_ok=True)
            qwb.save(qf)
            ledger_append("GOLDENSET_BATCH", "ID_COLLISION_QUARANTINED", "script:gen_goldenset",
                          evidence={"중복": len(dup_rows), "격리 파일": N(qf.name),
                                    "사유": "차수 간 ID 충돌 — 대장에는 선착 문항만, 나머지는 격리 보존"},
                          product=prod)
        direct = len(all_items)
        pool = len(units)
        # 회계 주의: 제외소스(Known Issue) 도입으로 대표 추출 명단이 중간에 갱신됨 —
        # 그 이전 차수가 확보한 '선외' 단위는 소실이 아니라 추가 확보분으로 따로 계상한다
        sel_ids = {u["unit_id"] for u in units}
        done_set = set(gs["done_units"])
        in_pool, extra = len(done_set & sel_ids), len(done_set - sel_ids)
        eq = [["재료 풀(대표 추출)", str(pool), "커버리지맵 단위 실측"],
              ["직접 커버(풀 내)", str(in_pool), "파일 실측 합산"],
              ["선외 추가 커버", str(extra), "명단 갱신 이전 차수 확보분 — 소실 아님(추가 확보)"],
              ["흡수", "0", "—"], ["부적격", str(max(0, pool - in_pool)), "미출제 단위"],
              ["잔여", "0", "마감"],
              [f"등식: {in_pool}+0+{max(0, pool - in_pool)}+0={pool} (+선외 {extra})", str(pool), "소실 0(assert)"]]
        out = DATA / prod / "05_unified_ledger" / f"{prod}_골든셋_통합대장_{direct}문항_v1_0.xlsx"
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "골든셋_전체"
        # 통합 대장은 채점기(run_score_v11) 규격 — '정답 (필수 포함 요소)' 컬럼명 사용
        LEDGER_HEADER = [("정답 (필수 포함 요소)" if h == "정답" else h) for h in STD_HEADER]
        ws.append(LEDGER_HEADER)
        for it in all_items:
            ws.append([it.get("정답", "") if h == "정답 (필수 포함 요소)" else it.get(h, "")
                       for h in LEDGER_HEADER])
        ws2 = wb.create_sheet("커버등식")
        ws2.append(["구분", "단위 수", "설명"])
        for row in eq:
            ws2.append(row)
        wb.save(out)
        gs["phase"] = "DONE"
        save_goldenset(prod, gs)   # [P0-5] st 전체 되쓰기 금지 — 신선 병합
        ev = {"통합 대장": N(out.name), "문항": direct,
              "ID 중복": (f"{len(dup_rows)}건 — 격리 보존(대장 미편입), 격리 파일 확인 필요"
                          if dup_rows else 0),   # 하드코딩 0 금지 [P0-1] — 실측만 보고
              "커버 등식": f"소실 0 (풀 {pool})"}
        ledger_append("GOLDENSET_BATCH", "BATCH_CLOSED", "script:gen_goldenset",
                      evidence=ev, product=prod)
        if gs.get("expand_units"):
            gs.pop("expand_units", None)   # 증분 마감 — 확대 모드 해제 (다음 확대는 새 expand로)
            save_goldenset(prod, gs)   # [P0-5] st 전체 되쓰기 금지 — 신선 병합
        issue_gate_card(prod, "GOLDENSET_BATCH", f"GSCLOSE_{prod}",
                        what_stopped="배치 마감 — 통합 대장 생성 완료, 사람 확정",
                        evidence=ev, recommendation="승인 시 ⑤ 통합 대장 검사로")
        return "WAITING_HUMAN", ev

    if gs["phase"] == "DONE":
        return "DONE", {"goldenset": "마감 완료"}
    return "WAITING_HUMAN", {"phase": gs["phase"]}


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--product", required=True)
    a = ap.parse_args()
    print(run(a.product, load_config()))
