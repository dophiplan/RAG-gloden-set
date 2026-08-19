#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_twin_exam.py — 쌍둥이 시험지(맥락 v3): 골든셋 문항의 '같은 질문, 다른 말투' 변형 생성

발주 (난희 2026-08-19): "1+1=2의 답을 찾는 게 아니다. 사과를 찾는데 빨간 사과면 맞고
  초록 사과면 틀리면 안 된다" — RAG가 표현이 달라진 질문도 알아듣는지 측정.

설계:
  - 통합대장 834문항에서 결정적 표본 100문항 (8문항마다 1개 + 수치·가부형 우선 보정 없음: 단순 등간)
  - 문항당 변형 2개: ①구어체(고객이 전화로 말하듯, 군더더기 포함) ②우회(용어를 안 쓰고 상황으로 설명)
  - 정답·필수 닻·근거는 원문항 것 재사용 → 채점 체계 그대로
  - 기계 검증: 변형은 원질문과 달라야 하고(정규화 불일치), 질문 속 수치·고유 표기는 보존
  - 산출: 시험지 xlsx (원ID-T1/T2) + 원문항 매핑 json — 성적은 별도 트랙 (기존과 안 섞임)

측정 지표(응시 후): 원문항 점수 vs 변형 점수의 격차 = '문자 매칭 의존도' (격차 0이 이상적)
사용: python3 tools/gen_twin_exam.py CI [--n 100]
"""
import argparse
import datetime
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import llm
from olib import ROOT, N, ledger_append, load_config

STRIP = re.compile(r"[^0-9A-Za-zぁ-ヿ一-鿿가-힣]")


def anorm(s):
    return STRIP.sub("", unicodedata.normalize("NFC", str(s or ""))).lower()


SYS = """[TASK:QUESTION_TWINS] 너는 콜센터 고객 발화 시뮬레이터다.
입력: [{"no":1,"q":"정제된 질문"}]. 각 질문을 **뜻은 동일하게** 두 가지로 다시 말하라:
- v1 (구어체): 고객이 전화로 묻듯 — 군더더기·머뭇거림 허용 ("あの、すみません、〜なんですけど…")
- v2 (우회): 핵심 용어를 직접 쓰지 않고 상황·증상으로 설명 (용어를 모르는 고객처럼)
규칙: ① 질문 속 수치·서비스명·제품명은 그대로 유지 ② 뜻이 바뀌거나 정보가 추가되면 안 됨
③ 일본어로 (원문이 일본어) ④ 출력: JSON만 — [{"no":1,"v1":"…","v2":"…"}]"""


def main(prod, n_target):
    cfg = load_config()
    led = sorted((ROOT / "data" / prod / "05_unified_ledger").glob("*통합대장*.xlsx"))[-1]
    ws = openpyxl.load_workbook(led, read_only=True, data_only=True).worksheets[0]
    hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    qi = hdr.index("질문")
    rows = [(N(r[0]), N(r[qi])) for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0] and r[qi]]
    stride = max(1, len(rows) // n_target)
    base = rows[::stride][:n_target]
    print(f"표본 {len(base)}문항 (등간 1/{stride}) → 변형 {len(base)*2}개 생성")
    ledger_append("EXTERNAL_QA", "TWIN_EXAM_START", "script:gen_twin_exam",
                  evidence={"표본": len(base), "변형": "구어체+우회 각 1", "발주": "난희 — 맥락 이해 측정(초록 사과)"},
                  product=prod)

    twins = {}
    for i in range(0, len(base), 10):
        part = base[i:i + 10]
        payload = [{"no": k + 1, "q": q} for k, (_, q) in enumerate(part)]
        out = llm.chat("generator", SYS, json.dumps(payload, ensure_ascii=False), cfg)
        got = llm.extract_json(out)
        for x in (got if isinstance(got, list) else []):
            if isinstance(x, dict) and "no" in x and 1 <= int(x["no"]) <= len(part):
                twins[part[int(x["no"]) - 1][0]] = (N(x.get("v1", "")), N(x.get("v2", "")))
        print(f"  {min(i+10, len(base))}/{len(base)}…")

    # 기계 검증: 원문과 상이 + 질문 속 수치 보존
    items, rej = [], []
    NUMS = re.compile(r"\d+")
    for iid, q in base:
        v = twins.get(iid)
        if not v:
            rej.append((iid, "생성 누락")); continue
        for tag, tq in (("T1", v[0]), ("T2", v[1])):
            if not tq or anorm(tq) == anorm(q):
                rej.append((f"{iid}-{tag}", "원문과 동일/비어 있음")); continue
            if set(NUMS.findall(q)) - set(NUMS.findall(tq)):
                rej.append((f"{iid}-{tag}", "질문 수치 소실")); continue
            items.append({"문항ID": f"{iid}-{tag}", "원문항ID": iid, "변형": "구어체" if tag == "T1" else "우회",
                          "질문": tq, "원질문": q})
    print(f"검증: 통과 {len(items)} / 반려 {len(rej)}")

    d = ROOT / "data" / prod / "맥락시험지"
    d.mkdir(parents=True, exist_ok=True)
    out_x = d / f"{prod}_맥락시험지_v3_쌍둥이_{len(items)}문항_v1_0.xlsx"
    wb = openpyxl.Workbook(); ws2 = wb.active; ws2.title = "쌍둥이_시험지"
    ws2.append(["문항ID", "질문"])                      # 발행 규격: 질문만 (정답 비공개)
    for it in items:
        ws2.append([it["문항ID"], it["질문"]])
    wb.save(out_x)
    (d / f"{prod}_쌍둥이_매핑_v1_0.json").write_text(
        json.dumps(items, ensure_ascii=False, indent=1), encoding="utf-8")
    ledger_append("EXTERNAL_QA", "TWIN_EXAM_DONE", "script:gen_twin_exam",
                  evidence={"시험지": N(out_x.name), "문항": len(items), "반려": len(rej),
                            "채점": "정답·닻은 원문항 재사용 (매핑 json) · 성적은 별도 트랙",
                            "지표": "원문항 vs 변형 점수 격차 = 문자 매칭 의존도"}, product=prod)
    print("산출:", out_x)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("product")
    ap.add_argument("--n", type=int, default=100)
    a = ap.parse_args()
    main(a.product, a.n)
