#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_qa.py — 외부 제작 Q&A 셋 인입 어댑터 (G19)

외부(타 회사 등)가 만든 질문·답변 파일을 받아 골든셋 재료로 쓰기 위한 옆문.
핵심 원칙: 옆문으로 들어와도 검문소는 같다 — 코퍼스 대조를 통과한 것만 편입 후보.

흐름: data/<제품>/external_qa/ 의 파일(xlsx·csv·json) 전부 파싱
  → 실측(건수·컬럼·중복) → 코퍼스 문자 대조(답변 절 단위 실재 확인)
  → 3갈래 분류: 편입 후보(근거 실재) / E형 후보(코퍼스 부재) / 부분 일치(사람 검토)
  → 분류결과 xlsx + QAIMP 게이트 카드 (사람 승인)
코퍼스가 비어 있으면 전량 '보류(코퍼스 대기)' — 코퍼스 인입 후 다시 실행하면 재대조.

사용: python3 tools/import_qa.py run --product CI
"""
import argparse
import csv
import json
import re
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from olib import ROOT, N, ledger_append, issue_gate_card, load_state, save_state, now

DATA = ROOT / "data"

Q_KEYS = ("질문", "문의", "question", "query", "q",
          "質問", "タイトル", "問い", "title")            # 일본어 FAQ 지원 (タイトル=FAQ 질문 제목)
A_KEYS = ("답변", "정답", "응답", "답", "answer", "a", "response",
          "回答", "本文", "答え")                          # 일본어 FAQ 지원 (本文=FAQ 답변 본문)


def _norm(s):
    """대조용 정규화 — 공백 전부 제거 (1축 문자 대조와 같은 원리)"""
    return re.sub(r"\s+", "", unicodedata.normalize("NFC", str(s or "")))


def _pick_col(headers, keys):
    hs = [N(str(h or "")).strip().lower() for h in headers]
    for k in keys:                       # 정확 일치 우선, 다음 부분 일치
        if k in hs:
            return hs.index(k)
    for i, h in enumerate(hs):
        if any(k in h for k in keys if len(k) > 1):
            return i
    return None


def _rows_from_xlsx(p):
    import openpyxl
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    out = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [[("" if c is None else str(c)) for c in (r or [])]
                for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        qi, ai = _pick_col(rows[0], Q_KEYS), _pick_col(rows[0], A_KEYS)
        if qi is None or ai is None:
            continue
        for r in rows[1:]:
            q = r[qi].strip() if qi < len(r) else ""
            a = r[ai].strip() if ai < len(r) else ""
            if q:
                out.append({"질문": q, "답변": a})
    wb.close()
    return out


def _ver_of(name):
    """파일명에서 vN 추출 — 문자열 정렬 금지 [P1-2]: '5문항_v1'과 '12문항_v2'는
    글자 정렬로는 v1이 뒤(최신 오인). 버전 숫자로만 비교한다."""
    m = re.search(r"_v(\d+)\.xlsx$", str(name))
    return int(m.group(1)) if m else 0


def _latest(paths):
    paths = list(paths)
    return max(paths, key=lambda p: _ver_of(p.name)) if paths else None


def _rows_from_csv(p):
    out = []
    rows = None
    for enc in ("utf-8-sig", "cp949", "euc-kr", "cp932"):   # [P1-4] 한국 엑셀=cp949, 일본 엑셀=cp932(Shift-JIS)
        try:
            with open(p, newline="", encoding=enc) as f:
                rows = list(csv.reader(f))
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if rows is None:
        return out
    if not rows:
        return out
    qi, ai = _pick_col(rows[0], Q_KEYS), _pick_col(rows[0], A_KEYS)
    if qi is None or ai is None:
        return out
    for r in rows[1:]:
        q = (r[qi] if qi < len(r) else "").strip()
        a = (r[ai] if ai < len(r) else "").strip()
        if q:
            out.append({"질문": q, "답변": a})
    return out


def _rows_from_json(p):
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []
    if isinstance(data, dict):           # {"qa": [...]} / {"items": [...]} 포장 해제
        for k in ("qa", "items", "data", "questions"):
            if isinstance(data.get(k), list):
                data = data[k]
                break
    if not isinstance(data, list):
        return []
    out = []
    for it in data:
        if not isinstance(it, dict):
            continue
        km = {N(str(k)).strip().lower(): v for k, v in it.items()}
        q = next((str(km[k]).strip() for k in Q_KEYS if km.get(k)), "")
        a = next((str(km[k]).strip() for k in A_KEYS if km.get(k)), "")
        if q:
            out.append({"질문": q, "답변": a})
    return out


def parse_qa_files(qa_dir):
    """external_qa/ 의 원본 파일 전부 → [{출처, 질문, 답변}] (분류결과·반려 파일 제외)"""
    items, skipped = [], []
    for p in sorted(qa_dir.glob("*")):
        # 외부QA_* = 우리가 만든 산출물(분류결과·시험지·정답키) — 원재료가 아니라 제외
        if (not p.is_file() or p.name.startswith((".", "반려_", "외부QA_"))
                or p.suffix == ".md"):
            continue
        try:   # [P1-4] 파일 하나(손상 xlsx·미지 인코딩)가 인입 전체를 죽이면 안 됨 — 건너뛰고 목록 보고
            rows = ({".xlsx": _rows_from_xlsx, ".csv": _rows_from_csv,
                     ".json": _rows_from_json}.get(p.suffix.lower(), lambda _: [])(p))
        except Exception as e:
            print(f"  ⚠ {N(p.name)} 읽기 실패({type(e).__name__}) — 건너뜀")
            rows = []
        if rows:
            for r in rows:
                r["출처 파일"] = N(p.name)
            items += rows
        else:
            skipped.append(N(p.name))
    return items, skipped


def load_corpus_text(prod):
    """코퍼스 원문 수집 — 읽을 수 있는 형식만 (pdf·zip 등은 미대조 목록으로 보고)"""
    d = DATA / prod / "corpus"
    corpus, used, unread = [], [], []
    if not d.is_dir():
        return "", used, unread
    for p in sorted(d.rglob("*")):
        if not p.is_file() or p.name.startswith("."):
            continue
        sfx = p.suffix.lower()
        try:
            if sfx in (".txt", ".md", ".csv", ".json", ".html", ".htm"):
                corpus.append(p.read_text(encoding="utf-8", errors="ignore"))
                used.append(N(p.name))
            elif sfx == ".xlsx":
                import openpyxl
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                for sn in wb.sheetnames:
                    for r in wb[sn].iter_rows(values_only=True):
                        corpus.append(" ".join(str(c) for c in (r or []) if c is not None))
                wb.close()
                used.append(N(p.name))
            else:
                unread.append(N(p.name))
        except Exception:
            unread.append(N(p.name))
    return _norm("\n".join(corpus)), used, unread


def _clauses(answer):
    """답변을 대조 가능한 절 단위로 — 문장·줄 단위로 쪼개 정규화 길이 10자 이상만"""
    parts = re.split(r"[.。!?\n;·•]|(?:니다|세요|합니다)\s", str(answer or ""))
    return [c.strip() for c in parts if len(_norm(c)) >= 10]


def classify(items, corpus_norm):
    """3갈래 분류 — 절 실재율로 판정. 코퍼스 없으면 전량 보류."""
    seen_q = set()
    for it in items:
        qn = _norm(it["질문"])
        it["비고"] = "중복 질문" if qn in seen_q else ""
        seen_q.add(qn)
        cl = _clauses(it["답변"])
        if not corpus_norm:
            it.update({"분류": "보류(코퍼스 대기)", "일치율": "", "확인 발췌": ""})
            continue
        if not cl:
            it.update({"분류": "부분 일치(사람 검토)", "일치율": "0/0",
                       "확인 발췌": "", "비고": (it["비고"] + " 답변이 짧아 대조 불가").strip()})
            continue
        hits = [c for c in cl if _norm(c) in corpus_norm]
        ratio = len(hits) / len(cl)
        it["일치율"] = f"{len(hits)}/{len(cl)}"
        it["확인 발췌"] = (hits[0][:120] if hits else "")
        it["분류"] = ("편입 후보(근거 실재)" if ratio >= 0.6
                     else "E형 후보(코퍼스 부재)" if ratio == 0
                     else "부분 일치(사람 검토)")
    return items


def write_report(prod, items, version):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "분류결과"
    cols = ["번호", "출처 파일", "질문", "답변", "분류", "일치율", "확인 발췌", "비고"]
    ws.append(cols)
    for i, it in enumerate(items, 1):
        ws.append([i] + [it.get(c, "") for c in cols[1:]])
    for c, w in zip("ABCDEFGH", (6, 18, 46, 46, 22, 8, 40, 16)):
        ws.column_dimensions[c].width = w
    out = DATA / prod / "external_qa" / f"외부QA_분류결과_{prod}_v{version}.xlsx"
    wb.save(out)
    return out


def run(prod, actor="난희"):
    qa_dir = DATA / prod / "external_qa"
    qa_dir.mkdir(parents=True, exist_ok=True)
    items, skipped = parse_qa_files(qa_dir)
    if not items:
        msg = ("외부 Q&A 파일에서 질문·답변을 찾지 못했어요 — "
               "질문/답변 컬럼(또는 question/answer 키)이 있는 xlsx·csv·json 인지 확인해 주세요."
               + (f" (읽지 못한 파일: {', '.join(skipped)})" if skipped else " (올라온 파일 없음)"))
        print(msg)
        return {"ok": False, "out": msg}
    corpus_norm, corpus_used, corpus_unread = load_corpus_text(prod)
    items = classify(items, corpus_norm)
    version = 1 + max((_ver_of(p.name) for p in
                       list(qa_dir.glob(f"외부QA_분류결과_{prod}_v*.xlsx"))
                       + list(qa_dir.glob(f"반려_외부QA_분류결과_{prod}_v*.xlsx"))), default=0)
    report = write_report(prod, items, version)
    cnt = {}
    for it in items:
        cnt[it["분류"]] = cnt.get(it["분류"], 0) + 1
    dup = sum(1 for it in items if "중복" in it.get("비고", ""))
    ev = {"총 문항": len(items),
          **{k: f"{v}건" for k, v in sorted(cnt.items(), key=lambda x: -x[1])},
          "중복 질문": f"{dup}건",
          "대조한 코퍼스": f"{len(corpus_used)}파일" if corpus_used else "없음 (전량 보류)",
          "분류결과 파일": N(report.name)}
    if corpus_unread:
        ev["대조 못한 코퍼스(형식)"] = ", ".join(corpus_unread[:5]) + \
            (f" 외 {len(corpus_unread)-5}" if len(corpus_unread) > 5 else "")
    if skipped:
        ev["읽지 못한 Q&A 파일"] = ", ".join(skipped)
    ledger_append("EXTERNAL_QA", "QA_IMPORTED", f"script:import_qa({actor})",
                  evidence=ev, product=prod)
    # 재대조면 이전 QAIMP 카드는 자동 교체 — 낡은 카드에 승인/반려하는 사고 방지
    from olib import close_gate
    st = load_state()
    for g in list(st["products"][prod].get("open_gates", [])):
        if g["id"].startswith("QAIMP_"):
            close_gate(prod, g["id"])
            ledger_append("EXTERNAL_QA", "GATE_SUPERSEDED", "script:import_qa",
                          gate_id=g["id"], evidence={"사유": "재대조로 새 분류 발행"}, product=prod)
    gate_id = f"QAIMP_{prod}_v{version}"
    if corpus_norm:
        what = (f"외부 Q&A {len(items)}문항을 코퍼스와 대조해 3갈래로 분류했어요. "
                f"[👁 실물 보고 결정] 팝업에서 분류결과({N(report.name)})를 확인해 주세요.\n"
                "- 편입 후보(근거 실재): 승인하면 골든셋 출제 재료로 쓰입니다 (검수 8종·사람 게이트는 그대로 통과해야 함)\n"
                "- E형 후보(코퍼스 부재): v3 함정 문항(E형 10%) 재료로 보관됩니다\n"
                "- 부분 일치: 코퍼스와 다르게 서술된 것 — 어느 쪽이 정본인지 사람 확인이 필요합니다\n"
                "- 반려하면 이번 분류는 폐기돼요 (원본 Q&A 파일은 보존 — 다시 올릴 필요 없음)")
    else:
        what = (f"외부 Q&A {len(items)}문항을 접수했어요. 아직 이 제품의 코퍼스가 없어서 "
                "대조를 못 했고 전량 '보류(코퍼스 대기)'예요.\n"
                "- 승인 = 접수 확정 (코퍼스를 올린 뒤 [📥 외부 Q&A 올리기] 옆 재대조를 누르면 3갈래 분류가 돌아요)\n"
                "- 골든셋 자격은 코퍼스 대조 통과가 조건이라, 코퍼스 없이 편입은 안 됩니다")
    issue_gate_card(prod, "EXTERNAL_QA", gate_id, what, ev)
    out = (f"외부 Q&A 인입 완료 — {len(items)}문항 · " +
           " · ".join(f"{k} {v}" for k, v in cnt.items()) +
           f" → 카드 {gate_id} 발행 (분류결과: {N(report.name)})")
    print(out)
    return {"ok": True, "out": out}


# ── G20 · 별도 트랙 — 외부 Q&A는 기존 골든셋과 절대 섞지 않는다 (난희 설계 2026-08-05)
#    승인 → 시험지(질문만)+정답키(봉인) 발행 → 꾸러미 전달 → 응답로그 → 자체 회차(qa r1, r2…)
#    합치기(v3 편입)는 필요해질 때 별도 공사.

def latest_report(prod):
    """최신 유효 분류결과 (반려_ 제외) — 버전 숫자 기준 [P1-2]"""
    return _latest((DATA / prod / "external_qa").glob(f"외부QA_분류결과_{prod}_v*.xlsx"))


def publish(prod, actor="난희"):
    """QAIMP 승인 후처리 — 편입 후보만으로 시험지(ID·질문, 정답 비공개)와 정답키(봉인) 발행."""
    import openpyxl
    rep = latest_report(prod)
    if not rep:
        print("발행 불가 — 분류결과 파일이 없어요 (먼저 외부 Q&A를 올려 분류부터)")
        return None
    ws = openpyxl.load_workbook(rep, read_only=True, data_only=True).active
    rows = [[("" if c is None else str(c)) for c in (r or [])] for r in ws.iter_rows(values_only=True)]
    hdr = rows[0]
    ix = {k: hdr.index(k) for k in ("출처 파일", "질문", "답변", "분류", "확인 발췌", "비고")}
    cand = [r for r in rows[1:]
            if r[ix["분류"]].startswith("편입 후보") and "중복" not in r[ix["비고"]]]
    if not cand:
        print("발행 생략 — 편입 후보(근거 실재) 문항이 0건이에요 (보류·E형·부분 일치만 있음)")
        return None
    d = DATA / prod / "external_qa"
    ver = 1 + max((_ver_of(p.name) for p in
                   list(d.glob(f"외부QA_시험지_{prod}_*문항_v*.xlsx"))
                   + list(d.glob(f"외부QA_정답키_{prod}_v*.xlsx"))), default=0)   # [P1-2] 개수 아닌 최대+1
    paper = openpyxl.Workbook()
    pw = paper.active
    pw.title = "시험지"
    pw.append(["문항ID", "질문"])
    key = openpyxl.Workbook()
    kw = key.active
    kw.title = "정답키(봉인)"
    kw.append(["문항ID", "질문", "정답(외부 제공)", "확인 발췌(코퍼스 실재)", "출처 파일"])
    for i, r in enumerate(cand, 1):
        # [P1-3] ID에 시험지 버전 내장 — v1 응답로그가 v2 정답키와 짝지어지면 ID가 안 맞아
        # '결측'으로 드러난다 (무증상 오채점 차단). 채점기는 ID로 정답키 버전을 역추적한다.
        qid = f"{prod}-Qv{ver}-{i:03d}"
        pw.append([qid, r[ix["질문"]]])
        kw.append([qid, r[ix["질문"]], r[ix["답변"]], r[ix["확인 발췌"]], r[ix["출처 파일"]]])
    pname = f"외부QA_시험지_{prod}_{len(cand)}문항_v{ver}.xlsx"
    paper.save(d / pname)
    key.save(d / f"외부QA_정답키_{prod}_v{ver}.xlsx")
    ledger_append("EXTERNAL_QA", "QA_PUBLISHED", f"script:import_qa({actor})",
                  evidence={"시험지": pname, "문항": len(cand), "분류결과": N(rep.name),
                            "트랙": "별도 (기존 골든셋과 미합류)"}, product=prod)
    print(f"📄 Q&A 시험지 발행 — {pname} ({len(cand)}문항, 정답키 봉인 보관). "
          f"[📦 Q&A 꾸러미]로 팀장님께 전달하세요.")
    return pname


def _hit_text(hit):
    """검색 hit 안의 모든 문자열을 모아 대조용으로 — 형식이 회사마다 달라도 내용만 있으면 잡는다"""
    if isinstance(hit, str):
        return hit
    if isinstance(hit, dict):
        return " ".join(str(v) for v in hit.values() if isinstance(v, (str, int, float)))
    return ""


def score(prod, log_name=None, actor="난희"):
    """Q&A 트랙 채점 — 검색축(내용 대조 기준): 정답 절이 hits 본문에 실재하는 첫 순위로 판정.
    기계 판정만 (LLM 없음) — 생성축은 미채점(별도 트랙 v0 규격)."""
    import hashlib
    import openpyxl
    d = DATA / prod / "external_qa"
    keys = list(d.glob(f"외부QA_정답키_{prod}_v*.xlsx"))
    if not keys:
        msg = "채점 불가 — Q&A 시험지가 아직 발행 전이에요 (분류 카드 승인이 먼저)"
        print(msg)
        return {"ok": False, "out": msg}
    ld = d / "로그"   # 응답로그 전용 폴더 — Q&A 원본과 섞이면 재대조가 로그를 문항으로 오인
    logs = ([ld / log_name] if log_name else
            sorted((p for p in ld.glob("*") if p.suffix.lower() == ".json"),
                   key=lambda p: p.stat().st_mtime) if ld.is_dir() else [])   # [P3] .JSON 대소문자 허용
    if not logs or not logs[-1].exists():
        msg = "채점 불가 — 응답로그(json)가 안 보여요. [⬆ Q&A 응답로그 채점]으로 올려주세요."
        print(msg)
        return {"ok": False, "out": msg}
    log_f = logs[-1]
    log_sha = hashlib.sha256(log_f.read_bytes()).hexdigest()[:16]
    # 같은 로그 재채점 차단 [P1-7] — 재클릭·재업로드로 중복 회차가 쌓이는 사고 방지
    for old_md in (ROOT / "results").glob(f"score_{prod}QA_r*/외부QA_r*_리포트.md"):
        if f"로그지문: {log_sha}" in old_md.read_text(encoding="utf-8"):
            msg = (f"이미 채점된 로그예요 ({old_md.parent.name}) — 같은 파일로는 회차를 더 만들지 않아요. "
                   f"새 응답로그를 받으면 올려주세요.")
            print(msg)
            return {"ok": False, "out": msg}
    data = json.loads(log_f.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        for k in ("responses", "items", "results", "data", "logs"):
            if isinstance(data.get(k), list):
                data = data[k]
                break
    if not isinstance(data, list):
        msg = f"채점 불가 — {N(log_f.name)} 이 응답 목록 형식이 아니에요"
        print(msg)
        return {"ok": False, "out": msg}
    # 로그-시험지 버전 바인딩 [P1-3]: 로그 ID의 Qv{N}으로 정답키 버전을 역추적 — 구판 로그를
    # 신판 키로 채점하는 무증상 오채점 차단. 버전 표기가 없는 로그(구형)는 최신 키 사용.
    vm = next((re.search(r"-Qv(\d+)-", str(it.get("id") or it.get("question_id") or ""))
               for it in data if isinstance(it, dict) and
               re.search(r"-Qv(\d+)-", str(it.get("id") or it.get("question_id") or ""))), None)
    key_f = None
    if vm:
        want = int(vm.group(1))
        key_f = next((k for k in keys if _ver_of(k.name) == want), None)
        if key_f is None:
            msg = f"채점 불가 — 로그가 시험지 v{want} 응답인데 그 버전 정답키가 없어요 (있는 버전: {sorted(_ver_of(k.name) for k in keys)})"
            print(msg)
            return {"ok": False, "out": msg}
    key_f = key_f or _latest(keys)
    ws = openpyxl.load_workbook(key_f, read_only=True, data_only=True).active
    answers = {}   # qid → {질문, 절 목록}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        qid, q, a = str(r[0] or ""), str(r[1] or ""), str(r[2] or "")
        cl = _clauses(a) or ([a] if _norm(a) else [])
        answers[qid] = {"질문": q, "절": cl, "qnorm": _norm(q)}
    by_id, by_q = {}, {}
    for it in data:
        if not isinstance(it, dict):
            continue
        iid = str(it.get("id") or it.get("question_id") or it.get("문항ID") or "")
        if iid:
            by_id[iid] = it
        qn = _norm(it.get("question") or it.get("질문") or "")
        if qn:
            by_q[qn] = it
    rows, cnt = [], {"hit_top1": 0, "hit_top5": 0, "miss": 0, "결측": 0}
    for qid, a in answers.items():
        it = by_id.get(qid) or by_q.get(a["qnorm"])
        verdict, rank, ev = "결측", "", ""
        if it is not None:
            hits = next((it[k] for k in ("hits", "results", "documents", "contexts", "chunks")
                         if isinstance(it.get(k), list)), [])
            verdict = "miss"
            for i, h in enumerate(hits[:5], 1):
                hn = _norm(_hit_text(h))
                found = next((c for c in a["절"] if _norm(c) in hn), None)
                if found:
                    verdict = "hit_top1" if i == 1 else "hit_top5"
                    rank, ev = i, found[:100]
                    break
        cnt[verdict] += 1
        rows.append({"문항ID": qid, "질문": a["질문"], "검색": verdict,
                     "명중순위": rank, "확인 절": ev})
    rnd = 1 + max((int(m.group(1)) for p in (ROOT / "results").glob(f"score_{prod}QA_r*")
                   if (m := re.fullmatch(rf"score_{re.escape(prod)}QA_r(\d+)", p.name))),
                  default=0)   # [P1-7] 개수 방식은 폴더 이동/삭제 시 기존 회차 덮어씀 — 최대+1
    out_d = ROOT / "results" / f"score_{prod}QA_r{rnd}"
    out_d.mkdir(parents=True, exist_ok=True)
    (out_d / "score_report.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=1), encoding="utf-8")
    wb = openpyxl.Workbook()
    xw = wb.active
    xw.title = "Q&A채점"
    xw.append(list(rows[0].keys()))
    for r in rows:
        xw.append(list(r.values()))
    wb.save(out_d / f"외부QA_채점_{prod}_r{rnd}.xlsx")
    n = len(rows)
    t1, t5 = cnt["hit_top1"], cnt["hit_top1"] + cnt["hit_top5"]
    warn = ""
    if cnt["결측"] > n / 2:
        warn = (f"\n⚠ 결측이 과반({cnt['결측']}/{n}) — 응답로그가 다른 버전 시험지의 것이거나 "
                f"ID 형식이 다를 수 있어요. 꾸러미를 다시 받아 재응시를 요청하세요.\n")
    (out_d / f"외부QA_r{rnd}_리포트.md").write_text(
        f"# 외부 Q&A 별도 트랙 — {prod} qa-r{rnd} (검색축만 · 내용 대조 기준)\n\n"
        f"- 응답로그: {N(log_f.name)} · 로그지문: {log_sha} · 정답키: {N(key_f.name)}\n"
        f"- top1 {t1}/{n} ({t1/n:.1%}) · top5 {t5}/{n} ({t5/n:.1%}) · "
        f"miss {cnt['miss']} · 결측 {cnt['결측']}\n{warn}\n"
        f"판정 방식: 정답(외부 제공)의 절이 검색 hits 본문에 실재하는 첫 순위 — 기계 판정(LLM 없음).\n"
        f"이 트랙은 기존 골든셋 성적과 **분리 집계**됩니다 (합류는 v3 편입 시 별도 공사).\n",
        encoding="utf-8")
    ledger_append("EXTERNAL_QA", "QA_SCORED", f"script:import_qa({actor})",
                  evidence={"회차": f"qa-r{rnd}", "n": n, "top1": t1, "top5": t5,
                            "결측": cnt["결측"], "로그": N(log_f.name)}, product=prod)
    msg = (f"✅ Q&A 트랙 채점 완료 — qa-r{rnd}: top1 {t1}/{n} ({t1/n:.0%}) · "
           f"top5 {t5}/{n} ({t5/n:.0%}) — 성적 히스토리 '외부 Q&A' 줄에서 확인")
    print(msg)
    return {"ok": True, "out": msg}


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    r = sub.add_parser("run")
    r.add_argument("--product", required=True)
    r.add_argument("--actor", default="난희")
    p = sub.add_parser("publish")
    p.add_argument("--product", required=True)
    p.add_argument("--actor", default="난희")
    s = sub.add_parser("score")
    s.add_argument("--product", required=True)
    s.add_argument("--log")
    s.add_argument("--actor", default="난희")
    a = ap.parse_args()
    if a.cmd == "run":
        run(a.product, a.actor)
    elif a.cmd == "publish":
        publish(a.product, a.actor)
    else:
        score(a.product, a.log, a.actor)


if __name__ == "__main__":
    main()
