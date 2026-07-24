#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
judge_run.py — ⑥ 캘리브레이션 판정 실행 + ⑦ 본판정 (사양서 §4-⑥⑦)

⑥: 통합 대장에서 30문항 추출(유형 비례·E 필수 포함) → judge 판정(규칙 B — 새 세션,
   [문항, 기준서]만) → 판정문 보존 → 사람 블라인드 판정 입력 시트 생성 →
   사람 판정 기입 후 대조표 완성(calibration.py 가 실측).
⑦: calibration_passed 확인(규칙 C) → 전건 판정 → 판정대장 xlsx + 무작위 재검
   {seed, 추출 ID 목록} 원장 기록 의무 [v1.1].
"""
import json
import random
import sys
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import llm
from olib import ROOT, N, load_config, load_state, ledger_append, issue_gate_card
from model_adapter import build_judge_request, effective_recheck_rate

DATA = ROOT / "data"


def load_items(prod):
    """정본 통합 대장 → 문항 리스트"""
    led = sorted((DATA / prod / "05_unified_ledger").glob("*.xlsx"))
    if not led:
        return [], None
    path = led[-1]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        try:
            hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        except StopIteration:
            continue
        if "ID" in hdr and "질문" in hdr:
            for r in ws.iter_rows(min_row=2, values_only=True):
                d = {h: N(v) for h, v in zip(hdr, r)}
                if d.get("ID"):
                    if "정답" not in d and "정답 (필수 포함 요소)" in d:
                        d["정답"] = d["정답 (필수 포함 요소)"]
                    items.append(d)
            break
    wb.close()
    return items, path


def load_rubric(prod):
    for pat in ("*프롬프트*vFinal*", "*기준서*", "*프롬프트*"):
        # FLAG-02: diff 파일(_diff_v2_to_vFinal_ 등) 오매칭 제외
        c = [p for p in sorted((DATA / prod / "07_stage2").glob(pat))
             if "diff" not in N(p.name).lower()]
        if c:
            return c[-1].read_text(encoding="utf-8"), N(c[-1].name)
    return "판정 기준서: 정답 필수 요소 전건 포함=합격 / 일부=부분 / 상충·창작=0점. E형은 부재 인정만 합격.", "(기본 기준서)"


SYSTEM_J = """[TASK:JUDGE_VERDICTS] 너는 골든셋 2축 판정관이다. 입력의 판정 기준서만 따르라.
출력: JSON 배열만 — [{ID, 판정(합격|부분|0점), 판정문}]. 판정문에 조항 근거를 인용하라(전건 보존됨)."""


RUBRIC_LIMIT = 4000


def judge_batch(items, rubric, cfg, batch=20, role="judge"):
    """규칙 B: build_judge_request 로 오염 키 제거 후 호출. 배치 단위 독립 호출(새 세션).
    role — 채점관 대결용: 'judge'(Kimi) 외에 'generator'(claude 새 세션)로도 판정 가능."""
    # FLAG-01: 기준서 절단 금지 — 초과 시 명시 정지 (조항이 잘린 채 판정되는 사고 방지)
    if len(rubric) > RUBRIC_LIMIT:
        raise RuntimeError(f"기준서 {len(rubric):,}자 — {RUBRIC_LIMIT:,}자 한도 초과. "
                           "분할 불가: 한도 상향 또는 기준서 축약 필요 (절단 판정 금지)")
    env = None
    if llm.is_mock():   # mock 모드 = 2키 구성으로 간주 (generator+judge)
        env = {m["api_key_env"]: "mock" for m in cfg["models"].values() if m and m.get("api_key_env")}
    out = []
    for i in range(0, len(items), batch):
        part = items[i:i + batch]
        reqs = [build_judge_request(it, rubric, cfg, env=env)["inputs"]["문항"]
                for it in part]
        resp = llm.chat(role, SYSTEM_J,
                        json.dumps({"items": reqs, "rubric": rubric}, ensure_ascii=False), cfg)
        out += llm.extract_json(resp)
    return out


def pick_calibration_set(items, n=30):
    """유형 비례 + E형 필수 포함 (결정적: ID 정렬 후 등간 추출)"""
    by_type = {}
    for it in sorted(items, key=lambda x: x["ID"]):
        by_type.setdefault(it.get("유형", "?"), []).append(it)
    total = len(items)
    picked = []
    for t, lst in sorted(by_type.items()):
        k = max(1, round(n * len(lst) / total)) if t != "E" else max(1, min(3, len(lst)))
        step = max(1, len(lst) // k)
        picked += lst[::step][:k]
    return picked[:n] if len(picked) >= n else picked


def run_calibration_judging(prod, cfg):
    items, src = load_items(prod)
    if not items:
        return "WAITING_INPUT", {"통합 대장": "없음"}
    rubric, rname = load_rubric(prod)
    cal = pick_calibration_set(items, 30)
    verdicts = judge_batch(cal, rubric, cfg)
    vmap = {N(v["ID"]): v for v in verdicts}
    out = DATA / prod / "06_calibration" / f"{prod}_judge_캘리브레이션_판정30_v1_0.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    # 사람용 기입 시트가 첫 화면 — 질문·정답·기준만 보이고 judge 판정은 숨김(블라인드 실체화)
    from openpyxl.worksheet.datavalidation import DataValidation
    ws0 = wb.active
    ws0.title = "판정_기입(여기만 채우세요)"
    ws0.append(["문항ID", "유형", "질문", "시스템 정답", "합격 기준", "사람 판정(합격/불합격)", "메모(선택)"])
    for it in cal:
        ws0.append([it["ID"], it.get("유형", ""), it["질문"],
                    N(it.get("정답 (필수 포함 요소)") or it.get("정답", "")),
                    N(it.get("합격 기준", "")), "", ""])
    dv = DataValidation(type="list", formula1='"합격,불합격"', allow_blank=True)
    ws0.add_data_validation(dv)
    dv.add(f"F2:F{len(cal) + 1}")
    for col, w in zip("ABCDEFG", (10, 5, 46, 46, 34, 16, 20)):
        ws0.column_dimensions[col].width = w
    ws0.freeze_panes = "A2"
    ws = wb.create_sheet("1_judge_판정_전건")
    ws.append(["문항ID", "유형", "질문", "judge 판정", "judge 판정문(전건 보존)"])
    for it in cal:
        v = vmap.get(it["ID"], {})
        ws.append([it["ID"], it.get("유형", ""), it["질문"], N(v.get("판정", "미판정")), N(v.get("판정문", ""))])
    ws2 = wb.create_sheet("2_대조표_판정완료")
    ws2.append(["문항ID", "judge 판정", "사람 판정(설계본부 기입)", "일치 여부", "불일치 사유 분류"])
    for it in cal:
        v = vmap.get(it["ID"], {})
        ws2.append([it["ID"], N(v.get("판정", "")), "", "", ""])   # 사람 판정은 기입 시트에서 취득
    ws.sheet_state = ws2.sheet_state = "hidden"   # 사람이 열면 기입 시트만 보인다
    wb.save(out)
    ev = {"세트": f"{len(cal)}문항 (유형 {dict(Counter(i.get('유형','?') for i in cal))})",
          "기준서": rname, "판정문 보존": "전건", "산출": N(out.name),
          "다음": "사람 블라인드 판정 30건 기입 → calibration measure"}
    ledger_append("CALIBRATION", "JUDGE30_EXECUTED", "script:judge_run", evidence=ev, product=prod)
    return "DONE", ev


def run_calibration_compare(prod, cfg, roles=("judge", "generator")):
    """채점관 대결 — 같은 30문항을 여러 후보(Kimi=judge / claude=generator)로 각각 판정해
    한 대조표에 나란히. 사람이 블라인드 판정 30건을 채우면 calibration 이 후보별 일치율을
    실측 → 높은 쪽을 채점관으로 (감이 아니라 우리 문제로 시험쳐서 결정)."""
    from model_adapter import detect_mode
    import os
    items, src = load_items(prod)
    if not items:
        return "WAITING_INPUT", {"통합 대장": "없음"}
    rubric, rname = load_rubric(prod)
    cal = pick_calibration_set(items, 30)
    have = detect_mode(cfg, os.environ)["have"]
    # 연결된 후보만 (mock 은 judge·generator 둘 다 가용으로 간주)
    cand = [r for r in roles if llm.is_mock() or have.get(r)]
    if len(cand) < 2:
        return "HALTED", {"halt": f"대결 불가 — 판정 후보 {len(cand)}명뿐 (2명 이상 필요). "
                                  "config에 judge(Kimi)·generator(claude) 둘 다 연결 필요"}
    eng = {"judge": "Kimi", "generator": "claude", "reviewer": "codex"}
    cols = {}
    for r in cand:
        cols[r] = {N(v["ID"]): v for v in judge_batch(cal, rubric, cfg, role=r)}
        ledger_append("CALIBRATION", "COMPARE_JUDGED", f"script:{r}",
                      evidence={"후보": eng.get(r, r), "문항": len(cal)}, product=prod)
    out = DATA / prod / "06_calibration" / f"{prod}_채점관대결_대조표_v1_0.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "채점관대결_대조표"
    head = ["문항ID", "유형", "질문"] + [f"{eng.get(r, r)} 판정" for r in cand] + ["사람 판정(블라인드 기입)"]
    ws.append(head)
    for it in cal:
        row = [it["ID"], it.get("유형", ""), it["질문"]]
        row += [N(cols[r].get(it["ID"], {}).get("판정", "미판정")) for r in cand]
        row.append("")   # 사람 판정 블라인드
        ws.append(row)
    wb.save(out)
    ev = {"후보": [eng.get(r, r) for r in cand], "문항": len(cal), "기준서": rname,
          "산출": N(out.name),
          "다음": "사람 30건 블라인드 판정 기입 → 후보별 일치율 실측 → 높은 쪽을 채점관으로"}
    ledger_append("CALIBRATION", "COMPARE_TABLE_BUILT", "script:judge_run", evidence=ev, product=prod)
    return "DONE", ev


def run_stage2(prod, cfg, batch_size=20):
    """[FIX-04] 배치 영속화 + resume (불변 단서 ②) — 사용량 한도·중단에도 기왕 판정 보존."""
    import os
    st = load_state()
    if not st["products"][prod]["calibration_passed"]:
        return "BLOCKED", {"사유": "규칙 C — calibration_passed=false"}
    items, src = load_items(prod)
    if not items:
        return "WAITING_INPUT", {"통합 대장": "없음"}
    rubric, rname = load_rubric(prod)

    # 진행 파일: 배치마다 즉시 append — 재실행 시 완료 ID 건너뜀
    prog = DATA / prod / "07_stage2" / "_stage2_progress.jsonl"
    prog.parent.mkdir(parents=True, exist_ok=True)
    done = {}
    if prog.exists():
        for line in prog.read_text(encoding="utf-8").splitlines():
            if line.strip():
                r = json.loads(line)
                done[N(r["ID"])] = r
    todo = [it for it in items if it["ID"] not in done]
    if done:
        print(f"  ↻ resume: 기완료 {len(done)} / 전체 {len(items)} — 잔여 {len(todo)}부터 재개")

    fail_at = os.environ.get("ORCH_FAIL_AT_BATCH")      # 회귀 T13 훅: N번째 배치 강제 예외
    drop_id = os.environ.get("ORCH_DROP_ID")            # 회귀 T13 훅: 판정 응답 ID 누락 주입
    try:
        for bi in range(0, len(todo), batch_size):
            bno = bi // batch_size + 1
            if fail_at and bno == int(fail_at):
                raise RuntimeError(f"[T13 주입] 배치 {bno} 강제 예외 (구독 한도 모의)")
            part = todo[bi:bi + batch_size]
            verdicts = judge_batch(part, rubric, cfg, batch=batch_size)
            vmap = {N(v.get("ID", "")): v for v in verdicts}
            if drop_id and drop_id in vmap:
                del vmap[drop_id]
            with open(prog, "a", encoding="utf-8") as f:
                for it in part:
                    v = vmap.get(it["ID"])
                    if v:   # 미수신 ID는 기록하지 않는다 — 침묵 "미판정" 집계 금지
                        row = {"ID": it["ID"], "유형": it.get("유형", ""),
                               "판정": N(v.get("판정", "")), "판정문": N(v.get("판정문", ""))}
                        f.write(json.dumps(row, ensure_ascii=False) + "\n")
                        done[it["ID"]] = row
    except Exception as e:
        ledger_append("STAGE2", "STAGE2_PARTIAL", "script:judge_run",
                      evidence={"완료": len(done), "전체": len(items), "기준서": rname,
                                "오류": str(e)[:200]},
                      reason="중단 — progress 보존됨", product=prod)
        print(f"⛔ 본판정 중단: {e}\n   진행 {len(done)}/{len(items)} 보존 — 같은 명령 재실행 시 이어서 판정한다.")
        raise

    # 전건 응답 확인 — 미판정 > 0 이면 DONE 금지, 사람 게이트 (FIX-04-4)
    missing = [it["ID"] for it in items if it["ID"] not in done]
    if missing:
        issue_gate_card(prod, "STAGE2", f"S2MISS_{prod}",
                        what_stopped=f"본판정 응답 ID 누락 {len(missing)}건 — judge 응답에서 미수신, 침묵 집계 금지",
                        evidence={"누락": missing[:20], "완료": len(done), "전체": len(items)},
                        flags=[{"type": "미판정", "id": i, "candidates": [], "ack_required": True}
                               for i in missing[:20]],
                        recommendation="재실행(resume)으로 재시도하거나, 반복 누락 시 문항/프롬프트 점검")
        return "WAITING_HUMAN", {"미판정": len(missing), "완료": len(done)}

    # 무작위 재검 — 시드+추출 목록 원장 기록 의무 [v1.1]
    seed = cfg["pipeline"].get("recheck_seed") or int.from_bytes(
        __import__("hashlib").sha256(f"{prod}{len(items)}{rname}".encode()).digest()[:4], "big")
    rate = effective_recheck_rate(cfg)
    rng = random.Random(seed)
    recheck_ids = sorted(rng.sample([i["ID"] for i in items], max(1, int(len(items) * rate))))
    out = DATA / prod / "07_stage2" / f"{prod}_본판정_판정대장_{len(items)}_v1_0.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "판정대장"
    ws.append(["문항ID", "유형", "판정", "판정문(전건 보존)", "재검 대상"])
    cnt = Counter()
    for it in items:
        v = done[it["ID"]]
        cnt[v["판정"]] += 1
        ws.append([it["ID"], it.get("유형", ""), v["판정"], v["판정문"],
                   "○" if it["ID"] in set(recheck_ids) else ""])
    wb.save(out)
    # progress 봉인 — 삭제 금지(증적), 판정문 원본으로 개명 보존
    sealed = prog.with_name(f"{prod}_본판정_판정문원본_{len(items)}건.jsonl")
    prog.rename(sealed)
    ev = {"판정": dict(cnt), "재검": f"{len(recheck_ids)}건 (rate {rate})",
          "seed": seed, "기준서": rname, "산출": N(out.name), "판정문 봉인": N(sealed.name)}
    ledger_append("STAGE2", "STAGE2_JUDGED", "script:judge_run",
                  evidence={**ev, "recheck_ids": recheck_ids}, product=prod)
    return "DONE", ev


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["calibrate", "stage2", "compare"])
    ap.add_argument("--product", required=True)
    a = ap.parse_args()
    if a.cmd == "compare":
        print(run_calibration_compare(a.product, load_config()))
    else:
        fn = run_calibration_judging if a.cmd == "calibrate" else run_stage2
        print(fn(a.product, load_config()))
