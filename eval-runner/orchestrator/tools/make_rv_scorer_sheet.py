#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_rv_scorer_sheet.py — 채점기 RV 이식 (인수인계서 v2 §3-②)

RV 통합 STAGE 대장(806)을 run_score 규격의 '채점용' 시트로 변환한다:
  1. '정답' → '정답 (필수 포함 요소)' 컬럼 개명
  2. 근거 출처에 URL 보강 — 커버리지맵 v1_3 join (Unit ID → source/resource).
     검색축 매칭용. 실측 데이터 결합이지 창작이 아니다.
  3. P-신규-2: 앵커 '(20xx-xx-xx 게시|시행|인덱싱|작성|서술 기준)' 를 정답에서 무조건 분리
     → '앵커(분리보존)' 컬럼에 보존. 삭제된 앵커의 날짜는 수치가 아니다.
산출: data/RV/08_scoring/RV_골든셋_채점용_806문항_v1_0.xlsx (+ 변환 계리 원장 기록)
"""
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from olib import ROOT, N, load_config, ledger_append

DATA = ROOT / "data"
LEDGER_XLSX = DATA / "RV/05_unified_ledger/RV_골든셋_통합STAGE대장_806문항_v1_0.xlsx"
MAP_XLSX = DATA / "RV/03_coverage_map/RV_커버리지맵_코퍼스판_v1_3.xlsx"
OUT = DATA / "RV/08_scoring/RV_골든셋_채점용_806문항_v1_0.xlsx"

ANCHOR = re.compile(r"\(20\d\d-\d\d-\d\d (게시|시행|인덱싱|작성|서술) 기준\)")


def load_unit_urls():
    cfg = load_config()
    prof = cfg["terrain"]["profiles"]["RV"]
    wb = openpyxl.load_workbook(MAP_XLSX, read_only=True, data_only=True)
    urls = {}
    for sn in wb.sheetnames:
        if not N(sn).endswith("커버리지맵"):
            continue
        ws = wb[sn]
        hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if "Unit ID" not in hdr or "source/resource" not in hdr:
            continue
        ui, si = hdr.index("Unit ID"), hdr.index("source/resource")
        for r in ws.iter_rows(min_row=2, values_only=True):
            u, s = N(r[ui]) if ui < len(r) else "", N(r[si]) if si < len(r) else ""
            if u and s.startswith("http"):
                urls[u] = s
    wb.close()
    return urls, re.compile(prof["citation_pattern"])


def main():
    urls, cit_re = load_unit_urls()
    wb = openpyxl.load_workbook(LEDGER_XLSX, data_only=True)
    ws = wb["1_통합대장_806"]
    hdr = [N(c.value) for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    out_wb = openpyxl.Workbook()
    ows = out_wb.active
    ows.title = "채점용_806"
    out_hdr = [("정답 (필수 포함 요소)" if h == "정답" else h) for h in hdr] + ["앵커(분리보존)"]
    ows.append(out_hdr)

    n_url, n_anchor, n_nourl = 0, 0, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [N(v) for v in row[:len(hdr)]]
        if not vals[idx["ID"]]:
            continue
        # P-신규-2: 앵커 무조건 분리 (정답 측)
        ans = vals[idx["정답"]]
        anchors = ANCHOR.findall(ans)
        anchors_full = [m.group(0) for m in ANCHOR.finditer(ans)]
        if anchors_full:
            n_anchor += 1
            ans = ANCHOR.sub("", ans)
        vals[idx["정답"]] = ans
        # 검색축 URL 보강: 근거 출처의 unit_id → 커버리지맵 URL
        src = vals[idx["근거 출처"]]
        units = cit_re.findall(src)
        joined = [urls[u] for u in units if u in urls]
        if joined:
            n_url += 1
            src = src + " ; " + " ; ".join(dict.fromkeys(joined))
        elif vals[idx["유형"]] != "E":
            n_nourl.append(vals[idx["ID"]])
        vals[idx["근거 출처"]] = src
        ows.append(vals + [" | ".join(anchors_full)])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(OUT)
    ev = {"행": ws.max_row - 1, "URL 보강": n_url, "URL 미해결(비E)": len(n_nourl),
          "앵커 분리(P-신규-2)": n_anchor, "산출": N(OUT.name),
          "미해결 예": n_nourl[:5]}
    ledger_append("SCORING", "RV_SCORER_SHEET_BUILT", "script:make_rv_scorer_sheet",
                  evidence=ev, product="RV")
    for k, v in ev.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
