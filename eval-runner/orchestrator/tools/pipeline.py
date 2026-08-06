#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pipeline.py — 오케스트레이터 게이트 CLI + 단계 실행기 (사양서 v1.3 §4·§5·§6·§7)

사용:
  python3 tools/pipeline.py status [--product RV]
  python3 tools/pipeline.py run --product RV          # 멈출 때까지 단계 진행
  python3 tools/pipeline.py approve <gate_id> [--ack-all]
  python3 tools/pipeline.py reject <gate_id> --reason "..."
  python3 tools/pipeline.py resume --after-fix <제품> --reason "..."   # HALT 해제
  python3 tools/pipeline.py appeal <gate_id> --evidence "..."          # 반론 채널
  python3 tools/pipeline.py set-stage --product RV --stage SCORING --reason "..."  # 관리(원장 기록)
  python3 tools/pipeline.py onboard --product RM --name 리모트미팅 [--base blank]

원칙: 반려는 일상, HALT는 사고다. HALT는 승인으로도 못 넘고 resume --after-fix 로만.
"""
import argparse
import glob
import hashlib
import json
import re
import subprocess
import sys
import unicodedata
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import olib
from olib import (N, STAGES, STAGE_KEYS, HUMAN_GATES, ROOT,
                  load_config, load_state, save_state, update_state, set_status,
                  ledger_append, issue_input_card, issue_gate_card,
                  close_gate, find_gate, advance_stage, now)

DATA = ROOT / "data"


def sha256_file(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def files_in(prod, subdir, pattern="*"):
    d = DATA / prod / subdir
    if not d.is_dir():
        return []
    return sorted(p for p in d.glob(pattern) if p.is_file() and not p.name.startswith("."))


def canonical_of(prod, stage_dir, keyword=None):
    """카탈로그에서 해당 단계 정본 파일을 찾는다."""
    cat = json.loads((ROOT / "catalog" / "manifest.json").read_text(encoding="utf-8"))
    best = None
    for r in cat["files"]:
        if r["product"] != prod or r.get("duplicate_of") or not r.get("canonical"):
            continue
        if r["stage_dir"] != stage_dir:
            continue
        if keyword and keyword not in r["file"]:
            continue
        best = r
    return (DATA / best["dest"]) if best and best.get("dest") else None


# ── 단계 실행기: 반환 = ("DONE"|"WAITING_INPUT"|"WAITING_HUMAN"|"HALTED"|"BLOCKED", evidence) ──
def st_corpus_audit(prod, cfg):
    corpus = files_in(prod, "01_corpus_audit") + files_in(prod, "corpus")
    if not corpus:
        issue_input_card(prod, "CORPUS_AUDIT",
                         what=f"{prod} 코퍼스 export (RAG 시스템 인입분 전량)",
                         where=f"data/{prod}/corpus/",
                         fmt="export 파일(xlsx/json/zip). 투입 시: 청크 실측 + match_corpus 중첩 검사 + 해시 매니페스트 등록 후 ① 시작")
        return "WAITING_INPUT", {"corpus_files": 0}
    # 입구 검사: 해시 매니페스트 등록 (규칙 B′의 벽 — 이름이 아니라 내용물)
    mf_path = ROOT / "catalog" / f"manifest_corpus_{prod}.json"
    entries = [{"file": N(p.name), "sha256": sha256_file(p), "size": p.stat().st_size,
                "registered": now()} for p in corpus]
    mf_path.write_text(json.dumps({"product": prod, "entries": entries},
                                  ensure_ascii=False, indent=2), encoding="utf-8")
    # 실측: 읽을 수 있는 청크 수 — "받으면 무조건 실측부터"(영역 0). 파일이 있어도 청크 0이면 통과 금지
    try:
        import gen_coverage
        chunks = gen_coverage.load_corpus(prod)
    except Exception as e:
        ledger_append("CORPUS_AUDIT", "CORPUS_PARSE_FAILED", "script:pipeline",
                      evidence={"err": str(e)[:200]}, product=prod)
        chunks = []
    if not chunks:
        issue_input_card(prod, "CORPUS_AUDIT",
                         what=f"{prod} 코퍼스 — 파일은 있으나 읽을 수 있는 청크 0 (형식 확인 필요)",
                         where=f"data/{prod}/corpus/",
                         fmt="json([{doc,chunk_id,text}] · {chunks:[…]} export) · md/txt · zip(내부 chunks-*.json)")
        return "WAITING_INPUT", {"corpus_files": len(corpus), "청크": 0}
    docs = len({c["doc"] for c in chunks})
    # [모달리티 실측 2026-08-05] 표·이미지·그래프 커버리지 — "시험지가 뭘 못 보는지"를 받자마자 알린다
    # (본부장님 질문 계기: 표/그래프/복합 질문 검증 여부 — 골든셋은 코퍼스에 있는 것만 출제 가능)
    import re as _re
    n_tbl = sum(1 for c in chunks if str(c.get("text", "")).count("|") >= 6)
    n_imgref = sum(1 for c in chunks if _re.search(r"!\[|<img|\.png|\.jpe?g", str(c.get("text", ""))))
    n_chart = sum(1 for c in chunks
                  if len(_re.findall(r"\d+(?:\.\d+)?\s*%", str(c.get("text", "")))) >= 3)
    ev = {"corpus_files": len(corpus), "청크": len(chunks), "문서": docs, "manifest": mf_path.name,
          "모달리티(표 청크)": f"{n_tbl}건 ({n_tbl/len(chunks):.0%})",
          "모달리티(이미지 참조)": f"{n_imgref}건 ({n_imgref/len(chunks):.0%}) — 참조≠내용",
          "모달리티(수치 나열=그래프 서술 추정)": f"{n_chart}건"}
    if n_imgref > len(chunks) * 0.3 and n_chart < len(chunks) * 0.02:
        ev["⚠ 모달리티 경고"] = ("이미지 참조는 많은데 그래프·수치 서술 텍스트가 희박 — "
                              "이미지 안의 내용(그래프·표 스크린샷)은 이 export에 없을 가능성. "
                              "그래프 능력을 시험하려면 이미지→텍스트 처리분 export 필요")
    ledger_append("CORPUS_AUDIT", "MODALITY_AUDIT", "script:pipeline", evidence=ev, product=prod)
    return "DONE", ev


def st_terrain(prod, cfg):
    prof = cfg["terrain"]["profiles"].get(prod, {})
    if prof.get("onboarding"):
        # 초안이 이미 완비면(제품명+citation 패턴) 형식 확인 게이트 — [계속 진행] 단일 버튼
        complete = bool(prof.get("product_name")) and bool(prof.get("citation_pattern"))
        ev = ({"프로파일": "완비된 초안", "제품명": prof.get("product_name"),
               "citation": prof.get("citation_pattern"), "앵커": len(prof.get("anchor_patterns") or [])}
              if complete else
              {"프로파일": "초안(복제/빈)", "확정 항목": "citation_pattern·anchor_patterns·appendix_switches"})
        issue_gate_card(prod, "TERRAIN", f"TERRAIN_{prod}",
                        what_stopped=(f"{prod} 전용 설정 확인 — 제품명·근거 ID 규칙이 이미 등록돼 있어 확인만 하면 됩니다"
                                      if complete else
                                      f"{prod} terrain 프로파일 확정 필요 — citation/앵커 패턴·부록 스위치를 사람이 판정"),
                        evidence=ev,
                        recommendation=("설정 완비 — 계속 진행하면 ③ 앙상블 추출이 시작됩니다"
                                        if complete else
                                        "기존 제품(RV/RC) 프로파일과 코퍼스 실측 결과를 대조해 결정"),
                        simple=complete)
        return "WAITING_HUMAN", {"profile": "onboarding"}
    return "DONE", {"profile": "확정", "citation_pattern": prof.get("citation_pattern", "")}


def _open_gate_ids(prod):
    st = load_state()
    return {g["id"] for g in st["products"][prod]["open_gates"]}


def st_coverage_map(prod, cfg):
    maps = files_in(prod, "03_coverage_map", "*.xlsx")
    st = load_state()
    ps = st["products"][prod]
    if maps and ps.get("expanding"):
        # G19 · 증분 확대: 추가 코퍼스의 새 문서만 보완 추출(맵 v+1) → 새 단위만 출제 대상으로
        import gen_coverage
        import gen_goldenset
        before = {u["unit_id"] for u in gen_goldenset.read_map_units(prod, cfg)[0]}
        try:
            gen_coverage.gapfill(prod)
        except gen_coverage.CoveragePaused as e:
            return "HALTED", {"halt": f"일시 중단(한도 추정) — 증분 추출 체크포인트 보존: {str(e)[:120]}"}
        after = gen_goldenset.read_map_units(prod, cfg)[0]
        new_ids = sorted({u["unit_id"] for u in after} - before)
        st = load_state()
        ps = st["products"][prod]
        ps.pop("expanding", None)
        gs = ps.setdefault("goldenset", {})
        gs["expand_units"] = new_ids
        gs["phase"] = "ROUNDS"
        save_state(st)
        ledger_append("COVERAGE_MAP", "EXPAND_MAP", "script:pipeline",
                      evidence={"신규 단위": len(new_ids), "예": new_ids[:5],
                                "처리": "새 단위만 ④ 차수 출제 — 기존 골든셋·판정 보존"}, product=prod)
        maps = files_in(prod, "03_coverage_map", "*.xlsx")
        return "DONE", {"map": N(maps[-1].name), "증분 신규 단위": len(new_ids)}
    if maps and f"COVMAP_{prod}" not in _open_gate_ids(prod):
        return "DONE", {"map": N(maps[-1].name), "files": len(maps)}
    # 맵이 없으면 생성 엔진 실행 (③ 생성→검수→반려 루프 + ③′ GAP_AUDIT)
    import gen_coverage
    corpus = files_in(prod, "corpus")
    if not corpus:
        issue_input_card(prod, "COVERAGE_MAP",
                         what=f"{prod} 코퍼스 export (커버리지맵 생성 재료)",
                         where=f"data/{prod}/corpus/",
                         fmt="json([{doc,chunk_id,text}]) 또는 md/txt (빈 줄 2개=청크 구분)")
        return "WAITING_INPUT", {}
    return gen_coverage.run(prod, cfg)


def st_goldenset_batch(prod, cfg):
    st = load_state()
    gs = st["products"][prod].get("goldenset", {})
    open_here = {g for g in _open_gate_ids(prod) if g.startswith(("GSPLAN", "GSBATCH", "GSCLOSE"))}
    if gs.get("phase") == "DONE" and not open_here:
        return "DONE", {"goldenset": "마감 완료", "batches": len(files_in(prod, "04_goldenset_batch", "*.xlsx"))}
    # 기존 제품(과거 수동 산출물 보유 + 통합 대장 존재)은 이력 인정
    if files_in(prod, "04_goldenset_batch", "*.xlsx") and files_in(prod, "05_unified_ledger", "*.xlsx") and not gs:
        return "DONE", {"batches": len(files_in(prod, "04_goldenset_batch", "*.xlsx")), "이력": "수동 생산분"}
    import gen_goldenset
    return gen_goldenset.run(prod, cfg)


def st_unified_ledger(prod, cfg):
    files = files_in(prod, "05_unified_ledger", "*.xlsx")
    if not files:
        issue_input_card(prod, "UNIFIED_LEDGER",
                         what=f"{prod} 통합 대장 (전 배치 합본 — ④ 마감이 자동 생성)",
                         where=f"data/{prod}/05_unified_ledger/",
                         fmt="통합 STAGE 대장 xlsx")
        return "WAITING_INPUT", {}
    u = files[-1]
    # 입구 검사: ID 중복 0 (전사 무결 최소선)
    import openpyxl
    wb = openpyxl.load_workbook(u, read_only=True, data_only=True)
    ids = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        try:
            hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        except StopIteration:
            continue
        if "ID" in hdr and "질문" in hdr:
            i = hdr.index("ID")
            for r in ws.iter_rows(min_row=2, values_only=True):
                v = N(r[i]) if i < len(r) else ""
                if v:
                    ids.append(v)
            break
    wb.close()
    dup = len(ids) - len(set(ids))
    if dup:
        return "HALTED", {"ledger": N(u.name), "rows": len(ids), "ID중복": dup,
                          "halt": "참조 집합 훼손 — ID 충돌"}
    return "DONE", {"ledger": N(u.name), "rows": len(ids), "ID중복": 0}


def st_calibration(prod, cfg):
    try:
        import calibration
        return calibration.run(prod, cfg)
    except ImportError:
        pass
    t = canonical_of(prod, "06_calibration", keyword="대조표")
    if not t:
        issue_input_card(prod, "CALIBRATION",
                         what=f"{prod} 캘리브레이션 대조표 (LLM 30 vs 사람 30)",
                         where=f"data/{prod}/06_calibration/",
                         fmt="대조표 xlsx — 문항별 양측 판정 + 불일치 사유 3분류")
        return "WAITING_INPUT", {}
    return "WAITING_HUMAN", {"대조표": N(t.name)}


def st_stage2(prod, cfg):
    st = load_state()
    if not st["products"][prod]["calibration_passed"]:
        return "BLOCKED", {"사유": "규칙 C — calibration_passed=false, 본판정 잠금"}
    v = files_in(prod, "07_stage2", "*판정대장*.xlsx")
    if v:
        return "DONE", {"판정대장": N(v[-1].name)}
    # 판정대장이 없으면 본판정 실행 (⑦ 전건 판정 + 재검 시드 원장 기록)
    import judge_run
    return judge_run.run_stage2(prod, cfg)


def st_scoring(prod, cfg):
    try:
        import scoring
        return scoring.run(prod, cfg)
    except ImportError:
        pass
    logs = files_in(prod, "08_scoring", "*응답로그*.json") + files_in(prod, "logs", "*.json")
    if not logs:
        issue_input_card(prod, "SCORING",
                         what=f"{prod} 응답 로그 (팀장님 시스템 응시 결과)",
                         where=f"data/{prod}/08_scoring/ 또는 data/{prod}/logs/",
                         fmt="json — meta(search_scope) + responses 전량. 투입 시 형식 게이트 자동 실행",
                         extra=f"질문셋 발행본이 이미 있다면 발행 게이트(ID·질문 2컬럼) 통과분인지 확인")
        return "WAITING_INPUT", {"logs": 0}
    return "WAITING_HUMAN", {"logs": len(logs)}


def st_maintenance(prod, cfg):
    issue_gate_card(prod, "MAINTENANCE", f"MAINT_{prod}",
                    what_stopped=f"{prod} 유지보수 큐 판단 — 전제 실효/오탐 되먹임은 사람 판정",
                    evidence={"되먹임 (a)": "전제 실효 후보 → 골든셋 vN+1 개정 큐",
                              "되먹임 (b)": "오탐·측정 불능 → 채점기 개정 큐 → 규칙 D"})
    return "WAITING_HUMAN", {}


RUNNERS = {
    "CORPUS_AUDIT": st_corpus_audit, "TERRAIN": st_terrain, "COVERAGE_MAP": st_coverage_map,
    "GOLDENSET_BATCH": st_goldenset_batch, "UNIFIED_LEDGER": st_unified_ledger,
    "CALIBRATION": st_calibration, "STAGE2": st_stage2, "SCORING": st_scoring,
    "MAINTENANCE": st_maintenance,
}


# ── 명령들 ──────────────────────────────────────────────────
def cmd_status(a):
    st = load_state()
    prods = [a.product] if a.product else list(st["products"])
    print(f"state.json 갱신: {st['updated']}")
    for p in prods:
        ps = st["products"][p]
        label = dict(STAGES).get(ps["stage"], ps["stage"])
        cal = "통과" if ps["calibration_passed"] else "미통과(규칙 C)"
        print(f"\n[{p}] {label} · {ps['status']} · 캘리브레이션 {cal}"
              + (f" · 비교차단(규칙 D)" if ps.get("compare_blocked") else ""))
        if ps.get("halt_reason"):
            print(f"   ⛔ HALT: {ps['halt_reason']}")
        for g in ps["open_gates"]:
            print(f"   ▸ 게이트 열림: {g['id']} (플래그 {len(g.get('flags', []))})")
        hist = ps.get("stage_history", {})
        if hist:
            print(f"   완료: {', '.join(dict(STAGES)[k].split(' ')[0] for k in STAGE_KEYS if k in hist)}")


def cmd_run(a):
    cfg = load_config()
    prod = a.product
    st = load_state()
    if prod not in st["products"]:
        sys.exit(f"미등록 제품: {prod} — onboard 먼저")
    for _ in range(len(STAGE_KEYS) + 1):
        st = load_state()
        ps = st["products"][prod]
        if ps["status"] == "HALTED":
            print(f"⛔ {prod} HALTED — {ps['halt_reason']} · resume --after-fix 로만 해제")
            return
        if ps["status"] == "WAITING_HUMAN":
            print(f"⏸  {prod} {ps['stage']} · WAITING_HUMAN — 카드 확인: 검수큐/")
            return
        # WAITING_INPUT / REJECTED 은 재검사 — 투입물이 들어왔으면 입구 검사 후 전이 (§5′)
        if ps["status"] == "DONE":
            print(f"✅ {prod} 전 단계 완료")
            return
        stage = ps["stage"]
        set_status(prod, "RUNNING", stage=stage)
        try:
            outcome, ev = RUNNERS[stage](prod, cfg)
        except Exception as e:
            # 단계 예외 = RUNNING 잔존 금지 — HALT로 명시 정지 (무인 워커 무한 재시도 방지)
            ledger_append(stage, "STAGE_CRASHED", "script:pipeline",
                          evidence={"err": f"{type(e).__name__}: {str(e)[:200]}"}, product=prod)
            set_status(prod, "HALTED", reason=f"단계 실행 예외 ({stage}): {str(e)[:150]}")
            print(f"⛔ {stage} 실행 예외 → HALT: {e}")
            return
        print(f"  {dict(STAGES)[stage]} → {outcome} {ev}")
        if outcome == "DONE":
            ledger_append(stage, "STAGE_DONE", "script:pipeline", evidence=ev, product=prod)
            # 소비된 INPUT 카드 정리 (입구 검사 통과 = 카드 해소)
            qdir = ROOT / cfg["paths"]["queue"]
            done_dir = qdir / "완료"
            for card in qdir.glob(f"INPUT_*_{prod}.md"):
                if f"INPUT_{stage}_" in card.name or stage in card.name:
                    done_dir.mkdir(exist_ok=True)
                    card.rename(done_dir / card.name)
            nxt = advance_stage(prod)
            if nxt == stage:
                break
        elif outcome == "WAITING_INPUT":
            set_status(prod, "WAITING_INPUT", reason="입력물 부재", evidence=ev)
            return
        elif outcome == "WAITING_HUMAN":
            set_status(prod, "WAITING_HUMAN", evidence=ev)
            return
        elif outcome == "BLOCKED":
            print(f"  🔒 {ev}")
            set_status(prod, "PENDING", reason=str(ev))
            return
        elif outcome == "HALTED":
            set_status(prod, "HALTED", reason=json.dumps(ev, ensure_ascii=False))
            return


def cmd_approve(a):
    prod, ps, g = find_gate(a.gate_id)
    if not g:
        sys.exit(f"열린 게이트 없음: {a.gate_id}")
    flags = g.get("flags", [])
    if flags and not a.ack_all:
        print(f"플래그 {len(flags)}건 — 항목별 ack 필요. 확인했으면 --ack-all 로 승인 (원장에 항목별 기록).")
        for f in flags:
            print(f"  · {f.get('type')} {f.get('id')} {f.get('candidates')}")
        sys.exit(1)
    if g["stage"] == "CALIBRATION" and a.gate_id.startswith("CAL_"):
        # 임계 미달 승인 차단 — 반드시 게이트를 닫기 '전에' 검사 (닫고 나서 거부하면 카드 증발 사고)
        import calibration as _cal
        _t = _cal.find_table(prod)
        _m = _cal.measure_agreement(_t) if _t else {"total": 0}
        _thr = load_config()["pipeline"].get("calibration_threshold", 0.90)
        if not _m["total"] or _m["agreement"] < _thr:
            sys.exit(f"승인 불가 — 일치율 실측 {_m.get('measured_match','?')}/{_m.get('total','?')} 가 "
                     f"임계(≥{_thr:.0%}) 미달이에요. 답안지에서 판정 수정/보완 후 다시 시도하세요. "
                     f"(카드는 그대로 유지됩니다)")
    # 게이트 선점이 먼저 [P0-5] — 대시보드·폰이 동시에 승인해도 한쪽만 통과 (이중 전진·이중 발행 차단)
    if not close_gate(prod, a.gate_id):
        sys.exit(f"이미 처리된 게이트예요: {a.gate_id} — 다른 곳(폰/대시보드)에서 먼저 승인/반려됐어요. 추가 조치 불필요.")
    for f in flags:
        ledger_append(g["stage"], "FLAG_ACK", f"사람:{a.actor}", gate_id=a.gate_id,
                      evidence=f, product=prod)
    ledger_append(g["stage"], "approve", f"사람:{a.actor}", gate_id=a.gate_id,
                  evidence={"flags_acked": len(flags)}, product=prod)
    # 승인 = 게이트만 닫는다. 단계 전진은 러너가 DONE을 보고할 때만 (다중 게이트 단계 대응).
    st = load_state()
    ps = st["products"][prod]
    if ps["stage"] == g["stage"]:
        if g["stage"] == "TERRAIN":
            # 승인 = 지형 확정 — terrain.d 오버레이의 onboarding 플래그 해제
            import yaml as _yaml
            tf = ROOT / "terrain.d" / f"{prod}.yaml"
            if tf.exists():
                d = _yaml.safe_load(tf.read_text(encoding="utf-8")) or {}
                if prod in d:
                    d[prod]["onboarding"] = False
                    tf.write_text(_yaml.safe_dump(d, allow_unicode=True, sort_keys=False),
                                  encoding="utf-8")
        if g["stage"] == "SCORING" and a.gate_id.startswith("SCORE_"):
            # 성적표 확정 = ⑧ 완료 → ⑨ 유지보수로 전진 (재채점 루프 방지)
            ledger_append("SCORING", "SCORE_CONFIRMED", f"사람:{a.actor}",
                          gate_id=a.gate_id, product=prod)
            advance_stage(prod)
            set_status(prod, "PENDING")
            print(f"✅ 성적표 확정 — {a.gate_id} · ⑨ 유지보수로 전진")
            return
        if g["stage"] == "MAINTENANCE":
            update_state(lambda s: s["products"][prod].__setitem__("status", "DONE"))
            ledger_append("MAINTENANCE", "PRODUCT_CYCLE_DONE", f"사람:{a.actor}", product=prod)
            print(f"✅ 승인 — {a.gate_id} · {prod} 사이클 완료")
            return

        def _mut(s):   # [P0-5] 낡은 st 되쓰기 금지 — 잠금 아래 신선 병합
            p = s["products"][prod]
            if g["stage"] == "CALIBRATION" and a.gate_id.startswith("CAL_"):
                p["calibration_passed"] = True   # 임계 통과 게이트 승인 시에만 (입구 실측 검사 통과)
            p["status"] = "PENDING"

        update_state(_mut)
    if a.gate_id.startswith("QAIMP_"):
        # 외부 Q&A 분류 승인 = 별도 트랙 시험지 자동 발행 (편입 후보 0건이면 발행 생략)
        import import_qa
        import_qa.publish(prod, a.actor)
        return
    print(f"✅ 승인 — {a.gate_id} (ack {len(flags)}건 원장 기록). 다음: run --product {prod}")


_UNIT_RE = re.compile(r"[A-Z][A-Z0-9]+(?:-[A-Z0-9가-힣]+)+")


def _gsbatch_partial_remove(prod, gs, lb, ids, reason, st):
    """부분 반려 — 사유에 문항 코드가 있으면 그 문항만 도려내고 나머지는 승인으로 처리.
    2문항 빼자고 75문항 전체를 재출제하는 낭비 방지 (13차 사고에서 학습)."""
    import openpyxl
    d = ROOT / "data" / prod / "04_goldenset_batch"
    f = d / lb["file"]
    if not f.exists():
        return None                      # 실물 없음 — 전체 반려 경로로
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    hdr = [str(c.value) for c in ws[1]]
    cid, csrc = hdr.index("ID") + 1, hdr.index("근거 출처") + 1
    hit = [r for r in range(2, ws.max_row + 1) if str(ws.cell(r, cid).value) in set(ids)]
    if not hit:
        return None                      # 코드가 이 배치에 없음 — 전체 반려 경로로
    for r in reversed(hit):
        ws.delete_rows(r)
    n = ws.max_row - 1
    kept_units = set()
    for r in range(2, ws.max_row + 1):
        kept_units |= set(_UNIT_RE.findall(str(ws.cell(r, csrc).value or "")))
    m = re.match(r"(.+?)(\d+)문항_v(\d+)_(\d+)\.xlsx$", lb["file"])
    new_name = (f"{m.group(1)}{n}문항_v{m.group(3)}_{int(m.group(4)) + 1}.xlsx"
                if m else f"{f.stem}_수리{f.suffix}")
    f.rename(d / f"반려_{f.stem}_부분{f.suffix}")   # 원본 보존 (마감 집계 제외)
    wb.save(d / new_name)
    # 삭제 문항만 인용하던 단위는 재료 풀로 반환 — 이후 차수가 자연 재커버
    returned = set(lb.get("units") or []) - kept_units
    gs["done_units"] = [u for u in gs["done_units"] if u not in returned]
    gs["last_batch"] = {"label": lb["label"], "file": N(new_name),
                        "units": sorted(set(lb.get("units") or []) & kept_units)}
    # 코드 말고도 이유가 적혀 있으면 다음 차수 출제에 피드백으로 반영
    text = _UNIT_RE.sub("", reason).strip(" ,.·-—은는이가도의를:;\n\t")
    if len(text) >= 8:
        gs["reject_feedback"] = reason
    save_state(st)
    ledger_append("GOLDENSET_BATCH", "ITEM_REMOVED", "script:pipeline(부분 반려)",
                  evidence={"삭제 문항": sorted(set(ids)), "잔존": n, "배치": N(new_name),
                            "반환 단위": len(returned), "사유": reason[:150]}, product=prod)
    if returned:
        ledger_append("GOLDENSET_BATCH", "UNITS_RETURNED", "script:pipeline(부분 반려)",
                      evidence={"반환": len(returned), "사유": "삭제 문항 단독 인용 단위"}, product=prod)
    print(f"✂ 부분 반려 — {len(hit)}문항 삭제({', '.join(sorted(set(ids)))}) · 잔존 {n}문항 승인 처리 · "
          f"{len(returned)}단위 반환. 다음 차수로 진행합니다.")
    return "partial"


def _gsbatch_reject_rollback(prod, gate_id, reason):
    """골든셋 배치 반려의 자동 원상복구 — 카드의 '피드백대로 재출제됩니다' 약속의 배관.
    사유에 문항 코드(예: RC2-539)가 있으면 부분 반려(그 문항만 삭제·나머지 승인),
    없으면 전체 반려: ① 단위 반환 ② 파일 반려_ 표시 ③ 피드백 전달 ④ 차수 되돌림."""
    st = load_state()
    gs = st["products"][prod].get("goldenset") or {}
    lb = gs.get("last_batch") or {}
    label = gate_id.rsplit("_", 1)[-1]
    if lb.get("label") != label:
        print(f"(자동 반환 생략 — 마지막 배치 장부와 불일치: {lb.get('label')} ≠ {label} — 사람 확인 필요)")
        return None
    ids = re.findall(rf"{re.escape(prod)}-[A-Za-z]{{0,2}}\d+\b", reason)   # RC2-539 · EE-A01 형식
    if ids:
        mode = _gsbatch_partial_remove(prod, gs, lb, ids, reason, st)
        if mode:
            return mode
        print(f"(사유의 문항 코드 {ids} 가 이 배치에 없음 — 전체 반려로 처리)")
        st = load_state()                # partial 시도가 손댄 적 없지만 명시 재로드
        gs = st["products"][prod].get("goldenset") or {}
        lb = gs.get("last_batch") or {}
    returned = set(lb.get("units") or [])
    gs["done_units"] = [u for u in gs["done_units"] if u not in returned]
    gs["reject_feedback"] = reason          # 재출제 프롬프트에 그대로 주입 (1회 반영 후 소거)
    gs["round"] = max(0, int(gs.get("round", 1)) - 1)   # 파일럿 반려면 다시 파일럿부터
    f = ROOT / "data" / prod / "04_goldenset_batch" / lb.get("file", "")
    if f.exists():
        f.rename(f.parent / f"반려_{f.name}")   # 접두로 마감 집계에서 제외 + 기록은 보존
    save_state(st)
    ledger_append("GOLDENSET_BATCH", "UNITS_RETURNED", "script:pipeline",
                  evidence={"반환": len(returned), "배치": lb.get("file"),
                            "사유": "사람 반려 — 피드백 반영해 재출제 예정"}, product=prod)
    print(f"↩ 배치 반환 — {len(returned)}단위 재료 풀 복귀 · 파일 반려_ 표시 · 피드백은 재출제에 반영")
    return "full"


def cmd_reject(a):
    if not a.reason:
        sys.exit("반려는 사유 필수 — --reason")
    prod, ps, g = find_gate(a.gate_id)
    if not g:
        sys.exit(f"이미 처리됐거나 닫힌 게이트예요: {a.gate_id} — 반려는 대기 중인 카드에서만 가능해요. "
                 f"입력한 사유는 반영되지 않았으니, 반려할 게 있으면 지금 열려 있는 카드에서 다시 눌러 주세요.")
    # 게이트 선점이 먼저 [P0-5] — 승인과 반려가 동시에 오면 한쪽만 통과
    if not close_gate(prod, a.gate_id):
        sys.exit(f"이미 처리된 게이트예요: {a.gate_id} — 다른 곳에서 먼저 승인/반려됐어요. "
                 f"입력한 사유는 반영되지 않았으니, 반려할 게 남았으면 새 카드에서 다시 눌러 주세요.")
    ledger_append(g["stage"], "reject", f"사람:{a.actor}", gate_id=a.gate_id,
                  reason=a.reason, product=prod)
    if a.gate_id.startswith("GSBATCH_"):
        mode = _gsbatch_reject_rollback(prod, a.gate_id, a.reason)
        if mode == "partial":
            # 부분 반려 = 지목 문항만 삭제 + 나머지 승인 — 다음 차수로 바로 진행
            set_status(prod, "PENDING")
            print(f"✂ 부분 반려 접수 — {a.gate_id} · 지목 문항만 빠지고 나머지는 승인 처리됐어요")
            return
        set_status(prod, "REJECTED", reason=a.reason)
        print(f"↩ 반려 접수 — {a.gate_id} · 피드백을 반영해 재출제합니다 (승인·실행 추가로 누를 필요 없음)")
        return
    if a.gate_id.startswith("QAIMP_"):
        # 외부 Q&A 분류 반려 = 이번 분류만 폐기 (원본 Q&A 보존 · 제품 상태는 건드리지 않음
        # — CI처럼 코퍼스 대기 중인 제품이 REJECTED로 오염되는 사고 방지)
        ver = a.gate_id.rsplit("_", 1)[-1]
        f = ROOT / "data" / prod / "external_qa" / f"외부QA_분류결과_{prod}_{ver}.xlsx"
        if f.exists():
            f.rename(f.parent / f"반려_{f.name}")
        print(f"↩ 반려 — {a.gate_id} · 분류결과 폐기(반려_ 표시). 원본 Q&A는 보존 — "
              f"파일을 고쳐 올리거나 [🔁 재대조]를 누르면 새 분류가 나와요.")
        return
    set_status(prod, "REJECTED", reason=a.reason)
    print(f"↩ 반려 — {a.gate_id} · 사유 원장 기록. (반려는 일상 — 수정 재제출 후 run)")


def cmd_expand(a):
    """G19 · 골든셋 증분 확대 — 추가 코퍼스만 추출·출제해 기존 골든셋에 잇는다.
    기존 문항·판정·캘리브레이션 보존, 새 재료만 ③′(보완 추출)→④ 차수→⑤⑦⑧ 재통과."""
    prod = a.product
    st = load_state()
    ps = st["products"].get(prod)
    if not ps:
        sys.exit(f"미등록 제품: {prod}")
    if ps["stage"] not in ("SCORING", "MAINTENANCE"):
        sys.exit(f"증분 확대는 발행 이후(⑧·⑨)에만 가능 — 현재 {ps['stage']}")
    d = ROOT / "data" / prod
    import datetime
    import shutil
    tag = datetime.datetime.now().strftime("%m%d")
    # 재생성/구세대 산출물은 구판/ 폴더로 격리 보존 — 기록은 절대 삭제하지 않는다
    # (응답로그도: 옛 시험지의 응답이라 새 발행본과 형식 게이트가 정당하게 충돌 — 새 로그를 받아야 함)
    for pat, sub_ in (("*판정대장*.xlsx", "07_stage2"), ("*질문셋*발행본*.xlsx", "08_scoring"),
                      ("*응답로그*.json", "08_scoring")):
        old = d / sub_ / "구판"
        for f in (d / sub_).glob(pat):
            old.mkdir(exist_ok=True)
            f.rename(old / f"{tag}_{f.name}")
    # 기존 본판정 봉인을 progress 로 복원 — 새 문항만 추가 판정 (기왕 판정 보존)
    seal = sorted((d / "07_stage2").glob("*본판정_판정문원본*.jsonl"))
    prog = d / "07_stage2" / "_stage2_progress.jsonl"
    if seal and not prog.exists():
        shutil.copy(seal[-1], prog)
    ps["expanding"] = True
    (ps.get("goldenset") or {}).pop("last_batch", None)
    save_state(st)
    set_status(prod, "WAITING_INPUT", stage="COVERAGE_MAP",
               reason="증분 확대 — 추가 코퍼스 투입 대기")
    issue_input_card(prod, "CORPUS_AUDIT",   # 카드 id에 CORPUS → 업로드가 corpus/ 로 (다중 업로드 흐름)
                     what=f"{prod} 추가 코퍼스 (증분 확대 — 기존 골든셋·성적 보존, 새 재료만 출제)",
                     where=f"data/{prod}/corpus/",
                     fmt="export 파일(xlsx/json/zip) — 다 올린 뒤 [다 올렸어요 — 검사 시작]")
    ledger_append("MAINTENANCE", "EXPAND_START", f"사람:{a.actor}",
                  evidence={"보존": "기존 골든셋·본판정·캘리브레이션·성적",
                            "재생성 예정": "판정대장·발행본 (구판 보존 개명)"}, product=prod)
    print(f"➕ 증분 확대 개시 — {prod}: 추가 코퍼스 업로드 → 새 재료만 추출·출제 → 대장·발행본 v+1")


def cmd_resume(a):
    st = load_state()
    prod = a.after_fix
    ps = st["products"].get(prod)
    if not ps or ps["status"] != "HALTED":
        sys.exit(f"{prod} 는 HALTED 상태가 아님")
    if not a.reason:
        sys.exit("HALT 해제는 사유 기록 필수 — --reason")
    ledger_append(ps["stage"], "RESUME_AFTER_FIX", f"사람:{a.actor}",
                  reason=a.reason, evidence={"halt_was": ps["halt_reason"]}, product=prod)
    ps["status"] = "PENDING"
    ps["halt_reason"] = None
    save_state(st)
    print(f"🔓 HALT 해제 — {prod} (원인 제거 확인 + 사유 원장 기록)")


def cmd_appeal(a):
    if not a.evidence:
        sys.exit("반론은 실측 증거 필수 — --evidence")
    ledger_append("APPEAL", "appeal", f"사람:{a.actor}", gate_id=a.gate_id,
                  evidence={"주장": a.evidence}, product=a.product)
    print(f"📮 반론 접수 — {a.gate_id}. 사람 재판정 → 철회 시 원장에 철회+판례화 행 기록 (FEAT-006/074)")


def cmd_set_stage(a):
    """관리용 — 실제 이력을 반영해 위치 보정. 반드시 원장 기록."""
    if not a.reason:
        sys.exit("set-stage 는 사유 필수 — --reason")
    st = load_state()
    ps = st["products"][a.product]
    if a.stage not in STAGE_KEYS:
        sys.exit(f"단계 오류: {a.stage} ∈ {STAGE_KEYS}")
    for k in STAGE_KEYS[:STAGE_KEYS.index(a.stage)]:
        ps["stage_history"].setdefault(k, {"done_at": "(이력 반영)"})
    ps["stage"] = a.stage
    ps["status"] = "PENDING"
    if a.calibration_passed:
        ps["calibration_passed"] = True
    save_state(st)
    ledger_append(a.stage, "SET_STAGE", f"사람:{a.actor}", reason=a.reason,
                  evidence={"calibration_passed": ps["calibration_passed"]}, product=a.product)
    print(f"🧭 {a.product} → {a.stage} (원장 기록)")


STRATEGY_KO = {"ensemble": "앙상블 (3-AI 각자 추출 → 병합)", "solo": "단독 (AI 1개)",
               "cross_check": "생성 + 타 AI 교차 검수", "self_check": "생성 + 자가 검수(새 세션)"}


def cmd_set_members(a):
    """AI 투입 선택 — 연결된 AI 중 누구를 쓸지 (원장 기록)"""
    st = load_state()
    ps = st["products"].get(a.product)
    if not ps:
        sys.exit(f"미등록 제품: {a.product}")
    use = [r.strip() for r in a.use.split(",") if r.strip()]
    prev = ps.get("ai_use")
    ps["ai_use"] = {r: (r in use) for r in ("generator", "judge", "reviewer")}
    save_state(st)
    ledger_append(ps["stage"], "AI_MEMBERS_SET", f"사람:{a.actor}",
                  evidence={"전": prev, "후": ps["ai_use"]}, product=a.product)
    print(f"AI 투입 변경: {a.product} → {ps['ai_use']}")


def cmd_new_round(a):
    """새 회차 채점 — 팀장님 응답로그는 반복해서 온다: ⑨/완료 상태에서 ⑧로 복귀 (원장 기록)"""
    st = load_state()
    ps = st["products"].get(a.product)
    if not ps:
        sys.exit(f"미등록 제품: {a.product}")
    if ps["stage"] not in ("SCORING", "MAINTENANCE"):
        sys.exit("새 회차는 ⑧ 실물 채점 도달 후부터 — 골든셋·질문셋 발행이 먼저다")
    prev = {"stage": ps["stage"], "status": ps["status"]}
    ps["stage"] = "SCORING"
    ps["status"] = "PENDING"
    save_state(st)
    ledger_append("SCORING", "NEW_ROUND_STARTED", f"사람:{a.actor}",
                  evidence={"전": prev}, product=a.product)
    print(f"🔁 새 회차 채점 — {a.product} ⑧ 복귀. 응답로그(json)를 data/{a.product}/08_scoring/ 에 올리고 run")


def cmd_set_strategy(a):
    """생성 전략 변경 — 커버리지맵 추출·골든셋 생성 리뷰 방식 (원장 기록)"""
    st = load_state()
    ps = st["products"].get(a.product)
    if not ps:
        sys.exit(f"미등록 제품: {a.product}")
    prev = ps.get("strategy", "ensemble")
    ps["strategy"] = a.strategy
    save_state(st)
    ledger_append(ps["stage"], "STRATEGY_SET", f"사람:{a.actor}",
                  evidence={"전": prev, "후": a.strategy, "설명": STRATEGY_KO[a.strategy]},
                  product=a.product)
    print(f"전략 변경: {a.product} → {STRATEGY_KO[a.strategy]}")


def cmd_onboard(a):
    """제품 온보딩 — §10.7: 제품은 화면 구조가 아니라 데이터다"""
    prod = a.product.upper()
    cfg = load_config()
    st = load_state()
    if prod in st["products"] and not a.force:
        sys.exit(f"{prod} 이미 존재")
    # ⓐ terrain.profiles.<코드> 자동 생성 — terrain.d/ 오버레이로 (config.yaml 본문 불변)
    if prod not in cfg["terrain"]["profiles"]:
        import yaml as _yaml
        base_prof = dict(cfg["terrain"]["profiles"].get(a.base, {})) if a.base in cfg["terrain"]["profiles"] else {}
        prof = {
            "citation_pattern": base_prof.get("citation_pattern", f"{prod}-[A-Z]+(?:-[A-Z0-9]+)*-\\d+").replace(a.base + "-", f"{prod}-") if base_prof else f"{prod}-[A-Z]+(?:-[A-Z0-9]+)*-\\d+",
            "question_id_pattern": f"{prod}-[A-Z]\\d+",
            "anchor_patterns": base_prof.get("anchor_patterns", []),
            "anchor_suspect_pattern": base_prof.get("anchor_suspect_pattern", "\\(20\\d\\d-\\d\\d-\\d\\d[^)]*\\)"),
            "anchor_required_prefixes": [],
            "appendix_switches": base_prof.get("appendix_switches", []),
            "map_fact_column": "이 단위가 말하는 사실", "map_unit_column": "Unit ID",
            "onboarding": True,   # ② 지형 판정 사람 게이트에서 확정 후 해제
        }
        tdir = ROOT / "terrain.d"
        tdir.mkdir(exist_ok=True)
        (tdir / f"{prod}.yaml").write_text(
            _yaml.safe_dump({prod: prof}, allow_unicode=True, sort_keys=False), encoding="utf-8")
        print(f"terrain.d/{prod}.yaml 생성 (base={a.base}) — ② 지형 판정 게이트에서 확정")
    # ⓑ data 스캐폴드
    for d in ["corpus", "01_corpus_audit", "03_coverage_map", "04_goldenset_batch",
              "05_unified_ledger", "06_calibration", "07_stage2", "08_scoring", "09_maintenance"]:
        (DATA / prod / d).mkdir(parents=True, exist_ok=True)
    # ⓒ 상태 생성 + ⓓ 규칙 C + 실행 전략(기본 앙상블)
    scoring_only = getattr(a, "start", "full") == "scoring"
    st["products"][prod] = {"stage": "SCORING" if scoring_only else "CORPUS_AUDIT",
                            "status": "PENDING",
                            "calibration_passed": False, "scorer_version": None,
                            "compare_blocked": False,
                            "stage_history": ({k: {"done_at": "(채점만 모드 — 보유 산출물 인정)"}
                                               for k in STAGE_KEYS[:STAGE_KEYS.index("SCORING")]}
                                              if scoring_only else {}),
                            "open_gates": [], "halt_reason": None,
                            "strategy": getattr(a, "strategy", "ensemble"),
                            "start_mode": "scoring" if scoring_only else "full"}
    save_state(st)
    # ⓔ 원장 ONBOARD 행
    ledger_append("SCORING" if scoring_only else "CORPUS_AUDIT", "ONBOARD", f"사람:{a.actor}",
                  evidence={"name": a.name or prod, "terrain_base": a.base, "dirs": "created",
                            "start": "채점만" if scoring_only else "처음부터",
                            "strategy": st["products"][prod]["strategy"]},
                  product=prod)
    if scoring_only:
        # 중간 시작: 골든셋(xlsx)과 응답 로그(json)만 받으면 바로 채점
        issue_input_card(prod, "SCORING",
                         what=f"{prod} 골든셋(xlsx)과 응답 로그(json) — 채점만 모드",
                         where=f"골든셋 → data/{prod}/05_unified_ledger/ · 로그 → data/{prod}/08_scoring/",
                         fmt="대시보드 📎 업로드에 두 파일 함께 올려도 됨 — xlsx는 골든셋 자리로, json은 로그 자리로 자동 분류")
        set_status(prod, "WAITING_INPUT", reason="채점만 모드 — 골든셋·로그 대기")
        print(f"🆕 {prod} 온보딩(채점만) — ⑧에서 골든셋·로그 대기")
    else:
        issue_input_card(prod, "CORPUS_AUDIT",
                         what=f"{prod} 코퍼스 export",
                         where=f"data/{prod}/corpus/",
                         fmt="export 파일. 투입 시: 청크 실측 + match_corpus 중첩 + 해시 매니페스트 등록 후 ① 시작")
        set_status(prod, "WAITING_INPUT", reason="온보딩 — 코퍼스 대기")
        print(f"🆕 {prod} 온보딩 — ① WAITING_INPUT 정지 · INPUT_CORPUS 카드 발행 · calibration_passed=false")


def cmd_qa_import(a):
    """G19 · 외부 제작 Q&A 셋 인입 — 코퍼스 대조·3갈래 분류 후 QAIMP 게이트 카드 발행."""
    st = load_state()
    if a.product not in st["products"]:
        sys.exit(f"미등록 제품: {a.product} — 먼저 제품을 추가(온보딩)해 주세요")
    import import_qa
    r = import_qa.run(a.product, a.actor)
    if not r.get("ok"):
        sys.exit(1)   # 화면 토스트가 실패 메시지를 그대로 보여주도록 (성공 오인 방지)


def cmd_qa_score(a):
    """G20 · 외부 Q&A 별도 트랙 채점 — 응답로그(내용 대조 기준) → results/score_<P>QA_rN"""
    import import_qa
    r = import_qa.score(a.product, getattr(a, "log", None), a.actor)
    if not r.get("ok"):
        sys.exit(1)


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    s = sub.add_parser("status"); s.add_argument("--product")
    s = sub.add_parser("run"); s.add_argument("--product", required=True)
    s = sub.add_parser("approve"); s.add_argument("gate_id"); s.add_argument("--ack-all", action="store_true"); s.add_argument("--actor", default="난희")
    s = sub.add_parser("reject"); s.add_argument("gate_id"); s.add_argument("--reason"); s.add_argument("--actor", default="난희")
    s = sub.add_parser("resume"); s.add_argument("--after-fix", required=True, metavar="제품"); s.add_argument("--reason"); s.add_argument("--actor", default="난희")
    s = sub.add_parser("appeal"); s.add_argument("gate_id"); s.add_argument("--evidence"); s.add_argument("--product"); s.add_argument("--actor", default="작업AI")
    s = sub.add_parser("set-stage"); s.add_argument("--product", required=True); s.add_argument("--stage", required=True); s.add_argument("--reason"); s.add_argument("--calibration-passed", action="store_true", dest="calibration_passed"); s.add_argument("--actor", default="난희")
    s = sub.add_parser("onboard"); s.add_argument("--product", required=True); s.add_argument("--name"); s.add_argument("--base", default="blank"); s.add_argument("--force", action="store_true"); s.add_argument("--actor", default="난희"); s.add_argument("--start", default="full", choices=["full", "scoring"]); s.add_argument("--strategy", default="ensemble", choices=["ensemble", "solo", "cross_check", "self_check"])
    s = sub.add_parser("set-strategy"); s.add_argument("--product", required=True); s.add_argument("--strategy", required=True, choices=["ensemble", "solo", "cross_check", "self_check"]); s.add_argument("--actor", default="난희")
    s = sub.add_parser("set-members"); s.add_argument("--product", required=True); s.add_argument("--use", required=True, help="쉼표 구분: generator,judge,reviewer"); s.add_argument("--actor", default="난희")
    s = sub.add_parser("new-round"); s.add_argument("--product", required=True); s.add_argument("--actor", default="난희")
    s = sub.add_parser("expand"); s.add_argument("--product", required=True); s.add_argument("--actor", default="난희")
    s = sub.add_parser("qa-import"); s.add_argument("--product", required=True); s.add_argument("--actor", default="난희")
    s = sub.add_parser("qa-score"); s.add_argument("--product", required=True); s.add_argument("--log"); s.add_argument("--actor", default="난희")
    a = ap.parse_args()
    {"status": cmd_status, "run": cmd_run, "approve": cmd_approve, "reject": cmd_reject,
     "resume": cmd_resume, "appeal": cmd_appeal, "set-stage": cmd_set_stage,
     "onboard": cmd_onboard, "set-strategy": cmd_set_strategy,
     "set-members": cmd_set_members, "new-round": cmd_new_round, "expand": cmd_expand,
     "qa-import": cmd_qa_import, "qa-score": cmd_qa_score}[a.cmd](a)


if __name__ == "__main__":
    main()
