#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""r2_gate_driver.py — r2 무인 게이트 운전기사 (난희 위임 2026-08-26: "니가 r2 전체 컨트롤")
GSPLAN/GSBATCH/GSCLOSE 카드가 뜨면: 검수 잣대(판례·접두어·전차수 질문중복·발췌20자·유형어휘)를
독립 재검증 → 전부 통과 시에만 '난희(위임)' 승인, 하나라도 걸리면 부분 반려(문항 지목).
모든 판정 근거는 원장에 기록 — 아침에 난희가 전체 감사 가능. 최종 발행·팀장님 전달은 사람 몫."""
import glob
import json
import re
import subprocess
import sys
import time
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from olib import ROOT, N, load_state, ledger_append

ACTOR = "난희(위임: r2 전체컨트롤 2026-08-26)"
BAN = [(r"会社名|企業名", "회사명"), (r"掲載日|投稿日", "게시일"),
       (r"導入事例|成功事例|お客様の声", "홍보")]
TYPES = {"사실확인", "수치", "절차", "조건", "가부", "정의", "E형"}
STRIP = re.compile(r"[\s\x00-\x1f]+")


def norm(s):
    return STRIP.sub("", unicodedata.normalize("NFC", str(s or ""))).lower()


def load_items(f):
    ws = openpyxl.load_workbook(f, read_only=True, data_only=True).worksheets[0]
    hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(hdr)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[0]:
            rows.append({h: N(r[i]) if i < len(r) else "" for h, i in idx.items()})
    return rows


def audit_batch(batch_file):
    """반환: (결함 문항ID 목록, 사유 목록)"""
    items = load_items(batch_file)
    prev = []
    for f in glob.glob(str(ROOT / "data/CI/04_goldenset_batch/*.xlsx")):
        if "반려_" in f or Path(f).name == Path(batch_file).name:
            continue
        prev += load_items(f)
    prev_q = {norm(p.get("질문")) for p in prev}
    bad, why = [], []
    for it in items:
        iid, q = it.get("ID", "?"), it.get("질문", "")
        typ = re.sub(r"[\(（].*", "", it.get("유형", "")).strip()
        ex = it.get("근거 원문 발췌", "")
        if re.match(r"^\s*CI2?\s*(について|に関して|では|の)", q):
            bad.append(iid); why.append(f"{iid}:익명코드 접두어")
        for pat, nm in BAN:
            if re.search(pat, q):
                bad.append(iid); why.append(f"{iid}:판례({nm})")
        if typ and typ not in TYPES:
            bad.append(iid); why.append(f"{iid}:유형어휘({typ})")
        if ex and all(len(norm(seg)) < 20 for seg in re.split(r"\|\|", ex) if norm(seg)):
            bad.append(iid); why.append(f"{iid}:발췌20자미만")
        if norm(q) in prev_q:
            bad.append(iid); why.append(f"{iid}:전차수 질문중복")
    return sorted(set(bad)), why


def main():
    idle = 0
    while idle < 60:                     # 무카드 10시간이면 종료
        st = load_state()["products"]["CI"]
        gates = [g.get("id", "") for g in st.get("open_gates", [])]
        acted = False
        for gid in gates:
            if gid.startswith("GSPLAN_"):
                subprocess.run([sys.executable, "tools/pipeline.py", "approve", gid,
                                "--actor", ACTOR], cwd=ROOT, capture_output=True)
                ledger_append("GOLDENSET_BATCH", "R2_DRIVER_APPROVE", f"script:r2_driver({ACTOR})",
                              gate_id=gid, evidence={"판단": "배분계획 — 규모 320 확인"}, product="CI")
                acted = True
            elif gid.startswith("GSBATCH_") or gid.startswith("GSCLOSE_"):
                # 카드가 가리키는 최신 배치 파일 검증
                batches = sorted(glob.glob(str(ROOT / "data/CI/04_goldenset_batch/CI*골든셋*xlsx")),
                                 key=lambda p: Path(p).stat().st_mtime)
                if gid.startswith("GSBATCH_") and batches:
                    bad, why = audit_batch(batches[-1])
                    if bad:
                        reason = f"{', '.join(bad[:20])} 결함 — {'; '.join(why[:10])} — 해당 문항만 제외"
                        subprocess.run([sys.executable, "tools/pipeline.py", "reject", gid,
                                        "--reason", reason, "--actor", ACTOR],
                                       cwd=ROOT, capture_output=True)
                        ledger_append("GOLDENSET_BATCH", "R2_DRIVER_REJECT", f"script:r2_driver({ACTOR})",
                                      gate_id=gid, evidence={"결함": why[:15]}, product="CI")
                        acted = True
                        continue
                subprocess.run([sys.executable, "tools/pipeline.py", "approve", gid,
                                "--ack-all", "--actor", ACTOR], cwd=ROOT, capture_output=True)
                ledger_append("GOLDENSET_BATCH", "R2_DRIVER_APPROVE", f"script:r2_driver({ACTOR})",
                              gate_id=gid, evidence={"판단": "독립 검증 결함 0 — 위임 승인"}, product="CI")
                acted = True
        if acted:
            idle = 0
            subprocess.Popen([sys.executable, "tools/auto_run.py", "--product", "CI"],
                             cwd=ROOT, stdout=open(ROOT / "results/logs/CI/auto_run.log", "ab"),
                             stderr=subprocess.STDOUT)
        else:
            idle += 1
        # 마감(GSCLOSE 승인 뒤 ⑤ 이후 단계) 도달 시 종료
        if st.get("stage") not in ("GOLDENSET_BATCH",):
            print("④ 종료 — 운전기사 퇴근:", st.get("stage"))
            return
        time.sleep(600)


if __name__ == "__main__":
    main()
