#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
regression_g1.py — verify_batch.py §9 회귀 테스트 (사양서 v1.3 §9·§9′)

회귀 원칙: 과거 합격 배치는 PASS, 과거 반려 사유는 FAIL로 재현돼야 한다.

[§9′ 대체 명기 — 각주]
* B2 57문항 "원판(반려당한 판)"은 미확보(P-09 정본 1파일 유지로 유실 추정).
  → 합격판 v1_0을 변조한 모의 파일로 반려 사유를 재현한다 (아래 T3~T7).
  테스트를 조용히 빼지 않고 대체 사실을 여기 명기한다.
* 80vs79 계리(B6 7차) → 행 1개 삭제 변조로 '선언 vs 실측 불일치' 유형 재현 (T3).
* RVB-G07 앵커 변형 → 합격판 B6 답변의 정상 앵커를 '(…게시)' 변형으로 치환 (T7).
  기대: FAIL 아님, '앵커 변형 의심' 플래그 (P-신규-1).
* RC 4차 오탐 3건(거절 패턴 노화)은 run_score 소관 — G5 회귀에서 다룬다.
* RVT-A22(앵커 삭제 시 날짜 미집계)는 judge/채점 소관 — G4/G5 회귀에서 다룬다.
"""
import re
import shutil
import subprocess
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
DATA = ROOT / "data"
TMP = ROOT / "results" / "_regression_tmp"
VB = ROOT / "tools" / "verify_batch.py"

B2 = DATA / "RV/04_goldenset_batch/RV_골든셋_B2_2차_마감_57문항_v1_0.xlsx"
B2_PILOT = DATA / "RV/04_goldenset_batch/RV_골든셋_B2_웹_파일럿_30문항_v1_0.xlsx"
B6_7 = DATA / "RV/04_goldenset_batch/RV_골든셋_B6_7차_서사구간_14문항_v1_0.xlsx"
B6_3 = DATA / "RV/04_goldenset_batch/RV_골든셋_B6_3차_시사구간_35문항_v1_1.xlsx"
MAP = DATA / "RV/03_coverage_map/RV_커버리지맵_코퍼스판_v1_3.xlsx"


def N(s):
    return unicodedata.normalize("NFC", str(s)) if s is not None else ""


def run_vb(batch, union=None, extra=None):
    cmd = [sys.executable, str(VB), "--batch", str(batch), "--map", str(MAP), "--product", "RV"]
    if union is not None:
        cmd += ["--union"] + [str(u) for u in union]
    if extra:
        cmd += extra
    p = subprocess.run(cmd, capture_output=True, text=True)
    return p.returncode, p.stdout + p.stderr


def tamper(src, name, fn):
    """합격판 사본을 변조해 모의 반려판 생성 (§9′ 대체 명기 규칙)"""
    dst = TMP / name
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    fn(wb)
    wb.save(dst)
    return dst


def qsheet(wb):
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdr = [N(c.value) for c in ws[1]]
        if {"ID", "질문", "근거 원문 발췌"} <= set(h for h in hdr if h):
            return ws, {h: i + 1 for i, h in enumerate(hdr)}
    raise RuntimeError("문항 시트 없음")


results = []


def check(tid, desc, expect_exit, got_exit, out, expect_in=None):
    ok = got_exit == expect_exit and (expect_in is None or expect_in in out)
    results.append((tid, desc, expect_exit, got_exit, ok))
    mark = "✅" if ok else "❌"
    print(f"{mark} {tid} {desc} — 기대 exit {expect_exit} / 실제 {got_exit}")
    if not ok:
        print(out[-1200:])


def main():
    TMP.mkdir(parents=True, exist_ok=True)

    # ── T1: 과거 합격 배치 PASS (B2 2차 마감 + 합집합)
    rc, out = run_vb(B2, union=[B2_PILOT])
    check("T1", "합격 배치 B2 2차(57) → PASS", 0, rc, out)

    # ── T2: 과거 합격 배치 PASS (B6 3차 v1_1 — 앵커 의무 RVB)
    rc, out = run_vb(B6_3, union=[])
    check("T2", "합격 배치 B6 3차 v1_1(35, 앵커 의무) → PASS", 0, rc, out)

    # ── T3: 80vs79 계리 유형 — 행 1개 삭제 → 선언 35 vs 실측 34 (① FAIL)
    def del_row(wb):
        ws, idx = qsheet(wb)
        ws.delete_rows(3)
    t = tamper(B6_3, "T3_계리불일치_35문항_v1_0.xlsx", del_row)  # 선언 35 유지 + 행 삭제 → 34
    rc, out = run_vb(t, union=[])
    check("T3", "계리 불일치(행 삭제 — 80vs79 유형) → REJECTED", 1, rc, out, "①")

    # ── T4: 1축 불일치 — 발췌 1자 변조 (② FAIL)
    def tamper_excerpt(wb):
        ws, idx = qsheet(wb)
        c = ws.cell(row=2, column=idx["근거 원문 발췌"])
        c.value = N(c.value) + "변조문자열추가됨"
    t = tamper(B6_3, "T4_발췌변조_1축불일치.xlsx", tamper_excerpt)
    rc, out = run_vb(t, union=[])
    check("T4", "1축 문자 불일치(발췌 변조) → REJECTED", 1, rc, out, "②")

    # ── T5: citation 위조 (③ FAIL)
    def fake_cit(wb):
        ws, idx = qsheet(wb)
        ws.cell(row=2, column=idx["근거 출처"]).value = "RV-FAKE-UNIT-999 ; 위조"
    t = tamper(B6_3, "T5_citation위조.xlsx", fake_cit)
    rc, out = run_vb(t, union=[])
    check("T5", "citation 맵 부재(위조) → REJECTED", 1, rc, out, "③")

    # ── T6: 질문 중복 (④ FAIL)
    def dup_q(wb):
        ws, idx = qsheet(wb)
        ws.cell(row=3, column=idx["질문"]).value = ws.cell(row=2, column=idx["질문"]).value
    t = tamper(B6_3, "T6_질문중복.xlsx", dup_q)
    rc, out = run_vb(t, union=[])
    check("T6", "질문 중복 → REJECTED", 1, rc, out, "④")

    # ── T7: RVB-G07 — 앵커 변형은 기각이 아니라 플래그 (P-신규-1)
    def variant_anchor(wb):
        ws, idx = qsheet(wb)
        c = ws.cell(row=2, column=idx["정답"])
        c.value = re.sub(r"\((20\d\d-\d\d-\d\d) (게시|시행|인덱싱|작성|서술) 기준\)",
                         r"(\1 게시)", N(c.value), count=1)
    t = tamper(B6_3, "T7_앵커변형_G07.xlsx", variant_anchor)
    rc, out = run_vb(t, union=[])
    check("T7", "앵커 변형 '(…게시)' → PASS + 플래그 (FAIL 금지)", 0, rc, out, "앵커 변형 의심")

    # ── T8: 커버 등식 소실 → HALT (등식 시트 '풀' 값 부풀리기)
    def lose_units(wb):
        for sn in wb.sheetnames:
            if "커버등식" in N(sn) or "등식" in N(sn):
                ws = wb[sn]
                for row in ws.iter_rows(min_row=2):
                    if row and "풀" in N(row[0].value):
                        row[1].value = int(row[1].value) + 5   # 풀 5 부풀림 = 소실 5
                return
    t = tamper(B2, "T8_등식소실_HALT.xlsx", lose_units)
    rc, out = run_vb(t, union=[B2_PILOT])
    check("T8", "커버 등식 소실 5 → HALT (반려 아님 — 사고)", 3, rc, out, "소실")

    # ── T9: union 인자 누락 시 실행 거부 (합집합 원칙 — FAIL 아닌 사용 오류)
    rc, out = run_vb(B2, union=None)
    check("T9", "흡수 대장 有 + --union 누락 → 실행 거부 exit 2", 2, rc, out, "합집합 원칙")

    # ── 결과 요약
    npass = sum(1 for r in results if r[4])
    print(f"\n{'='*54}\n회귀 결과: {npass}/{len(results)} 통과")
    if npass < len(results):
        sys.exit(1)


if __name__ == "__main__":
    main()
