#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
regression_g19.py — 외부 Q&A 트랙(G19/G20) 회귀 (전수검수 P2: 커버리지 공백 해소)

Q1 파싱: xlsx/csv/json 혼합 + cp949 CSV 폴백 + 손상 파일 격리(전체 크래시 금지)
Q2 분류: 3갈래(편입/E형/부분) + 코퍼스 없으면 전량 보류 + 중복 질문 표시
Q3 발행: 버전 = 최대+1, 문항 ID에 Qv{버전} 내장, 편입 후보만 편입
Q4 채점: 로그-정답키 버전 바인딩(v1 로그→v1 키), 미존재 버전 친절 거절,
         같은 로그 재채점 차단(로그지문), 회차 번호 최대+1
Q5 최신 선택: 문자열 정렬 함정(2문항_v1 vs 1문항_v2)에서 버전 숫자 기준 승리

원장 오염 금지: ledger_append를 no-op으로 대체(모의), 산출물은 finally 전량 정리.
"""
import json
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "tools"))
import import_qa as iq   # noqa: E402

iq.ledger_append = lambda *a, **k: None   # 회귀는 실전 원장에 기록하지 않는다 [P1-5 원칙]

PROD = "EE"
D = ROOT / "data" / PROD / "external_qa"
results = []


def check(tid, desc, ok, detail=""):
    results.append((tid, ok))
    print(f"{'✅' if ok else '❌'} {tid} {desc}" + (f" — {detail}" if detail else ""))


def q1_parse():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["질문", "답변"])
    ws.append(["엑셀 질문?", "엑셀 답변입니다 길이확보용 문장"])
    wb.save(D / "src.xlsx")
    (D / "src_cp949.csv").write_bytes("question,answer\ncp949 질문?,cp949 답변입니다 길이확보\n".encode("cp949"))
    (D / "깨진.xlsx").write_bytes(b"NOT-A-ZIP")   # 손상 파일 — 전체 크래시 금지, skipped 목록으로
    items, skipped = iq.parse_qa_files(D)
    check("Q1", "혼합 파싱 + cp949 폴백 + 손상 파일 격리",
          len(items) == 2 and "깨진.xlsx" in skipped,
          f"items={len(items)} skipped={skipped}")


def q2_classify():
    items = [{"질문": "있는 것?", "답변": "코퍼스에 실재하는 문장입니다"},
             {"질문": "없는 것?", "답변": "코퍼스 어디에도 없는 내용입니다"},
             {"질문": "있는 것?", "답변": "코퍼스에 실재하는 문장입니다"}]
    corpus = iq._norm("안내: 코퍼스에 실재하는 문장입니다. 끝.")
    r = iq.classify([dict(x) for x in items], corpus)
    ok = (r[0]["분류"].startswith("편입 후보") and r[1]["분류"].startswith("E형 후보")
          and "중복" in r[2]["비고"])
    hold = iq.classify([dict(x) for x in items], "")
    check("Q2", "3갈래 분류 + 중복 표시 + 무코퍼스 전량 보류",
          ok and all(x["분류"].startswith("보류") for x in hold))


def q3_publish():
    def mk_report(ver, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["번호", "출처 파일", "질문", "답변", "분류", "일치율", "확인 발췌", "비고"])
        for i, (q, a, cls) in enumerate(rows, 1):
            ws.append([i, "src.xlsx", q, a, cls, "1/1", a[:20], ""])
        wb.save(D / f"외부QA_분류결과_{PROD}_v{ver}.xlsx")
    mk_report(1, [("질문A?", "답변A 길이확보용 문장입니다", "편입 후보(근거 실재)"),
                  ("질문A2?", "답변A2 길이확보용 문장입니다", "편입 후보(근거 실재)"),
                  ("질문B?", "답변B 길이확보용 문장입니다", "E형 후보(코퍼스 부재)")])
    p1 = iq.publish(PROD, "회귀")   # 2문항_v1 — Q5의 문자 정렬 함정 재료
    mk_report(2, [("질문C?", "답변C 길이확보용 문장입니다", "편입 후보(근거 실재)")])
    p2 = iq.publish(PROD, "회귀")   # 1문항_v2
    ws = openpyxl.load_workbook(D / p1, read_only=True).active
    ids = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)]
    check("Q3", "발행 버전 최대+1 + Qv{버전} ID + 편입 후보만",
          p1.endswith("v1.xlsx") and p2.endswith("v2.xlsx")
          and ids == [f"{PROD}-Qv1-001", f"{PROD}-Qv1-002"] and "2문항" in p1,
          f"{p1} ids={ids}")


def q4_score():
    ld = D / "로그"
    ld.mkdir(exist_ok=True)
    log1 = {"responses": [{"id": f"{PROD}-Qv1-001",
                           "hits": [{"rank": 1, "content": "안내: 답변A 길이확보용 문장입니다"}],
                           "answer": None},
                          {"id": f"{PROD}-Qv1-002",
                           "hits": [{"rank": 1, "content": "무관한 문서입니다"},
                                    {"rank": 2, "content": "부록: 답변A2 길이확보용 문장입니다"}],
                           "answer": None}]}   # 목록 위치=순위 (본선 형식 게이트 규격과 동일)
    (ld / "회귀_log1.json").write_text(json.dumps(log1, ensure_ascii=False), encoding="utf-8")
    r1 = iq.score(PROD, "회귀_log1.json", "회귀")
    r_dup = iq.score(PROD, "회귀_log1.json", "회귀")
    log9 = {"responses": [{"id": f"{PROD}-Qv9-001", "hits": [], "answer": None}]}
    (ld / "회귀_log9.json").write_text(json.dumps(log9, ensure_ascii=False), encoding="utf-8")
    r9 = iq.score(PROD, "회귀_log9.json", "회귀")
    check("Q4", "버전 바인딩 채점 + 재채점 차단 + 미존재 버전 거절",
          r1["ok"] and "top1 1/2" in r1["out"] and "top5 2/2" in r1["out"]
          and (not r_dup["ok"]) and "이미 채점" in r_dup["out"]
          and (not r9["ok"]) and "버전 정답키가 없" in r9["out"])


def q5_latest():
    papers = list(D.glob(f"외부QA_시험지_{PROD}_*문항_v*.xlsx"))
    latest = max(papers, key=lambda p: iq._ver_of(p.name))
    lexi = sorted(papers)[-1]
    check("Q5", "최신 시험지 = 버전 숫자 기준 (문자 정렬은 실제로 구판을 집는 함정)",
          latest.name.endswith("v2.xlsx") and lexi.name.endswith("v1.xlsx"),
          f"버전기준={latest.name} / 문자정렬={lexi.name}")


def main():
    if D.parent.exists():
        shutil.rmtree(D.parent)
    D.mkdir(parents=True)
    try:
        q1_parse()
        q2_classify()
        q3_publish()
        q4_score()
        q5_latest()
    finally:
        shutil.rmtree(ROOT / "data" / PROD, ignore_errors=True)
        for f in (ROOT / "results").glob(f"score_{PROD}QA_r*"):
            shutil.rmtree(f, ignore_errors=True)
    n_ok = sum(1 for _, ok in results if ok)
    print("=" * 54)
    print(f"G19 회귀: {n_ok}/{len(results)} 통과")
    sys.exit(0 if n_ok == len(results) else 1)


if __name__ == "__main__":
    main()
