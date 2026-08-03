#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""report_gen.py — 회차 평가 리포트 초안 자동 생성 (성적리포트 규격 v1.0 준용)

채점 직후 자동 호출 — 엄격/관대 병기, 정답 소스군별 miss 분해, 출처 결측 진단,
검색축만/전체 회차 구분. 산출: results/score_<prod>_<r>/<prod>_평가리포트_<r>_초안_v0_9.md
(확정은 사람: 발주자 검증 후 외부 전달 — 규격 §A-3)
"""
import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from olib import ROOT, N


def _cause_section(prod, rep, log_path, n):
    """§3 저조 원인 분석 (표준 섹션) — 억울한 miss 분해 + 합집합 지표"""
    ca = content_analysis(prod, rep, log_path)
    if not ca:
        return "## 3. 저조 원인 분석\n(내용 기준 재채점 불가 — 코퍼스/발췌 미확보)"
    return f"""## 3. 저조 원인 분석 — 점수가 왜 이렇게 나왔나 (전건 실측·매 회차 표준)

검증: 시스템이 가져온 청크 **본문**에 골든셋 '근거 원문 발췌'가 실재하는지 {ca['n']}건 전건 대조.

| 기준 | top1 | top5 |
|---|---|---|
| 공식 (URL 일치) | {summ_cell(rep,'hit_top1',n)} | {summ_cell(rep,'top5',n)} |
| **합집합 (URL 또는 내용 실재 — v12 병기)** | **{ca['u1']} ({ca['u1']/ca['n']:.0%})** | **{ca['u5']} ({ca['u5']/ca['n']:.0%})** |

### 공식 miss {ca['miss']}건의 속사정
| 분류 | 건수 | 뜻 |
|---|---|---|
| 내용은 맞는데 miss ① 출처 URL 결측 | {ca['rescue_nourl']} | 골든셋 데이터 문제 — 보정 후 소급 재채점 대상 |
| 내용은 맞는데 miss ② 다른 문서에 같은 내용 | {ca['rescue_alt']} | 모호 질문/중복 문서 — 채점 기준의 한계 (v12가 구제) |
| **진짜 검색 실패** (top5 어디에도 정답 내용 없음) | **{ca['true_fail']}** | 검색기/인입 개선 대상 — §2 소스군 분해 참조 |

→ 요약: 낮은 점수의 {ca['rescued']/max(1,ca['miss']):.0%}는 채점·데이터 요인, {ca['true_fail']/max(1,ca['miss']):.0%}는 실제 검색 실패."""


def summ_cell(rep, kind, n):
    t1 = sum(1 for r in rep if r.get("검색") == "hit_top1")
    v = t1 if kind == "hit_top1" else t1 + sum(1 for r in rep if r.get("검색") == "hit_top5")
    return f"{v} ({v/max(1,n):.0%})"


def _golden(prod):
    d = sorted((ROOT / "data" / prod / "05_unified_ledger").glob("*통합대장*.xlsx"))
    return d[-1] if d else None


def _norm(s):
    return re.sub(r"[\s\x00-\x1f]+", "", str(s or "")).lower()


def content_analysis(prod, rep, log_path):
    """왜 낮은가 — 내용 기준 재채점: 시스템이 가져온 청크 본문에 정답 원문이 실재하나.
    억울한 miss(내용은 맞는데 URL 불일치/결측) 분해 — 보고서 표준 섹션 재료."""
    import zipfile
    g = _golden(prod)
    if not g or not log_path:
        return None
    ws = openpyxl.load_workbook(g, read_only=True, data_only=True).active
    hdr = [N(c) for c in next(ws.iter_rows(max_row=1, values_only=True))]
    if "근거 원문 발췌" not in hdr:
        return None
    ei = hdr.index("근거 원문 발췌")
    ui = next((i for i, h in enumerate(hdr) if h.startswith("근거 출처")), None)
    segs_of, url_of = {}, {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        segs = [re.sub(r"^\[[^\]]*\]\s*", "", s).strip("\"'“”‘’ ")
                for s in re.split(r"\|\|", N(r[ei]))]
        segs_of[N(r[0])] = [_norm(s) for s in segs if len(s) >= 10]
        url_of[N(r[0])] = N(r[ui]) if ui is not None else ""
    ctext = {}
    corpus_dir = ROOT / "data" / prod / "corpus"
    for zp in sorted(corpus_dir.glob("*.zip*")) if corpus_dir.is_dir() else []:
        try:
            with zipfile.ZipFile(zp) as z:
                for name in z.namelist():
                    if name.startswith("chunks") and name.endswith(".json"):
                        d = json.load(z.open(name))
                        for c in d.get("chunks", []):
                            if c.get("point_id"):
                                ctext[c["point_id"]] = _norm(c.get("text", ""))
        except Exception:
            continue
    log = json.loads(Path(log_path).read_text(encoding="utf-8"))
    rmap = {N(r.get("id", "")): r for r in rep}
    out = {"n": 0, "u1": 0, "u5": 0, "miss": 0, "rescued": 0, "rescue_nourl": 0,
           "rescue_alt": 0, "true_fail": 0}
    for resp in log.get("responses", []):
        iid = N(resp.get("id", ""))
        segs = segs_of.get(iid, [])
        if not segs:
            continue
        out["n"] += 1
        texts = [ctext.get(h.get("chunk_id")) or _norm(h.get("snippet", ""))
                 for h in (resp.get("hits") or [])[:5]]
        rank = next((i for i, t in enumerate(texts) if any(s in t for s in segs)), None)
        v = rmap.get(iid, {}).get("검색")
        if v == "hit_top1" or rank == 0:
            out["u1"] += 1
        if v in ("hit_top1", "hit_top5") or rank is not None:
            out["u5"] += 1
        if v == "miss":
            out["miss"] += 1
            if rank is not None:
                out["rescued"] += 1
                if "http" not in url_of.get(iid, ""):
                    out["rescue_nourl"] += 1
                else:
                    out["rescue_alt"] += 1
            else:
                out["true_fail"] += 1
    return out if out["n"] else None


def draft(prod, round_label, rep, summ, search_only, sys_ver="", log_path=None):
    out_dir = ROOT / "results" / f"score_{prod}_{round_label}"
    out = out_dir / f"{prod}_평가리포트_{round_label}_초안_v0_9.md"
    n = len(rep)
    top1 = sum(1 for r in rep if r.get("검색") == "hit_top1")
    top5 = top1 + sum(1 for r in rep if r.get("검색") == "hit_top5")
    # 정답 출처 조인 — 소스군 분해 + URL 결측(채점 불가) 진단
    rmap = {N(r.get("id", "")): r for r in rep}
    sect, sect_miss, no_url, no_url_miss = Counter(), Counter(), 0, 0
    g = _golden(prod)
    if g:
        gw = openpyxl.load_workbook(g, read_only=True, data_only=True).active
        gh = [N(c) for c in next(gw.iter_rows(max_row=1, values_only=True))]
        si = next((i for i, h in enumerate(gh) if h.startswith("근거 출처")), None)
        for row in gw.iter_rows(min_row=2, values_only=True):
            if not row[0] or si is None:
                continue
            src = N(row[si])
            v = rmap.get(N(row[0]), {}).get("검색", "?")
            m = re.search(r"https?://([^/]+)(?:/(?:kr|ko)/([^/]+))?", src)
            if not m:
                no_url += 1
                no_url_miss += v == "miss"
                continue
            key = m.group(2) or m.group(1)
            sect[key] += 1
            if v == "miss":
                sect_miss[key] += 1
    scoreable = n - no_url
    lens = (f"| top1 (관대*) | **{top1} / {scoreable} = {top1/scoreable:.1%}** |\n"
            f"| top5 (관대*) | **{top5} / {scoreable} = {top5/scoreable:.1%}** |\n"
            if no_url else "")
    sect_rows = "\n".join(f"| {k} | {v} | {sect_miss[k]} | {sect_miss[k]/v:.0%} |"
                          for k, v in sect.most_common(12))
    dead = [k for k, v in sect.items() if v >= 5 and sect_miss[k] == v]
    scope = "검색축만 (answer 미제출 — 생성축·E형 미응시)" if search_only else "전체 (검색+생성)"
    gen_row = ("| 생성/E형 | — 미응시 | 본 회차 answer 미제출 |" if search_only else
               f"| 생성 | 실측 | pass {summ.get('pass')} · partial {summ.get('partial')} |")
    body = f"""# {prod} RAG 평가 리포트 — {round_label} (초안 v0.9 · 자동 생성)

> ⚠ 초안 — 발주자(난희) 검증 전 외부 전달 금지 (규격 v1.0 §A-3)

## 0. 성적 요약
| 항목 | 값 |
|---|---|
| 회차 | {round_label} · 응시 범위: {scope} |
| 시스템 | {sys_ver or '(로그 meta 참조)'} |
| 모수 | {n}문항 전량 · 형식 게이트 PASS |
| top1 (엄격) | **{top1} / {n} = {top1/n:.1%}** |
| top5 (엄격) | **{top5} / {n} = {top5/n:.1%}** |
{lens}
{'* 관대 = 채점 가능 문항만 (정답 출처 URL 결측 ' + str(no_url) + '건 제외 — §2-③)' if no_url else ''}

## 1. 시스템 현재 위치
| 축 | 상태 | 근거 |
|---|---|---|
| 검색 | {'🟡' if top1/max(1,scoreable) < 0.5 else '✅'} | top1 {top1/max(1,scoreable):.1%} · top5 {top5/max(1,scoreable):.1%} (관대) |
{gen_row}
| 실행 안정성 | ✅ | {n}/{n} 전량 응답 · 형식 결함 0 |

## 2. 검색 분해 (실측)
### ① 정답 소스군별 miss율
| 소스군 | 문항 | miss | miss율 |
|---|---|---|---|
{sect_rows}

### ② 전멸 소스군 (인입/인덱싱 점검 요청)
{('· ' + ' · '.join(dead)) if dead else '(없음)'}

### ③ 방법론 이슈
{f'- 정답 출처 URL 결측 {no_url}건 (miss {no_url_miss}) — 구조적 채점 불가, 골든셋 보정 후 규칙 D 소급 재채점 권고' if no_url else '- 없음'}

{_cause_section(prod, rep, log_path, n)}

## E형
{'미응시 (검색축만 회차)' if search_only else '원시/실질 병기 — E형 문맥확인 자료 참조'}

## 부록
- score_report.json / score_report.xlsx (재현용 원자료)
"""
    out_dir.mkdir(parents=True, exist_ok=True)
    out.write_text(body, encoding="utf-8")
    return out
