#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scoring.py — ⑧ 실물 채점 + ⑨ MAINTENANCE / 규칙 D (사양서 v1.3 §4-⑧⑨, 규칙 D)

⑧ 흐름: 발행 게이트(질문셋 컬럼 화이트리스트 — 유출 시 HALT) → 질문셋 발행
  → 로그 수령 → 형식 게이트(전량·ID 중복·rank 연속·meta 코퍼스 버전·source 실재)
  → 채점(run_score_v11, selftest 선행) → 성적표 [GATE:사람 확정]

규칙 D: run_score 버전(내용 해시) 변경 감지 → 전 차수 소급 재채점 잡 자동 생성
  → 완료 전 신규 차수 "회차 비교" 출력 차단 (단일 차수는 '비교 불가' 라벨)

사용:
  python3 tools/scoring.py publish-gate <질문셋.xlsx>
  python3 tools/scoring.py format-gate --product RC --log <로그.json>
  python3 tools/scoring.py score --product RC --log <로그.json> --round r2
  python3 tools/scoring.py rule-d --product RC        # 채점기 버전 감시
  python3 tools/scoring.py selftest
"""
import argparse
import hashlib
import json
import re
import subprocess
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from olib import (ROOT, N, load_config, load_state, save_state,
                  ledger_append, issue_gate_card, issue_input_card, now)

DATA = ROOT / "data"
RUN_SCORE = (ROOT.parent / "tools" / "run_score_v11.py").resolve()

# 발행 게이트 화이트리스트 — ID·질문 2컬럼 강제 (유출 = HALT)
PUBLISH_WHITELIST = {"문항ID", "ID", "질문"}
LEAK_MARKERS = ("정답", "합격", "발췌", "근거", "기준", "채점", "출제")


# ── 발행 게이트 ─────────────────────────────────────────────
def publish_gate(xlsx_path, product=None):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)) if c]
    nrows = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r))
    wb.close()
    extra = [h for h in hdr if h not in PUBLISH_WHITELIST]
    leaks = [h for h in extra if any(m in h for m in LEAK_MARKERS)]
    if leaks:
        ledger_append("SCORING", "PUBLISH_GATE_LEAK", "script:발행게이트",
                      evidence={"file": N(Path(xlsx_path).name), "leak_columns": leaks},
                      reason="발행 게이트 유출 감지 — HALT (정답·합격 기준·발췌는 응시측에 절대 미노출)",
                      product=product)
        return "HALTED", {"leak_columns": leaks, "rows": nrows}
    if extra:
        return "FAIL", {"extra_columns": extra, "rows": nrows}
    ledger_append("SCORING", "PUBLISH_GATE_PASS", "script:발행게이트",
                  evidence={"file": N(Path(xlsx_path).name), "rows": nrows, "columns": hdr, "leak": 0},
                  product=product)
    return "PASS", {"rows": nrows, "columns": hdr}


# ── 형식 게이트 ─────────────────────────────────────────────
def _ledger_candidates(prod):
    """⑤ 후보들을 버전 내림차순으로 — 파일시스템 실측 (신규 생성분 포함)"""
    d = DATA / prod / "05_unified_ledger"
    cands = [p for p in d.glob("*.xlsx") if not p.name.startswith(".")] if d.is_dir() else []

    def vkey(p):
        m = re.search(r"_v(\d+(?:_\d+)+)", p.name)
        v = tuple(int(x) for x in m.group(1).split("_")) if m else (0,)
        return (v, p.stat().st_mtime)   # 동일 버전이면 최신 생성분 (증분 확대 대장 v+0 대응)
    return sorted(cands, key=vkey, reverse=True)


def _ids_from_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ids = set()
    for sn in wb.sheetnames:
        ws = wb[sn]
        try:
            hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        except StopIteration:
            continue
        if "ID" in hdr and "질문" in hdr:
            i = hdr.index("ID")
            for r in ws.iter_rows(min_row=2, values_only=True):
                v = N(r[i]) if i < len(r) else ""
                if v:
                    ids.add(v)
            break
    wb.close()
    return ids


def load_golden_ids(prod):
    """전량 검사 참조 집합 = 채점기가 쓰는 바로 그 골든셋 (참조 집합 분열 금지 —
    게이트와 채점기가 다른 기준을 보면 assert 기준 검증 실패 사고가 된다)"""
    try:
        p = golden_xlsx(prod)
    except FileNotFoundError:
        return set(), None
    ids = _ids_from_xlsx(p)
    if ids:
        return ids, N(p.name)
    # 채점용/최고버전에서 ID가 안 나오면 후보 순회 (구형 산출물 호환)
    for p in _ledger_candidates(prod):
        ids = _ids_from_xlsx(p)
        if ids:
            return ids, N(p.name)
    return set(), None


def parse_corpus_meta(cv):
    """corpus_version 문자열에서 (문서수, 청크수) 추출 — '문서 N건·청크 M건' 마지막 쌍(계) 채택"""
    pairs = re.findall(r"문서\s*([\d,]+)\s*건[·\s]*청크\s*([\d,]+)\s*건", N(cv))
    if not pairs:
        return None, None
    d, c = pairs[-1]
    return int(d.replace(",", "")), int(c.replace(",", ""))


def _meta_history_path(prod):
    return ROOT / "results" / f"meta_history_{prod}.json"


def meta_history_compare(prod, round_label, cv):
    """[v2] meta 스코프·청크 수를 전 회차와 대조 — r5 스코프 결손(−19문서) 적발 게이트.
    반환: (경고 목록, 이번 회차 기록)"""
    docs, chunks = parse_corpus_meta(cv)
    hp = _meta_history_path(prod)
    hist = json.loads(hp.read_text(encoding="utf-8")) if hp.exists() else {}
    warns = []
    if docs is not None and hist:
        prev_r = sorted(hist)[-1]
        pd_, pc = hist[prev_r].get("docs"), hist[prev_r].get("chunks")
        if pd_ is not None and (docs != pd_ or chunks != pc):
            warns.append({"type": "스코프·청크 변동", "id": f"{prev_r}→{round_label}",
                          "candidates": [f"문서 {pd_}→{docs} ({docs-pd_:+d})",
                                         f"청크 {pc}→{chunks} ({chunks-pc:+d})"],
                          "ack_required": True,
                          "note": "전 회차 대비 코퍼스 변동 — 의도 여부 확인 의무 (r5 결손 −19문서 판례)"})
    hist[round_label] = {"docs": docs, "chunks": chunks, "corpus_version": N(cv)[:200]}
    hp.parent.mkdir(exist_ok=True)
    hp.write_text(json.dumps(hist, ensure_ascii=False, indent=2), encoding="utf-8")
    return warns


def format_gate(log_path, prod, round_label=None):
    d = json.loads(Path(log_path).read_text(encoding="utf-8"))
    meta = d.get("meta", {})
    resp = d.get("responses", d.get("results", []))
    checks = []
    warns = []
    ok = True
    # 1) meta 코퍼스 버전 실재
    cv = N(meta.get("corpus_version", ""))
    good = bool(cv)
    ok &= good
    checks.append(("meta 코퍼스 버전", good, cv[:60] or "부재"))
    # 1′) [v2] answer null 검사 + 응시 범위 대조 — 업로드 시 사람이 선언한 범위(사이드카)와
    # 실물(answer null 여부)을 대조. 선언 없으면 자동 감지(전건 null=검색축만). (2026-08)
    nulls = sum(1 for r in resp if not N(r.get("answer") or ""))
    all_null = bool(resp) and nulls == len(resp)
    declared = ""
    _sc = Path(str(log_path) + ".scope")
    if _sc.exists():
        declared = N(_sc.read_text(encoding="utf-8")).strip()
    if declared == "search":
        good = all_null
        ok &= good
        checks.append(("응시 범위 대조", good,
                       "선언=검색축만 · 전건 null 일치 — 검색축만 회차 접수" if good else
                       f"선언=검색축만인데 answer 있는 응답 {len(resp) - nulls}건 — 선언과 실물 불일치 (재확인 필요)"))
    elif declared == "full":
        good = nulls == 0
        ok &= good
        checks.append(("응시 범위 대조", good,
                       "선언=전체 · answer 전건 존재 일치" if good else
                       f"선언=전체인데 answer null {nulls}건 — 결손 또는 선언 오류 (재확인 필요)"))
    elif all_null:
        checks.append(("answer null", True, "전건 null — 검색축만 응시 회차로 자동 접수 (생성축·E형 미응시)"))
    else:
        good = nulls == 0
        ok &= good
        checks.append(("answer null", good, f"{nulls}건 (검색축만 회차는 전건 null이어야 함)"))
    # 1″) [v2] meta 스코프·청크 전 회차 대조 (변동 = 플래그, 사람 ack)
    if round_label and cv:
        had_history = _meta_history_path(prod).exists()
        warns += meta_history_compare(prod, round_label, cv)
        checks.append(("meta 전 회차 대조", True,
                       f"변동 {len(warns)}건 (플래그)" if warns
                       else ("전 회차와 동일" if had_history else "첫 회차 — 기준 등재")))
    # 2) ID 중복 0
    ids = [N(r.get("id", "")) for r in resp]
    dup = len(ids) - len(set(ids))
    good = dup == 0
    ok &= good
    checks.append(("ID 중복", good, f"{dup}"))
    # 3) 전량 — 골든셋 ID 전건 커버 (참조 집합 자체 검증: 골든셋 0이면 그것도 사고)
    gids, gsrc = load_golden_ids(prod)
    if not gids:
        ok = False
        checks.append(("전량(골든셋 커버)", False, "참조 집합 로드 실패 — 골든셋 0 (assert 기준 검증 실패)"))
    else:
        missing = gids - set(ids)
        extra = set(ids) - gids
        good = not missing
        ok &= good
        checks.append(("전량(골든셋 커버)", good,
                       f"기준 {gsrc} {len(gids)} · 로그 {len(ids)} · 누락 {len(missing)} · 잉여 {len(extra)}"))
    # 4) rank 연속
    bad_rank = 0
    for r in resp:
        ranks = [h.get("rank") for h in r.get("hits", [])]
        if ranks and ranks != list(range(1, len(ranks) + 1)):
            bad_rank += 1
    good = bad_rank == 0
    ok &= good
    checks.append(("rank 연속", good, f"불연속 {bad_rank}"))
    # 5) source_url/name 최소 1 실재 (전체 히트 중)
    n_src = sum(1 for r in resp for h in r.get("hits", [])
                if N(h.get("source_url") or "") or N(h.get("source_name") or ""))
    n_hits = sum(len(r.get("hits", [])) for r in resp)
    good = n_src > 0 or n_hits == 0
    ok &= good
    checks.append(("source 실재", good, f"{n_src}/{n_hits} 히트에 source_url/name"))
    status = "PASS" if ok else "HALTED"   # 형식 게이트 실패 = 킬스위치 (§6)
    ledger_append("SCORING", f"FORMAT_GATE_{status}", "script:형식게이트",
                  evidence={"log": N(Path(log_path).name),
                            "checks": [{"name": n, "ok": g, "detail": det} for n, g, det in checks],
                            "warns": warns},
                  product=prod)
    return status, checks, warns


# ── 채점 ───────────────────────────────────────────────────
def scorer_fingerprint():
    return hashlib.sha256(RUN_SCORE.read_bytes()).hexdigest()


def golden_xlsx(prod):
    # '채점용' 시트가 있으면 우선 (채점기 이식 산출물 — 예: RV run_score 규격 변환본)
    override = sorted((DATA / prod / "08_scoring").glob("*채점용*.xlsx"))
    if override:
        return override[-1]
    for p in _ledger_candidates(prod):
        if _ids_from_xlsx(p):
            return p
    raise FileNotFoundError(f"{prod} 정본 골든셋(ID·질문 시트) 없음")


def make_question_set(prod):
    """⑧ 질문셋 발행 — 골든셋에서 ID·질문 2컬럼만 추출 (시험세트는 ⑧의 산출물, §2′)"""
    src = golden_xlsx(prod)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        try:
            hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        except StopIteration:
            continue
        if "ID" in hdr and "질문" in hdr:
            ii, qi = hdr.index("ID"), hdr.index("질문")
            for r in ws.iter_rows(min_row=2, values_only=True):
                if ii < len(r) and N(r[ii]):
                    rows.append((N(r[ii]), N(r[qi]) if qi < len(r) else ""))
            break
    wb.close()
    out = DATA / prod / "08_scoring" / f"{prod}_질문셋_발행본_v1_0.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    nwb = openpyxl.Workbook()
    ws = nwb.active
    ws.title = f"질문셋_{len(rows)}"
    ws.append(["문항ID", "질문"])          # 화이트리스트 2컬럼 — 그 외 전부 유출
    for r in rows:
        ws.append(list(r))
    nwb.save(out)
    ledger_append("SCORING", "QUESTION_SET_PUBLISHED", "script:publish",
                  evidence={"file": N(out.name), "rows": len(rows), "columns": ["문항ID", "질문"],
                            "source": N(src.name)}, product=prod)
    return out


def _secondary_scorer():
    """[v2] 병기 채점기 — run_score_v12.py 가 tools/ 에 놓이면 자동 인식.
    공식 회차 비교는 v1.1 기준(팀장님 합의), v1.2.1 은 병기."""
    p = RUN_SCORE.parent / "run_score_v12.py"
    return p if p.exists() else None


def _run_scorer(scorer, prod, log_path, out_dir):
    st = subprocess.run([sys.executable, str(scorer), "--selftest"],
                        capture_output=True, text=True, cwd=str(scorer.parent))
    if "모두 통과" not in st.stdout:
        return None, {"selftest": "실패", "tail": st.stdout[-300:]}
    out_dir.mkdir(parents=True, exist_ok=True)
    p = subprocess.run([sys.executable, str(scorer),
                        "--golden", str(golden_xlsx(prod)),
                        "--log", str(Path(log_path).resolve()),
                        "--out-dir", str(out_dir)],
                       capture_output=True, text=True, cwd=str(scorer.parent))
    rep_path = out_dir / "score_report.json"
    if not rep_path.exists():
        return None, {"오류": "score_report.json 미생성", "tail": (p.stdout + p.stderr)[-400:]}
    return json.loads(rep_path.read_text(encoding="utf-8")), None


def aggregate(rep):
    from collections import Counter
    c = Counter(r.get("검색") for r in rep)
    g = Counter(r.get("생성") for r in rep)
    top1 = c.get("hit_top1", 0)
    return {"top1": top1, "top5": top1 + c.get("hit_top5", 0), "pass": g.get("pass", 0),
            "partial": g.get("partial", 0), "unparsed": g.get("unparsed", 0),
            "E환각(원시)": sum(1 for r in rep if r.get("E형환각")),
            "E거절": sum(1 for r in rep if r.get("E형거절")), "n": len(rep)}


def e_context_review(prod, round_label, rep, log_path, out_dir):
    """[v2 의무] E형 환각 판정 전건 — 답변 원문을 열람 자료로 발행 + 사람 ack 플래그.
    채점기 원시 수치를 그대로 믿지 않는다 (P-신규-3). 실질 수치는 사람 확인 후 확정."""
    d = json.loads(Path(log_path).read_text(encoding="utf-8"))
    answers = {N(r.get("id", "")): N(r.get("answer") or "") for r in d.get("responses", [])}
    items = [r for r in rep if r.get("E형환각")]
    if not items:
        return [], None
    lines = [f"# E형 문맥 확인 — {prod} {round_label} (원시 {len(items)}건)",
             "", "> 채점기 원시 판정이다. 오탐(신형 거절 표현)·전제 실효(신규 문서)를 사람이 분리하라.",
             "> 자동 기각 코드 작성 금지 (P-04). 실질 수치는 확인 후 리포트에 병기.", ""]
    for r in items:
        qid = N(r.get("id", ""))
        lines += [f"## {qid}", "```", answers.get(qid, "(답변 원문 없음)")[:800], "```", ""]
    md = out_dir / "E형_문맥확인.md"
    md.write_text("\n".join(lines), encoding="utf-8")
    flags = [{"type": "E형 환각(원시) 문맥 확인", "id": N(r.get("id", "")),
              "candidates": [answers.get(N(r.get("id", "")), "")[:60]],
              "ack_required": True} for r in items]
    return flags, md


def score(prod, log_path, round_label, warns=None):
    out_dir = ROOT / "results" / f"score_{prod}_{round_label}"
    # 검색축만 응시 회차 감지 (전건 answer null) — 생성축·E형은 '미응시'로 정직 표기
    try:
        _d = json.loads(Path(log_path).read_text(encoding="utf-8"))
        _resp = _d.get("responses", [])
        search_only = bool(_resp) and all(not N(r.get("answer") or "") for r in _resp)
    except Exception:
        search_only = False
    rep, err = _run_scorer(RUN_SCORE, prod, log_path, out_dir)   # 공식(v1.1 기준점)
    if rep is None:
        return "HALTED", err
    summ = aggregate(rep)
    if search_only:
        for k in ("pass", "partial", "unparsed", "E환각(원시)", "E거절"):
            summ[k] = "미응시"
        summ["응시 범위"] = "검색축만 (answer 미제출 — 생성축·E형 채점 없음)"
    # [v2] 병기 채점기 (v12 존재 시)
    sec = _secondary_scorer()
    sec_summ = None
    if sec:
        rep2, err2 = _run_scorer(sec, prod, log_path, ROOT / "results" / f"score_{prod}_{round_label}_v12")
        if rep2 is not None:
            sec_summ = aggregate(rep2)
    # [v2] E형 문맥 확인 의무 — 전건 열람 자료 + ack (검색축만 회차는 답변이 없어 생략)
    e_flags, e_md = ([], None) if search_only else e_context_review(prod, round_label, rep, log_path, out_dir)
    stt = load_state()
    ps = stt["products"][prod]
    compare_blocked = ps.get("compare_blocked", False)
    ev = {"round": round_label, **{f"공식(v11).{k}": v for k, v in summ.items()},
          "회차 비교": "차단(규칙 D — 소급 재채점 미완)" if compare_blocked else "허용(공식 v1.1 기준)",
          "out": str(out_dir.relative_to(ROOT))}
    if ps.get("key_released"):
        ev["⚠ 세대 라벨"] = (f"키 공개 후 참고치 — 정답키 공개({ps['key_released']}). "
                            "공식 측정은 차세대(신규 출제·봉인) 골든셋 담당")
    if sec_summ:
        ev.update({f"병기(v12).{k}": v for k, v in sec_summ.items()})
    if e_md:
        ev["E형 문맥확인 자료"] = str(e_md.relative_to(ROOT))
    ledger_append("SCORING", "SCORED", "script:scoring", evidence=ev, product=prod)
    flags = (warns or []) + e_flags
    if search_only:
        what = f"{round_label} 성적표 확정 — 검색축만 회차 (top1·top5) · 표본 확인 후 사람 확정"
        rec = ("검색축만 회차 — E형 원문 확인은 해당 없음 (answer 미제출).\n"
               + _search_guide(prod, rep, log_path, summ))
    else:
        what = f"{round_label} 성적표 확정 — E형 원문 열람({len(e_flags)}건) + 스코프 변동 확인 후 사람 확정"
        rec = ("E형 원시 판정은 그대로 믿지 말 것(P-신규-3) — 원문 확인 후 실질 수치 병기. "
               + f"열람 자료: {ev.get('E형 문맥확인 자료','—')}")
    issue_gate_card(prod, "SCORING", f"SCORE_{prod}_{round_label}",
                    what_stopped=what, evidence=ev, flags=flags, recommendation=rec)
    return "WAITING_HUMAN", ev


def _search_guide(prod, rep, log_path, summ, n_show=3):
    """검색축 회차 사람 확인 가이드 — '채점이 말이 되는가'를 표본으로 보여준다:
    성공/실패 사례마다 정답 출처 vs 시스템이 가져온 출처를 나란히."""
    try:
        d = json.loads(Path(log_path).read_text(encoding="utf-8"))
        hits = {N(r.get("id", "")): [N(h.get("source") or h.get("source_url") or h.get("name") or "")
                                     for h in (r.get("hits") or [])[:3]]
                for r in d.get("responses", [])}
        import openpyxl
        gw = openpyxl.load_workbook(golden_xlsx(prod), read_only=True, data_only=True).active
        gh = [N(c) for c in next(gw.iter_rows(max_row=1, values_only=True))]
        si = next((i for i, h in enumerate(gh) if h.startswith("근거 출처")), None)
        gold = ({N(r[0]): N(r[si]) for r in gw.iter_rows(min_row=2, values_only=True) if r[0]}
                if si is not None else {})
        def url(s):
            m = re.search(r"https?://\S+", s or "")
            return m.group(0) if m else (s or "?")
        oks = [r for r in rep if r.get("검색") == "hit_top1"][:n_show]
        miss = [r for r in rep if r.get("검색") not in ("hit_top1", "hit_top5")][:n_show]
        L = ["", "### 👀 사람 확인 가이드 — 검색축만 회차: 표본으로 '채점이 말이 되는지'만 보면 됩니다",
             "", "**① top1 성공 표본 — 1순위 출처가 정답 출처와 같은 문서인가요?**"]
        for r in oks:
            i = N(r["id"])
            L += [f"- {i} · 정답: {url(gold.get(i))[:75]}",
                  f"  ↳ 시스템 1순위: {(hits.get(i) or ['?'])[0][:75]}"]
        L += ["", "**② 실패 표본 — 정답과 가져온 출처가 정말 다른가요? (사실 같은 문서인데 실패 처리면 반려)**"]
        for r in miss:
            i = N(r["id"])
            L += [f"- {i} · 정답: {url(gold.get(i))[:75]}",
                  f"  ↳ 시스템 상위: {' | '.join((hits.get(i) or ['?'])[:2])[:95]}"]
        L += ["", f"**③ 수치 감**: top1 {summ.get('top1')} · top5 {summ.get('top5')} / {summ.get('n')}문항 — "
                  "표본과 모순 없으면 승인하세요. (생성축·E형은 미응시 — 이번 회차 확인 대상 아님)"]
        return "\n".join(L)
    except Exception as e:
        return f"(표본 가이드 생성 실패 — 수치만으로 판단: {str(e)[:80]})"


def extract_summary(rep):
    """score_report.json 에서 핵심 수치 추출 (구조 유연 대응)"""
    txt = json.dumps(rep, ensure_ascii=False)
    out = {}
    if isinstance(rep, dict):
        for k in ("search", "검색", "summary", "요약", "totals"):
            if k in rep and isinstance(rep[k], dict):
                out.update({f"{k}.{k2}": v for k2, v in rep[k].items()
                            if isinstance(v, (int, float, str)) and len(str(v)) < 40})
        for k, v in rep.items():
            if isinstance(v, (int, float)) :
                out[k] = v
    if not out:
        out["raw_keys"] = list(rep)[:8] if isinstance(rep, dict) else type(rep).__name__
    return out


# ── 규칙 D ─────────────────────────────────────────────────
SRC_MANIFEST = ROOT / "catalog" / "manifest_원문데이터.json"
TEST_PRODUCTS = {"EE"}   # 자동 테스트 전용 — 원문이 매회 재생성되는 모의 데이터 (등재 비대상)


def source_hash_gate(paths, register_new=False):
    """[T18] 원문 데이터 해시 대조 (규칙 B′의 데이터 확장) — 응답로그·질문셋 원본 변조를 기계가 잡는다.
    등재된 파일: sha256 불일치 → 위반 목록 반환 (호출측 정지).
    미등재 파일: register_new=True 일 때만 첫 등재 (업로드 원본용 — 생성 산출물은 등재 금지)."""
    import datetime
    man = json.loads(SRC_MANIFEST.read_text(encoding="utf-8")) if SRC_MANIFEST.exists() else {"files": {}}
    bad, newly = [], []
    dirty = False
    for p in paths:
        p = Path(p)
        if not p.exists():
            continue
        h = hashlib.sha256(p.read_bytes()).hexdigest()
        try:
            key = str(p.resolve().relative_to(ROOT.resolve()))
        except ValueError:
            key = str(p)
        ent = man["files"].get(key)
        if ent is None:
            if register_new:
                man["files"][key] = {"sha256": h,
                                     "등재": datetime.datetime.now().isoformat(timespec="seconds")}
                newly.append(key)
                dirty = True
        elif ent["sha256"] != h:
            bad.append({"file": key, "등재": ent["sha256"][:16], "현재": h[:16]})
    if dirty:
        SRC_MANIFEST.parent.mkdir(exist_ok=True)
        SRC_MANIFEST.write_text(json.dumps(man, ensure_ascii=False, indent=1), encoding="utf-8")
    return bad, newly


def rule_d_check(prod, actor="script:rule_d"):
    fp = scorer_fingerprint()
    st = load_state()
    ps = st["products"][prod]
    stored = ps.get("scorer_version")
    if stored is None:
        ps["scorer_version"] = fp
        save_state(st)
        return False
    if stored != fp:
        ps["scorer_version"] = fp
        ps["compare_blocked"] = True
        save_state(st)
        # 전 차수 소급 재채점 잡 큐 적치
        rounds = sorted(p.name for p in (ROOT / "results").glob(f"score_{prod}_*"))
        q = ROOT / load_config()["paths"]["queue"]
        q.mkdir(exist_ok=True)
        card = q / f"MAINT_RESCORE_{prod}.md"
        card.write_text(f"""# MAINT_RESCORE_{prod} — 규칙 D 소급 재채점

- 발행: {now()} · 채점기 버전 변경 감지 (내용 해시 {stored[:12]}… → {fp[:12]}…)
- 시스템이 진화하면 채점기도 늙는다 (P-신규-3)

## 소급 재채점 대상 (전 차수 동일 자)
{chr(10).join('- [ ] ' + r for r in rounds) or '- (기록된 차수 없음 — 과거 성적표 원본 기준 수동 지정)'}

## 차단
- 완료 전 신규 차수 성적의 **회차 비교 출력 차단** (단일 차수는 '비교 불가' 라벨)
- 전 차수 재채점 완료 후: `python3 tools/scoring.py rule-d-clear --product {prod}`
""", encoding="utf-8")
        ledger_append("MAINTENANCE", "RULE_D_TRIGGER", actor,
                      evidence={"scorer_was": stored[:12], "scorer_now": fp[:12],
                                "rescore_queue": rounds},
                      reason="채점기 버전 변경 — 전 차수 소급 재채점 전 회차 비교 차단",
                      product=prod)
        print(f"⚠ 규칙 D 발동 — {prod} 소급 재채점 큐 적치 + 회차 비교 차단")
        return True
    return False


def rule_d_clear(prod, actor):
    st = load_state()
    ps = st["products"][prod]
    ps["compare_blocked"] = False
    save_state(st)
    ledger_append("MAINTENANCE", "RULE_D_CLEAR", f"사람:{actor}",
                  reason="전 차수 소급 재채점 완료 확인 — 회차 비교 차단 해제", product=prod)
    print(f"규칙 D 해제 — {prod} 회차 비교 허용")


# ── pipeline 인터페이스 (stage=SCORING) ─────────────────────
def run(prod, cfg):
    rule_d_check(prod)
    # 1) 질문셋 발행 — 없으면 골든셋에서 생성(2컬럼), 있으면 그대로 게이트
    qs = sorted((DATA / prod / "08_scoring").glob("*질문셋*발행본*.xlsx")) \
        or sorted((DATA / prod / "08_scoring").glob("*질문셋*.xlsx"))
    pub_ev = {}
    if not qs:
        try:
            qs = [make_question_set(prod)]
        except FileNotFoundError:
            qs = []
    if qs and prod not in TEST_PRODUCTS:
        # [T18] 발행 전 원문 대조 — 등재된 질문셋 원본과 불일치 = 정지
        bad, _ = source_hash_gate([qs[-1]])
        if bad:
            ledger_append("SCORING", "SOURCE_HASH_MISMATCH", "script:scoring",
                          evidence={"위반": bad, "판정": "원문 변조 의심 — 발행·채점 중단 (T18)"},
                          product=prod)
            return "HALTED", {"halt": "원문 해시 불일치 (질문셋) — T18 게이트", "위반": bad}
    if qs:
        status, ev = publish_gate(qs[-1], product=prod)
        pub_ev = {"발행 게이트": f"{status} — {N(qs[-1].name)}", **{f"발행.{k}": v for k, v in ev.items()}}
        if status == "HALTED":
            return "HALTED", pub_ev
        if status == "FAIL":
            return "BLOCKED", pub_ev
    # 2) 응답 로그
    logs = sorted((DATA / prod / "08_scoring").glob("*응답로그*.json")) \
        + sorted((DATA / prod / "logs").glob("*.json")) if (DATA / prod / "logs").is_dir() \
        else sorted((DATA / prod / "08_scoring").glob("*응답로그*.json"))
    if not logs:
        issue_input_card(prod, "SCORING",
                         what=f"{prod} 응답 로그 (팀장님 시스템 응시 결과)",
                         where=f"data/{prod}/08_scoring/",
                         fmt="json — meta(corpus_version) + responses 전량 (id/hits/answer)",
                         extra=("발행 게이트 통과분: " + pub_ev.get("발행 게이트", "발행본 미확인")))
        return "WAITING_INPUT", {**pub_ev, "logs": 0}
    # 로그 도착 = INPUT 카드 해소
    qdir = ROOT / load_config()["paths"]["queue"]
    for card in qdir.glob(f"INPUT_SCORING_{prod}.md"):
        (qdir / "완료").mkdir(exist_ok=True)
        card.rename(qdir / "완료" / card.name)
    # 3) 형식 게이트 → 채점 — 회차 반복: 가장 최근 올라온 로그를 채점
    log = max(logs, key=lambda p: p.stat().st_mtime)
    if prod not in TEST_PRODUCTS:
        # [T18] 채점 전 원문 대조 — 업로드 원본은 첫 인입 시 등재, 이후 변조 = 정지
        bad, newly = source_hash_gate([log], register_new=True)
        if newly:
            ledger_append("SCORING", "SOURCE_DATA_REGISTERED", "script:scoring",
                          evidence={"등재": newly}, product=prod)
        if bad:
            ledger_append("SCORING", "SOURCE_HASH_MISMATCH", "script:scoring",
                          evidence={"위반": bad, "판정": "원문 변조 의심 — 채점 중단 (T18)"},
                          product=prod)
            return "HALTED", {"halt": "원문 해시 불일치 (응답로그) — T18 게이트", "위반": bad}
    rnd_m = re.search(r"r\d+", N(log.name))
    if rnd_m:
        rnd = rnd_m.group()
    else:
        # 파일명에 회차가 없으면 자동 부여 — 기존 성적 폴더의 다음 번호
        used = [int(m.group(1)) for d in (ROOT / "results").glob(f"score_{prod}_r*")
                for m in [re.fullmatch(rf"score_{re.escape(prod)}_r(\d+)", d.name)] if m]
        rnd = f"r{max(used) + 1 if used else 1}"
    fstatus, checks, warns = format_gate(log, prod, round_label=rnd)
    if fstatus == "HALTED":
        return "HALTED", {"형식 게이트": [f"{n}:{'OK' if g else 'FAIL'}" for n, g, _ in checks]}
    return score(prod, log, rnd, warns=warns)


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    s = sub.add_parser("publish-gate"); s.add_argument("xlsx"); s.add_argument("--product")
    s = sub.add_parser("format-gate"); s.add_argument("--product", required=True); s.add_argument("--log", required=True); s.add_argument("--round", default=None)
    s = sub.add_parser("score"); s.add_argument("--product", required=True); s.add_argument("--log", required=True); s.add_argument("--round", required=True)
    s = sub.add_parser("rule-d"); s.add_argument("--product", required=True)
    s = sub.add_parser("rule-d-clear"); s.add_argument("--product", required=True); s.add_argument("--actor", default="난희")
    sub.add_parser("selftest")
    a = ap.parse_args()
    if a.cmd == "publish-gate":
        st, ev = publish_gate(a.xlsx, a.product)
        print(f"발행 게이트: {st} {ev}")
        sys.exit(0 if st == "PASS" else (3 if st == "HALTED" else 1))
    elif a.cmd == "format-gate":
        st, checks, warns = format_gate(a.log, a.product, round_label=a.round if hasattr(a, "round") else None)
        for n, g, det in checks:
            print(f"  [{'PASS' if g else 'FAIL':^4}] {n}: {det}")
        for w in warns:
            print(f"  ⚑ {w['type']} {w['id']}: {w['candidates']}")
        print(f"▶ {st}")
        sys.exit(0 if st == "PASS" else 3)
    elif a.cmd == "score":
        fstatus, checks, warns = format_gate(a.log, a.product, round_label=a.round)
        if fstatus != "PASS":
            for n, g, det in checks:
                print(f"  [{'PASS' if g else 'FAIL':^4}] {n}: {det}")
            sys.exit(3)
        st, ev = score(a.product, a.log, a.round, warns=warns)
        print(f"채점: {st}")
        for k, v in ev.items():
            print(f"  {k}: {v}")
    elif a.cmd == "rule-d":
        if not rule_d_check(a.product):
            print(f"{a.product} — 채점기 버전 변경 없음")
    elif a.cmd == "rule-d-clear":
        rule_d_clear(a.product, a.actor)
    elif a.cmd == "selftest":
        selftest()


def selftest():
    import tempfile
    ok = True
    # 1) 발행 게이트: RV 발행본(2컬럼) → PASS
    st, ev = publish_gate(DATA / "RV/08_scoring/RV_질문셋_발행본_v1_1.xlsx", "RV")
    good = st == "PASS" and ev["rows"] == 806
    ok &= good
    print(f"  {'✅' if good else '❌'} 발행 게이트 RV 발행본(806·2컬럼) → {st}")
    # 2) 발행 게이트: 유출 컬럼 → HALT
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["문항ID", "질문", "정답"])   # 유출!
    ws.append(["X-1", "q?", "leak"])
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        st, ev = publish_gate(f.name, "TEST")
    good = st == "HALTED" and "정답" in ev["leak_columns"]
    ok &= good
    print(f"  {'✅' if good else '❌'} 발행 게이트 정답 컬럼 유출 → {st}")
    # 3) 형식 게이트: RC r2 실로그 (회차 라벨 없이 — meta 이력 오염 방지)
    st, checks, warns = format_gate(DATA / "RC/08_scoring/RC_골든셋_r2_응답로그_2026-07-13.json", "RC")
    good = st == "PASS"
    ok &= good
    print(f"  {'✅' if good else '❌'} 형식 게이트 RC r2 실로그 → {st}")
    for n, g, det in checks:
        print(f"        {'·' if g else '✗'} {n}: {det}")
    print(f"selftest: {'전건 통과' if ok else '실패 있음'}")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
