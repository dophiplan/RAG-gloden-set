#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
smoke_cli_real.py — 실제 구독 CLI(claude)로 TT 전 트랙 관통 (첫 실전 가동 스모크)

mock 이 아니라 실제 claude CLI 가 커버리지맵·골든셋·판정을 생성한다.
config.yaml 을 CLI 앙상블로 잠깐 교체 → 실행 → 원복. 사람/팀장 파트는 시뮬레이션.

사용: python3 tools/smoke_cli_real.py
전제: `claude` CLI 로그인됨, ANTHROPIC_API_KEY 미설정(구독 경로).
"""
import json
import os
import shutil
import subprocess
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT / "tools"))
os.environ.pop("ORCH_MOCK", None)           # 실모델 경로
os.environ.pop("ANTHROPIC_API_KEY", None)   # 구독 과금 방지(이중 안전)
PROD = "EE"   # 자동 테스트 전용 (TT는 난희 연습장 — 건드리지 않음)
DATA = ROOT / "data" / PROD
CONFIG = ROOT / "config.yaml"

CLI_MODELS = """models:
  generator: {provider: "cli", command: ["claude", "-p"], model: "claude-opus-4-8"}
  judge:     {provider: "cli", command: ["claude", "-p"], model: "claude-opus-4-8"}
"""

CORPUS = {
    "manual": ["티티뷰어는 원격지 PC 화면을 실시간으로 표시하는 기능을 제공한다. 최대 4대까지 동시 모니터링이 가능하다.",
               "화면 잠금 기능을 켜면 원격 제어 중 로컬 모니터가 검게 표시된다. 개인정보 보호를 위해 권장된다.",
               "파일 전송은 드래그 앤 드롭으로 동작하며 1회 최대 2GB까지 지원한다.",
               "모바일 앱은 iOS 15 이상과 Android 10 이상에서 설치할 수 있다."],
    "policy": ["무료 평가판은 14일 동안 모든 기능을 제한 없이 사용할 수 있다.",
               "라이선스는 동시 접속 세션 수 기준으로 과금되며 연간 계약이 기본이다.",
               "개인정보는 수집일로부터 3년 후 자동 파기된다. 파기 내역은 관리 콘솔에서 확인 가능하다."],
}


def N(s):
    return unicodedata.normalize("NFC", str(s)) if s is not None else ""


def run(*args, real=True):
    env = {**os.environ}
    env.pop("ORCH_MOCK", None)
    p = subprocess.run([sys.executable, str(ROOT / "tools/pipeline.py"), *args],
                       capture_output=True, text=True, cwd=str(ROOT), env=env)
    if p.returncode != 0:
        # [P3] onboard/approve 실패가 침묵하면 스모크가 헛것을 측정 — 즉시 드러낸다
        print(f"  ⚠ pipeline {args[0]} 실패(exit {p.returncode}): {(p.stderr or p.stdout)[:200]}")
    return p


def state():
    return json.loads((ROOT / "state.json").read_text(encoding="utf-8"))["products"][PROD]


def gates():
    return [g["id"] for g in state()["open_gates"]]


def reset():
    shutil.rmtree(DATA, ignore_errors=True)
    (ROOT / "terrain.d" / f"{PROD}.yaml").unlink(missing_ok=True)
    st = json.loads((ROOT / "state.json").read_text(encoding="utf-8"))
    st["products"].pop(PROD, None)
    (ROOT / "state.json").write_text(json.dumps(st, ensure_ascii=False, indent=2), encoding="utf-8")
    for base in [ROOT / "검수큐", ROOT / "검수큐/완료"]:
        for f in base.glob(f"*{PROD}*.md"):
            f.unlink()
    for r in (ROOT / "results").glob(f"score_{PROD}_*"):
        shutil.rmtree(r)


def main():
    import e2e_fulltrack as e2e   # 사람/팀장 시뮬 재사용
    backup = CONFIG.read_text(encoding="utf-8")
    try:
        # config → CLI 앙상블 (models 섹션만 교체, 나머지 유지)
        rest = backup[backup.index("pipeline:"):]
        CONFIG.write_text(CLI_MODELS + "\n" + rest, encoding="utf-8")
        reset()
        print("═══ 실제 claude CLI 전 트랙 스모크 (TT) ═══")
        run("onboard", "--product", PROD, "--name", "자동테스트", "--base", "RC", "--force")
        cdir = DATA / "corpus"
        cdir.mkdir(parents=True, exist_ok=True)
        for doc, chunks in CORPUS.items():
            (cdir / f"{doc}.md").write_text("\n\n\n".join(chunks), encoding="utf-8")
        print(f"코퍼스: 문서 {len(CORPUS)} · 청크 {sum(len(v) for v in CORPUS.values())}")

        stages_seen = []
        for _ in range(40):
            ps = state()
            stg = f"{ps['stage']}·{ps['status']}"
            if not gates():
                p = run("run", "--product", PROD)
                last = (p.stdout + p.stderr).strip().splitlines()
                print(f"  run → {last[-1] if last else ''}"[:160])
                ps = state()
                if ps["status"] == "HALTED":
                    print(f"⛔ HALT: {ps['halt_reason']}"); break
                if ps["stage"] == "SCORING" and ps["status"] == "WAITING_HUMAN":
                    break
                if ps["stage"] == "SCORING" and ps["status"] == "WAITING_INPUT":
                    e2e.make_mock_response_log()
                    continue
                if ps["status"] == "DONE":
                    break
                if not gates() and ps["status"] == "WAITING_HUMAN":
                    print("교착"); break
                continue
            for g in gates():
                if g.startswith("CALIN_"):
                    e2e.fill_human_judgments()
                run("approve", g, "--ack-all", "--actor", "난희(시뮬)")
                print(f"  승인: {g}")

        # 결과
        print("\n═══ 실모델 산출물 실측 ═══")
        checks = [
            ("커버리지맵", list((DATA / "03_coverage_map").glob("*.xlsx"))),
            ("골든셋 배치", list((DATA / "04_goldenset_batch").glob("*.xlsx"))),
            ("통합 대장", list((DATA / "05_unified_ledger").glob("*.xlsx"))),
            ("캘리브 판정30", list((DATA / "06_calibration").glob("*판정*.xlsx"))),
            ("판정대장", list((DATA / "07_stage2").glob("*판정대장*.xlsx"))),
            ("성적표", list((ROOT / "results").glob(f"score_{PROD}_*/score_report.json"))),
        ]
        ok = True
        for name, fs in checks:
            good = bool(fs)
            ok &= good
            print(f"  {'✅' if good else '❌'} {name}: {[N(f.name) for f in fs][:1]}")
        ps = state()
        print(f"  최종: {ps['stage']}·{ps['status']} (cal={ps['calibration_passed']})")
        # 커버리지맵 실내용 표본
        cm = list((DATA / "03_coverage_map").glob("*.xlsx"))
        if cm:
            import openpyxl
            wb = openpyxl.load_workbook(cm[0], read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
            print(f"\n  [실제 생성 커버리지맵 표본] {ws.max_row-1}개 단위")
            for r in rows[1:]:
                print(f"    · {N(r[1])} | {N(r[4])[:50]}")
            wb.close()
        print(f"\n실제 CLI 스모크: {'전 구간 관통 ✅' if ok else '일부 실패 — 위 참조'}")
        return 0 if ok else 1
    finally:
        CONFIG.write_text(backup, encoding="utf-8")
        print("(config.yaml 원복 완료)")


if __name__ == "__main__":
    sys.exit(main())
