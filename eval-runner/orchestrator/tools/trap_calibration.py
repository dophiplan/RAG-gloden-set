#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
trap_calibration.py — ⑥′ 함정 면접: '구성상 정답이 알려진' 변조 답안으로 채점관 판별력 검증

왜 (난희 고지 2026-08-18): CI는 타사 도메인이라 사람이 내용 판정을 할 수 없다 —
  CAL_CI 30/30=100%는 사람 측이 일괄 합격이라 눈높이 검증이 아니었다.
  사람 지식 대신 **기계가 정답을 아는 시험지**를 만든다: 골든셋 정답을
  ① 의역(뜻 보존 — 합격이 정답) ② 수치/가부 뒤집기(기계 변조 — 0점이 정답)로 변형해
  채점관(Kimi, ⑦과 동일 호출 경로)에게 블라인드로 채점시키고 기대 판정과 대조한다.
  일본어를 몰라도 결과 해석 가능: "뒤집은 걸 뒤집었다고 잡아냈는가"는 구성상 자명하다.

판정 기준서(유도리 v1.0) 검증 항목:
  - 조항 1 (의미 동등 인정): 의역 답안을 합격 처리하는가 → 유도리
  - 조항 2 (수치·가부 엄격): 뒤집은 답안을 0점 처리하는가 → 판별력
사용: python3 tools/trap_calibration.py CI [--n 12]
"""
import argparse
import datetime
import json
import random
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
import llm
from judge_run import SYSTEM_J, load_rubric, judge_batch
from olib import ROOT, N, ledger_append, load_config

NUM = re.compile(r"(\d+)")
FLIPS = [("できます", "できません"), ("できません", "できます"),
         ("可能です", "できません"), ("必要です", "不要です"),
         ("無料", "有料"), ("有料", "無料"), ("かかりません", "かかります")]

PARA_SYS = """[TASK:PARAPHRASE_JA] 일본어 답변을 **뜻을 완전히 보존**하며 다른 표현으로 바꿔라.
규칙: ① 수치·고유명사·가부(가능/불가)는 절대 변경 금지 ② 어순·어휘만 바꿈 ③ '필수:' 줄은 제거.
입력: [{"no":1,"ja":"…"}] 출력: JSON만 [{"no":1,"para":"…"}]"""


def flip_answer(ans):
    """기계 변조 — 수치 +1 또는 가부 반전. 성공하면 (변조답, 변조내용), 못하면 None."""
    body = ans.split("필수:")[0].split("必須:")[0].strip()
    for a, b in FLIPS:
        if a in body:
            return body.replace(a, b, 1), f"가부 반전: {a}→{b}"
    m = NUM.search(body)
    if m and len(m.group(1)) <= 4:
        n = m.group(1)
        return body[:m.start()] + str(int(n) + 1) + body[m.end():], f"수치 변조: {n}→{int(n)+1}"
    return None


def main(prod, n_items):
    cfg = load_config()
    led = sorted((ROOT / "data" / prod / "05_unified_ledger").glob("*통합대장*.xlsx"))[-1]
    ws = openpyxl.load_workbook(led, read_only=True, data_only=True).worksheets[0]
    hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    qi = hdr.index("질문")
    ai = hdr.index("정답") if "정답" in hdr else hdr.index("정답 (필수 포함 요소)")
    rows = [(N(r[0]), N(r[qi]), N(r[ai])) for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[0] and r[qi] and r[ai]]
    # 결정적 선정(난수 시드 고정) — 변조 가능한(수치·가부 포함) 문항 우선
    random.seed(20260818)
    flippable = [r for r in rows if flip_answer(r[2])]
    random.shuffle(flippable)
    base = flippable[:n_items]
    if len(base) < n_items:
        sys.exit(f"[중단] 변조 가능 문항 {len(base)} < {n_items}")

    # 의역 생성 (claude — 구성자·채점관 분리: 채점은 Kimi)
    payload = [{"no": i + 1, "ja": b[2].split("필수:")[0].split("必須:")[0].strip()} for i, b in enumerate(base)]
    out = llm.chat("generator", PARA_SYS, json.dumps(payload, ensure_ascii=False), cfg)
    got = llm.extract_json(out)
    paras = {int(x["no"]): N(x.get("para", "")) for x in got if isinstance(x, dict) and "no" in x}

    # 시험 항목 조립 (기대 판정 봉인) — 의역은 수치 보존 기계 검증 후 채택
    trials, expected = [], {}
    for i, (iid, q, ans) in enumerate(base):
        gold_nums = NUM.findall(ans.split("필수:")[0].split("必須:")[0])
        para = paras.get(i + 1, "")
        if para and NUM.findall(para) == gold_nums and N(para) != N(ans):
            tid = f"T{i+1:02d}-P"
            trials.append({"ID": tid, "질문": q, "정답": ans, "응답": para})
            expected[tid] = ("합격", "의역(뜻 보존) — 조항1 유도리")
        fl = flip_answer(ans)
        if fl:
            tid = f"T{i+1:02d}-F"
            trials.append({"ID": tid, "질문": q, "정답": ans, "응답": fl[0]})
            expected[tid] = ("0점", fl[1] + " — 조항2 엄격")
    random.shuffle(trials)   # 패턴 학습 방지 (시드 고정 = 재현 가능)

    # 채점 — ⑦과 동일 경로(judge_batch, Kimi)
    rubric, rname = load_rubric(prod)
    verdicts = judge_batch(trials, rubric, cfg)
    vmap = {N(v.get("ID", "")): N(v.get("판정", "")) for v in verdicts}

    ok_p = ok_f = n_p = n_f = 0
    detail = []
    for t in trials:
        tid = t["ID"]
        exp, why = expected[tid]
        got_v = vmap.get(tid, "(무응답)")
        # 유도리 채점: 의역은 '합격'이어야, 변조는 '0점'이어야 (부분은 각각 불충분/과관대)
        hit = (got_v == exp)
        if tid.endswith("-P"):
            n_p += 1; ok_p += hit
        else:
            n_f += 1; ok_f += hit
        detail.append((tid, exp, got_v, "✅" if hit else "❌", why))

    print(f"\n■ 함정 면접 결과 (채점관: Kimi · 기준서: {rname})")
    print(f"  유도리(의역→합격): {ok_p}/{n_p}")
    print(f"  판별력(변조→0점): {ok_f}/{n_f}")
    for d in detail:
        print(f"  {d[3]} {d[0]} 기대 {d[1]} / 실제 {d[2]} — {d[4]}")
    verdict = "합격" if (ok_p >= n_p * 0.8 and ok_f >= n_f * 0.8) else "불합격 — 채점관 재검토 필요"
    print(f"  ▶ 종합: {verdict} (임계 각 80%)")

    out_p = ROOT / "data" / prod / "06_calibration" / f"{prod}_함정면접_{datetime.date.today():%Y%m%d}.json"
    out_p.write_text(json.dumps({"trials": trials, "expected": {k: v[0] for k, v in expected.items()},
                                 "verdicts": vmap, "유도리": f"{ok_p}/{n_p}", "판별력": f"{ok_f}/{n_f}",
                                 "종합": verdict}, ensure_ascii=False, indent=1), encoding="utf-8")
    ledger_append("CALIBRATION", "TRAP_CALIBRATION", "script:trap_calibration",
                  evidence={"방식": "구성상 정답이 알려진 변조 답안 (사람 도메인 지식 불요)",
                            "유도리(의역→합격)": f"{ok_p}/{n_p}", "판별력(변조→0점)": f"{ok_f}/{n_f}",
                            "종합": verdict, "산출": N(out_p.name),
                            "맥락": "CAL_CI 100%가 사람 일괄합격이라 판별력 미검증 — 그 보강"},
                  product=prod)
    return 0 if verdict == "합격" else 1


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("product")
    ap.add_argument("--n", type=int, default=12)
    a = ap.parse_args()
    sys.exit(main(a.product, a.n))
