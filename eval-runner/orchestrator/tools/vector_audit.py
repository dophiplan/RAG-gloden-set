#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
vector_audit.py — 인입 무결성 감사 (벡터 export vs 원본 코퍼스/커버리지맵)

팀장님 벡터 DB의 export를 받아 "인입 과정에서 뭐가 빠졌나"를 기계로 대조한다.
(판례: RC r5 스코프 결손 — 문서가 소리 없이 빠진 채 응시돼 성적이 왜곡)

1단계 — 실측 대조: 원본 코퍼스 vs export 의 문서 수·청크 수, 통째 누락 문서 목록
2단계 — 단위 실재 대조: 커버리지맵의 '사실(fact)' 문장이 export 본문에 원문 실재하는지
        전수 문자 대조(1축과 동일 원리) → "벡터에서 증발한 지식 단위" 목록

출력: results/벡터감사_<제품>_<라벨>.md + 증발단위 xlsx (제품 데이터 경로에 저장)

사용: python3 tools/vector_audit.py --product CI --export <팀장님_export.json> [--label r1]
      export 형식: 코퍼스와 동일 허용 범위 ({chunks:[{doc,chunk_id,text,…}]} / [{...}] / md·txt·zip)
"""
import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from olib import ROOT, N, load_config, ledger_append

DATA = ROOT / "data"


def _norm(s):
    return re.sub(r"\s+", "", unicodedata.normalize("NFC", str(s or "")))


def load_export(path):
    """export → [{doc, text}] — gen_coverage.load_corpus와 같은 허용 범위"""
    import gen_coverage
    p = Path(path)
    if p.suffix == ".json":
        rows = gen_coverage._rows_from_json(json.loads(p.read_text(encoding="utf-8")))
        return [{"doc": N(r.get("doc", "?")), "text": N(r.get("text", ""))} for r in rows]
    if p.suffix == ".xlsx":
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        out = []
        for sn in wb.sheetnames:
            for r in wb[sn].iter_rows(min_row=2, values_only=True):
                txt = " ".join(str(c) for c in (r or []) if c is not None)
                if txt.strip():
                    out.append({"doc": N(sn), "text": N(txt)})
        wb.close()
        return out
    raise SystemExit(f"지원하지 않는 export 형식: {p.suffix} (json/xlsx)")


def load_map_units(prod):
    import gen_goldenset
    cfg = load_config()
    units, path = gen_goldenset.read_map_units(prod, cfg)
    return units, path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--product", required=True)
    ap.add_argument("--export", required=True)
    ap.add_argument("--label", default="v1")
    a = ap.parse_args()
    prod = a.product

    # ── 1단계: 실측 대조
    import gen_coverage
    corpus = gen_coverage.load_corpus(prod)
    export = load_export(a.export)
    c_docs = {c["doc"] for c in corpus}
    e_norm = _norm("\n".join(e["text"] for e in export))
    missing_docs = []
    for d in sorted(c_docs):
        # 문서 대표 표본(첫 청크 앞부분)이 export 전체에 실재하는지로 통째 누락 판정
        probe = next((_norm(c["text"])[:80] for c in corpus if c["doc"] == d and len(_norm(c["text"])) >= 80), None)
        if probe and probe not in e_norm:
            missing_docs.append(d)
    print(f"1단계 — 원본 {len(c_docs)}문서/{len(corpus)}청크 vs export {len(export)}행 · 통째 누락 의심 문서 {len(missing_docs)}건")

    # ── 2단계: 커버리지맵 단위 실재 대조
    units, map_path = load_map_units(prod)
    lost = []
    if units:
        for u in units:
            f = _norm(u.get("fact", ""))
            if len(f) >= 10 and f not in e_norm:
                lost.append(u)
        print(f"2단계 — 맵 단위 {len(units)}개 중 export에서 증발 {len(lost)}개 ({len(lost)/len(units):.1%})")
    else:
        print("2단계 — 커버리지맵 없음 (맵 완성 후 재실행하면 단위 대조까지 수행)")

    # ── 리포트
    out_dir = DATA / prod / "벡터감사"
    out_dir.mkdir(parents=True, exist_ok=True)
    md = out_dir / f"벡터감사_{prod}_{a.label}.md"
    lines = [f"# 인입 무결성 감사 — {prod} · {a.label}",
             "",
             f"- export: {N(Path(a.export).name)} ({len(export)}행)",
             f"- 원본 코퍼스: {len(c_docs)}문서 · {len(corpus)}청크",
             f"- 1단계 통째 누락 의심 문서: **{len(missing_docs)}건**"]
    lines += [f"  - {d}" for d in missing_docs[:50]]
    if units:
        lines += ["", f"- 2단계 증발 지식 단위: **{len(lost)}/{len(units)} ({len(lost)/len(units):.1%})** — 상세는 xlsx"]
    lines += ["", "> 판정 안내: 누락·증발이 0에 가까우면 벡터 인입이 정본과 정합 — 맵 재제작(3단계) 불필요.",
              "> 차이가 크면 해당 목록을 들고 인입 파이프라인 점검 후 재export 권장."]
    md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    if lost:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "증발단위"
        ws.append(["unit_id", "문서", "사실(fact)"])
        for u in lost:
            ws.append([u["unit_id"], u.get("source", u.get("title", "")), u.get("fact", "")[:500]])
        wb.save(out_dir / f"증발단위_{prod}_{a.label}_{len(lost)}건.xlsx")
    ledger_append("MAINTENANCE", "VECTOR_AUDIT", "script:vector_audit",
                  evidence={"label": a.label, "export행": len(export), "누락 문서": len(missing_docs),
                            "증발 단위": len(lost), "리포트": N(md.name)}, product=prod)
    print(f"리포트: {md}")


if __name__ == "__main__":
    main()
