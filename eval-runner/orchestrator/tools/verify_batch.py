#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
verify_batch.py — 스크립트 검수 7종 (사양서 v1.3 §8 · 최우선 구현물)

사용:
  python3 verify_batch.py --batch <xlsx> --map <커버리지맵.xlsx> --product RV \
      [--union <이전 차수 파일들...>] [--config config.yaml] [--json 결과.json]

검수 7종:
  ① 행수·유형 분포 = 보고 일치 (파일명 N문항 선언 대조)
  ② 1축 문자 대조: 발췌 세그(|| 분리, 선두 [라벨]·따옴표 제거, len<6 스킵)
     → norm(공백 제거·소문자) → 맵 '사실' 필드 포함 대조 — 불일치 0
  ③ citation 실재 — 패턴은 terrain 설정에서 로드 (하드코딩 금지)
  ④ 질문 중복 0
  ⑤ 커버 등식 (직접+흡수+부적격+잔여=풀, 소실 0)  ※ 소실≠0 은 HALT
  ⑥ 흡수 대표 ∈ 전 차수 합집합 직접 집합 — union 인자 누락 시 실행 거부(사용 오류)
  ⑦ 앵커 — terrain 패턴 로드, 미포착 변형은 FAIL 아닌 '앵커 변형 의심' 플래그
     [v1.2] 플래그>0 이면 게이트 카드 필수 ack 항목으로 승격

종료 코드: 0=PASS(플래그 가능) · 1=REJECTED(반려) · 2=사용 오류(union 누락 등) · 3=HALT급
원칙: 반려는 일상, HALT는 사고다 (P-신규-4).
한글 파일명 NFD → NFC 정규화. openpyxl read_only+data_only.
"""
import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
import yaml

STD_COLS = {"ID", "유형", "질문", "정답", "근거 출처", "근거 원문 발췌"}


def N(s):
    return unicodedata.normalize("NFC", str(s).strip()) if s is not None and str(s).strip() else ""


def norm(s):
    return re.sub(r"\s+", "", N(s)).lower()


def load_terrain(config_path, product):
    # 공용 load_config 사용 — terrain.d/ 오버레이(온보딩 제품 RM/HR/TT…)를 포함해야
    # 온보딩 제품의 검수가 지형 프로파일을 찾는다. (config.yaml 직접 파싱은 오버레이를 놓침)
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        from olib import load_config
        cfg = load_config()
    except Exception:
        cfg = yaml.safe_load(Path(config_path).read_text(encoding="utf-8"))
    profiles = cfg.get("terrain", {}).get("profiles", {})
    if product not in profiles:
        sys.exit(f"[사용 오류] terrain.profiles.{product} 가 config/terrain.d 에 없음 — --product 확인")
    return profiles[product]


def find_question_sheet(wb):
    """표준 컬럼(ID·질문·근거 출처…)을 가진 첫 시트를 찾는다."""
    for sn in wb.sheetnames:
        ws = wb[sn]
        try:
            hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        except StopIteration:
            continue
        if STD_COLS <= set(hdr):
            return sn, hdr
    return None, None


def read_batch(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn, hdr = find_question_sheet(wb)
    if not sn:
        wb.close()
        sys.exit(f"[사용 오류] 표준 문항 시트를 찾지 못함: {path}")
    ws = wb[sn]
    idx = {h: i for i, h in enumerate(hdr)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rid = N(r[idx["ID"]]) if idx["ID"] < len(r) else ""
        if not rid:
            continue
        rows.append({
            "ID": rid,
            "유형": N(r[idx["유형"]]) if idx["유형"] < len(r) else "",
            "질문": N(r[idx["질문"]]) if idx["질문"] < len(r) else "",
            "정답": N(r[idx["정답"]]) if idx["정답"] < len(r) else "",
            "출처": N(r[idx["근거 출처"]]) if idx["근거 출처"] < len(r) else "",
            "발췌": N(r[idx["근거 원문 발췌"]]) if idx["근거 원문 발췌"] < len(r) else "",
        })
    # 커버등식/흡수 시트 존재 여부
    eq_sheet = next((s for s in wb.sheetnames if "커버등식" in N(s) or "등식" in N(s)), None)
    absorb_sheet = next((s for s in wb.sheetnames if "흡수" in N(s)), None)
    eq_rows, absorb_reps = [], []
    if eq_sheet:
        for r in wb[eq_sheet].iter_rows(min_row=1, values_only=True):
            eq_rows.append([N(c) for c in r if c is not None])
    if absorb_sheet:
        ws2 = wb[absorb_sheet]
        try:
            hdr2 = [N(c) for c in next(ws2.iter_rows(min_row=1, max_row=1, values_only=True))]
            # 대표 컬럼 추정: '대표' 포함 헤더
            rep_i = next((i for i, h in enumerate(hdr2) if "대표" in h), None)
            if rep_i is not None:
                for r in ws2.iter_rows(min_row=2, values_only=True):
                    v = N(r[rep_i]) if rep_i < len(r) else ""
                    if v:
                        absorb_reps.append(v)
        except StopIteration:
            pass
    wb.close()
    return rows, sn, eq_rows, absorb_reps


def read_map_facts(path, unit_col, fact_col):
    """'…커버리지맵'으로 끝나는 전 시트에서 {Unit ID: 사실}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    facts = {}
    for sn in wb.sheetnames:
        if not N(sn).endswith("커버리지맵"):
            continue
        ws = wb[sn]
        try:
            hdr = [N(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        except StopIteration:
            continue
        if unit_col not in hdr or fact_col not in hdr:
            continue
        ui, fi = hdr.index(unit_col), hdr.index(fact_col)
        for r in ws.iter_rows(min_row=2, values_only=True):
            u = N(r[ui]) if ui < len(r) else ""
            f = N(r[fi]) if fi < len(r) else ""
            if u:
                facts[u] = facts.get(u, "") + " " + f
    wb.close()
    return facts


def split_segments(excerpt):
    """|| 분리 → 선두 [라벨] 제거 → 따옴표 제거 → len<6 스킵"""
    segs = []
    for seg in re.split(r"\|\|", excerpt):
        s = N(seg)
        s = re.sub(r"^\[[^\]]*\]\s*", "", s)      # 선두 [라벨]
        s = s.strip("\"'“”‘’ ")
        if len(s) >= 6:
            segs.append(s)
    return segs


def collect_direct_union(union_paths):
    """이전 차수 파일들의 직접 커버 단위(citation) + 문항ID 합집합"""
    ids, cits = set(), set()
    for p in union_paths:
        try:
            rows, _, _, _ = read_batch(p)
        except SystemExit:
            continue
        for r in rows:
            ids.add(r["ID"])
            for c in re.findall(r"[A-Z]{2,}(?:-[A-Z0-9]+)+", r["출처"]):
                cits.add(c)
    return ids, cits


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--batch", required=True)
    ap.add_argument("--map", required=True)
    ap.add_argument("--product", required=True)
    ap.add_argument("--union", nargs="*", default=None)
    ap.add_argument("--config", default=str(Path(__file__).parent.parent / "config.yaml"))
    ap.add_argument("--json", default=None, help="결과 JSON 저장 경로")
    a = ap.parse_args()

    terrain = load_terrain(a.config, a.product)
    cit_re = re.compile(terrain["citation_pattern"])
    anchor_res = [re.compile(p) for p in terrain.get("anchor_patterns", [])]
    suspect_re = re.compile(terrain.get("anchor_suspect_pattern", r"\(20\d\d-\d\d-\d\d[^)]*\)"))

    rows, sheet, eq_rows, absorb_reps = read_batch(a.batch)
    facts = read_map_facts(a.map, terrain["map_unit_column"], terrain["map_fact_column"])
    facts_norm = {u: norm(f) for u, f in facts.items()}

    checks, flags = [], []
    halted = False
    fname = N(Path(a.batch).name)

    # ① 행수·유형 분포
    m = re.search(r"(\d+)\s*문항", fname)
    declared = int(m.group(1)) if m else None
    from collections import Counter
    dist = Counter(r["유형"] for r in rows)
    if not rows:
        checks.append(("① 행수·유형", "HALT", "행 0 — 카운트 미실측(빈 값)"))
        halted = True
    elif declared is not None and declared != len(rows):
        checks.append(("① 행수·유형", "FAIL", f"실측 {len(rows)} ≠ 선언 {declared} (파일명)"))
    else:
        d = f" = 선언 {declared}" if declared is not None else " (선언 없음 — 실측만)"
        checks.append(("① 행수·유형", "PASS", f"실측 {len(rows)}{d} · 분포 {dict(dist)}"))

    # ② 1축 문자 대조
    mism = []
    for r in rows:
        cits = cit_re.findall(r["출처"])
        # [P3] 구분자 없는 join은 인접 두 fact의 경계에 걸친 위조 발췌가 우연 매치될 수 있다
        pool = "\x1f".join(facts_norm.get(c, "") for c in cits)
        for seg in split_segments(r["발췌"]):
            if norm(seg) not in pool:
                mism.append((r["ID"], seg[:40]))
    if mism:
        checks.append(("② 1축 문자대조", "FAIL", f"불일치 {len(mism)} — 예: {mism[:3]}"))
    else:
        nseg = sum(len(split_segments(r["발췌"])) for r in rows)
        checks.append(("② 1축 문자대조", "PASS", f"세그 {nseg}건 전건 포함 대조 · 불일치 0"))

    # ③ citation 실재
    missing_cit = []
    for r in rows:
        for c in cit_re.findall(r["출처"]):
            if c not in facts:
                missing_cit.append((r["ID"], c))
    ncit = sum(len(cit_re.findall(r["출처"])) for r in rows)
    if missing_cit:
        checks.append(("③ citation 실재", "FAIL", f"맵 부재 {len(missing_cit)} — 예: {missing_cit[:3]}"))
    else:
        checks.append(("③ citation 실재", "PASS", f"citation {ncit}건 전건 맵 실재"))

    # ④ 질문 중복 — 배치 내 + [수리 2026-08-18] 전 차수(union)와도 대조.
    # 실측: 15차 CI-793이 8차 CI-480과 동일 질문인데 7종 PASS — 배치 내부만 보고 있었음
    # (같은 지식이 성적에 이중 반영되는 왜곡). union 질문 합집합과 교차 검사 추가.
    qn = Counter(norm(r["질문"]) for r in rows if r["질문"])
    dups = {q: c for q, c in qn.items() if c > 1}
    prev_q = set()
    for up in (a.union or []):
        try:
            urows, _, _, _ = read_batch(up)
        except SystemExit:
            continue
        prev_q |= {norm(r["질문"]) for r in urows if r["질문"]}
    cross = [r["ID"] for r in rows if r["질문"] and norm(r["질문"]) in prev_q]
    fail = bool(dups) or bool(cross)
    checks.append(("④ 질문 중복", "FAIL" if fail else "PASS",
                   f"배치 내 {len(dups)} · 전 차수와 {len(cross)}"
                   + (f" — 예: {(list(dups)[:1] + cross[:2])[:2]}" if fail else " · 0")))

    # ⑧ 제품명 내부코드 검사 [2026-07-23 설계본부 지적: 결함본도 7종 PASS — 제품명 결함을
    #    구조적으로 못 잡았음. 질문·정답에 내부 트랙 코드(RC2 등)가 노출되면 응시 불가 시험지]
    prod_name = terrain.get("product_name")
    if prod_name and prod_name != a.product:
        leak = [r["ID"] for r in rows
                if re.search(rf"(?<![A-Za-z0-9]){re.escape(a.product)}(?![A-Za-z0-9-])", f"{r.get('질문','')} {r.get('정답','')}")]
        checks.append(("⑧ 제품명 내부코드", "FAIL" if leak else "PASS",
                       f"질문·정답에 '{a.product}' 노출 {len(leak)}건" + (f" — 예: {leak[:3]}" if leak else f" · 0 (제품명={prod_name})")))
    else:
        checks.append(("⑧ 제품명 내부코드", "SKIP", "terrain product_name 미설정 — 등록 권장"))

    # ⑤ 커버 등식
    if eq_rows:
        nums = {}
        for er in eq_rows:
            if len(er) >= 2 and re.fullmatch(r"\d+", er[1] if len(er) > 1 else ""):
                key = er[0]
                for k, tag in [("직접", "직접"), ("흡수", "흡수"), ("부적격", "부적격"), ("잔여", "잔여"), ("풀", "풀")]:
                    if k in key:
                        nums[tag] = int(er[1])
        if {"직접", "흡수", "부적격", "잔여", "풀"} <= set(nums):
            s = nums["직접"] + nums["흡수"] + nums["부적격"] + nums["잔여"]
            if s == nums["풀"]:
                checks.append(("⑤ 커버 등식", "PASS",
                               f"{nums['직접']}+{nums['흡수']}+{nums['부적격']}+{nums['잔여']}={nums['풀']} · 소실 0"))
            else:
                lost = nums["풀"] - s
                checks.append(("⑤ 커버 등식", "HALT" if lost > 0 else "FAIL",
                               f"등식 불일치: 합 {s} vs 풀 {nums['풀']} (소실 {lost})"))
                if lost > 0:
                    halted = True
        else:
            checks.append(("⑤ 커버 등식", "WARN", f"등식 시트 있으나 5항 파싱 불가 — 수동 확인 (파싱: {nums})"))
    else:
        checks.append(("⑤ 커버 등식", "SKIP", "등식 시트 없음 (파일럿/부분 차수) — 마감 배치는 필수"))

    # ⑥ 흡수 대표 ∈ 합집합 (union 누락 시 실행 거부 — 합집합 원칙)
    if absorb_reps:
        if a.union is None:
            print("[사용 오류] 흡수 대장이 있는 배치는 --union <전 차수 파일들> 필수 — 합집합 원칙(오반려 사고 2회 재발 방지). 실행 거부.")
            sys.exit(2)
        uids, ucits = collect_direct_union(list(a.union) + [a.batch])
        bad = [rep for rep in absorb_reps
               if rep not in uids and rep not in ucits
               and not any(c in uids or c in ucits for c in re.findall(r"[A-Z]{2,}(?:-[A-Z0-9]+)+", rep))]
        checks.append(("⑥ 흡수 합집합", "FAIL" if bad else "PASS",
                       f"흡수 대표 {len(absorb_reps)} 중 합집합 밖 {len(bad)}" + (f" — 예: {bad[:3]}" if bad else "")))
    else:
        checks.append(("⑥ 흡수 합집합", "SKIP", "흡수 대장 없음"))

    # ⑦ 앵커 — 의무 범위는 terrain.anchor_required_prefixes (실측 근거로 확정)
    req_prefixes = tuple(terrain.get("anchor_required_prefixes", []))
    if anchor_res and req_prefixes:
        miss_anchor, suspects = [], []
        n_req = 0
        for r in rows:
            if r["유형"] == "E" or not r["ID"].startswith(req_prefixes):
                continue
            n_req += 1
            body = r["정답"]
            if any(p.search(body) for p in anchor_res):
                continue
            cand = suspect_re.findall(body)
            if cand:
                suspects.append((r["ID"], cand[:2]))
            else:
                miss_anchor.append(r["ID"])
        if n_req == 0:
            checks.append(("⑦ 앵커", "SKIP", f"앵커 의무 접두({','.join(req_prefixes)}) 해당 문항 없음"))
        elif miss_anchor:
            checks.append(("⑦ 앵커", "FAIL", f"의무 대상 {n_req} 중 앵커 부재 {len(miss_anchor)} — 예: {miss_anchor[:5]}"))
        else:
            checks.append(("⑦ 앵커", "PASS", f"의무 대상 {n_req} 전건 앵커 (변형 의심 {len(suspects)} — 플래그)"))
        for sid, cand in suspects:
            flags.append({"type": "앵커 변형 의심", "id": sid, "candidates": cand,
                          "ack_required": True, "note": "FAIL 아님 — 사람 ack 필수 (P-신규-1, RVB-G07)"})
    else:
        checks.append(("⑦ 앵커", "SKIP", f"{a.product} terrain에 앵커 패턴 없음"))

    # ── 판정
    has_fail = any(s == "FAIL" for _, s, _ in checks)
    verdict = "HALTED" if halted else ("REJECTED" if has_fail else "PASS")

    W = 16
    print(f"\n=== verify_batch — {fname} (product={a.product}, sheet={sheet}) ===")
    for name, st, msg in checks:
        print(f"  {name:<{W}} [{st:^6}] {msg}")
    if flags:
        print(f"  ── 플래그 {len(flags)}건 (게이트 카드 필수 ack 항목으로 승격):")
        for f in flags:
            print(f"     · {f['type']} {f['id']} {f['candidates']}")
    print(f"  ▶ 판정: {verdict}" + ("  (반려는 일상, HALT는 사고)" if verdict == "REJECTED" else ""))

    result = {"batch": fname, "product": a.product, "sheet": sheet, "rows": len(rows),
              "checks": [{"name": n, "status": s, "msg": m} for n, s, m in checks],
              "flags": flags, "verdict": verdict}
    if a.json:
        Path(a.json).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    sys.exit(3 if halted else (1 if has_fail else 0))


if __name__ == "__main__":
    main()
