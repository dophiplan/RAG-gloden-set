"""2축 검증용 배치 분할기 - kit/ 경로 참조"""
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from common.log import log

BATCH_SIZE = 15
KIT_DIR = Path(__file__).parent.parent / "kit"
JUDGE_PROMPT = KIT_DIR / "judge_prompt_2axis_표준.md"


def load_goldenset(goldenset_path: Path) -> list:
    """골든셋 xlsx 로드"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl이 필요합니다")
    
    wb = openpyxl.load_workbook(goldenset_path, data_only=True)
    ws = wb['골든셋본문'] if '골든셋본문' in wb.sheetnames else wb.active
    
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    col_map = {}
    for standard_name, aliases in {
        'ID': ['ID', 'id'], '유형': ['유형', 'type'], '질문': ['질문', 'question'],
        '정답': ['정답', '정답(필수 포함 요소)'], '근거 출처': ['근거 출처', 'citation'],
        '합격 기준': ['합격 기준'], '난이도': ['난이도'], '기대 라우팅': ['기대 라우팅'],
        '검증 상태': ['검증 상태'], '근거 원문 발췌': ['근거 원문 발췌'],
        'acl_level': ['acl_level'], 'answer_type': ['answer_type']
    }.items():
        for idx, header in enumerate(headers):
            if any(alias in header for alias in aliases):
                col_map[standard_name] = headers[header]
                break
    
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_map.get('ID', 0)]:
            continue
        
        answer_type = 'single'
        if 'answer_type' in col_map and row[col_map['answer_type']]:
            at_val = str(row[col_map['answer_type']]).strip().lower()
            answer_type = 'multi_equivalent' if 'multi' in at_val else 'single'
        
        item = {
            'id': str(row[col_map['ID']]),
            'type': str(row[col_map.get('유형', 0)]) if '유형' in col_map else '',
            'difficulty': str(row[col_map.get('난이도', 0)]) if '난이도' in col_map else '',
            'question': str(row[col_map['질문']]),
            'gold_answer_required': str(row[col_map.get('정답', 0)]),
            'answer_type': answer_type,
            'pass_criteria': str(row[col_map.get('합격 기준', 0)]) if '합격 기준' in col_map else '',
            'citation': str(row[col_map.get('근거 출처', 0)]) if '근거 출처' in col_map else '',
            'source_code': '',
            'evidence_excerpt': str(row[col_map.get('근거 원문 발췌', 0)]) if '근거 원문 발췌' in col_map else '',
            'acl_level': str(row[col_map.get('acl_level', 0)]) if 'acl_level' in col_map else 'all',
            'axis1_status': str(row[col_map.get('검증 상태', 0)]) if '검증 상태' in col_map else '초안'
        }
        
        if item['citation']:
            parts = item['citation'].split(';')
            if len(parts) > 1:
                item['source_code'] = parts[1].strip()
        
        items.append(item)
    
    wb.close()
    return items


def get_judge_prompt_content() -> str:
    """kit/에서 judge 프롬프트 내용 읽기"""
    if JUDGE_PROMPT.exists():
        with open(JUDGE_PROMPT, 'r', encoding='utf-8') as f:
            return f.read()
    return "[judge_prompt_2axis_표준.md 파일 없음 - kit/ 폴터 확인]"


def make_batches(product: str, goldenset_path: Path, handoff_dir: Path, batch_size: int = BATCH_SIZE) -> dict:
    """배치 분할"""
    items = load_goldenset(goldenset_path)
    handoff_dir.mkdir(parents=True, exist_ok=True)
    
    # 전체 jsonl
    jsonl_path = handoff_dir / 'goldenset_2axis.jsonl'
    with open(jsonl_path, 'w', encoding='utf-8') as f:
        for item in items:
            f.write(json.dumps(item, ensure_ascii=False) + '\n')
    
    # 배치 분할
    batch_files = []
    for idx, start in enumerate(range(0, len(items), batch_size), start=1):
        batch = items[start:start + batch_size]
        end = min(start + batch_size, len(items))
        
        batch_path = handoff_dir / f"batch{idx:02d}_{start+1:03d}-{end:03d}.txt"
        
        content_lines = ["다음 문항들을 판정해줘:", "```"]
        for item in batch:
            content_lines.append(json.dumps(item, ensure_ascii=False))
        content_lines.append("```")
        
        with open(batch_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(content_lines) + '\n')
        
        batch_files.append(str(batch_path.name))
    
    # 프롬프트 내용 읽기
    judge_content = get_judge_prompt_content()
    
    log(product, '4_2axis_handoff_prep', goldenset_path.name,
        f"배치 분할 완료: {len(items)}건 → {len(batch_files)}배치", "§2.1/HANDOFF")
    
    return {
        'product': product,
        'total_items': len(items),
        'batch_count': len(batch_files),
        'jsonl_file': str(jsonl_path),
        'batch_files': batch_files,
        'handoff_dir': str(handoff_dir),
        'judge_prompt': judge_content[:500] + "..." if len(judge_content) > 500 else judge_content
    }


def main():
    import argparse
    parser = argparse.ArgumentParser(description='2축 배치 분할')
    parser.add_argument('--product', '-p', required=True, choices=['RC', 'RV', 'RM', 'HR'])
    parser.add_argument('--goldenset', '-g', required=True, type=Path)
    parser.add_argument('--handoff-dir', '-d', type=Path, default=None)
    parser.add_argument('--batch-size', '-b', type=int, default=BATCH_SIZE)
    
    args = parser.parse_args()
    
    if args.handoff_dir is None:
        args.handoff_dir = Path(__file__).parent.parent / 'products' / args.product / 'handoff'
    
    result = make_batches(args.product, args.goldenset, args.handoff_dir, args.batch_size)
    
    print(f"\n=== 배치 분할 완료 ({args.product}) ===")
    print(f"총 문항: {result['total_items']}건")
    print(f"배치 수: {result['batch_count']}개")
    print(f"\n[Judge 프롬프트 - kit/]")
    print(result['judge_prompt'][:300] + "...")
    print(f"\n[핸드오프 안내]")
    print(f"1. kit/judge_prompt_2axis_표준.md를 생성과 다른 AI 세션에 로드")
    print(f"2. 각 batch 파일을 순서대로 복사해 검증 요청")
    print(f"3. 결과 JSON을 {result['handoff_dir']}/에 저장")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
