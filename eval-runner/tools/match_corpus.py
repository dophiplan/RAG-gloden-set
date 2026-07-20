"""
match_corpus.py — 골든셋 ↔ RAG 코퍼스 export 대조 도구

역할 분리 (v2 설계):
[작업1] a/b 대조 = 문서 수준, 커버리지맵 경유
  골든셋 citation_id → 커버리지맵 Unit ID 조인 → 단위의 source/resource(URL)
  → canonical_url+url_variants로 doc_key 접기 → export docs와 대조
  - source/resource가 URL이 아닌 것(파일 경로 등) → "대조불가(파일원천)" (사람확인 대상, b 아님)
  - 커버리지맵에 Unit 자체가 없는 것 → "대조불가(조인실패)" (사람확인 대상, b 아님)
[작업2] 발췌 매칭 = 청크 수준 매핑 전용
  발췌→청크 공백무시 매칭 (corpus_loader.py의 normalize_text/verify_excerpt 로직 재사용)
  - 발췌를 못 찾은 것은 "매핑불가"이지 인입누락(b)이 아님 — 두 실패를 섞지 않는다
  - 문서는 (a)인데 발췌가 매핑불가 → 청킹 경계 문제 후보 (별도 시트)

출력 4종 (전부 xlsx):
  1. 대조표.xlsx        문항ID/근거문서/doc_key/분류
  2. 인입누락_리포트.xlsx  (b) 문서 단위 집계 (시스템팀 전달용) + 대조불가 시트(사람확인)
  3. 매핑시트.xlsx       발췌→point_id 매핑 + 청킹경계후보 시트
  4. 파일럿후보.xlsx      (a)+검증완료 층화 추출 25±5 + RAG답변/RAG출처 기록칸

사용:
  python3 tools/match_corpus.py -g 골든셋.xlsx -e export폴더/ -m 커버리지맵.xlsx -o 출력폴더/

주의: 골든셋/커버리지맵은 읽기 전용 — 절대 수정하지 않는다.
"""
import argparse
import csv
import json
import math
import random
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("오류: openpyxl이 필요합니다. pip install openpyxl")
    sys.exit(1)


# ============================================================
# 공통: 텍스트 정규화 (corpus_loader.normalize_text 로직 재사용)
# ============================================================

def normalize_text(text) -> str:
    """공백 제거하여 정규화 (공백무시 비교용)"""
    if text is None:
        return ""
    return ''.join(str(text).split())


def normalize_url(url: str) -> str:
    """URL 정규화: 스킴/트레일링 슬래시/대소문자 무시 (쿼리는 보존 — terms?20170403 같은 변형 구분)"""
    u = str(url).strip()
    u = re.sub(r'^https?://', '', u, flags=re.I)
    u = u.rstrip('/')
    return u.lower()


# ============================================================
# export 로더 (docs / chunks — JSON 우선, CSV 폴백)
# ============================================================

def _find_export_file(export_dir: Path, kind: str):
    """export 폴더에서 docs-*/chunks-* 파일 자동 감지 (경로 하드코딩 금지)"""
    for ext in ('json', 'csv'):
        cands = sorted(export_dir.glob(f'{kind}-*.{ext}')) or sorted(export_dir.glob(f'{kind}*.{ext}'))
        if cands:
            return cands[-1], ext  # 최신(사전순 마지막) 선택
    return None, None


def load_docs(export_dir: Path) -> list:
    path, ext = _find_export_file(export_dir, 'docs')
    if not path:
        print(f"오류: export 폴더에서 docs 파일을 찾을 수 없습니다: {export_dir}")
        sys.exit(1)
    print(f"docs 로드: {path.name}")
    if ext == 'json':
        data = json.loads(path.read_text(encoding='utf-8'))
        docs = data['documents'] if isinstance(data, dict) else data
    else:
        with open(path, encoding='utf-8-sig') as f:
            docs = list(csv.DictReader(f))
        for d in docs:  # CSV의 url_variants는 문자열 → 리스트화
            v = d.get('url_variants') or ''
            d['url_variants'] = [x.strip() for x in re.split(r'[;|,]\s*', v) if x.strip()] if isinstance(v, str) else v
    # 필수 컬럼 확인 (추측 금지 — 실제 키 검증)
    need = {'doc_key', 'canonical_url', 'url_variants'}
    missing = need - set(docs[0].keys())
    if missing:
        print(f"오류: docs 파일에 필수 컬럼 없음: {missing} / 실제: {list(docs[0].keys())}")
        sys.exit(1)
    return docs


def load_chunks(export_dir: Path) -> list:
    path, ext = _find_export_file(export_dir, 'chunks')
    if not path:
        print(f"오류: export 폴더에서 chunks 파일을 찾을 수 없습니다: {export_dir}")
        sys.exit(1)
    print(f"chunks 로드: {path.name}")
    if ext == 'json':
        data = json.loads(path.read_text(encoding='utf-8'))
        chunks = data['chunks'] if isinstance(data, dict) else data
    else:
        with open(path, encoding='utf-8-sig') as f:
            chunks = list(csv.DictReader(f))
    need = {'point_id', 'doc_key', 'text'}
    missing = need - set(chunks[0].keys())
    if missing:
        print(f"오류: chunks 파일에 필수 컬럼 없음: {missing} / 실제: {list(chunks[0].keys())}")
        sys.exit(1)
    return chunks


def build_url_index(docs: list) -> dict:
    """canonical_url + url_variants → doc_key 접기용 인덱스"""
    idx = {}
    for d in docs:
        dk = d['doc_key']
        urls = [d.get('canonical_url') or ''] + list(d.get('url_variants') or [])
        for u in urls:
            if u:
                idx[normalize_url(u)] = dk
    return idx


def fold_url_to_dockey(url: str, url_index: dict):
    """URL → doc_key 접기 (문자열 직접 비교 금지 원칙의 구현부)"""
    return url_index.get(normalize_url(url))


# ============================================================
# 커버리지맵 로더 (Unit ID → source/resource)
# ============================================================

def load_coverage_map(path: Path) -> dict:
    """Unit ID 컬럼을 가진 모든 시트에서 단위 수집 → {unit_id: {sheet, source, status}}"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    units = {}
    for ws in wb.worksheets:
        hdr = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        low = [h.lower() for h in hdr]
        uid_i = src_i = st_i = None
        for i, h in enumerate(low):
            if uid_i is None and 'unit' in h and 'id' in h:
                uid_i = i
            if src_i is None and 'source' in h:
                src_i = i
            if st_i is None and h == '상태':
                st_i = i
        if uid_i is None or src_i is None:
            continue  # 단위 시트 아님 (Summary/안내 등)
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            uid = row[uid_i] if len(row) > uid_i else None
            if not uid or not str(uid).strip():
                continue
            uid = str(uid).strip()
            if uid in units:
                continue  # 최초 등장 시트 우선
            units[uid] = {
                'sheet': ws.title,
                'source': str(row[src_i]).strip() if src_i < len(row) and row[src_i] else '',
                'status': str(row[st_i]).strip() if st_i is not None and st_i < len(row) and row[st_i] else '',
            }
            n += 1
        print(f"  커버리지맵 시트 '{ws.title}': 단위 {n}건")
    wb.close()
    if not units:
        print(f"오류: 커버리지맵에서 Unit ID/source 컬럼을 가진 시트를 찾지 못했습니다: {path}")
        sys.exit(1)
    return units


# ============================================================
# 골든셋 로더 (컬럼 자동 감지 — run_1axis.py 방식)
# ============================================================

def load_goldenset(path: Path) -> tuple:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]  # 본문 시트 = 첫 시트
    hdr = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def find_col(pred, exclude=None):
        for i, h in enumerate(hdr):
            hl = h.lower()
            if exclude and any(x in hl for x in exclude):
                continue
            if pred(hl):
                return i
        return None

    col = {
        'id':      find_col(lambda h: h == 'id' or '문항' in h or h in ('no', '번호')),
        'type':    find_col(lambda h: '유형' in h or h == 'type'),
        'question': find_col(lambda h: '질문' in h or 'question' in h),
        'citation': find_col(lambda h: '출처' in h and ('citation' in h or '근거' in h)),
        'excerpt': find_col(lambda h: '발췌' in h or 'excerpt' in h),
        'verify':  find_col(lambda h: '검증' in h and '상태' in h),
        'answer':  find_col(lambda h: '정답' in h),
        'difficulty': find_col(lambda h: '난이도' in h),
        'srcset':  find_col(lambda h: '출처셋' in h),
    }
    missing = [k for k in ('id', 'type', 'question', 'citation', 'excerpt') if col[k] is None]
    if missing:
        print(f"오류: 골든셋 필수 컬럼 감지 실패: {missing} / 헤더: {hdr}")
        sys.exit(1)
    print(f"골든셋 컬럼 감지: " + ", ".join(f"{k}='{hdr[v]}'" for k, v in col.items() if v is not None))

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rid = row[col['id']] if len(row) > col['id'] else None
        if not rid or not str(rid).strip():
            continue
        def get(k):
            i = col.get(k)
            return str(row[i]).strip() if i is not None and i < len(row) and row[i] is not None else ''
        items.append({
            'id': str(rid).strip(), 'type': get('type').upper(), 'question': get('question'),
            'citation': get('citation'), 'excerpt': get('excerpt'), 'verify': get('verify'),
            'answer': get('answer'), 'difficulty': get('difficulty'), 'srcset': get('srcset'),
        })
    wb.close()
    return items, hdr


def extract_citation_ids(citation: str) -> list:
    """'RM-ERRDEF-30001 ; [RM-ERRDEF]Network Engine 30001' → citation_id 목록.
    ';' 분리 후 Unit ID 패턴(RM-XXX-…)인 토큰만 채택. 첫 토큰은 항상 후보."""
    ids = []
    for i, part in enumerate(str(citation).split(';')):
        tok = part.strip()
        if not tok:
            continue
        if i == 0 or re.fullmatch(r'[A-Z]{2,5}-[A-Z0-9]+(-[A-Za-z0-9]+)+', tok):
            if tok not in ids:
                ids.append(tok)
    return ids


# ============================================================
# 작업1: a/b 대조 (문서 수준, 커버리지맵 경유)
# ============================================================

CLS_A = 'a(인입됨)'
CLS_B = 'b(인입누락)'
CLS_FILE = '대조불가(파일원천)'
CLS_NOJOIN = '대조불가(조인실패)'
CLS_E = 'E형(해당없음)'


def classify_item(item, units, url_index, doc_title_index):
    """반환: (분류, doc_keys, sources, sheets, note)"""
    if item['type'] == 'E' and not item['excerpt']:
        return CLS_E, [], [], [], 'E형 — 의도된 코퍼스 부재 (작업2에서 키워드 검색)'

    cids = extract_citation_ids(item['citation'])
    joined = [(cid, units[cid]) for cid in cids if cid in units]
    if not joined:
        return CLS_NOJOIN, [], [], [], f"커버리지맵에 Unit 없음: {', '.join(cids) if cids else '(citation_id 파싱 실패)'}"

    url_sources, file_sources, sheets = [], [], []
    for cid, u in joined:
        (url_sources if re.match(r'^https?://', u['source'], re.I) else file_sources).append(u['source'])
        sheets.append(u['sheet'])
    url_sources = list(dict.fromkeys(url_sources))
    file_sources = list(dict.fromkeys(file_sources))
    sheets = list(dict.fromkeys(sheets))

    if url_sources:
        doc_keys = list(dict.fromkeys(dk for u in url_sources if (dk := fold_url_to_dockey(u, url_index))))
        if doc_keys:
            return CLS_A, doc_keys, url_sources, sheets, ''
        note = 'URL이 export docs에 없음'
        if file_sources:
            note += f" (파일원천 병존: {len(file_sources)}건)"
        return CLS_B, [], url_sources, sheets, note

    # 파일 경로 원천 → b로 단정 금지, 사람확인 대상
    hints = []
    for fs in file_sources:
        base = normalize_text(Path(fs).name).lower()
        for t_norm, dk in doc_title_index.items():
            if base and (base in t_norm or t_norm in base):
                hints.append(f"유사 파일명 export 존재? {dk}")
    note = '파일 경로 원천 — URL 접기 불가, 사람확인 필요' + (f" / {'; '.join(sorted(set(hints)))}" if hints else '')
    return CLS_FILE, [], file_sources, sheets, note


# ============================================================
# 작업2: 발췌 → 청크 매핑 (corpus_loader.verify_excerpt 로직 재사용)
# ============================================================

def match_excerpt_to_chunks(excerpt: str, chunks_norm: list):
    """strict → 공백무시 순서로 청크 매칭.
    반환: (매칭방식, [(point_id, doc_key), ...])  / 못 찾으면 ('매핑불가', [])"""
    ex = str(excerpt)
    strict = [(pid, dk) for pid, dk, raw, _ in chunks_norm if ex in raw]
    if strict:
        return '정확일치', strict
    nex = normalize_text(ex)
    if len(nex) < 3:  # corpus_loader와 동일: 너무 짧으면 정확도 낮음
        return '매핑불가', []
    ws = [(pid, dk) for pid, dk, _, norm in chunks_norm if nex in norm]
    if ws:
        return '공백무시', ws
    # 부분 매칭(앞/뒤 20자) — point_id 단정에는 쓰지 않고 참고 정보로만 (추측 금지)
    pre, suf = nex[:min(20, len(nex))], nex[-min(20, len(nex)):]
    partial = [(pid, dk) for pid, dk, _, norm in chunks_norm if pre in norm and suf in norm]
    if partial:
        return '매핑불가(부분일치참고)', partial
    return '매핑불가', []


E_STOPWORDS = {
    '리모트미팅', 'remotemeeting', '기능', '기능이', '있나요', '있습니까', '있는지', '가능한가요',
    '지원하나요', '지원되나요', '지원', '제공하나요', '제공되나요', '제공', '되나요', '하나요',
    '회의', '화상회의', '무엇인가요', '어떻게', '합니까', '인가요', '수', '중', '및', '또는', '때',
}


def e_type_keyword_search(question: str, chunks_norm: list, rare_df=0.02, max_df=0.10):
    """E형: 질문 핵심 키워드를 chunks 전체에서 검색.

    노이즈 억제 규칙 (실데이터 캘리브레이션 완료):
    - 문서빈도(df) > max_df(10%)인 범용 키워드(예: '제한')는 판정에서 제외
    - '발견됨' 판정 조건: (1) 희귀 키워드(df ≤ rare_df=2%) 단독 히트,
      또는 (2) 동일 청크 내 키워드 2개 이상 공존
    - 그 외 범용 키워드 단독 히트는 '코퍼스부재확인' + 참고 표기 (사람확인 부담 축소)
    반환: (판정 문자열, 발견 point_id 목록, 사용 키워드 목록)"""
    n_total = len(chunks_norm) or 1
    lows = [(pid, norm.lower()) for pid, dk, _, norm in chunks_norm]
    tokens = re.findall(r'[가-힣A-Za-z0-9]{2,}', question)
    cand = list(dict.fromkeys(t for t in tokens if t.lower() not in E_STOPWORDS))

    kws = []  # (keyword, df_ratio, hit_pids)
    weak = []
    for kw in cand:
        nk = normalize_text(kw).lower()
        hits = [pid for pid, low in lows if nk in low]
        if not hits:
            continue
        ratio = len(hits) / n_total
        if ratio > max_df:
            continue  # 범용어 — 판정 제외
        kws.append((kw, ratio, hits))

    strong = [(kw, hits) for kw, ratio, hits in kws if ratio <= rare_df]
    cooccur = []
    if len(kws) >= 2:
        for pid, low in lows:
            present = [kw for kw, _, _ in kws if normalize_text(kw).lower() in low]
            if len(present) >= 2:
                cooccur.append((pid, present))
    weak = [(kw, hits) for kw, ratio, hits in kws if ratio > rare_df]

    if strong or cooccur:
        parts = [f"{kw}: {','.join(h[:5])}" for kw, h in strong]
        parts += [f"공존[{'+'.join(pres)}]: {pid}" for pid, pres in cooccur[:5]]
        pids = list(dict.fromkeys([p for _, h in strong for p in h[:5]] + [pid for pid, _ in cooccur]))
        return f"발견됨({'; '.join(parts)})", pids, [k for k, _, _ in kws]
    note = f" (약한 단독히트 참고: {', '.join(kw for kw, _ in weak)})" if weak else ''
    return f"코퍼스부재확인{note}", [], [k for k, _, _ in kws]


# ============================================================
# 출력 4종
# ============================================================

HDR_FILL = PatternFill('solid', fgColor='DDDDDD')


def _write_sheet(ws, header, rows, widths=None):
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = HDR_FILL
    for r in rows:
        ws.append([str(x) if x is not None else '' for x in r])
    for i, w in enumerate(widths or []):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = w
    ws.freeze_panes = 'A2'


def write_matrix(out: Path, results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '대조표'
    rows = [(r['id'], r['type'], r['verify'], '; '.join(extract_citation_ids(r['citation'])),
             '; '.join(r['sources']), '; '.join(r['sheets']), '; '.join(r['doc_keys']),
             r['cls'], r['note']) for r in results]
    _write_sheet(ws, ['문항ID', '유형', '검증상태', 'citation_id', '근거문서(source/resource)',
                      '커버리지맵시트', 'doc_key', '분류', '비고'], rows,
                 [12, 6, 10, 24, 55, 14, 30, 18, 45])
    wb.save(out)


def write_missing_report(out: Path, results):
    wb = openpyxl.Workbook()
    # (b) 문서 단위 집계 — 시스템팀 전달용
    by_doc = defaultdict(list)
    for r in results:
        if r['cls'] == CLS_B:
            for s in (r['sources'] or ['(원천 미상)']):
                by_doc[s].append(r['id'])
    ws = wb.active
    ws.title = '인입누락(b)_문서단위'
    rows = [(doc, len(ids), '; '.join(ids)) for doc, ids in
            sorted(by_doc.items(), key=lambda kv: -len(kv[1]))]
    _write_sheet(ws, ['문서(source/resource URL)', '영향 문항수', '문항ID'], rows, [70, 12, 60])

    # 대조불가 — 사람확인 대상 (b와 절대 섞지 않음)
    for cls, title in ((CLS_FILE, '대조불가_파일원천(사람확인)'), (CLS_NOJOIN, '대조불가_조인실패(사람확인)')):
        by_src = defaultdict(list)
        for r in results:
            if r['cls'] == cls:
                key = '; '.join(r['sources']) if r['sources'] else r['note']
                by_src[key].append(r['id'])
        ws2 = wb.create_sheet(title)
        rows2 = [(src, len(ids), '; '.join(ids)) for src, ids in
                 sorted(by_src.items(), key=lambda kv: -len(kv[1]))]
        _write_sheet(ws2, ['원천/사유', '영향 문항수', '문항ID'], rows2, [70, 12, 60])
    wb.save(out)


def write_mapping(out: Path, results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '매핑시트'
    rows = []
    for r in results:
        if r['type'] == 'E' and not r['excerpt']:
            rows.append((r['id'], r['cls'], '; '.join(r['doc_keys']), r['e_result'],
                         '; '.join(r['e_pids']), 'E형 키워드검색', ''))
        else:
            pids = '; '.join(p for p, _ in r['map_hits']) if r['map_method'] in ('정확일치', '공백무시') else '매핑불가'
            chunk_dks = '; '.join(dict.fromkeys(dk for _, dk in r['map_hits']))
            note = ''
            if r['map_method'] == '매핑불가(부분일치참고)':
                note = f"부분일치(앞/뒤 20자) 참고: {chunk_dks} — point_id 단정 금지"
            elif not r['excerpt']:
                pids, note = '매핑불가', '발췌 공란(유형≠E)'
            rows.append((r['id'], r['cls'], '; '.join(r['doc_keys']), pids, chunk_dks,
                         r['map_method'] if r['excerpt'] else '발췌없음', note))
    _write_sheet(ws, ['문항ID', '문서분류(작업1)', '문서doc_key(작업1)', 'point_id(작업2)',
                      '청크doc_key(작업2)', '매칭방식', '비고'], rows, [12, 18, 30, 45, 30, 20, 45])

    # 문서는 (a)인데 발췌 매핑불가 → 청킹 경계 문제 후보
    ws2 = wb.create_sheet('청킹경계후보(a인데_매핑불가)')
    rows2 = [(r['id'], '; '.join(r['doc_keys']), r['excerpt'][:200], r['map_method'])
             for r in results
             if r['cls'] == CLS_A and r['excerpt'] and r['map_method'].startswith('매핑불가')]
    _write_sheet(ws2, ['문항ID', '문서doc_key', '발췌(앞 200자)', '매칭방식'], rows2, [12, 30, 80, 22])
    wb.save(out)


def select_pilot(results, size, tol, seed):
    """(a)+검증완료 층화 추출: 유형 비율 원본 유사, E형 2~3개, 같은 문서 편중 금지"""
    rng = random.Random(seed)
    pool = [r for r in results if r['cls'] == CLS_A and r['verify'] == '검증완료']
    e_pool = [r for r in results if r['cls'] == CLS_E and r['verify'] == '검증완료']
    e_pool_absent = [r for r in e_pool if r['e_result'].startswith('코퍼스부재확인')] or e_pool

    e_take = min(3, max(2, 3), len(e_pool_absent)) if e_pool_absent else 0
    e_take = min(e_take, 3)
    target_rest = size - e_take
    warns = []

    if len(pool) < target_rest:
        warns.append(f"파일럿 풀 부족: (a)+검증완료 {len(pool)}건 < 목표 {target_rest}건 "
                     f"— 가용분 전부 선정. 코퍼스 인입 보강 후 재실행하면 자동으로 {size}±{tol} 충족")
        target_rest = len(pool)

    # 유형 비율 배분
    by_type = defaultdict(list)
    for r in pool:
        by_type[r['type']].append(r)
    alloc = {t: (len(v) / len(pool)) * target_rest for t, v in by_type.items()} if pool else {}
    take = {t: int(a) for t, a in alloc.items()}
    rem = target_rest - sum(take.values())
    for t in sorted(alloc, key=lambda t: -(alloc[t] - take[t]))[:rem]:
        take[t] += 1

    # 문서 편중 상한: 동일 doc_key 최대 ceil(목표/문서수)+1
    n_docs = len(set(dk for r in pool for dk in r['doc_keys'])) or 1
    cap = max(2, math.ceil(target_rest / n_docs) + 1)

    selected, doc_load = [], Counter()
    for t, n in sorted(take.items()):
        cands = sorted(by_type[t], key=lambda r: r['id'])
        rng.shuffle(cands)
        got = 0
        for r in cands:
            if got >= n:
                break
            if any(doc_load[dk] >= cap for dk in r['doc_keys']):
                continue
            selected.append(r)
            for dk in r['doc_keys']:
                doc_load[dk] += 1
            got += 1
        if got < n:
            warns.append(f"유형 {t}: 문서편중 상한(cap={cap})으로 {n}건 중 {got}건만 선정")
    e_sel = sorted(e_pool_absent, key=lambda r: r['id'])
    rng.shuffle(e_sel)
    selected += e_sel[:e_take]
    if len(selected) < size - tol:
        warns.append(f"최종 파일럿 {len(selected)}건 — 목표 하한 {size - tol}건 미달")
    return selected, warns


def write_pilot(out: Path, selected):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '파일럿후보'
    rows = [(r['id'], r['type'], r['difficulty'], r['question'], r['answer'],
             '; '.join(extract_citation_ids(r['citation'])), '; '.join(r['doc_keys']),
             ('; '.join(p for p, _ in r['map_hits']) if r['map_method'] in ('정확일치', '공백무시')
              else (r['e_result'] if r['type'] == 'E' else '매핑불가')),
             '', '') for r in sorted(selected, key=lambda r: r['id'])]
    _write_sheet(ws, ['문항ID', '유형', '난이도', '질문', '정답(필수 포함 요소)', 'citation_id',
                      'doc_key', 'point_id/E형결과', 'RAG답변', 'RAG출처'],
                 rows, [12, 6, 8, 50, 50, 24, 30, 40, 50, 40])
    wb.save(out)


# ============================================================
# main
# ============================================================

def main():
    p = argparse.ArgumentParser(description='골든셋 ↔ RAG 코퍼스 export 대조 (문서수준 a/b + 청크수준 매핑)')
    p.add_argument('--golden', '-g', required=True, type=Path, help='골든셋 xlsx (읽기 전용)')
    p.add_argument('--export-dir', '-e', required=True, type=Path, help='export 폴더 (docs-*/chunks-* CSV/JSON)')
    p.add_argument('--coverage-map', '-m', required=True, type=Path,
                   help='커버리지맵 xlsx (citation_id→Unit ID→source URL 다리)')
    p.add_argument('--output-dir', '-o', required=True, type=Path, help='출력 폴더')
    p.add_argument('--pilot-size', type=int, default=25, help='파일럿 목표 문항수 (기본 25)')
    p.add_argument('--pilot-tol', type=int, default=5, help='파일럿 허용 편차 (기본 ±5)')
    p.add_argument('--seed', type=int, default=42, help='층화 추출 시드 (재현성)')
    args = p.parse_args()

    for path, name in ((args.golden, '골든셋'), (args.export_dir, 'export 폴더'), (args.coverage_map, '커버리지맵')):
        if not path.exists():
            print(f"오류: {name}을(를) 찾을 수 없습니다: {path}")
            return 1

    # 입력 로드
    docs = load_docs(args.export_dir)
    chunks = load_chunks(args.export_dir)
    print(f"  export: 문서 {len(docs)}건 / 청크 {len(chunks)}건")
    url_index = build_url_index(docs)
    doc_title_index = {normalize_text(d.get('title', '')).lower(): d['doc_key'] for d in docs if d.get('title')}
    units = load_coverage_map(args.coverage_map)
    print(f"  커버리지맵 단위 합계: {len(units)}건")
    items, _ = load_goldenset(args.golden)
    print(f"  골든셋 문항: {len(items)}건")

    chunks_norm = [(c['point_id'], c['doc_key'], str(c.get('text') or ''), normalize_text(c.get('text')))
                   for c in chunks]

    # 작업1 + 작업2
    results = []
    for it in items:
        cls, doc_keys, sources, sheets, note = classify_item(it, units, url_index, doc_title_index)
        r = {**it, 'cls': cls, 'doc_keys': doc_keys, 'sources': sources, 'sheets': sheets,
             'note': note, 'map_method': '', 'map_hits': [], 'e_result': '', 'e_pids': []}
        if cls == CLS_E:
            r['e_result'], r['e_pids'], _kw = e_type_keyword_search(it['question'], chunks_norm)
        elif it['excerpt']:
            r['map_method'], r['map_hits'] = match_excerpt_to_chunks(it['excerpt'], chunks_norm)
        else:
            r['map_method'] = '매핑불가'
            r['note'] = (r['note'] + ' / ' if r['note'] else '') + '발췌 공란(유형≠E)'
        results.append(r)

    # 출력
    args.output_dir.mkdir(parents=True, exist_ok=True)
    paths = {k: args.output_dir / f'{k}.xlsx' for k in ('대조표', '인입누락_리포트', '매핑시트', '파일럿후보')}
    write_matrix(paths['대조표'], results)
    write_missing_report(paths['인입누락_리포트'], results)
    write_mapping(paths['매핑시트'], results)
    pilot, pilot_warns = select_pilot(results, args.pilot_size, args.pilot_tol, args.seed)
    write_pilot(paths['파일럿후보'], pilot)

    # 요약 숫자
    n = Counter(r['cls'] for r in results)
    map_ok = sum(1 for r in results if r['map_method'] in ('정확일치', '공백무시'))
    map_fail_a = sum(1 for r in results if r['cls'] == CLS_A and r['excerpt']
                     and r['map_method'].startswith('매핑불가'))
    map_fail_all = sum(1 for r in results if r['cls'] != CLS_E and r['map_method'].startswith('매핑불가'))
    e_absent = sum(1 for r in results if r['cls'] == CLS_E and r['e_result'].startswith('코퍼스부재확인'))
    e_found = sum(1 for r in results if r['cls'] == CLS_E and r['e_result'].startswith('발견됨'))

    print("\n=== 대조 완료 ===")
    print(f"총 문항: {len(results)}")
    print(f"[작업1: 문서 수준 a/b — 커버리지맵 경유]")
    print(f"  (a) 인입됨:            {n[CLS_A]}")
    print(f"  (b) 인입누락:          {n[CLS_B]}")
    print(f"  대조불가(파일원천):     {n[CLS_FILE]}  ← 사람확인 대상 (b 아님)")
    print(f"  대조불가(조인실패):     {n[CLS_NOJOIN]}  ← 사람확인 대상 (b 아님)")
    print(f"  E형(해당없음):         {n[CLS_E]}")
    print(f"[작업2: 청크 수준 매핑 — 발췌 공백무시 매칭]")
    print(f"  매핑성공:              {map_ok}")
    print(f"  매핑불가(전체):         {map_fail_all}")
    print(f"  매핑불가(문서는 a인데): {map_fail_a}  ← 청킹 경계 문제 후보")
    print(f"  E형 부재확인:          {e_absent} / 발견됨(사람확인): {e_found}")
    print(f"[파일럿]")
    print(f"  파일럿 선정:           {len(pilot)} (목표 {args.pilot_size}±{args.pilot_tol}, seed={args.seed})")
    for w in pilot_warns:
        print(f"  ⚠ {w}")
    print("\n요약: (a) {a} / (b) {b} / 대조불가(파일원천) {f} / 매핑성공 {m} / 매핑불가(문서는 a인데) {mf} / E형 부재확인 {e} / 파일럿 {p}".format(
        a=n[CLS_A], b=n[CLS_B], f=n[CLS_FILE], m=map_ok, mf=map_fail_a, e=e_absent, p=len(pilot)))
    print(f"출력: {', '.join(str(v) for v in paths.values())}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
