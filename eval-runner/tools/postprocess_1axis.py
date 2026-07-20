"""
1축 검증 결과 후처리기 - ok_uncited는 커버리지맵에서 실제 Unit ID 찾기
"""
import json
import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from common.log import log, log_batch


def find_citation_id_for_excerpt(excerpt: str, coverage_map_path: Path) -> str:
    """
    발췌 내용으로부터 coverage_map.xlsx에서 진짜 citation_id 찾기
    
    원칙: "유일하게 확실할 때만 자동, 애매하면 사람."
    - 매칭 후보를 전부 모아서 세어
    - 정확히 1개일 때만 그 Unit ID를 채움
    - 2개 이상이면 어느 단위인지 모호하니 None 반환 (사람확인으로)
    - 발췌가 10자 미만이면 신뢰도 낮으니 None 반환 (사람확인으로)
    
    Returns:
        찾은 citation_id 또는 None (못 찾거나 애매하면 None 반환 - 가짜 값 금지)
    """
    if not coverage_map_path.exists():
        return None
    
    # 발췌가 10자 미만이면 신뢰도 낮음 → 사람확인
    if len(str(excerpt).strip()) < 10:
        return None
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(coverage_map_path, data_only=True)
        ws = wb.active
        
        # 헤더 찾기
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        citation_col = headers.get('citation_id')
        
        # 발췌가 있는 컬럼 찾기
        excerpt_col = None
        for key in headers:
            if any(k in key for k in ['발췌', 'excerpt', '내용', 'verbatim']):
                excerpt_col = headers[key]
                break
        
        if not citation_col:
            wb.close()
            return None
        
        # 정규화된 발췌 (공백 제거)
        normalized_excerpt = ''.join(str(excerpt).split())
        
        # 모든 매칭 후보 모으기
        matches = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            citation_id = row[citation_col - 1] if len(row) >= citation_col else None
            
            # 발췌 컬럼이 있으면 비교
            if excerpt_col and row[excerpt_col - 1]:
                cell_excerpt = str(row[excerpt_col - 1])
                normalized_cell = ''.join(cell_excerpt.split())
                
                # 포함 관계 확인
                if normalized_excerpt in normalized_cell or normalized_cell in normalized_excerpt:
                    if citation_id:
                        matches.append(str(citation_id))
            
            # 발췌 컬럼이 없으면 비교 불가
            elif not excerpt_col:
                wb.close()
                return None  # 안전하게 None
        
        wb.close()
        
        # 원칙: 유일하게 확실할 때만 자동
        if len(matches) == 1:
            # 정확히 1개만 매칭 → 자동 채움
            return matches[0]
        elif len(matches) >= 2:
            # 2개 이상 매칭 → 모호함, 사람확인으로
            return None
        # 0개 매칭 → 못 찾음
        
    except Exception as e:
        print(f"커버리지맵 조회 오류: {e}")
        return None
    
    return None  # 못 찾았거나 애매하면 None


def postprocess_1axis(product: str, verify_json_path: Path, coverage_map_path: Path, output_dir: Path) -> dict:
    """1축 검증 결과 후처리"""
    
    with open(verify_json_path, 'r', encoding='utf-8') as f:
        verify_data = json.load(f)
    
    decisions = verify_data.get('decisions', [])
    
    processed = {
        'meta': {
            'product': product,
            'total': len(decisions),
            'auto_completed': 0,
            'human_review': 0,
            'by_action': {
                'e_trap_completed': 0,
                'ok_cited_completed': 0,
                'ok_uncited_filled': 0,
                'partial_to_human': 0,
                'missing_to_human': 0
            }
        },
        'completed': [],
        'human_review': []
    }
    
    log_entries = []
    
    for decision in decisions:
        item_id = decision.get('id', 'unknown')
        status = decision.get('status', '')
        
        if status == 'e_trap':
            completed_item = {
                'id': item_id,
                'status': 'completed',
                'original_status': 'e_trap',
                'note': 'E형 의도된 부재'
            }
            processed['completed'].append(completed_item)
            processed['meta']['auto_completed'] += 1
            processed['meta']['by_action']['e_trap_completed'] += 1
            log_entries.append((f"{item_id} e_trap → 완료", "DEC-010"))
        
        elif status == 'ok_cited':
            # 발췌 확인됨 + 출처 있음 → 자동완료 (재매핑 불필요)
            citation = decision.get('citation', '')
            completed_item = {
                'id': item_id,
                'status': 'completed',
                'original_status': 'ok_cited',
                'note': '발췌+출처 모두 확인됨',
                'citation_filled': False,  # 원래 있었으므로 채울 필요 없음
                'original_citation': citation
            }
            processed['completed'].append(completed_item)
            processed['meta']['auto_completed'] += 1
            processed['meta']['by_action']['ok_cited_completed'] += 1
            log_entries.append((f"{item_id} ok_cited → 완료 (출처: {citation})", "§3.1/이미출처있음"))
        
        elif status == 'ok_uncited':
            # 진짜 커버리지맵에서 citation_id 찾기
            excerpt = decision.get('excerpt', '')
            found_citation = find_citation_id_for_excerpt(excerpt, coverage_map_path)
            
            if found_citation:
                # 찾았으면 채우고 완료
                completed_item = {
                    'id': item_id,
                    'status': 'completed',
                    'original_status': 'ok_uncited',
                    'note': '출처 자동 채움',
                    'citation_filled': True,
                    'filled_citation': found_citation  # 진짜 값
                }
                processed['completed'].append(completed_item)
                processed['meta']['auto_completed'] += 1
                processed['meta']['by_action']['ok_uncited_filled'] += 1
                log_entries.append((f"{item_id} ok_uncited → 완료 (citation: {found_citation})", "§3.1"))
            else:
                # 못 찾았으면 사람확인으로 (가짜 값 금지!)
                human_item = {
                    'id': item_id,
                    'status': 'human_review',
                    'original_status': 'ok_uncited',
                    'reason': '발췌 확인됨 but citation_id 매핑 불가',
                    'note': '커버리지맵에서 Unit ID 수동 매핑 필요',
                    'found_in': decision.get('found_in', [])
                }
                processed['human_review'].append(human_item)
                processed['meta']['human_review'] += 1
                log_entries.append((f"{item_id} ok_uncited → 사람확인 (citation 매핑 실패)", "§3.1/매핑실패"))
        
        elif status == 'partial':
            human_item = {
                'id': item_id,
                'status': 'human_review',
                'original_status': 'partial',
                'reason': '원문 표현 약간 다름'
            }
            processed['human_review'].append(human_item)
            processed['meta']['human_review'] += 1
            processed['meta']['by_action']['partial_to_human'] += 1
            log_entries.append((f"{item_id} partial → 사람확인", "DEC-010"))
        
        elif status == 'missing':
            human_item = {
                'id': item_id,
                'status': 'human_review',
                'original_status': 'missing',
                'reason': '원문에 발췌 없음 (환각 의심)',
                'note': '§3.2: missing을 원문 대조 없이 완료로 올리지 마라'
            }
            processed['human_review'].append(human_item)
            processed['meta']['human_review'] += 1
            processed['meta']['by_action']['missing_to_human'] += 1
            log_entries.append((f"{item_id} missing → 사람확인 (원문대조 필수)", "§3.2/DEC-002"))
    
    # 출력 저장
    output_dir.mkdir(parents=True, exist_ok=True)
    
    output_json = output_dir / 'verify_1axis_processed.json'
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(processed, f, ensure_ascii=False, indent=2)
    
    csv_path = output_dir / 'human_review_1axis.csv'
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        if processed['human_review']:
            fieldnames = ['id', 'original_status', 'reason', 'note']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for item in processed['human_review']:
                writer.writerow({
                    'id': item['id'],
                    'original_status': item['original_status'],
                    'reason': item['reason'],
                    'note': item.get('note', '')
                })
    
    # 로그
    log_batch(product, '3_verify_1axis_post', verify_json_path.name, log_entries)
    
    return processed


def main():
    import argparse
    parser = argparse.ArgumentParser(description='1축 검증 결과 후처리')
    parser.add_argument('--product', '-p', required=True, choices=['RC', 'RV', 'RM', 'HR'])
    parser.add_argument('--verify-json', '-v', required=True, type=Path)
    parser.add_argument('--coverage-map', '-c', required=True, type=Path)
    parser.add_argument('--output-dir', '-o', type=Path, default=None)
    
    args = parser.parse_args()
    
    if args.output_dir is None:
        args.output_dir = Path(__file__).parent.parent / 'products' / args.product
    
    result = postprocess_1axis(args.product, args.verify_json, args.coverage_map, args.output_dir)
    
    print(f"\n=== 1축 후처리 완료 ({args.product}) ===")
    print(f"총 처리: {result['meta']['total']}건")
    print(f"자동완료: {result['meta']['auto_completed']}건")
    print(f"사람확인: {result['meta']['human_review']}건")
    
    # 핵심 검증
    if result['meta']['by_action']['missing_to_human'] > 0:
        print(f"✓ missing {result['meta']['by_action']['missing_to_human']}건 → 사람확인 (자동완료 0건)")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
