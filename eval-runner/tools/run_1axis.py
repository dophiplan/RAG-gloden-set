"""
1축 검증기 - 골든셋 발췌를 원문 코퍼스와 대조

입력: 골든셋 xlsx(근거 원문 발췌 컬럼) + 원문 코퍼스 폴터
출력: products/{제품}/verify_1axis.json

status 정의:
- e_trap: 유형=E이고 발췌 없음 (의도된 부재)
- ok_cited: 발췌 찾음 + 출처 있음 → 자동완료 (재매핑 불필요)
- ok_uncited: 발췌 찾음 + 출처 없음 → 재매핑 대상
- partial: 원문에서 부분적으로 찾음
- missing: 원문에서 찾지 못함 (유형≠E인데 발췌 없음도 포함)
"""
import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from kit.verifier.corpus_loader import load_corpus, verify_excerpt

sys.path.insert(0, str(Path(__file__).parent))
from common.log import log, log_batch


def verify_1axis(product: str, golden_xlsx: Path, corpus_dir: Path, output_dir: Path) -> dict:
    """1축 검증 실행"""
    
    # 골든셋 xlsx 읽기
    try:
        import openpyxl
        wb = openpyxl.load_workbook(golden_xlsx, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"골든셋 파일 읽기 오류: {e}")
        sys.exit(1)
    
    # 헤더 찾기
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    
    # 필요한 컬럼 찾기
    id_col = None
    type_col = None      # 유형 컬럼
    excerpt_col = None   # 발췌 컬럼
    citation_col = None  # 근거 출처 컬럼
    
    # ID 컬럼 후보
    for key in headers:
        key_lower = key.lower()
        if any(k in key_lower for k in ['id', '문항', 'item', '번호', 'no', 'num']):
            id_col = headers[key]
            break
    
    # 유형 컬럼 후보
    for key in headers:
        key_lower = key.lower()
        if any(k in key_lower for k in ['유형', 'type', '형식', '종류']):
            type_col = headers[key]
            break
    
    # 발췌 컬럼 후보
    for key in headers:
        key_lower = key.lower()
        if any(k in key_lower for k in ['발췌', 'excerpt', '근거', 'citation', '원문']):
            # '근거 출처'는 제외 (출처 컬럼과 구분)
            if '출처' not in key_lower:
                excerpt_col = headers[key]
                break
    
    # 근거 출처 컬럼 후보
    for key in headers:
        key_lower = key.lower()
        if any(k in key_lower for k in ['출처', 'source', 'cite', 'unit']):
            citation_col = headers[key]
            break
    
    if not id_col:
        print("오류: ID 컬럼을 찾을 수 없습니다. (문항ID, 번호 등)")
        wb.close()
        sys.exit(1)
    
    if not excerpt_col:
        print("오류: 발췌 컬럼을 찾을 수 없습니다. (근거 원문 발췌, excerpt 등)")
        wb.close()
        sys.exit(1)
    
    print(f"ID 컬럼: {id_col}, 유형 컬럼: {type_col}, 발췌 컬럼: {excerpt_col}, 출처 컬럼: {citation_col}")
    
    # 코퍼스 로드
    print(f"코퍼스 로드 중: {corpus_dir}")
    corpus = load_corpus(corpus_dir)
    print(f"  로드된 파일: {len(corpus)}개")
    
    # 검증 실행
    decisions = []
    log_entries = []
    
    stats = {
        'total': 0,
        'e_trap': 0,
        'ok_cited': 0,
        'ok_uncited': 0,
        'partial': 0,
        'missing': 0
    }
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item_id = row[id_col - 1] if id_col and len(row) >= id_col else None
        item_type = row[type_col - 1] if type_col and len(row) >= type_col else None
        excerpt = row[excerpt_col - 1] if excerpt_col and len(row) >= excerpt_col else None
        citation = row[citation_col - 1] if citation_col and len(row) >= citation_col else None
        
        if not item_id:
            continue
        
        item_id = str(item_id).strip()
        item_type_str = str(item_type).strip() if item_type else ''
        excerpt_str = str(excerpt).strip() if excerpt else ''
        citation_str = str(citation).strip() if citation else ''
        
        stats['total'] += 1
        
        # E형 체크: 유형='E'이고 발췌 없음 → e_trap
        is_type_e = item_type_str.upper() == 'E'
        has_excerpt = len(excerpt_str) > 0
        has_citation = len(citation_str) > 0
        
        if is_type_e and not has_excerpt:
            # E형 의도된 부재
            decision = {
                'id': item_id,
                'status': 'e_trap',
                'excerpt': excerpt_str
            }
            stats['e_trap'] += 1
            log_entries.append((f"{item_id} → e_trap", "E형부재"))
        
        elif not has_excerpt:
            # 유형≠E인데 발췌 없음 → missing (결함)
            decision = {
                'id': item_id,
                'status': 'missing',
                'excerpt': '',
                'reason': '유형≠E인데 발췌 없음 (결함)',
                'found_in': []
            }
            stats['missing'] += 1
            log_entries.append((f"{item_id} → missing", "발췌누락"))
        
        else:
            # 원문 코퍼스 검색
            match_status, found_in = verify_excerpt(excerpt_str, corpus)
            
            if match_status == 'ok':
                # 발췌 찾음
                if has_citation:
                    # 출처 있음 → ok_cited (자동완료, 재매핑 불필요)
                    decision = {
                        'id': item_id,
                        'status': 'ok_cited',
                        'excerpt': excerpt_str,
                        'citation': citation_str,
                        'found_in': found_in
                    }
                    stats['ok_cited'] += 1
                    log_entries.append((f"{item_id} → ok_cited", f"찾음+출처:{citation_str[:20]}"))
                else:
                    # 출처 없음 → ok_uncited (재매핑 대상)
                    decision = {
                        'id': item_id,
                        'status': 'ok_uncited',
                        'excerpt': excerpt_str,
                        'found_in': found_in
                    }
                    stats['ok_uncited'] += 1
                    log_entries.append((f"{item_id} → ok_uncited", f"찾음:{','.join(found_in[:3])}"))
            
            elif match_status == 'partial':
                decision = {
                    'id': item_id,
                    'status': 'partial',
                    'excerpt': excerpt_str,
                    'found_in': found_in
                }
                stats['partial'] += 1
                log_entries.append((f"{item_id} → partial", f"부분:{','.join(found_in[:3])}"))
            
            else:  # missing
                decision = {
                    'id': item_id,
                    'status': 'missing',
                    'excerpt': excerpt_str,
                    'reason': '원문에 발췌 없음 (환각 의심)',
                    'found_in': []
                }
                stats['missing'] += 1
                log_entries.append((f"{item_id} → missing", "미발견"))
        
        decisions.append(decision)
    
    wb.close()
    
    # 결과 JSON 구성 (postprocess_1axis.py가 읽는 형식과 일치)
    result = {
        'meta': {
            'product': product,
            'golden_set': str(golden_xlsx),
            'corpus_dir': str(corpus_dir),
            'stats': stats
        },
        'decisions': decisions
    }
    
    # 출력 저장
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / 'verify_1axis.json'
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    # 로그
    log_batch(product, '2_verify_1axis', golden_xlsx.name, log_entries)
    
    return result


def main():
    parser = argparse.ArgumentParser(description='1축 검증 - 골든셋 발췌를 원문과 대조')
    parser.add_argument('--product', '-p', required=True, choices=['RC', 'RV', 'RM', 'HR'],
                       help='제품 코드')
    parser.add_argument('--golden', '-g', required=True, type=Path,
                       help='골든셋 xlsx 파일 경로')
    parser.add_argument('--corpus', '-c', required=True, type=Path,
                       help='원문 코퍼스 폴더 경로')
    parser.add_argument('--output-dir', '-o', type=Path, default=None,
                       help='출력 디렉토리 (기본: products/{제품}/)')
    
    args = parser.parse_args()
    
    # 입력 검증
    if not args.golden.exists():
        print(f"오류: 골든셋 파일을 찾을 수 없습니다: {args.golden}")
        return 1
    
    if not args.corpus.exists():
        print(f"오류: 코퍼스 폴더를 찾을 수 없습니다: {args.corpus}")
        return 1
    
    # 출력 디렉토리 설정
    if args.output_dir is None:
        args.output_dir = Path(__file__).parent.parent / 'products' / args.product
    
    # 검증 실행
    result = verify_1axis(args.product, args.golden, args.corpus, args.output_dir)
    
    # 결과 출력
    print(f"\n=== 1축 검증 완료 ({args.product}) ===")
    print(f"총 문항: {result['meta']['stats']['total']}")
    print(f"  - e_trap: {result['meta']['stats']['e_trap']}")
    print(f"  - ok_cited: {result['meta']['stats']['ok_cited']}")
    print(f"  - ok_uncited: {result['meta']['stats']['ok_uncited']}")
    print(f"  - partial: {result['meta']['stats']['partial']}")
    print(f"  - missing: {result['meta']['stats']['missing']}")
    print(f"\n출력: {args.output_dir / 'verify_1axis.json'}")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
