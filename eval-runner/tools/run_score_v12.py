#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""run_score_v12.py — 병기 채점기: 검색 판정 = 합집합(URL 일치 OR 내용 실재)

배경 (2026-08-03 r1 분석): 공식(v11)은 URL 일치로만 검색 히트를 인정 —
① 골든셋 근거 출처 URL 결측(PDF 등) 문항은 구조적 miss ② 같은 내용이 다른 문서에도
있으면(모호 질문) 억울한 miss. v12는 "시스템이 가져온 청크 본문에 골든셋 근거 원문
발췌가 실재하는가"를 추가 인정한다. 생성/E축은 v11 결과 그대로 (검색만 승격).

인터페이스: v11과 동일 — --golden --log --out-dir (오케스트레이터 병기 슬롯 규격)
"""
import argparse
import json
import re
import subprocess
import sys
import zipfile
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
V11 = HERE / "run_score_v11.py"


def norm(s):
    return re.sub(r"[\s\x00-\x1f]+", "", str(s or "")).lower()


def load_gold_segs(golden):
    ws = openpyxl.load_workbook(golden, read_only=True, data_only=True).active
    hdr = [str(c) for c in next(ws.iter_rows(max_row=1, values_only=True))]
    ei = hdr.index("근거 원문 발췌") if "근거 원문 발췌" in hdr else None
    out = {}
    if ei is None:
        return out
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        segs = [re.sub(r"^\[[^\]]*\]\s*", "", s).strip("\"'“”‘’ ")
                for s in re.split(r"\|\|", str(r[ei] or ""))]
        out[str(r[0])] = [norm(s) for s in segs if len(s) >= 10]
    return out


def load_chunk_texts(golden):
    """골든셋 경로에서 제품 코퍼스 zip을 찾아 chunk_id → 본문 (없으면 빈 dict — snippet 폴백)"""
    try:
        corpus_dir = Path(golden).resolve().parent.parent / "corpus"
        texts = {}
        for zp in sorted(corpus_dir.glob("*.zip*")):
            with zipfile.ZipFile(zp) as z:
                for name in z.namelist():
                    if name.startswith("chunks") and name.endswith(".json"):
                        d = json.load(z.open(name))
                        for c in d.get("chunks", d if isinstance(d, list) else []):
                            pid = c.get("point_id") or c.get("chunk_id")
                            if pid:
                                texts[pid] = norm(c.get("text", ""))
        return texts
    except Exception:
        return {}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--golden", required=True)
    ap.add_argument("--log", required=True)
    ap.add_argument("--out-dir", required=True)
    a = ap.parse_args()
    out_dir = Path(a.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1) 공식(v11) 먼저 — 생성/E축과 검색 기준선
    p = subprocess.run([sys.executable, str(V11), "--golden", a.golden, "--log", a.log,
                        "--out-dir", str(out_dir)], capture_output=True, text=True, cwd=str(HERE))
    rep_path = out_dir / "score_report.json"
    if not rep_path.exists():
        sys.exit(f"v11 선행 실패: {(p.stdout + p.stderr)[-300:]}")
    rep = json.loads(rep_path.read_text(encoding="utf-8"))

    # 2) 내용 기준 — 시스템이 가져온 청크에 근거 원문 발췌가 실재하나
    gold = load_gold_segs(a.golden)
    ctext = load_chunk_texts(a.golden)
    log = json.loads(Path(a.log).read_text(encoding="utf-8"))
    content_rank = {}
    for resp in log.get("responses", []):
        iid = str(resp.get("id", ""))
        segs = gold.get(iid, [])
        if not segs:
            continue
        texts = [ctext.get(h.get("chunk_id")) or norm(h.get("snippet", ""))
                 for h in (resp.get("hits") or [])[:5]]
        content_rank[iid] = next((i for i, t in enumerate(texts)
                                  if any(s in t for s in segs)), None)

    # 3) 합집합 승격 — hit_top1 > hit_top5 > miss (v11 판정과 내용 판정 중 좋은 쪽)
    order = {"hit_top1": 2, "hit_top5": 1}
    upgraded = 0
    for r in rep:
        cr = content_rank.get(str(r.get("id", "")))
        cv = "hit_top1" if cr == 0 else ("hit_top5" if cr is not None else "miss")
        r["검색_v11"] = r.get("검색")
        r["검색_내용"] = cv
        best = max([r.get("검색"), cv], key=lambda x: order.get(x, 0))
        if order.get(best, 0) > order.get(r.get("검색"), 0):
            upgraded += 1
        r["검색"] = best
    rep_path.write_text(json.dumps(rep, ensure_ascii=False, indent=1), encoding="utf-8")
    t1 = sum(1 for r in rep if r["검색"] == "hit_top1")
    t5 = t1 + sum(1 for r in rep if r["검색"] == "hit_top5")
    print(f"v12(합집합): top1 {t1}/{len(rep)} · top5 {t5}/{len(rep)} · 내용 기준 승격 {upgraded}건 "
          f"(코퍼스 본문 {'사용' if ctext else '미발견 — snippet 폴백'})")


if __name__ == "__main__":
    main()
