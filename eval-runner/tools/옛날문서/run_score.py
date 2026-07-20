#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RAG 채점기 (STAGE3 - 실물 채점)

사용법:
    python3 run_score.py --golden 골든셋.xlsx --log 응답로그.json [--mapping 매핑시트.xlsx]
    python3 run_score.py --selftest
"""

import argparse
import json
import re
import unicodedata
import urllib.parse
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill


# ---------------------------------------------------------------------------
# 상수 (확장 가능)
# ---------------------------------------------------------------------------
REJECTION_PATTERNS = [
    "자료에 없",
    "찾을 수 없",
    "확인되지 않",
    "제공되지 않",
    "알 수 없",
    "문서에 없",
    "관련 정보가 없",
    "답변할 수 없",
    "응답할 수 없",
    "Unable to",
    "cannot find",
    "not found",
    "no information",
    "does not contain",
]

# 인입 누락 의심 판정: 골든셋 문서가 hits 전체에 한 번도 없는 문항 수 임계치
OMISSION_THRESHOLD = 1

# URL 정규화용
URL_RE = re.compile(
    r"https?://"                # 프로토콜
    r"(www\.)?"                 # 선택적 www
    r"([^\s/?#]+)"              # 호스트
    r"(:\d+)?"                  # 선택적 포트
    r"([^?#\s]*?)"              # 경로
    r"(?:/|\?|$)",              # 말미 슬래시 또는 쿼리 시작 또는 문자열 끝
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# 유틸리티
# ---------------------------------------------------------------------------
def normalize_text(text: str) -> str:
    """공백 제거, 소문자, NFC 정규화."""
    if not isinstance(text, str):
        text = str(text) if pd.notna(text) else ""
    text = unicodedata.normalize("NFC", text)
    text = re.sub(r"\s+", "", text)
    text = text.lower()
    # 수치 콤마 흡수: 1,000 -> 1000
    text = text.replace(",", "")
    return text


def normalize_url(raw_url: str) -> str:
    """프로토콜, www, 말미 슬래시, 쿼리스트링 제거 후 반환."""
    if not isinstance(raw_url, str) or not raw_url.strip():
        return ""
    url = raw_url.strip()
    parsed = urllib.parse.urlparse(url)
    netloc = parsed.netloc or parsed.path
    path = parsed.path if parsed.netloc else ""

    # www 제거
    netloc = re.sub(r"^www\.", "", netloc, flags=re.IGNORECASE)
    # 포트 제거
    netloc = netloc.split(":")[0]
    # 경로의 말미 슬래시 제거
    path = path.rstrip("/")
    # 쿼리/프래그먼트 제거는 urlparse로 이미 분리됨
    return f"{netloc}{path}".lower()


def extract_domain(raw_url: str) -> str:
    """URL 정규화 후 도메인만 반환 (인입 누락 의심용)."""
    norm = normalize_url(raw_url)
    return norm.split("/")[0]


def extract_urls(text: str) -> list[str]:
    """텍스트에서 http/https URL 목록 추출."""
    if not isinstance(text, str):
        return []
    return re.findall(r"https?://[^\s\)\]\>\"]+", text)


def parse_required_elements(answer_field: str) -> list[str]:
    """'정답 (필수 포함 요소)' 컬럼에서 '필수:' 뒤 요소들을 파싱.

    줄 중간 삽입 형식("정답 텍스트 / 필수: 요소1, 요소2")도 지원하며,
    줄 시작 형식 및 후속 줄 연속 작성도 그대로 지원한다.
    """
    if not isinstance(answer_field, str):
        return []
    elements = []
    in_required = False
    for raw_line in answer_field.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        # 다른 섹션 시작이면 필수 구역 종료
        if re.match(r"^(선택|선택사항|참고|※)", line, re.IGNORECASE):
            in_required = False
            continue
        # 줄 중간/시작 "필수:" 모두 인식
        m = re.search(r"필수\s*[:：]\s*(.*)$", line)
        if m:
            in_required = True
            body = m.group(1)
        elif in_required:
            body = line
        else:
            continue
        # 구분자(·, /, |, , ;)로 분리
        for part in re.split(r"[·/|，,;；]", body):
            part = part.strip().strip("-•*")
            if part:
                elements.append(part)
    return elements


def contains_normalized(haystack: str, needle: str) -> bool:
    """haystack에 needle이 정규화 상태로 포함되는지 검사."""
    return normalize_text(needle) in normalize_text(haystack)


def is_rejection(answer: str) -> bool:
    """거절/부재 응답인지 판정."""
    if not isinstance(answer, str) or not answer.strip():
        return True
    return any(pat in answer for pat in REJECTION_PATTERNS)


def suffix_from_id(qid: str) -> str:
    """ID에 @ 접미가 있으면 반환, 없으면 '없음'."""
    if isinstance(qid, str) and "@" in qid:
        return qid.split("@", 1)[1]
    return "없음"


# ---------------------------------------------------------------------------
# 데이터 로딩
# ---------------------------------------------------------------------------
def load_golden(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, engine="openpyxl")
    required_cols = ["ID", "유형", "질문", "정답 (필수 포함 요소)", "합격 기준", "근거 출처", "근거 원문 발췌"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"골든셋에 누락된 컬럼: {missing}")
    # ID를 문자열로
    df["ID"] = df["ID"].astype(str).str.strip()
    return df


def load_log(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if "responses" not in data:
        raise ValueError("응답 로그에 'responses' 키가 없습니다.")
    return data


def load_mapping(path: str | None) -> dict[str, dict] | None:
    if not path:
        return None
    df = pd.read_excel(path, engine="openpyxl")
    # 예상 컬럼: 문항ID, doc_key, point_id
    id_col = next((c for c in df.columns if "문항" in c or "ID" in c), None)
    doc_col = next((c for c in df.columns if "doc" in c.lower() or "문서" in c), None)
    point_col = next((c for c in df.columns if "point" in c.lower() or "포인트" in c or "chunk" in c.lower()), None)
    if not id_col:
        raise ValueError("매핑시트에 문항ID 컬럼이 없습니다.")
    mapping = {}
    for _, row in df.iterrows():
        qid = str(row[id_col]).strip()
        mapping[qid] = {
            "doc_key": str(row[doc_col]).strip() if doc_col and pd.notna(row[doc_col]) else None,
            "point_id": str(row[point_col]).strip() if point_col and pd.notna(row[point_col]) else None,
        }
    return mapping


# ---------------------------------------------------------------------------
# 채점 함수
# ---------------------------------------------------------------------------
def score_search(
    row: pd.Series,
    response: dict,
    mapping: dict[str, dict] | None,
) -> tuple[str, bool]:
    """
    검색 채점.
    반환: ("hit_top1" | "hit_top5" | "miss", strict_match)
    """
    qid = row["ID"]
    golden_sources = row["근거 출처"]
    golden_urls = [normalize_url(u) for u in extract_urls(golden_sources) if normalize_url(u)]
    hits = response.get("hits", [])

    # 매핑시트 기준 (우선)
    if mapping and qid in mapping:
        golden_doc_key = mapping[qid].get("doc_key")
        golden_point_id = mapping[qid].get("point_id")
        for idx, hit in enumerate(hits):
            hit_chunk = str(hit.get("chunk_id", "")).strip()
            # chunk_id 일치 -> 엄격 일치
            if golden_point_id and hit_chunk == golden_point_id:
                rank_label = "hit_top1" if idx == 0 else "hit_top5" if idx < 5 else "miss"
                return rank_label, True
        # doc_key 기준
        if golden_doc_key:
            for idx, hit in enumerate(hits):
                hit_doc = str(hit.get("doc_key", "")).strip()
                if hit_doc and hit_doc == golden_doc_key:
                    rank_label = "hit_top1" if idx == 0 else "hit_top5" if idx < 5 else "miss"
                    return rank_label, False
        return "miss", False

    # 폰백: URL 정규화 비교
    for idx, hit in enumerate(hits):
        src_url = normalize_url(hit.get("source_url", ""))
        if src_url and src_url in golden_urls:
            rank_label = "hit_top1" if idx == 0 else "hit_top5" if idx < 5 else "miss"
            return rank_label, False
    return "miss", False


def score_generation(row: pd.Series, response: dict) -> tuple[str, list[str]]:
    """
    생성 채점.
    반환: ("pass" | "partial" | "fail" | "unparsed", 누락 요소 목록)
    """
    answer = response.get("answer", "")
    if not isinstance(answer, str):
        answer = str(answer) if pd.notna(answer) else ""
    required = parse_required_elements(row["정답 (필수 포함 요소)"])
    if not required:
        # 필수 요소가 0개로 파싱되면 공허한 pass를 금지하고 unparsed로 분리 집계
        return "unparsed", []
    missing = [el for el in required if not contains_normalized(answer, el)]
    if not missing:
        return "pass", []
    if len(missing) == len(required):
        return "fail", missing
    return "partial", missing


def score_e_type(row: pd.Series, response: dict) -> dict:
    """E형 전용 무응답 함정 채점."""
    answer = response.get("answer", "")
    if not isinstance(answer, str):
        answer = str(answer) if pd.notna(answer) else ""
    rejected = is_rejection(answer)
    return {
        "e_type": True,
        "rejection": rejected,
        "hallucination": not rejected,  # 거절 없이 구체적 답 -> 환각
    }


def failure_tag(search_score: str, gen_score: str, e_hallucination: bool) -> str:
    if e_hallucination:
        return "E형 환각"
    if gen_score == "unparsed":
        return "필수 파싱 실패"
    if search_score == "miss" and gen_score == "fail":
        return "검색 실패 기인"
    if search_score.startswith("hit") and gen_score == "fail":
        return "생성 실패"
    if search_score == "miss":
        return "검색 실패"
    if gen_score in ("partial", "fail"):
        return "생성 실패"
    return ""


def compute_omission_docs(
    golden_df: pd.DataFrame,
    responses: list[dict],
) -> list[dict]:
    """
    골든셋 근거 출처 문서가 로그 hits 전체에 한 번도 없는 문항이 다수인 문서 목록.
    """
    # 로그에 등장한 모든 문서(도메인)
    hit_domains = set()
    for resp in responses:
        for hit in resp.get("hits", []):
            domain = extract_domain(hit.get("source_url", ""))
            if domain:
                hit_domains.add(domain)

    # 문서별로 누락 문항 수 집계
    doc_missing_counts = defaultdict(int)
    doc_examples = defaultdict(list)
    for _, row in golden_df.iterrows():
        qid = row["ID"]
        golden_domains = {extract_domain(u) for u in extract_urls(row["근거 출처"]) if extract_domain(u)}
        for dom in golden_domains:
            if dom not in hit_domains:
                doc_missing_counts[dom] += 1
                if len(doc_examples[dom]) < 3:
                    doc_examples[dom].append(qid)

    result = []
    for doc, count in sorted(doc_missing_counts.items(), key=lambda x: -x[1]):
        if count >= OMISSION_THRESHOLD:
            result.append({
                "문서(도메인)": doc,
                "누락문항수": count,
                "예시문항IDs": ", ".join(doc_examples[doc]),
            })
    return result


# ---------------------------------------------------------------------------
# 보고서 작성
# ---------------------------------------------------------------------------
def score_all(
    golden_df: pd.DataFrame,
    log_data: dict,
    mapping: dict[str, dict] | None,
) -> list[dict]:
    responses = {str(r.get("id")).strip(): r for r in log_data.get("responses", [])}
    results = []
    for _, row in golden_df.iterrows():
        qid = row["ID"]
        qtype = str(row["유형"]).strip()
        response = responses.get(qid, {"hits": [], "answer": ""})

        # E형 여부 (첫 글자가 E)
        is_e = bool(qtype) and qtype[0].upper() == "E"

        # 검색 채점 (E형 제외)
        if is_e:
            search_score, strict = "N/A", False
        else:
            search_score, strict = score_search(row, response, mapping)

        # 생성 채점
        gen_score, missing_elements = score_generation(row, response)

        # E형 전용
        if is_e:
            e_info = score_e_type(row, response)
            # E형은 거절 여부를 생성 점수로 반영
            gen_score = "pass" if e_info["rejection"] else "fail"
            missing_elements = []
        else:
            e_info = {"e_type": False, "rejection": False, "hallucination": False}

        # 실패 태그
        tag = ""
        if is_e and e_info["hallucination"]:
            tag = "E형 환각"
        elif gen_score == "unparsed":
            tag = "필수 파싱 실패"
        else:
            tag = failure_tag(search_score, gen_score, False)

        results.append({
            "id": qid,
            "유형": qtype,
            "검색": search_score,
            "생성": gen_score,
            "누락요소": missing_elements,
            "E형거절": e_info["rejection"],
            "E형환각": e_info["hallucination"],
            "엄격일치": strict,
            "실패태그": tag,
        })
    return results


def build_summary(results: list[dict]) -> pd.DataFrame:
    rows = []

    def _stats(items: list[dict]) -> dict:
        total = len(items)
        hit_top1 = sum(1 for r in items if r["검색"] == "hit_top1")
        hit_top5 = sum(1 for r in items if r["검색"] == "hit_top5")
        hit_any = hit_top1 + hit_top5
        gen_pass = sum(1 for r in items if r["생성"] == "pass")
        gen_partial = sum(1 for r in items if r["생성"] == "partial")
        gen_unparsed = sum(1 for r in items if r["생성"] == "unparsed")
        e_hallu = sum(1 for r in items if r["E형환각"])
        return {
            "문항수": total,
            "검색_top1율": round(hit_top1 / total, 4) if total else 0,
            "검색_top5율": round(hit_any / total, 4) if total else 0,
            "생성_pass율": round(gen_pass / total, 4) if total else 0,
            "생성_partial율": round(gen_partial / total, 4) if total else 0,
            "생성_unparsed수": gen_unparsed,
            "E형환각수": e_hallu,
        }

    # 전체
    rows.append({"구분": "전체", **_stats(results)})

    # 유형별 (첫 글자 A~G)
    type_groups = defaultdict(list)
    for r in results:
        t = r["유형"][0].upper() if r["유형"] else "기타"
        type_groups[t].append(r)
    for t in sorted(type_groups):
        rows.append({"구분": f"유형_{t}", **_stats(type_groups[t])})

    # 원천별 (@접미)
    src_groups = defaultdict(list)
    for r in results:
        src_groups[suffix_from_id(r["id"])].append(r)
    for src in sorted(src_groups):
        rows.append({"구분": f"원천_{src}", **_stats(src_groups[src])})

    return pd.DataFrame(rows)


def build_defects(results: list[dict]) -> pd.DataFrame:
    defects = []
    for r in results:
        if r["실패태그"]:
            defects.append({
                "id": r["id"],
                "유형": r["유형"],
                "검색": r["검색"],
                "생성": r["생성"],
                "실패태그": r["실패태그"],
                "누락요소": "; ".join(r["누락요소"]),
            })
    return pd.DataFrame(defects)


def write_json_report(results: list[dict], output_path: str):
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)


def write_excel_report(
    summary_df: pd.DataFrame,
    detail_results: list[dict],
    defects_df: pd.DataFrame,
    omission_docs: list[dict],
    output_path: str,
):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 요약
        summary_df.to_excel(writer, sheet_name="요약", index=False)

        # 문항별
        detail_df = pd.DataFrame(detail_results)
        detail_df["누락요소"] = detail_df["누락요소"].apply(lambda x: "; ".join(x))
        detail_df.to_excel(writer, sheet_name="문항별", index=False)

        # 결함목록
        if defects_df.empty:
            pd.DataFrame({"메시지": ["결함 항목이 없습니다."]}).to_excel(
                writer, sheet_name="결함목록", index=False
            )
        else:
            defects_df.to_excel(writer, sheet_name="결함목록", index=False)

        # 인입누락의심
        omission_df = pd.DataFrame(omission_docs)
        if omission_df.empty:
            pd.DataFrame({"메시지": ["인입 누락 의심 문서가 없습니다."]}).to_excel(
                writer, sheet_name="인입누락의심", index=False
            )
        else:
            omission_df.to_excel(writer, sheet_name="인입누락의심", index=False)

    # 간단한 서식
    wb = openpyxl.load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
    wb.save(output_path)


# ---------------------------------------------------------------------------
# 자가검증 (칵나리아)
# ---------------------------------------------------------------------------
def run_selftest():
    print("[selftest] 가상 골든셋/로그 생성 및 채점 검증\n")

    # 가짜 골든셋 (실제 형식 반영: "정답 텍스트 / 필수: 요소1, 요소2")
    golden_df = pd.DataFrame({
        "ID": ["RC-A001", "RC-B001", "RC-C001", "RC-D001", "RC-E001"],
        "유형": ["A_사실", "B_수치", "C_정책", "D_키워드", "E_없음"],
        "질문": ["Q1", "Q2", "Q3", "Q4", "Q5"],
        "정답 (필수 포함 요소)": [
            "서울특례시 / 필수: 서울, 특례시",
            "1000명, 2024년 데이터 / 필수: 1000, 2024년",
            "회사정책 / 필수: 정책, 2024년",
            "핵심키워드 필요 / 필수: 핵심키워드",
            "(해당 없음) / 필수: (해당 없음)",
        ],
        "합격 기준": ["", "", "", "", ""],
        "근거 출처": [
            "https://www.example.com/page/",
            "https://example.com/data?q=1",
            "https://example.com/policy",
            "https://example.com/keyword",
            "https://example.com/e-type",
        ],
        "근거 원문 발췌": ["", "", "", "", ""],
    })

    # 정답 로그: 전부 pass
    correct_log = {
        "meta": {"run_date": "2024-01-01", "system_version": "v1"},
        "responses": [
            {
                "id": "RC-A001",
                "hits": [{"rank": 1, "source_url": "https://example.com/page", "chunk_id": "c1", "snippet": "서울"}],
                "answer": "서울특례시입니다.",
            },
            {
                "id": "RC-B001",
                "hits": [{"rank": 1, "source_url": "https://example.com/data", "chunk_id": "c2", "snippet": "1,000"}],
                "answer": "1000명이며 2024년 데이터입니다.",
            },
            {
                "id": "RC-C001",
                "hits": [{"rank": 1, "source_url": "https://example.com/policy", "chunk_id": "c3", "snippet": "정책"}],
                "answer": "회사정책이며 2024년에 개정되었습니다.",
            },
            {
                "id": "RC-D001",
                "hits": [{"rank": 1, "source_url": "https://example.com/keyword", "chunk_id": "c4", "snippet": "핵심키워드"}],
                "answer": "핵심키워드입니다.",
            },
            {
                "id": "RC-E001",
                "hits": [],
                "answer": "자료에 없어 답변할 수 없습니다.",
            },
        ],
    }

    # 요소 절반 포함 -> partial + 누락 목록
    partial_log = {
        "meta": {"run_date": "2024-01-01", "system_version": "v1"},
        "responses": [
            {
                "id": "RC-C001",
                "hits": [{"rank": 1, "source_url": "https://example.com/policy", "chunk_id": "c3", "snippet": "정책"}],
                "answer": "회사정책입니다.",  # "2024년" 누락
            },
        ],
    }

    # hit + 요소 전무 -> 생성 실패 태그
    fail_hit_log = {
        "meta": {"run_date": "2024-01-01", "system_version": "v1"},
        "responses": [
            {
                "id": "RC-D001",
                "hits": [{"rank": 1, "source_url": "https://example.com/keyword", "chunk_id": "c4", "snippet": ""}],
                "answer": "무관한 답변입니다.",  # "핵심키워드" 누락
            },
        ],
    }

    # 오답 로그: 검색/생성 모두 실패 + E형 환각
    wrong_log = {
        "meta": {"run_date": "2024-01-01", "system_version": "v1"},
        "responses": [
            {
                "id": "RC-A001",
                "hits": [{"rank": 1, "source_url": "https://other.com/page", "chunk_id": "c9", "snippet": ""}],
                "answer": "부산입니다.",
            },
            {
                "id": "RC-B001",
                "hits": [{"rank": 2, "source_url": "https://other.com/data", "chunk_id": "c8", "snippet": ""}],
                "answer": "모르겠습니다.",
            },
            {
                "id": "RC-C001",
                "hits": [{"rank": 3, "source_url": "https://other.com/policy", "chunk_id": "c7", "snippet": ""}],
                "answer": "정책입니다.",  # 2024년 누락, 검색도 miss
            },
            {
                "id": "RC-D001",
                "hits": [{"rank": 1, "source_url": "https://other.com/keyword", "chunk_id": "c6", "snippet": ""}],
                "answer": "잘 모르겠습니다.",
            },
            {
                "id": "RC-E001",
                "hits": [],
                "answer": "E형 질문에 대한 구체적인 답변입니다.",
            },
        ],
    }

    expected_correct = {
        "RC-A001": ("hit_top1", "pass", ""),
        "RC-B001": ("hit_top1", "pass", ""),
        "RC-C001": ("hit_top1", "pass", ""),
        "RC-D001": ("hit_top1", "pass", ""),
        "RC-E001": ("N/A", "pass", ""),
    }
    expected_partial = {
        "RC-C001": ("hit_top1", "partial", "생성 실패"),
    }
    expected_fail_hit = {
        "RC-D001": ("hit_top1", "fail", "생성 실패"),
    }
    expected_wrong = {
        "RC-A001": ("miss", "fail", "검색 실패 기인"),
        "RC-B001": ("miss", "fail", "검색 실패 기인"),
        # 검색이 miss이면 생성 partial이라도 태그는 검색 실패가 우선
        "RC-C001": ("miss", "partial", "검색 실패"),
        "RC-D001": ("miss", "fail", "검색 실패 기인"),
        "RC-E001": ("N/A", "fail", "E형 환각"),
    }

    cases = [
        ("정답로그", correct_log, expected_correct),
        ("부분누락로그", partial_log, expected_partial),
        ("hit요소전무로그", fail_hit_log, expected_fail_hit),
        ("오답로그", wrong_log, expected_wrong),
    ]

    table = []
    for name, log_data, expected in cases:
        results = score_all(golden_df, log_data, None)
        all_ok = True
        for r in results:
            if r["id"] not in expected:
                continue
            exp = expected[r["id"]]
            ok = (r["검색"], r["생성"], r["실패태그"]) == exp
            status = "PASS" if ok else "FAIL"
            if not ok:
                all_ok = False
            table.append({
                "로그": name,
                "ID": r["id"],
                "검색예상": exp[0],
                "검색실제": r["검색"],
                "생성예상": exp[1],
                "생성실제": r["생성"],
                "태그예상": exp[2],
                "태그실제": r["실패태그"],
                "누락요소": "; ".join(r["누락요소"]),
                "결과": status,
            })
        print(f"[{name}] 전체 {'통과' if all_ok else '실패'}")

    df = pd.DataFrame(table)
    print("\n" + df.to_string(index=False))

    if any(r["결과"] == "FAIL" for r in table):
        raise SystemExit("[selftest] 실패: 예상과 다른 결과가 존재합니다.")
    print("\n[selftest] 모두 통과")


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="RAG 채점기 (STAGE3 - 실물 채점)")
    parser.add_argument("--golden", help="골든셋 Excel 경로")
    parser.add_argument("--log", help="응답 로그 JSON 경로")
    parser.add_argument("--mapping", help="매핑시트 Excel 경로 (선택)")
    parser.add_argument("--selftest", action="store_true", help="자가검증 모드")
    parser.add_argument("--out-dir", default=".", help="출력 디렉토리 (기본: 현재 디렉토리)")
    args = parser.parse_args()

    if args.selftest:
        run_selftest()
        return

    if not args.golden or not args.log:
        parser.error("--golden 과 --log 는 필수입니다. (또는 --selftest 사용)")

    golden_df = load_golden(args.golden)
    log_data = load_log(args.log)
    mapping = load_mapping(args.mapping)

    results = score_all(golden_df, log_data, mapping)
    omission_docs = compute_omission_docs(golden_df, log_data.get("responses", []))

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    json_path = out_dir / "score_report.json"
    xlsx_path = out_dir / "score_report.xlsx"

    write_json_report(results, str(json_path))
    summary_df = build_summary(results)
    defects_df = build_defects(results)
    write_excel_report(summary_df, results, defects_df, omission_docs, str(xlsx_path))

    print(f"[완료] JSON: {json_path}")
    print(f"[완료] Excel: {xlsx_path}")
    print(f"- 문항 수: {len(results)}")
    print(f"- 검색 top1: {sum(1 for r in results if r['검색']=='hit_top1')} / {len(results)}")
    print(f"- 검색 top5: {sum(1 for r in results if r['검색'] in ('hit_top1','hit_top5'))} / {len(results)}")
    print(f"- 생성 pass: {sum(1 for r in results if r['생성']=='pass')} / {len(results)}")
    print(f"- 생성 unparsed: {sum(1 for r in results if r['생성']=='unparsed')} / {len(results)}")
    print(f"- E형 환각: {sum(1 for r in results if r['E형환각'])}")
    print(f"- 인입 누락 의심 문서: {len(omission_docs)}개")


if __name__ == "__main__":
    main()
